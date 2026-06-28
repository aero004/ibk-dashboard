from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd


import calendar

TNVED_PATH = Path(os.environ.get("IM70_74_TNVED", r"D:\Эгамбердиев\Ижро хужжатлари\2026\6.Iyun\04.06.2026\Справичник ТН ВЕД 2025.xlsx"))

# HS boblar: iste'mol muddati bo'lishi mumkin bo'lgan tovarlar (1-24 oziq-ovqat, 30 farmatsiya, 33 kosmetika)
_SL_CHAPTERS: frozenset[str] = frozenset([
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
    "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
    "21", "22", "23", "24",  # barcha oziq-ovqat/ichimlik/tamaki
    "30",                     # farmatsevtika mahsulotlari
    "33",                     # kosmetika, parfyumeriya
    "34",                     # sovun, yuvish vositalari
    "38",                     # kimyoviy preparatlar (ba'zilari)
])

# Regex: 7-band markeri yoki kalit so'zlardan keyin sana
# "7. 01.01.2027", "7. 06.2026", "srok godnosti 01.01.2027", "goden do 06.2026", "годен до 01.2027"
_SL_RE = re.compile(
    r"""
    (?:
        (?<![0-9])7[.)]\s*
        (?:srok\s*godnosti\s*|srок\s*godnosti\s*|срок\s*годн[^\s]{0,6}\s*|
           goden\s*do\s*|годен\s*до\s*|yaroqlilik\s*(?:muddati\s*)?)?
    |
        (?:srok\s+godnosti|срок\s+годности|srok\s+gdn|
           goden\s+do|годен\s+до|
           yaroqlilik\s+muddati|yaroqliligi|
           годен|годности)
        \s*[:\-]?\s*
    )
    (?:
        (\d{2})[./](\d{2})[./](\d{4})
    |   (\d{2})[./](\d{4})
    |   (\d{4})
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)


def hs_has_shelf_life(hs_code: str) -> bool:
    """HS kodi iste'mol muddatiga ega tovar bobiga kirish-kirmasligini tekshiradi."""
    code = digits(str(hs_code or ""))
    return bool(code) and code[:2] in _SL_CHAPTERS


def extract_shelf_life_date(text: str) -> pd.Timestamp | None:
    """Tovar tavsifidagi matndan iste'mol muddatini ajratib oladi."""
    if not text:
        return None
    m = _SL_RE.search(str(text))
    if not m:
        return None
    g = m.groups()  # (dd, mm1, yyyy1, mm2, yyyy2, yyyy3)
    try:
        if g[2]:  # DD.MM.YYYY
            return pd.Timestamp(int(g[2]), int(g[1]), int(g[0]))
        if g[4]:  # MM.YYYY → oyning oxirgi kuni
            y, mo = int(g[4]), int(g[3])
            return pd.Timestamp(y, mo, calendar.monthrange(y, mo)[1])
        if g[5]:  # YYYY → 31 dekabr
            return pd.Timestamp(int(g[5]), 12, 31)
    except Exception:
        pass
    return None


def shelf_life_from_goods(df: pd.DataFrame, report_date: datetime) -> list[dict]:
    """
    Asos faylning 9-ustunidan (tovar tavsifi) iste'mol muddatini ajratib chiqaradi.
    Faqat HS bob 1-24, 30, 33, 34 ga mansub iste'mol tovarlari tekshiriladi.
    94-modda bo'yicha 180 kunlik qoida baholanadi.
    """
    ref = pd.Timestamp(report_date)
    mask = df[SRC["hs"]].map(hs_has_shelf_life)
    cd = df[mask].copy()
    if cd.empty:
        return []
    cd["_sl_date"] = cd[SRC["goods"]].map(extract_shelf_life_date)
    cd = cd[cd["_sl_date"].notna()].copy()
    if cd.empty:
        return []
    cd["_sl_days"] = (cd["_sl_date"] - ref).dt.days.astype(int)

    def _holat(days: int) -> str:
        if days < 0:
            return "Muddati o'tgan"
        if days <= 180:
            return "180 kun qoidasi (buzilish)"
        if days <= 365:
            return "Diqqat talab"
        return "Normal"

    rows = []
    for _, r in cd.sort_values("_sl_days").iterrows():
        days = int(r["_sl_days"])
        rows.append({
            "stir": clean(r[SRC["stir"]]),
            "korxona": to_latin(clean(r[SRC["company"]])),
            "hs": clean(r[SRC["hs"]]),
            "tnved": to_latin(clean(r.get("_tnved_name", ""))),
            "goods": clean(r[SRC["goods"]])[:300],
            "yaroqlilik": r["_sl_date"].strftime("%Y-%m-%d"),
            "qolgan_kun": days,
            "holat": _holat(days),
            "warehouse": to_latin(clean(r[SRC["warehouse"]])),
            "regime": clean(r[SRC["regime"]]),
            "vazn": float(r["_weight_tn"]),
            "qiymat": float(r["_value_usd_k"]),
        })
    return rows


SRC = {
    "tr": 0, "decl_type": 1, "regime": 2, "decl_no": 3, "decl_date": 4,
    "stir": 5, "company": 6, "goods_no": 7, "hs": 8, "goods": 9,
    "weight_kg": 10, "value_usd_k": 11, "reason": 12, "post": 13,
    "warehouse": 14, "country": 15, "customs_value_k": 16, "pay_total_k": 17,
    "pay_20_k": 18, "pay_27_k": 19, "pay_29_k": 20, "priv_total_k": 21,
    "priv_20_k": 22, "priv_27_k": 23, "priv_29_k": 24,
}


def clean(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def digits(value) -> str:
    return re.sub(r"\D", "", clean(value))


def num(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace("\u00a0", "", regex=False)
    s = s.str.replace(" ", "", regex=False).str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce").fillna(0.0)


def parse_date(value):
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    return "" if pd.isna(parsed) else parsed.strftime("%d.%m.%Y")


def post_group(value: str) -> str:
    p = clean(value).lower()
    if "00101" in p:
        return "Toshkent xalqaro aeroporti CHBP"
    if "00102" in p:
        return "Avia yuklar TIF"
    if "00107" in p:
        return "Elektron tijorat TIF"
    return clean(value) or "Boshqa"


def post_code(value: str) -> str:
    match = re.search(r"\b(00101|00102|00107)\b", clean(value))
    return match.group(1) if match else ""


def regime_group(value: str) -> str:
    r = clean(value).upper()
    if "70" in r:
        return "Вақтинча сақлаш"
    if "74" in r:
        return "Божхона омбори"
    if "80" in r:
        return "Транзит"
    return clean(value) or "Бошқа"


def _pick_best_table(tables: list) -> pd.DataFrame:
    best, best_count = tables[0], 0
    for t in tables:
        if len(t.columns) <= SRC["decl_no"]:
            continue
        candidate = t.iloc[2:].copy()
        valid = candidate[SRC["decl_no"]].map(lambda v: bool(clean(str(v)) and not pd.isna(v))).sum()
        if valid > best_count:
            best, best_count = candidate, valid
    return best if best_count > 0 else tables[0].iloc[2:].copy()


def read_source(path: Path) -> pd.DataFrame:
    tables = pd.read_html(path, encoding="utf-8")
    df = _pick_best_table(tables)
    df = df[df[SRC["decl_no"]].notna()].copy()
    df = df[~df[SRC["reason"]].map(clean).str.lower().str.startswith("сабаби:")].copy()
    df["_decl_sort"] = df[SRC["decl_no"]].map(clean)
    df["_goods_sort"] = num(df[SRC["goods_no"]])
    df.sort_values(["_decl_sort", "_goods_sort"], kind="mergesort", inplace=True)
    df.reset_index(drop=True, inplace=True)
    df["_partiya"] = (df["_decl_sort"] != df["_decl_sort"].shift(-1)).astype(int)
    for k in ["weight_kg", "value_usd_k", "customs_value_k", "pay_total_k", "pay_20_k", "pay_27_k", "pay_29_k", "priv_total_k", "priv_20_k", "priv_27_k", "priv_29_k"]:
        df[f"_{k}"] = num(df[SRC[k]])
    df["_weight_tn"] = df["_weight_kg"] / 1000
    df["_pay_total_mln"] = df["_pay_total_k"] / 1000
    df["_priv_total_mln"] = df["_priv_total_k"] / 1000
    parsed_dates = pd.to_datetime(df[SRC["decl_date"]], errors="coerce", dayfirst=True)
    df["_date_text"] = df[SRC["decl_date"]].map(parse_date)
    df["_date"] = parsed_dates
    df["_year"] = parsed_dates.dt.year
    tnved = read_tnved()
    df["_post_code"] = df[SRC["post"]].map(post_code)
    df["_post_group"] = df[SRC["post"]].map(post_group)
    df["_regime_group"] = df[SRC["regime"]].map(regime_group)
    df["_regime_code"] = df[SRC["regime"]].map(regime_code)
    hs_names = df[SRC["hs"]].map(lambda x: tnved_name(x, tnved))
    fallback = hs_names == "Boshqa tovarlar"
    if fallback.any():
        hs_names = hs_names.copy()
        hs_names[fallback] = df.loc[fallback, SRC["goods"]].map(lambda x: clean(x) or "Boshqa tovarlar")
    df["_tnved_name"] = hs_names
    return df


def report_date_from_name(name: str) -> datetime | None:
    patterns = [
        r"(?P<y>20\d{2})[.\-_](?P<m>\d{2})[.\-_](?P<d>\d{2})",
        r"(?P<d>\d{2})[.,\-_](?P<m>\d{2})[.,\-_](?P<y>20\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, name)
        if match:
            parts = {k: int(v) for k, v in match.groupdict().items()}
            return datetime(parts["y"], parts["m"], parts["d"])
    return None


def read_deposit(path: Path | None) -> tuple[dict[str, float], datetime | None, float]:
    if not path:
        return {}, None, 0.0
    raw = pd.read_excel(path, sheet_name=0, header=None)
    header_idx = None
    stir_col = 1
    amount_col = 8
    inn_words = {"inn", "stir", "\u0438\u043d\u043d", "\u0441\u0442\u0438\u0440"}
    amount_words = {"kredit", "summa", "qoldiq", "\u043a\u0440\u0435\u0434\u0438\u0442", "\u0441\u0443\u043c\u043c\u0430", "\u043e\u0441\u0442\u0430\u0442\u043e\u043a"}
    for idx, row in raw.head(20).iterrows():
        labels = [clean(v).lower() for v in row.tolist()]
        for i, label in enumerate(labels):
            if label in inn_words:
                stir_col = i
                header_idx = idx
            if label in amount_words:
                amount_col = i
                header_idx = idx if header_idx is None else header_idx
    if header_idx is None:
        header_idx = 3
    out: dict[str, float] = {}
    for _, row in raw.iloc[header_idx + 1:].iterrows():
        stir = re.sub(r"\D", "", str(row.iloc[stir_col] if len(row) > stir_col else ""))
        if not stir:
            continue
        raw_amount = row.iloc[amount_col] if len(row) > amount_col else 0
        amount = pd.to_numeric(str(raw_amount).replace(" ", "").replace("\xa0", "").replace(",", "."), errors="coerce")
        if pd.isna(amount):
            continue
        out[stir] = out.get(stir, 0.0) + float(amount) / 1_000_000
    return out, report_date_from_name(path.name), sum(out.values())

def read_tnved(path: Path | None = None) -> dict[str, str]:
    path = path or TNVED_PATH
    if not path or not path.exists():
        return {}
    raw = pd.read_excel(path, sheet_name="Асос", header=1, dtype=str)
    mapping: dict[str, str] = {}
    for _, row in raw.iterrows():
        name = clean(row.get("Spr_T2_uz")) or clean(row.get("Spr_T1_uz"))
        if not name:
            continue
        for col in ["Kod", "9 хона", "8 хона", "7 хона", "6 хона"]:
            code = digits(row.get(col))
            if code and code not in mapping:
                mapping[code] = name
    return mapping


def tnved_name(code, mapping: dict[str, str]) -> str:
    code = digits(code)
    if not code:
        return "Boshqa tovarlar"
    for length in [10, 9, 8, 7, 6, 4, 2]:
        key = code[:length]
        if key in mapping:
            return to_latin(clean(mapping[key]))
    return "Boshqa tovarlar"


def regime_code(value: str) -> str:
    r = clean(value).upper()
    if "80" in r or "TD80" in r or "TR80" in r or "ТР80" in r:
        return "TR80"
    if "74" in r:
        return "IM74"
    if "70" in r:
        return "IM70"
    return clean(value) or "Boshqa"


def expiry_date(row) -> pd.Timestamp:
    d = row["_date"]
    if pd.isna(d):
        return pd.NaT
    code = regime_code(row[SRC["regime"]])
    if code == "IM74":
        return d + pd.DateOffset(years=3)
    if code == "IM70":
        return d + pd.DateOffset(months=2)
    if code == "TR80":
        return d + pd.DateOffset(days=61)
    return pd.NaT


def summarize_goods(items: pd.Series) -> str:
    cleaned = [clean(x) for x in items if clean(x)]
    if not cleaned:
        return ""
    top = cleaned[:3]
    text = "; ".join(t[:70] for t in top)
    return text + ("..." if len(cleaned) > 3 else "")


def goods_category(items: pd.Series) -> str:
    text = " ".join(clean(x).lower() for x in items if clean(x))
    rules = [
        ("qurilish mollari", ["строител", "цемент", "кабель", "transformator", "бетон", "плит", "eshik", "deraza", "armatura", "сантех"]),
        ("yoqilg'i mahsulotlari", ["топливо", "yoqilg", "бензин", "дизел", "керосин", "jet", "ts-1", "нефт"]),
        ("avtomobillar va ehtiyot qismlar", ["автомоб", "легковой", "bmw", "chevrolet", "liхiang", "lixiang", "запас", "ehtiyot", "двигател", "шина"]),
        ("lift va uskunalar", ["лифт", "оборуд", "ускуна", "станок", "machine", "apparat"]),
        ("oziq-ovqat mahsulotlari", ["oziq", "озиқ", "овқат", "food", "сахар", "шакар", "рис", "guruch", "масло", "yog", "молок", "sut", "мяс", "go'sht"]),
        ("kimyo mahsulotlari", ["хим", "kimyo", "лак", "краска", "поли", "смола", "реагент"]),
        ("farmatsevtika mahsulotlari", ["фарма", "dori", "лекар", "medic", "препарат"]),
        ("elektrotexnika mahsulotlari", ["электр", "кабель", "lamp", "аккумулятор", "battery"]),
    ]
    for label, needles in rules:
        if any(n in text for n in needles):
            return label
    return summarize_goods(items)


def tnved_category(items: pd.Series) -> str:
    vals = [clean(x) for x in items if clean(x)]
    if not vals:
        return "Boshqa tovarlar"
    counts = pd.Series(vals).value_counts()
    return counts.index[0]


def base_rows(df: pd.DataFrame) -> list[list]:
    rows = []
    for _, r in df.iterrows():
        rows.append([
            r[SRC["tr"]], r[SRC["decl_type"]], r[SRC["regime"]], r[SRC["decl_no"]],
            int(r["_partiya"]), r["_date_text"], clean(r[SRC["stir"]]), r[SRC["company"]],
            r[SRC["goods_no"]], r[SRC["hs"]], r[SRC["goods"]], r["_weight_kg"],
            r["_weight_tn"], r["_value_usd_k"], r[SRC["reason"]], r[SRC["post"]],
            r[SRC["warehouse"]], r[SRC["country"]], r["_customs_value_k"],
            r["_pay_total_k"], r["_pay_total_mln"], r["_pay_20_k"], r["_pay_27_k"],
            r["_pay_29_k"], r["_priv_total_k"], r["_priv_total_mln"], r["_priv_20_k"],
            r["_priv_27_k"], r["_priv_29_k"],
        ])
    return rows


def agg(df: pd.DataFrame, by: list[str]) -> pd.DataFrame:
    return df.groupby(by, dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
        imtiyoz=("_priv_total_mln", "sum"),
    ).reset_index()


def metric(row, include_imtiyoz=True):
    vals = [int(row.partiya), float(row.vazn), float(row.qiymat), float(row.tolov)]
    if include_imtiyoz:
        vals.append(float(row.imtiyoz))
    return vals


def row_or_zero(grouped: pd.DataFrame, mask, include_imtiyoz=True):
    hit = grouped[mask]
    if hit.empty:
        return [0, 0, 0, 0, 0] if include_imtiyoz else [0, 0, 0, 0]
    return metric(hit.iloc[0], include_imtiyoz)


def jami_rows(df: pd.DataFrame) -> list[list]:
    g = agg(df, ["_regime_group", "_post_group"])
    order = [
        ("Вақтинча сақлаш", "Toshkent xalqaro aeroporti CHBP"), ("Вақтинча сақлаш", "Avia yuklar TIF"), ("Вақтинча сақлаш", "Elektron tijorat TIF"),
        ("Божхона омбори", "Toshkent xalqaro aeroporti CHBP"), ("Божхона омбори", "Avia yuklar TIF"), ("Божхона омбори", "Elektron tijorat TIF"),
        ("Транзит", "Toshkent xalqaro aeroporti CHBP"), ("Транзит", "Avia yuklar TIF"), ("Транзит", "Elektron tijorat TIF"),
    ]
    return [row_or_zero(g, (g["_regime_group"].eq(r) & g["_post_group"].eq(p))) for r, p in order]


def combined_jami_rows(df: pd.DataFrame) -> list[list]:
    labels = [
        ("IM70", "Toshkent xalqaro aeroporti CHBP"),
        ("IM70", "Avia yuklar TIF"),
        ("IM70", "Elektron tijorat TIF"),
        ("IM74", "Toshkent xalqaro aeroporti CHBP"),
        ("IM74", "Avia yuklar TIF"),
        ("IM74", "Elektron tijorat TIF"),
        ("TR80", "Toshkent xalqaro aeroporti CHBP"),
        ("TR80", "Avia yuklar TIF"),
        ("TR80", "Elektron tijorat TIF"),
    ]
    rows = []
    for i, ((regime, post), vals) in enumerate(zip(labels, jami_rows(df)), start=1):
        rows.append([i, regime, post, *vals])
    return rows


def by_year_rows(df: pd.DataFrame) -> list[list]:
    g = agg(df, ["_post_group", "_regime_group", "_year"])
    order = [
        ("Avia yuklar TIF", "Вақтинча сақлаш", 2026),
        ("Avia yuklar TIF", "Божхона омбори", 2023), ("Avia yuklar TIF", "Божхона омбори", 2024),
        ("Avia yuklar TIF", "Божхона омбори", 2025), ("Avia yuklar TIF", "Божхона омбори", 2026),
        ("Avia yuklar TIF", "Транзит", 2025), ("Avia yuklar TIF", "Транзит", 2026),
        ("Elektron tijorat TIF", "Божхона омбори", 2024), ("Elektron tijorat TIF", "Божхона омбори", 2025),
        ("Elektron tijorat TIF", "Божхона омбори", 2026), ("Elektron tijorat TIF", "Транзит", 2026),
        ("Toshkent xalqaro aeroporti CHBP", "Божхона омбори", 2024), ("Toshkent xalqaro aeroporti CHBP", "Божхона омбори", 2025),
        ("Toshkent xalqaro aeroporti CHBP", "Божхона омбори", 2026), ("Toshkent xalqaro aeroporti CHBP", "Транзит", 2025),
    ]
    rows = []
    for post, regime, year in order:
        rows.append(row_or_zero(g, g["_post_group"].eq(post) & g["_regime_group"].eq(regime) & g["_year"].eq(year), False))
    return rows


def company_rows(df: pd.DataFrame, deposits: dict[str, float] | None = None) -> list[list]:
    deposits = deposits or {}
    first_names = df.groupby(SRC["stir"], dropna=False)[SRC["company"]].first()
    g = df.groupby(SRC["stir"], dropna=False).agg(
        partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    rows = []
    for i, r in enumerate(g.itertuples(index=False), start=1):
        stir = clean(r[0])
        rows.append([i, first_names.get(r[0], ""), stir, int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov), float(deposits.get(stir, 0.0))])
    return rows


def warehouse_rows(df: pd.DataFrame) -> list[list]:
    df = df.copy()
    df["_warehouse_clean"] = df[SRC["warehouse"]].map(clean).replace("", "Ўз омбор")
    g = df.groupby("_warehouse_clean", dropna=False).agg(
        partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    return [[i, r[0], int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def reason_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby(SRC["reason"], dropna=False).agg(
        partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["partiya", "qiymat"], ascending=False)
    return [[i, r[0], int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def month_rows(df: pd.DataFrame, months: int) -> list[list]:
    max_date = df["_date"].max()
    cutoff = max_date - pd.DateOffset(months=months)
    order = [
        ("Вақтинча сақлаш", None), ("Вақтинча сақлаш", "ТХА"), ("Вақтинча сақлаш", "Авиа юклар"), ("Вақтинча сақлаш", "Бош почтамт"),
        ("Божхона омбори", None), ("Божхона омбори", "ТХА"), ("Божхона омбори", "Авиа юклар"), ("Божхона омбори", "Бош почтамт"),
    ]
    rows = []
    for regime, post in order:
        mask = df["_regime_group"].eq(regime)
        if post:
            mask &= df["_post_group"].eq(post)
        total = df[mask]
        over = df[mask & (df["_date"] <= cutoff)]
        under_partiya = int(total["_partiya"].sum() - over["_partiya"].sum())
        under_qiymat = float(total["_value_usd_k"].sum() - over["_value_usd_k"].sum())
        over_partiya = int(over["_partiya"].sum())
        over_qiymat = float(over["_value_usd_k"].sum())
        total_partiya = int(total["_partiya"].sum())
        total_qiymat = float(total["_value_usd_k"].sum())
        pct = 0 if total_qiymat == 0 else over_qiymat / total_qiymat
        rows.append([under_partiya, under_qiymat, over_partiya, over_qiymat, total_partiya, total_qiymat, pct])
    grand = [
        rows[0][0] + rows[4][0], rows[0][1] + rows[4][1], rows[0][2] + rows[4][2],
        rows[0][3] + rows[4][3], rows[0][4] + rows[4][4], rows[0][5] + rows[4][5], 0,
    ]
    grand[6] = 0 if grand[5] == 0 else grand[3] / grand[5]
    return [grand] + rows


def own_warehouse_rows(df: pd.DataFrame, report_date: datetime, over_3_months: bool = False) -> list[list]:
    own = df[df[SRC["warehouse"]].map(clean).eq("")].copy()
    if over_3_months:
        cutoff = pd.Timestamp(report_date) - pd.DateOffset(months=3)
        own = own[own["_date"].le(cutoff)].copy()
    if own.empty:
        return []
    g = own.groupby(SRC["stir"], dropna=False).agg(
        company=(SRC["company"], "first"),
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
        first_date=("_date", "min"),
        goods=("_tnved_name", tnved_category),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    rows = []
    for i, r in enumerate(g.itertuples(index=False), start=1):
        first_date = "" if pd.isna(r.first_date) else pd.Timestamp(r.first_date).strftime("%d.%m.%Y")
        rows.append([i, r.company, clean(r[0]), int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov), first_date, r.goods])
    return rows


def with_expiry(df: pd.DataFrame, report_date: datetime) -> pd.DataFrame:
    out = df.copy()
    out["_regime_code"] = out[SRC["regime"]].map(regime_code)
    out["_expiry"] = out.apply(expiry_date, axis=1)
    out["_expired"] = out["_expiry"].notna() & out["_expiry"].lt(pd.Timestamp(report_date))
    return out


def expired_summary_rows(df: pd.DataFrame, report_date: datetime) -> list[list]:
    d = with_expiry(df, report_date)
    posts = [
        ('"Toshkent" xalqaro aeroporti CHBP', "00101"),
        ("Avia yuklar TIF", "00102"),
        ("Elektron tijorat TIF", "00107"),
    ]
    rows = []
    for label, code in posts:
        part = d[d["_post_code"].eq(code)]
        expired = part[part["_expired"]]
        row = [
            label,
            int(part["_partiya"].sum()), float(part["_value_usd_k"].sum()),
            int(expired["_partiya"].sum()), float(expired["_value_usd_k"].sum()),
            0 if part["_value_usd_k"].sum() == 0 else float(expired["_value_usd_k"].sum() / part["_value_usd_k"].sum()),
        ]
        for code in ["IM70", "IM74", "TR80"]:
            p = expired[expired["_regime_code"].eq(code)]
            row += [int(p["_partiya"].sum()), float(p["_value_usd_k"].sum())]
        rows.append(row)
    return rows


def expired_block_total_row(df: pd.DataFrame, report_date: datetime, regimes: set[str] | None = None) -> list:
    """expired_block_rows uchun IBK bo'yicha Jami total qatorini qaytaradi."""
    d = with_expiry(df, report_date)
    d = d[d["_expired"]].copy()
    if regimes:
        d = d[d["_regime_code"].isin(regimes)].copy()
    if d.empty:
        return []
    return [
        "IBK bo'yicha Jami", "", "",
        int(d["_partiya"].sum()), float(d["_value_usd_k"].sum()),
        float(d["_weight_tn"].sum()), float(d["_pay_total_mln"].sum()), "",
    ]


def expired_block_rows(df: pd.DataFrame, report_date: datetime, regimes: set[str] | None = None) -> list[list]:
    d = with_expiry(df, report_date)
    d = d[d["_expired"]].copy()
    if regimes:
        d = d[d["_regime_code"].isin(regimes)].copy()
    if d.empty:
        return []

    posts = [
        ("00101 - \"Toshkent\" xalqaro aeroporti ChBP", "00101"),
        ("00102 - \"Avia yuklar\" TIF", "00102"),
        ("00107 - Elektron tijorat TIF", "00107"),
    ]
    regime_order = ["IM70", "IM74", "TR80"]
    rows: list[list] = []
    idx = 1

    for post_label, post in posts:
        post_df = d[d["_post_code"].eq(post)].copy()
        if post_df.empty:
            continue
        rows.append([post_label, "", "", int(post_df["_partiya"].sum()), float(post_df["_value_usd_k"].sum()), float(post_df["_weight_tn"].sum()), float(post_df["_pay_total_mln"].sum()), ""])

        for regime in regime_order:
            if regimes and regime not in regimes:
                continue
            r_df = post_df[post_df["_regime_code"].eq(regime)].copy()
            if r_df.empty:
                continue
            rows.append([regime, "", "", int(r_df["_partiya"].sum()), float(r_df["_value_usd_k"].sum()), float(r_df["_weight_tn"].sum()), float(r_df["_pay_total_mln"].sum()), ""])

            g = r_df.groupby(SRC["stir"], dropna=False).agg(
                company=(SRC["company"], "first"),
                partiya=("_partiya", "sum"),
                qiymat=("_value_usd_k", "sum"),
                vazn=("_weight_tn", "sum"),
                tolov=("_pay_total_mln", "sum"),
                reason=(SRC["reason"], "first"),
            ).reset_index().sort_values(["partiya", "qiymat", "tolov"], ascending=False)

            for item in g.itertuples(index=False):
                rows.append([
                    idx,
                    item.company,
                    clean(item[0]),
                    int(item.partiya),
                    float(item.qiymat),
                    float(item.vazn),
                    float(item.tolov),
                    clean_reason_display(item.reason),
                ])
                idx += 1
    return rows


def expired_detail_rows(df: pd.DataFrame, report_date: datetime, regimes: set[str] | None = None) -> list[list]:
    d = with_expiry(df, report_date)
    d = d[d["_expired"]].copy()
    if regimes:
        d = d[d["_regime_code"].isin(regimes)].copy()
    if d.empty:
        return []
    g = d.groupby(["_post_group", SRC["stir"], "_regime_code"], dropna=False).agg(
        company=(SRC["company"], "first"),
        partiya=("_partiya", "sum"),
        qiymat=("_value_usd_k", "sum"),
        vazn=("_weight_tn", "sum"),
        tolov=("_pay_total_mln", "sum"),
        first_date=("_date", "min"),
        expiry=("_expiry", "min"),
        reason=(SRC["reason"], "first"),
        goods=(SRC["goods"], summarize_goods),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    rows = []
    for i, r in enumerate(g.itertuples(index=False), start=1):
        rows.append([i, r[0], r.company, clean(r[1]), int(r.partiya), float(r.qiymat), float(r.vazn), float(r.tolov), r.reason])
    return rows


def expired_7074_rows(df: pd.DataFrame, report_date: datetime) -> list[list]:
    d = with_expiry(df, report_date)
    d = d[d["_expired"] & d["_regime_code"].isin({"IM70", "IM74"})].copy()
    if d.empty:
        return []
    g = d.groupby([SRC["decl_no"], "_regime_code", SRC["stir"]], dropna=False).agg(
        company=(SRC["company"], "first"),
        goods=("_tnved_name", tnved_category),
        qiymat=("_value_usd_k", "sum"),
    ).reset_index().sort_values("qiymat", ascending=False)
    rows = []
    for i, r in enumerate(g.itertuples(index=False), start=1):
        rows.append([i, clean(r[0]), r[1], r.company, r.goods, float(r.qiymat)])
    return rows


def food_rows(df: pd.DataFrame) -> list[list]:
    categories = [
        ("Ichimliklar (alkogolsiz)", "ичимлик|напит|drink|сок|water|вода"),
        ("Shakar va qandolat mahsulotlari", "шакар|сахар|sugar|qand|конфет|шоколад"),
        ("Sut mahsulotlari, tuxum va asal", "сут|молок|milk|tuxum|яйц|egg|asal|мед"),
        ("Go'sht va go'sht mahsulotlari", "go'sht|мяс|meat|колбас|tovuq|куриц"),
        ("Yog' va moy mahsulotlari", "yog|мой|масло|oil"),
        ("Don, un va yorma mahsulotlari", "don|un|мука|круп|guruch|рис|bug'doy|пшениц"),
        ("Meva-sabzavot mahsulotlari", "мева|сабзавот|овощ|фрукт|картоф|томат|пиёз|лук"),
        ("Boshqa oziq-ovqatlar", "oziq|овқат|озиқ|food|confection|кондитер|озуқ|пищ"),
    ]
    max_date = df["_date"].max()
    cutoff = max_date - pd.DateOffset(months=3)
    rows = []
    used = pd.Series(False, index=df.index)
    for label, pattern in categories:
        hay = (df["_tnved_name"].map(clean) + " " + df[SRC["goods"]].map(clean))
        mask = hay.str.contains(pattern, case=False, na=False, regex=True)
        used |= mask
        part = df[mask]
        over = part[part["_date"].le(cutoff)]
        total_q = float(part["_value_usd_k"].sum())
        over_q = float(over["_value_usd_k"].sum())
        rows.append([len(rows) + 1, label, float(part["_weight_tn"].sum()), total_q, float(over["_weight_tn"].sum()), over_q, 0 if total_q == 0 else over_q / total_q * 100])
    other = df[~used & df["_tnved_name"].map(clean).str.contains("озиқ|овқат|пищ|food|озуқ", case=False, na=False, regex=True)]
    if not other.empty:
        over = other[other["_date"].le(cutoff)]
        total_q = float(other["_value_usd_k"].sum())
        over_q = float(over["_value_usd_k"].sum())
        rows.append([len(rows) + 1, "Boshqa mahsulotlar", float(other["_weight_tn"].sum()), total_q, float(over["_weight_tn"].sum()), over_q, 0 if total_q == 0 else over_q / total_q * 100])
    return rows


def goods_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby("_tnved_name", dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
        korxona=(SRC["stir"], "nunique"),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    return [[i, clean(r[0]), int(r.partiya), int(r.korxona), float(r.vazn), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def reason_year_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby([SRC["reason"], "_year"], dropna=False).agg(
        partiya=("_partiya", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["_year", "qiymat"], ascending=[True, False])
    return [[i, r[0], "" if pd.isna(r[1]) else int(r[1]), int(r.partiya), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def company_regime_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby([SRC["stir"], "_regime_group"], dropna=False).agg(
        company=(SRC["company"], "first"),
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    return [[i, r.company, clean(r[0]), r[1], int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def post_country_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby([SRC["post"], SRC["country"]], dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    return [[i, r[0], r[1], int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def jami_year_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby(["_year", "_regime_group"], dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["_year", "qiymat"], ascending=[True, False])
    rows = []
    for i, r in enumerate(g.itertuples(index=False), start=1):
        year = "" if pd.isna(r[0]) else int(r[0])
        rows.append([i, year, r[1], int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)])
    return rows


def product_analysis_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby("_tnved_name", dropna=False).agg(
        partiya=("_partiya", "sum"),
        korxona=(SRC["stir"], "nunique"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
        imtiyoz=("_priv_total_mln", "sum"),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    return [[i, r[0], int(r.partiya), int(r.korxona), float(r.vazn), float(r.qiymat), float(r.tolov), float(r.imtiyoz)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def post_product_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby(["_post_group", "_tnved_name"], dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["_post_group", "qiymat"], ascending=[True, False])
    return [[i, r[0], r[1], int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def warehouse_product_rows(df: pd.DataFrame) -> list[list]:
    d = df.copy()
    d["_warehouse_clean"] = d[SRC["warehouse"]].map(clean).replace("", "O'z ombor")
    g = d.groupby(["_warehouse_clean", "_tnved_name"], dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    return [[i, r[0], r[1], int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)] for i, r in enumerate(g.itertuples(index=False), start=1)]


def country_rows(df: pd.DataFrame) -> list[list]:
    g = df.groupby(SRC["country"], dropna=False).agg(
        partiya=("_partiya", "sum"),
        korxona=(SRC["stir"], "nunique"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index().sort_values(["qiymat", "partiya"], ascending=False)
    rows = []
    for i, r in enumerate(g.itertuples(index=False), start=1):
        rows.append([i, clean(r[0]) or "Ko'rsatilmagan", int(r.partiya), int(r.korxona), float(r.vazn), float(r.qiymat), float(r.tolov)])
    return rows


def age_plus_rows(df: pd.DataFrame, report_date: datetime) -> list[list]:
    buckets = [
        ("3 oy+", pd.Timestamp(report_date) - pd.DateOffset(months=3)),
        ("6 oy+", pd.Timestamp(report_date) - pd.DateOffset(months=6)),
        ("1 yil+", pd.Timestamp(report_date) - pd.DateOffset(years=1)),
    ]
    grouped: dict[tuple[str, str], dict[str, tuple[int, float]]] = {}
    for label, cutoff in buckets:
        d = df[df["_date"].le(cutoff)].copy()
        if d.empty:
            continue
        g = d.groupby(["_regime_code", "_post_group"], dropna=False).agg(
            partiya=("_partiya", "sum"),
            qiymat=("_value_usd_k", "sum"),
        ).reset_index()
        for r in g.itertuples(index=False):
            grouped.setdefault((r[0], r[1]), {})[label] = (int(r.partiya), float(r.qiymat))

    rows = []
    for idx, ((regime, post), values) in enumerate(
        sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1])),
        start=1,
    ):
        row = [idx, regime, post]
        for label, _ in buckets:
            partiya, qiymat = values.get(label, (0, 0.0))
            row += [partiya, qiymat]
        rows.append(row)
    return rows


def write_json(path: Path, rows: list[list]):
    def convert(value):
        if isinstance(value, str):
            return to_latin(value)
        if isinstance(value, list):
            return [convert(v) for v in value]
        return value

    path.write_text(json.dumps(convert(rows), ensure_ascii=False), encoding="utf-8-sig")


CYR_TO_LAT = str.maketrans({
    "А": "A", "а": "a", "Б": "B", "б": "b", "В": "V", "в": "v", "Г": "G", "г": "g",
    "Д": "D", "д": "d", "Е": "E", "е": "e", "Ё": "Yo", "ё": "yo", "Ж": "J", "ж": "j",
    "З": "Z", "з": "z", "И": "I", "и": "i", "Й": "Y", "й": "y", "К": "K", "к": "k",
    "Л": "L", "л": "l", "М": "M", "м": "m", "Н": "N", "н": "n", "О": "O", "о": "o",
    "П": "P", "п": "p", "Р": "R", "р": "r", "С": "S", "с": "s", "Т": "T", "т": "t",
    "У": "U", "у": "u", "Ф": "F", "ф": "f", "Х": "X", "х": "x", "Ц": "S", "ц": "s",
    "Ч": "Ch", "ч": "ch", "Ш": "Sh", "ш": "sh", "Щ": "Sh", "щ": "sh", "Ъ": "", "ъ": "",
    "Ы": "I", "ы": "i", "Ь": "", "ь": "", "Э": "E", "э": "e", "Ю": "Yu", "ю": "yu",
    "Я": "Ya", "я": "ya", "Қ": "Q", "қ": "q", "Ғ": "G'", "ғ": "g'", "Ў": "O'", "ў": "o'",
    "Ҳ": "H", "ҳ": "h",
})
CYR_TO_LAT.update(str.maketrans({
    "Ё": "YO", "Ч": "CH", "Ш": "SH", "Щ": "SH", "Ю": "YU", "Я": "YA",
}))


def to_latin(value: str) -> str:
    text = value.translate(CYR_TO_LAT)
    fixes = {
        "Engil": "Yengil",
        "engil": "yengil",
        "Yyengil": "Yengil",
        "yyengil": "yengil",
        "Medisina": "Meditsina",
        "medisina": "meditsina",
    }
    for old, new in fixes.items():
        text = text.replace(old, new)
    return text


def clean_reason_display(value: str) -> str:
    text = to_latin(clean(value))
    match = re.search(r"\bSababi:\s*(.+)$", text, flags=re.IGNORECASE)
    if match:
        text = match.group(1)
    text = re.sub(r"\bMaqsadi:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bMaksadi:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bSababi:\s*", "", text, flags=re.IGNORECASE)
    if "=>" in text:
        text = text.split("=>", 1)[0]
    return clean(text).strip(" ,;=>")


def fmt_uz_number(value: float) -> str:
    return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")


def transliterate_workbook(path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import Border, Side

    wb = load_workbook(path)
    used_titles: set[str] = set()
    for ws in wb.worksheets:
        title = to_latin(ws.title).strip()[:31] or ws.title
        base_title = title
        suffix = 2
        while title in used_titles:
            marker = f" {suffix}"
            title = (base_title[: 31 - len(marker)] + marker).strip()
            suffix += 1
        ws.title = title
        used_titles.add(title)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    cell.value = to_latin(cell.value)

    def fmt(sheet: str, cell_range: str, number_format: str):
        if sheet not in wb.sheetnames:
            return
        ws = wb[sheet]
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="center", vertical="center")

    int_fmt = "# ##0"
    num_fmt = "# ##0.00"
    for sheet, ranges in {
        "Tahlil tovar guruhlari": [("C4:D5000", int_fmt), ("E4:H5000", num_fmt)],
        "Tahlil umumiy": [("D4:D5000", int_fmt), ("E4:H5000", num_fmt)],
        "Tahlil post-tovar": [("D4:D5000", int_fmt), ("E4:G5000", num_fmt)],
        "Tahlil ombor-tovar": [("D4:D5000", int_fmt), ("E4:G5000", num_fmt)],
        "Tahlil davlatlar": [("C4:D5000", int_fmt), ("E4:G5000", num_fmt)],
        "Oziq ovqat": [("B4:B5000", int_fmt), ("D4:H5000", num_fmt)],
        "Korxonalar kesimida": [("C6:C5000", "0"), ("D6:D5000", int_fmt), ("E6:H5000", num_fmt)],
        "Omborlar kesimida": [("C6:C5000", int_fmt), ("D6:F5000", num_fmt)],
        "O'z ombor jami": [("C5:C5000", "0"), ("D5:D5000", int_fmt), ("E5:G5000", num_fmt)],
        "O'z ombor 3 oy+": [("C5:C5000", "0"), ("D5:D5000", int_fmt), ("E5:G5000", num_fmt)],
        "Jami muddati o'tgan": [("B7:B20", int_fmt), ("C7:C20", num_fmt), ("D7:D20", int_fmt), ("E7:E20", num_fmt), ("F7:F20", '0.00"%"'), ("G7:G20", int_fmt), ("H7:H20", num_fmt), ("I7:I20", int_fmt), ("J7:J20", num_fmt), ("K7:K20", int_fmt), ("L7:L20", num_fmt)],
        "muddati o`tgan": [("C5:C5000", "0"), ("D5:D5000", int_fmt), ("E5:G5000", num_fmt)],
        "Muddati o'tgan 80": [("C5:C5000", "0"), ("D5:D5000", int_fmt), ("E5:G5000", num_fmt)],
        "Muddati o'tgan 70-74": [("C5:C5000", "0"), ("D5:D5000", int_fmt), ("F5:F5000", num_fmt)],
    }.items():
        for cell_range, number_format in ranges:
            fmt(sheet, cell_range, number_format)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            if any(cell.value not in (None, "") for cell in row):
                for cell in row:
                    if cell.value not in (None, ""):
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(path)


def ps_quote(path: Path) -> str:
    return "'" + str(path).replace("'", "''") + "'"


def run_excel_writer(template: Path, output: Path, files: dict, date_label: str, short_date: str, deposit_label: str, company_count: int):
    shutil.copy2(template, output)
    script = f"""
$ErrorActionPreference = 'Stop'
# Kill any lingering Excel processes from prior failed runs.
# Broken Excel instances leave the COM type library in an unloadable state
# which causes TYPE_E_CANTLOADLIBRARY on the NEXT attempt.
Get-Process -Name EXCEL -ErrorAction SilentlyContinue | ForEach-Object {{
  try {{ $_.Kill() }} catch {{}}
}}
$kw = 0
while ((Get-Process -Name EXCEL -ErrorAction SilentlyContinue) -and $kw -lt 10) {{
  Start-Sleep -Milliseconds 400; $kw++
}}
[System.GC]::Collect(); [System.GC]::WaitForPendingFinalizers()
Start-Sleep -Milliseconds 800
$xlCLSID = [System.Guid]::new("00024500-0000-0000-C000-000000000046")
$excel = $null
for ($attempt = 1; $attempt -le 3; $attempt++) {{
  try {{
    $excel = [System.Activator]::CreateInstance([System.Type]::GetTypeFromCLSID($xlCLSID, $true))
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    break
  }} catch {{
    if ($null -ne $excel) {{
      try {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }} catch {{}}
      $excel = $null
    }}
    [System.GC]::Collect(); [System.GC]::WaitForPendingFinalizers()
    Get-Process -Name EXCEL -ErrorAction SilentlyContinue | ForEach-Object {{ try {{ $_.Kill() }} catch {{}} }}
    if ($attempt -lt 3) {{ Start-Sleep -Seconds 2 }}
  }}
}}
if ($null -eq $excel) {{ throw "Excel COM yuklanmadi" }}
try {{
  $wb = $excel.Workbooks.Open({ps_quote(output)})
  function Write-Json($jsonPath, $sheetName, $row, $col, $clearRows, $clearCols) {{
    $ws = $wb.Worksheets.Item($sheetName)
    if ($clearRows -gt 0 -and $clearCols -gt 0) {{
      try {{
        $ws.Range($ws.Cells($row, $col), $ws.Cells($row + $clearRows - 1, $col + $clearCols - 1)).ClearContents() | Out-Null
      }} catch {{
      }}
    }}
    $data = Get-Content -LiteralPath $jsonPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $rows = @($data).Count
    if ($rows -eq 0) {{ return }}
    $cols = @($data[0]).Count
    $arr = New-Object 'object[,]' $rows, $cols
    for ($r = 0; $r -lt $rows; $r++) {{
      $line = @($data[$r])
      for ($c = 0; $c -lt $cols; $c++) {{
        $arr[$r, $c] = $line[$c]
      }}
    }}
    $ws.Range($ws.Cells($row, $col), $ws.Cells($row + $rows - 1, $col + $cols - 1)).Value2 = $arr
  }}
  function Ensure-Sheet($sheetName) {{
    try {{
      return $wb.Worksheets.Item($sheetName)
    }} catch {{
      $ws = $wb.Worksheets.Add()
      $ws.Name = $sheetName
      return $ws
    }}
  }}
  function Setup-Analysis-Sheet($sheetName, $title, $headers) {{
    $ws = Ensure-Sheet $sheetName
    $ws.Cells.Clear()
    $ws.Range('A1').Value2 = $title
    $ws.Range('A1').Font.Bold = $true
    $ws.Range('A1').Font.Size = 14
    for ($i = 0; $i -lt $headers.Count; $i++) {{
      $cell = $ws.Cells(3, $i + 1)
      $cell.Value2 = $headers[$i]
      $cell.Font.Bold = $true
      $cell.Interior.Color = 8210719
      $cell.Font.Color = 16777215
      $cell.HorizontalAlignment = -4108
      $cell.VerticalAlignment = -4108
    }}
    return $ws
  }}
  Write-Json {ps_quote(files["base"])} 'БАЗА' 3 1 50000 29
  Write-Json {ps_quote(files["jami"])} 'Жами' 8 2 3 5
  Write-Json {ps_quote(files["jami2"])} 'Жами' 12 2 3 5
  Write-Json {ps_quote(files["jami3"])} 'Жами' 16 2 3 5
  Write-Json {ps_quote(files["year1"])} '70-74-80' 8 2 1 4
  Write-Json {ps_quote(files["year2"])} '70-74-80' 10 2 4 4
  Write-Json {ps_quote(files["year3"])} '70-74-80' 15 2 2 4
  Write-Json {ps_quote(files["year4"])} '70-74-80' 20 2 3 4
  Write-Json {ps_quote(files["year5"])} '70-74-80' 24 2 1 4
  Write-Json {ps_quote(files["year6"])} '70-74-80' 28 2 3 4
  Write-Json {ps_quote(files["year7"])} '70-74-80' 32 2 1 4
  Write-Json {ps_quote(files["company"])} 'Корхоналар кесимида' 6 1 2000 8
  Write-Json {ps_quote(files["warehouse"])} 'Омборлар кесимида' 6 1 500 6
  Write-Json {ps_quote(files["reason"])} 'Сақланиш сабаблари' 6 1 500 6
  Write-Json {ps_quote(files["month3"])} '3 ой' 6 2 9 7
  Write-Json {ps_quote(files["month6"])} '6 ой' 6 2 9 7
  Write-Json {ps_quote(files["own_all"])} 'Ўз омбор жами' 5 1 2000 9
  Write-Json {ps_quote(files["own_3m"])} 'Ўз омбор 3 ой+' 5 1 2000 9
  Write-Json {ps_quote(files["sklad_company"])} '70-74-80 склад Корхонабай пар' 6 1 2000 8
  Write-Json {ps_quote(files["expired_summary"])} 'Жами муддати ўтган' 8 1 20 12
  $wb.Worksheets.Item('muddati o`tgan ').Range('A4:H5000').UnMerge()
  $wb.Worksheets.Item('Муддати ўтган 80').Range('A4:H5000').UnMerge()
  Write-Json {ps_quote(files["expired_detail"])} 'muddati o`tgan ' 6 1 5000 8
  Write-Json {ps_quote(files["expired80"])} 'Муддати ўтган 80' 6 1 5000 8
  Write-Json {ps_quote(files["expired7074"])} 'Муддати ўтган 70-74' 8 1 5000 6
  Write-Json {ps_quote(files["food"])} 'Озиқ овқат' 5 2 5000 7
  $wb.Worksheets.Item('Товар номи').Range('A5:G5000').UnMerge()
  Write-Json {ps_quote(files["goods"])} 'Товар номи' 5 1 5000 7
  Write-Json {ps_quote(files["reason_year"])} 'Сақланиш сабаблари йиллар кесим' 5 1 5000 6
  Write-Json {ps_quote(files["gtc"])} 'ГТК' 5 1 5000 7
  Write-Json {ps_quote(files["jami_year"])} 'Жами йил' 6 1 5000 7
  Setup-Analysis-Sheet 'Tahlil tovar guruhlari' 'Tovar guruhlari boyicha tahlil' @('№','Tovar guruhi','Partiya','Korxona','Vazni (tn)','Qiymati (ming $)','Kutilayotgan (mln som)','Imtiyoz (mln som)') | Out-Null
  Write-Json {ps_quote(files["product_analysis"])} 'Tahlil tovar guruhlari' 4 1 5000 8
  Setup-Analysis-Sheet 'Tahlil umumiy' 'Rejim va postlar boyicha umumiy jamlanma' @('№','Rejim','Bojxona posti','Partiya','Vazni (tn)','Qiymati (ming $)','Kutilayotgan (mln som)','Imtiyoz (mln som)') | Out-Null
  Write-Json {ps_quote(files["combined_jami"])} 'Tahlil umumiy' 4 1 5000 8
  Setup-Analysis-Sheet 'Tahlil post-tovar' 'Bojxona postlari va tovar guruhlari kesimida' @('№','Bojxona posti','Tovar guruhi','Partiya','Vazni (tn)','Qiymati (ming $)','Kutilayotgan (mln som)') | Out-Null
  Write-Json {ps_quote(files["post_product"])} 'Tahlil post-tovar' 4 1 5000 7
  Setup-Analysis-Sheet 'Tahlil ombor-tovar' 'Omborlar va tovar guruhlari kesimida' @('№','Ombor','Tovar guruhi','Partiya','Vazni (tn)','Qiymati (ming $)','Kutilayotgan (mln som)') | Out-Null
  Write-Json {ps_quote(files["warehouse_product"])} 'Tahlil ombor-tovar' 4 1 5000 7
  Setup-Analysis-Sheet 'Tahlil davlatlar' 'Yuk jo`natuvchi davlatlar kesimida' @('№','Davlat','Partiya','Korxona','Vazni (tn)','Qiymati (ming $)','Kutilayotgan (mln som)') | Out-Null
  Write-Json {ps_quote(files["country"])} 'Tahlil davlatlar' 4 1 5000 7
  Setup-Analysis-Sheet 'Tahlil muddatlar' '3 oy+, 6 oy+, 1 yil+ muddatlar kesimida' @('№','Rejim','Bojxona posti','3 oy+ partiya','3 oy+ qiymat','6 oy+ partiya','6 oy+ qiymat','1 yil+ partiya','1 yil+ qiymat') | Out-Null
  Write-Json {ps_quote(files["age_plus"])} 'Tahlil muddatlar' 4 1 5000 9
  function Clear-Tail($sheet, $startRow, $jsonPath, $lastCol) {{
    $ws = $wb.Worksheets.Item($sheet)
    $data = Get-Content -LiteralPath $jsonPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $count = @($data).Count
    $first = $startRow + $count
    if ($first -le 5000) {{
      $rng = $ws.Range($ws.Cells($first, 1), $ws.Cells(5000, $lastCol))
      $rng.ClearContents() | Out-Null
      $rng.Borders.LineStyle = -4142
    }}
  }}
  function Apply-Data-Range-Format($sheet, $startRow, $jsonPath, $lastCol) {{
    $ws = $wb.Worksheets.Item($sheet)
    $data = Get-Content -LiteralPath $jsonPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $count = @($data).Count
    if ($count -eq 0) {{ return }}
    $rng = $ws.Range($ws.Cells($startRow, 1), $ws.Cells($startRow + $count - 1, $lastCol))
    $rng.Borders.LineStyle = 1
    $rng.HorizontalAlignment = -4108
    $rng.VerticalAlignment = -4108
  }}
  function Apply-Block-Sheet-Format($sheet, $startRow, $jsonPath, $lastCol) {{
    $ws = $wb.Worksheets.Item($sheet)
    $data = Get-Content -LiteralPath $jsonPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $count = @($data).Count
    if ($count -eq 0) {{ return }}
    $fullRng = $ws.Range($ws.Cells($startRow, 1), $ws.Cells($startRow + $count - 1, $lastCol))
    $fullRng.Font.Bold = $false
    $fullRng.Font.Size = 9
    $fullRng.Borders.LineStyle = 1
    $fullRng.VerticalAlignment = -4108
    $compIdx = 0
    for ($i = 0; $i -lt $count; $i++) {{
      $rowData = @($data[$i])
      $rowNum = $startRow + $i
      $rng = $ws.Range($ws.Cells($rowNum, 1), $ws.Cells($rowNum, $lastCol))
      $col2 = if ($rowData.Count -gt 1) {{ [string]$rowData[1] }} else {{ '' }}
      if ($col2 -eq '') {{
        $rng.Font.Bold = $true
        $rng.Interior.Color = 16247257
        $rng.HorizontalAlignment = -4108
      }} else {{
        $compIdx++
        $rng.Font.Bold = $false
        if ($compIdx % 2 -eq 0) {{ $rng.Interior.Color = 16446960 }} else {{ $rng.Interior.Color = 16777215 }}
        $rng.HorizontalAlignment = -4108
      }}
    }}
  }}
  Clear-Tail 'Корхоналар кесимида' 6 {ps_quote(files["company"])} 8
  Clear-Tail 'Омборлар кесимида' 6 {ps_quote(files["warehouse"])} 6
  Clear-Tail 'Ўз омбор жами' 5 {ps_quote(files["own_all"])} 9
  Clear-Tail 'Ўз омбор 3 ой+' 5 {ps_quote(files["own_3m"])} 9
  Clear-Tail 'Жами муддати ўтган' 8 {ps_quote(files["expired_summary"])} 12
  Clear-Tail 'muddati o`tgan ' 6 {ps_quote(files["expired_detail"])} 8
  Clear-Tail 'Озиқ овқат' 5 {ps_quote(files["food"])} 8
  Apply-Data-Range-Format 'Корхоналар кесимида' 6 {ps_quote(files["company"])} 8
  Apply-Data-Range-Format 'Омборлар кесимида' 6 {ps_quote(files["warehouse"])} 6
  Apply-Data-Range-Format 'Ўз омбор жами' 5 {ps_quote(files["own_all"])} 9
  Apply-Data-Range-Format 'Ўз омбор 3 ой+' 5 {ps_quote(files["own_3m"])} 9
  Apply-Data-Range-Format 'Жами муддати ўтган' 8 {ps_quote(files["expired_summary"])} 12
  Apply-Block-Sheet-Format 'muddati o`tgan ' 6 {ps_quote(files["expired_detail"])} 8
  Apply-Data-Range-Format 'Озиқ овқат' 5 {ps_quote(files["food"])} 8
  Clear-Tail 'Ўз омбор жами' 5 {ps_quote(files["own_all"])} 9
  Clear-Tail 'Ўз омбор 3 ой+' 5 {ps_quote(files["own_3m"])} 9
  $wb.Worksheets.Item('Жами').Range('B3').Value2 = {json.dumps(date_label, ensure_ascii=False)}
  $wb.Worksheets.Item('70-74-80').Range('B2').Value2 = {json.dumps(date_label, ensure_ascii=False)}
  $wb.Worksheets.Item('Корхоналар кесимида').Range('G3').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Корхоналар кесимида').Range('H4').Value2 = {json.dumps(deposit_label, ensure_ascii=False)}
  $companyEnd = 5 + {company_count}
  $wb.Worksheets.Item('Корхоналар кесимида').Range('A5').Value2 = "{company_count} ta korxona bo'yicha Jami:"
  foreach ($col in @('D','E','F','G','H')) {{
    $wb.Worksheets.Item('Корхоналар кесимида').Range($col + '5').Formula = '=SUM(' + $col + '6:' + $col + $companyEnd + ')'
  }}
  $wb.Worksheets.Item('Омборлар кесимида').Range('F3').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Сақланиш сабаблари').Range('F3').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Ўз омбор жами').Range('I2').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Ўз омбор 3 ой+').Range('I2').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Жами муддати ўтган').Range('A2').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Жами муддати ўтган').Range('F6').Value2 = 'Partiyadagi ulushi (%)'
  $wb.Worksheets.Item('Жами муддати ўтган').Range('B7').Formula = '=SUM(B8:B10)'
  $wb.Worksheets.Item('Жами муддати ўтган').Range('C7').Formula = '=SUM(C8:C10)'
  $wb.Worksheets.Item('Жами муддати ўтган').Range('D7').Formula = '=SUM(D8:D10)'
  $wb.Worksheets.Item('Жами муддати ўтган').Range('E7').Formula = '=SUM(E8:E10)'
  $wb.Worksheets.Item('Жами муддати ўтган').Range('F7').Formula = '=IFERROR(D7/B7*100,0)'
  foreach ($r in 8..10) {{
    $wb.Worksheets.Item('Жами муддати ўтган').Range('F' + $r).Formula = '=IFERROR(D' + $r + '/B' + $r + '*100,0)'
  }}
  $wb.Worksheets.Item('muddati o`tgan ').Range('G3').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Муддати ўтган 80').Range('G3').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Муддати ўтган 70-74').Range('A2').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Озиқ овқат').Range('H2').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  $wb.Worksheets.Item('Товар номи').Range('H2').Value2 = {json.dumps(short_date, ensure_ascii=False)}
  foreach ($sheet in @('muddati o`tgan ', 'Муддати ўтган 80')) {{
    $headers = @('№','Korxona nomi','STIR','Partiya','Qiymati (ming $)','Vazni (tn)','Kutilayotgan (mln som)','Saqlanish sababi')
    for ($i = 0; $i -lt $headers.Count; $i++) {{
      $wb.Worksheets.Item($sheet).Cells(4, $i + 1).Value2 = $headers[$i]
    }}
  }}
  function Merge-Expired-Blocks($sheet) {{
    $ws = $wb.Worksheets.Item($sheet)
    for ($r = 6; $r -le 5000; $r++) {{
      $a = $ws.Cells($r, 1).Value2
      $b = $ws.Cells($r, 2).Value2
      $c = $ws.Cells($r, 3).Value2
      if ($a -ne $null -and "$a" -ne "" -and ($b -eq $null -or "$b" -eq "") -and ($c -eq $null -or "$c" -eq "")) {{
        $rng = $ws.Range($ws.Cells($r, 1), $ws.Cells($r, 3))
        $rng.Merge() | Out-Null
        $rng.Font.Bold = $true
        $rng.HorizontalAlignment = -4108
        $rng.VerticalAlignment = -4108
      }}
    }}
  }}
  Merge-Expired-Blocks 'muddati o`tgan '
  Merge-Expired-Blocks 'Муддати ўтган 80'
  function Fmt($sheet, $range, $fmt, $fmtLocal = $null) {{
    try {{
      $wb.Worksheets.Item($sheet).Range($range).NumberFormat = $fmt
    }} catch {{
      if ($fmtLocal -ne $null) {{
        try {{
          $wb.Worksheets.Item($sheet).Range($range).NumberFormatLocal = $fmtLocal
        }} catch {{
        }}
      }}
    }}
  }}
  $intFmt = '#,##0;[Red]-#,##0;0'
  $intFmtLocal = '# ##0;[Red]-# ##0;0'
  $numFmt = '#,##0.00;-#,##0.00;0'
  $numFmtLocal = '# ##0,00;-# ##0,00;0'
  $pctFmt = '0"%"'
  $pctFmtLocal = '0"%"'
  Fmt 'Жами' 'B8:B24' $intFmt $intFmtLocal; Fmt 'Жами' 'C8:F24' $numFmt $numFmtLocal
  Fmt '70-74-80' 'B8:B32' $intFmt $intFmtLocal; Fmt '70-74-80' 'C8:E32' $numFmt $numFmtLocal
  Fmt 'Корхоналар кесимида' 'C6:C5000' '0'; Fmt 'Корхоналар кесимида' 'D6:D5000' $intFmt $intFmtLocal; Fmt 'Корхоналар кесимида' 'E6:H5000' $numFmt $numFmtLocal; Fmt 'Корхоналар кесимида' 'D5:H5' $numFmt $numFmtLocal; Fmt 'Корхоналар кесимида' 'D5:D5' $intFmt $intFmtLocal
  Fmt 'Омборлар кесимида' 'C6:C5000' $intFmt $intFmtLocal; Fmt 'Омборлар кесимида' 'D6:F5000' $numFmt $numFmtLocal
  Fmt 'Ўз омбор жами' 'C5:C5000' '0'; Fmt 'Ўз омбор жами' 'D5:D5000' $intFmt $intFmtLocal; Fmt 'Ўз омбор жами' 'E5:G5000' $numFmt $numFmtLocal
  Fmt 'Ўз омбор 3 ой+' 'C5:C5000' '0'; Fmt 'Ўз омбор 3 ой+' 'D5:D5000' $intFmt $intFmtLocal; Fmt 'Ўз омбор 3 ой+' 'E5:G5000' $numFmt $numFmtLocal
  Fmt 'Жами муддати ўтган' 'B7:B20' $intFmt $intFmtLocal; Fmt 'Жами муддати ўтган' 'D7:D20' $intFmt $intFmtLocal; Fmt 'Жами муддати ўтган' 'G7:G20' $intFmt $intFmtLocal; Fmt 'Жами муддати ўтган' 'I7:I20' $intFmt $intFmtLocal; Fmt 'Жами муддати ўтган' 'K7:K20' $intFmt $intFmtLocal
  Fmt 'Жами муддати ўтган' 'C7:C20' $numFmt $numFmtLocal; Fmt 'Жами муддати ўтган' 'E7:E20' $numFmt $numFmtLocal; Fmt 'Жами муддати ўтган' 'F7:F20' $pctFmt $pctFmtLocal; Fmt 'Жами муддати ўтган' 'H7:H20' $numFmt $numFmtLocal; Fmt 'Жами муддати ўтган' 'J7:J20' $numFmt $numFmtLocal; Fmt 'Жами муддати ўтган' 'L7:L20' $numFmt $numFmtLocal
  Fmt 'muddati o`tgan ' 'C5:C5000' '0'; Fmt 'muddati o`tgan ' 'D5:D5000' $intFmt $intFmtLocal; Fmt 'muddati o`tgan ' 'E5:G5000' $numFmt $numFmtLocal
  Fmt 'Муддати ўтган 80' 'C5:C5000' '0'; Fmt 'Муддати ўтган 80' 'D5:D5000' $intFmt $intFmtLocal; Fmt 'Муддати ўтган 80' 'E5:G5000' $numFmt $numFmtLocal
  Fmt 'Муддати ўтган 70-74' 'C5:C5000' '0'; Fmt 'Муддати ўтган 70-74' 'D5:D5000' $intFmt $intFmtLocal; Fmt 'Муддати ўтган 70-74' 'F5:F5000' $numFmt $numFmtLocal
  Fmt 'Озиқ овқат' 'B4:B5000' $intFmt $intFmtLocal; Fmt 'Озиқ овқат' 'D4:G5000' $numFmt $numFmtLocal; Fmt 'Озиқ овқат' 'H4:H5000' '0.00' '0,00'
  Fmt 'Озиқ овқат' 'D4:H4' '#,##0.00'
  $wb.Worksheets.Item('Озиқ овқат').Range('D4:H4').HorizontalAlignment = -4108
  $wb.Worksheets.Item('Озиқ овқат').Range('D4:H4').VerticalAlignment = -4108
  Fmt 'БАЗА' 'G3:G50000' '0'; Fmt 'БАЗА' 'E3:E50000' 'dd.mm.yyyy'; Fmt 'БАЗА' 'I3:I50000' $intFmt $intFmtLocal; Fmt 'БАЗА' 'L3:M50000' $numFmt $numFmtLocal; Fmt 'БАЗА' 'S3:AC50000' $numFmt $numFmtLocal
  Fmt 'Tahlil umumiy' 'D4:D5000' $intFmt $intFmtLocal; Fmt 'Tahlil umumiy' 'E4:H5000' $numFmt $numFmtLocal
  Fmt 'Tahlil tovar guruhlari' 'C4:D5000' $intFmt $intFmtLocal; Fmt 'Tahlil tovar guruhlari' 'E4:H5000' $numFmt $numFmtLocal
  Fmt 'Tahlil post-tovar' 'D4:D5000' $intFmt $intFmtLocal; Fmt 'Tahlil post-tovar' 'E4:G5000' $numFmt $numFmtLocal
  Fmt 'Tahlil ombor-tovar' 'D4:D5000' $intFmt $intFmtLocal; Fmt 'Tahlil ombor-tovar' 'E4:G5000' $numFmt $numFmtLocal
  Fmt 'Tahlil davlatlar' 'C4:D5000' $intFmt $intFmtLocal; Fmt 'Tahlil davlatlar' 'E4:G5000' $numFmt $numFmtLocal
  Fmt 'Tahlil muddatlar' 'D4:D5000' $intFmt $intFmtLocal; Fmt 'Tahlil muddatlar' 'F4:F5000' $intFmt $intFmtLocal; Fmt 'Tahlil muddatlar' 'H4:H5000' $intFmt $intFmtLocal
  Fmt 'Tahlil muddatlar' 'E4:E5000' $numFmt $numFmtLocal; Fmt 'Tahlil muddatlar' 'G4:G5000' $numFmt $numFmtLocal; Fmt 'Tahlil muddatlar' 'I4:I5000' $numFmt $numFmtLocal
  function Center($sheet, $range) {{
    $r = $wb.Worksheets.Item($sheet).Range($range)
    $r.HorizontalAlignment = -4108
    $r.VerticalAlignment = -4108
  }}
  Center 'Жами' 'B7:F24'
  Center '70-74-80' 'B8:E32'
  Center 'Корхоналар кесимида' 'C5:H5000'
  Center 'Омборлар кесимида' 'C6:F5000'
  Center 'Ўз омбор жами' 'C5:G5000'
  Center 'Ўз омбор 3 ой+' 'C5:G5000'
  Center 'Жами муддати ўтган' 'B7:L20'
  Center 'muddati o`tgan ' 'C5:G5000'
  Center 'Муддати ўтган 80' 'C5:G5000'
  Center 'Муддати ўтган 70-74' 'C5:F5000'
  Center 'Озиқ овқат' 'B5:H5000'
  Center 'БАЗА' 'A3:AC50000'
  Center 'Tahlil umumiy' 'A4:H5000'
  Center 'Tahlil tovar guruhlari' 'A4:H5000'
  Center 'Tahlil post-tovar' 'A4:G5000'
  Center 'Tahlil ombor-tovar' 'A4:G5000'
  Center 'Tahlil davlatlar' 'A4:G5000'
  Center 'Tahlil muddatlar' 'A4:I5000'
  $wb.Worksheets.Item('Корхоналар кесимида').Range('F6:F5000').Font.Bold = $true
  $wb.Worksheets.Item('Корхоналар кесимида').Range('G6:G5000').Font.Bold = $false
  function Move-Sheets-To-Front($names) {{
    for ($i = $names.Count - 1; $i -ge 0; $i--) {{
      try {{
        $ws = $wb.Worksheets.Item($names[$i])
        $ws.Move($wb.Worksheets.Item(1))
      }} catch {{
      }}
    }}
  }}
  Move-Sheets-To-Front @(
    'Жами',
    '70-74-80',
    'Корхоналар кесимида',
    'Омборлар кесимида',
    'Ўз омбор жами',
    'Ўз омбор 3 ой+',
    'Жами муддати ўтган',
    'muddati o`tgan '
  )
  $UpperLatinPairs = @(
      @('Ё','YO'), @('Ч','CH'), @('Ш','SH'), @('Щ','SH'), @('Ю','YU'), @('Я','YA')
  )
  $LatinPairs = @(
      @('Ў','O'''), @('ў','o'''), @('Қ','Q'), @('қ','q'), @('Ғ','G'''), @('ғ','g'''), @('Ҳ','H'), @('ҳ','h'),
      @('А','A'), @('а','a'), @('Б','B'), @('б','b'), @('В','V'), @('в','v'), @('Г','G'), @('г','g'),
      @('Д','D'), @('д','d'), @('Е','E'), @('е','e'), @('Ё','Yo'), @('ё','yo'), @('Ж','J'), @('ж','j'),
      @('З','Z'), @('з','z'), @('И','I'), @('и','i'), @('Й','Y'), @('й','y'), @('К','K'), @('к','k'),
      @('Л','L'), @('л','l'), @('М','M'), @('м','m'), @('Н','N'), @('н','n'), @('О','O'), @('о','o'),
      @('П','P'), @('п','p'), @('Р','R'), @('р','r'), @('С','S'), @('с','s'), @('Т','T'), @('т','t'),
      @('У','U'), @('у','u'), @('Ф','F'), @('ф','f'), @('Х','X'), @('х','x'), @('Ц','S'), @('ц','s'),
      @('Ч','Ch'), @('ч','ch'), @('Ш','Sh'), @('ш','sh'), @('Щ','Sh'), @('щ','sh'), @('Ъ',''), @('ъ',''),
      @('Ы','I'), @('ы','i'), @('Ь',''), @('ь',''), @('Э','E'), @('э','e'), @('Ю','Yu'), @('ю','yu'),
      @('Я','Ya'), @('я','ya')
    )
  function Convert-ToLatin($text) {{
    if ($null -eq $text) {{ return $text }}
    $s = [string]$text
    foreach ($pair in $UpperLatinPairs) {{ $s = $s.Replace($pair[0], $pair[1]) }}
    foreach ($pair in $LatinPairs) {{ $s = $s.Replace($pair[0], $pair[1]) }}
    $s = $s.Replace('Engil', 'Yengil').Replace('engil', 'yengil')
    $s = $s.Replace('Medisina', 'Meditsina').Replace('medisina', 'meditsina')
    return $s
  }}
  foreach ($ws in @($wb.Worksheets)) {{
    $used = $ws.UsedRange
    $textRange = $null
    try {{
      $textRange = $used.SpecialCells(2, 2)
    }} catch {{
      $textRange = $null
    }}
    if ($null -eq $textRange) {{ continue }}
    foreach ($pair in $UpperLatinPairs) {{
      try {{
        $textRange.Replace($pair[0], $pair[1], 2, 1, $false) | Out-Null
      }} catch {{
      }}
    }}
    foreach ($pair in $LatinPairs) {{
      try {{
        $textRange.Replace($pair[0], $pair[1], 2, 1, $false) | Out-Null
      }} catch {{
      }}
    }}
    foreach ($pair in @(@('Engil','Yengil'), @('engil','yengil'), @('Medisina','Meditsina'), @('medisina','meditsina'))) {{
      try {{
        $textRange.Replace($pair[0], $pair[1], 2, 1, $false) | Out-Null
      }} catch {{
      }}
    }}
  }}
  function Left-Text($sheet, $range, $fontSize = 12) {{
    try {{
      $rng = $wb.Worksheets.Item($sheet).Range($range)
      $rng.HorizontalAlignment = -4131
      $rng.VerticalAlignment = -4108
      $rng.Font.Name = 'Times New Roman'
      $rng.Font.Size = $fontSize
    }} catch {{
    }}
  }}
  Left-Text 'Корхоналар кесимида' 'B6:B5000'
  Left-Text 'Омборлар кесимида' 'B6:B5000'
  Left-Text 'Ўз омбор жами' 'B5:B5000'
  Left-Text 'Ўз омбор жами' 'H5:H5000'
  Left-Text 'Ўз омбор 3 ой+' 'B5:B5000'
  Left-Text 'Ўз омбор 3 ой+' 'H5:H5000'
  Left-Text 'muddati o`tgan ' 'B5:B5000'
  Left-Text 'muddati o`tgan ' 'H5:H5000'
  Left-Text 'Муддати ўтган 80' 'B5:B5000'
  Left-Text 'Муддати ўтган 80' 'H5:H5000'
  Left-Text 'Муддати ўтган 70-74' 'D5:E5000'
  Left-Text 'Озиқ овқат' 'C5:C5000'
  Left-Text 'Товар номи' 'B5:C5000'
  Left-Text 'БАЗА' 'F3:F50000'
  Left-Text 'БАЗА' 'I3:J50000'
  Left-Text 'Tahlil tovar guruhlari' 'B4:B5000'
  Left-Text 'Tahlil umumiy' 'C4:C5000'
  Left-Text 'Tahlil post-tovar' 'B4:C5000'
  Left-Text 'Tahlil ombor-tovar' 'B4:C5000'
  Left-Text 'Tahlil davlatlar' 'B4:B5000'
  function Sentence-Case-Text($value) {{
    if ($null -eq $value) {{ return $value }}
    $s = [string]$value
    if ($s.StartsWith('=')) {{ return $value }}
    if ($s.Length -eq 0) {{ return $s }}
    $lower = $s.ToLower()
    return $lower.Substring(0, 1).ToUpper() + $lower.Substring(1)
  }}
  function Sentence-Case-Range($sheet, $range) {{
    try {{
      $rng = $wb.Worksheets.Item($sheet).Range($range)
      foreach ($cell in @($rng.Cells)) {{
        try {{
          if (-not $cell.HasFormula -and $cell.Value2 -is [string]) {{
            $cell.Value2 = Sentence-Case-Text $cell.Value2
          }}
        }} catch {{
        }}
      }}
    }} catch {{
    }}
  }}
  Sentence-Case-Range 'Жами' 'A1:F7'
  Sentence-Case-Range 'Жами' 'A8:A24'
  Sentence-Case-Range '70-74-80' 'A1:E7'
  Sentence-Case-Range '70-74-80' 'A8:A32'
  Sentence-Case-Range 'Корхоналар кесимида' 'A1:H5'
  Sentence-Case-Range 'Омборлар кесимида' 'A1:F5'
  Sentence-Case-Range 'Ўз омбор жами' 'A1:I4'
  Sentence-Case-Range 'Ўз омбор 3 ой+' 'A1:I4'
  Sentence-Case-Range 'Жами муддати ўтган' 'A1:L7'
  Sentence-Case-Range 'muddati o`tgan ' 'A1:H5'
  Sentence-Case-Range 'Муддати ўтган 80' 'A1:H5'
  Sentence-Case-Range 'Муддати ўтган 70-74' 'A1:F7'
  Sentence-Case-Range 'Озиқ овқат' 'A1:H4'
  Sentence-Case-Range 'Товар номи' 'A1:H4'
  Sentence-Case-Range 'Tahlil tovar guruhlari' 'A1:H3'
  Sentence-Case-Range 'Tahlil umumiy' 'A1:H3'
  Sentence-Case-Range 'Tahlil post-tovar' 'A1:G3'
  Sentence-Case-Range 'Tahlil ombor-tovar' 'A1:G3'
  Sentence-Case-Range 'Tahlil davlatlar' 'A1:G3'
  Sentence-Case-Range 'Tahlil davlatlar' 'B4:B5000'
  Sentence-Case-Range 'Tahlil muddatlar' 'A1:I3'
  Sentence-Case-Range 'Жами йил' 'A1:I5'
  Sentence-Case-Range '70-74-80 склад Корхонабай пар' 'A1:G5'
  Sentence-Case-Range 'Сақланиш сабаблари' 'A1:F5'
  Sentence-Case-Range '3 ой' 'A1:H5'
  Sentence-Case-Range '6 ой' 'A1:H5'
  Sentence-Case-Range 'Сақланиш сабаблари йиллар кесим' 'A1:F4'
  Sentence-Case-Range 'ГТК' 'A1:G4'
  $usedNames = @{{}}
  foreach ($ws in @($wb.Worksheets)) {{
    $newName = Convert-ToLatin $ws.Name
    if ($newName.Length -gt 31) {{ $newName = $newName.Substring(0, 31) }}
    $baseName = $newName
    $i = 2
    while ($usedNames.ContainsKey($newName)) {{
      $suffix = " $i"
      $maxLen = 31 - $suffix.Length
      if ($baseName.Length -gt $maxLen) {{ $newName = $baseName.Substring(0, $maxLen) + $suffix }} else {{ $newName = $baseName + $suffix }}
      $i += 1
    }}
    $usedNames[$newName] = $true
    $ws.Name = $newName
  }}
  $wb.ForceFullCalculation = $true
  $wb.RefreshAll()
  $wb.Save()
  $wb.Close($true)
}} finally {{
  if ($null -ne $excel) {{
    try {{ $excel.Quit() }} catch {{}}
    try {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }} catch {{}}
  }}
}}
"""
    ps1 = output.with_suffix(".writer.ps1")
    ps1.write_text(script, encoding="utf-8-sig")
    subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-File", str(ps1)], check=True)


def build(source: Path, template: Path, output: Path, report_date: datetime, deposit: Path | None = None):
    df = read_source(source)
    deposits, deposit_date, deposit_total = read_deposit(deposit)
    months = {
        1: "yanvar", 2: "fevral", 3: "mart", 4: "aprel", 5: "may", 6: "iyun",
        7: "iyul", 8: "avgust", 9: "sentyabr", 10: "oktyabr", 11: "noyabr", 12: "dekabr",
    }
    date_label = f"{report_date:%Y} yil {report_date.day:02d} {months[report_date.month]} holatiga"
    short_date = f"{report_date:%d.%m.%Y} holatiga"
    if deposit_date:
        deposit_label = f"Depozit {deposit_date:%d.%m.%Y} holatiga: {fmt_uz_number(deposit_total)} mln so'm"
    else:
        deposit_label = "Depozit: 0 mln so'm"
    output.parent.mkdir(parents=True, exist_ok=True)
    # Use a persistent directory so JSON files survive PS1 retries
    tmp = output.parent / "artifacts_json"
    tmp.mkdir(parents=True, exist_ok=True)
    if True:
        files = {}
        data = {
            "base": base_rows(df),
            "jami": jami_rows(df)[0:3],
            "jami2": jami_rows(df)[3:6],
            "jami3": jami_rows(df)[6:9],
            "company": company_rows(df, deposits),
            "warehouse": warehouse_rows(df),
            "reason": reason_rows(df),
            "month3": month_rows(df, 3),
            "month6": month_rows(df, 6),
            "own_all": own_warehouse_rows(df, report_date, False),
            "own_3m": own_warehouse_rows(df, report_date, True),
            "sklad_company": company_regime_rows(df),
            "expired_summary": expired_summary_rows(df, report_date),
            "expired_detail": ([t] + expired_block_rows(df, report_date) if (t := expired_block_total_row(df, report_date)) else expired_block_rows(df, report_date)),
            "expired80": ([t80] + expired_block_rows(df, report_date, {"TR80"}) if (t80 := expired_block_total_row(df, report_date, {"TR80"})) else expired_block_rows(df, report_date, {"TR80"})),
            "expired7074": expired_7074_rows(df, report_date),
            "food": food_rows(df),
            "goods": goods_rows(df),
            "reason_year": reason_year_rows(df),
            "gtc": post_country_rows(df),
            "jami_year": jami_year_rows(df),
            "combined_jami": combined_jami_rows(df),
            "product_analysis": product_analysis_rows(df),
            "post_product": post_product_rows(df),
            "warehouse_product": warehouse_product_rows(df),
            "country": country_rows(df),
            "age_plus": age_plus_rows(df, report_date),
        }
        yr = by_year_rows(df)
        data.update({
            "year1": yr[0:1], "year2": yr[1:5], "year3": yr[5:7],
            "year4": yr[7:10], "year5": yr[10:11], "year6": yr[11:14], "year7": yr[14:15],
        })
        for name, rows in data.items():
            files[name] = tmp / f"{name}.json"
            write_json(files[name], rows)
        run_excel_writer(template, output, files, date_label, short_date, deposit_label, len(data["company"]))
        return len(df)


def main():
    parser = argparse.ArgumentParser(description="Nazoratdagi tovarlar faylidan IM70-74 workbookini Excel COM orqali to'ldirish.")
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--date", required=True, help="Masalan: 03.06.2026")
    parser.add_argument("--deposit", type=Path, help="Depozit xlsx fayli. Berilmasa depozit 0 bo'ladi.")
    args = parser.parse_args()
    rows = build(args.source, args.template, args.output, datetime.strptime(args.date, "%d.%m.%Y"), args.deposit)
    print(f"OK: {rows} qator asosida yaratildi: {args.output}")


if __name__ == "__main__":
    main()
