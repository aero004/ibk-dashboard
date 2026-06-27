from __future__ import annotations

import hashlib
import importlib.util
import json
import os
import re
import secrets
import shutil
import sys
import threading
import time
import traceback
from datetime import datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse
from urllib.request import Request, urlopen

import socket

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border


APP_DIR = Path(__file__).resolve().parent
DASHBOARD_DIR = Path(os.environ.get("IBK_GENERATOR_DIR", str(APP_DIR)))
sys.path.insert(0, str(DASHBOARD_DIR))

import im70_74_excel_com as core  # noqa: E402
from im70_74_analysis_report import build as build_png_report  # noqa: E402
from ibk_store import IBKStore  # noqa: E402


HOST = os.environ.get("IBK_HOST", "0.0.0.0")
PORT = int(os.environ.get("IBK_PORT", "8788"))

def _lan_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

LAN_IP = _lan_ip()
DATA_DIR = APP_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
REPORT_DIR = DATA_DIR / "reports"
CHUNK_DIR  = DATA_DIR / "chunks"
ASSET_DIR = APP_DIR / "assets"
USER_PATH = DATA_DIR / "users.json"
SESSION_PATH = DATA_DIR / "sessions.json"
INDEX_PATH = DATA_DIR / "archive.json"
DB_PATH = DATA_DIR / "ibk_dashboard.sqlite3"
TEMPLATE_PATH = Path(os.environ.get("IM70_74_TEMPLATE", DASHBOARD_DIR / "template_im70_74.xlsx"))
TOLOV_OUTPUT_DIR = Path(os.environ.get("IBK_TOLOV_OUTPUT_DIR", str(APP_DIR / "data" / "tolov_generated")))
UI_CONFIG_PATH = DATA_DIR / "ui_config.json"
TOLOV_GENERATOR_PATH = Path(os.environ.get("IBK_TOLOV_GENERATOR", str(APP_DIR / "tolov_generator.py")))
TOLOV_FILES = {
    "1. Sbor11-12.xlsx": "Sbor 11-12: yig'imlar",
    "2. Sbor 44.xlsx": "Sbor 44: TR80 yig'imi",
    "3. Sbor im40.xlsx": "Sbor 10: IM40/ND40",
    "4. Sbor ek10.xlsx": "Sbor 10: EK10",
    "5. Sbor dr.xlsx": "Sbor 10: boshqa rejimlar",
    "6. 29.xlsx": "29-kod to'lovi",
    "7. 20.xlsx": "20-kod to'lovi",
    "8. 27.xlsx": "27-kod to'lovi",
    "9. 25.xlsx": "25-kod to'lovi",
    "10. 30.xlsx": "30-kod to'lovi",
    "11. 74.xlsx": "74-kod to'lovi",
    "11. 79.xlsx": "79-kod to'lovi",
    "12. im42.xlsx": "IM42 bo'yicha ro'yxat",
}

JOBS: dict[str, dict] = {}
SESSIONS: dict[str, dict] = {}
STORE: IBKStore | None = None

# Login rate limiting: 5 urinishdan keyin 15 daqiqa blok
LOGIN_FAILS: dict[str, dict] = {}
LOGIN_BLOCK_AFTER = 5
LOGIN_BLOCK_SECS = 900

# CORS: faqat ibkinfo.uz dan cross-origin so'rovlarga ruxsat
ALLOWED_ORIGINS = {"https://ibkinfo.uz"}

TAS_ICAO = "UTTT"  # Toshkent (Islom Karimov) xalqaro aeroporti
FLIGHTS_CACHE: dict = {"ts": 0.0, "data": None}
FLIGHTS_CACHE_TTL = 60  # 1 daqiqa (FR24 schedule)

WAREHOUSE_REGISTRY_PATH: Path | None = None
WAREHOUSE_REGISTRY_CACHE: list[dict] | None = None

_WAREHOUSE_TYPE_MAP = {
    "очик турдаги": "ochiq",
    "очик": "ochiq",
    "yopiq": "yopiq",
    "ёпик турдаги": "yopiq",
    "ёпик": "yopiq",
    "бож олинмайдиган": "dutyfree",
    "savdo": "dutyfree",
    "эркин омбор": "erkin",
    "erkin": "erkin",
}


def _parse_date(val) -> str | None:
    if val is None or str(val).strip() in ("", "nan", "None"):
        return None
    try:
        return pd.to_datetime(str(val), dayfirst=True).strftime("%d.%m.%Y")
    except Exception:
        return str(val).strip()


def _days_until(date_str: str | None) -> int | None:
    if not date_str:
        return None
    try:
        d = datetime.strptime(date_str, "%d.%m.%Y").date()
        return (d - datetime.today().date()).days
    except Exception:
        return None


def _wr_type(raw: str) -> str:
    r = str(raw or "").lower().strip()
    for key, val in _WAREHOUSE_TYPE_MAP.items():
        if key in r:
            return val
    return "ochiq"


def _wr_file_date(path: Path) -> str:
    name = path.stem
    import re as _re
    m = _re.search(r"(\d{2}[.\-]\d{2}[.\-]\d{4})", name)
    return m.group(1).replace("-", ".") if m else datetime.today().strftime("%d.%m.%Y")


def load_warehouse_registry(path: Path | None = None) -> dict:
    global WAREHOUSE_REGISTRY_CACHE, WAREHOUSE_REGISTRY_PATH
    try:
        target = path or WAREHOUSE_REGISTRY_PATH
        if target is None:
            found = sorted(
                DASHBOARD_DIR.glob("omborlarReestri*.xlsx"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            if not found:
                return {"warehouses": [], "file_date": "", "loaded": False}
            target = found[0]
        WAREHOUSE_REGISTRY_PATH = target
        file_date = _wr_file_date(target)
        df = pd.read_excel(target, header=0)
        rows = []
        today = datetime.today().date()
        for _, row in df.iterrows():
            lat = row.iloc[15] if len(row) > 15 else None
            lon = row.iloc[16] if len(row) > 16 else None
            if pd.isna(lat) or pd.isna(lon):
                continue
            tur = _wr_type(str(row.iloc[7]) if len(row) > 7 else "")
            ins_sum = float(row.iloc[28]) if len(row) > 28 and not pd.isna(row.iloc[28]) else 0
            ins_exp = _parse_date(row.iloc[29] if len(row) > 29 else None)
            lic_exp = _parse_date(row.iloc[10] if len(row) > 10 else None)
            area_open = float(row.iloc[21]) if len(row) > 21 and not pd.isna(row.iloc[21]) else 0
            area_closed = float(row.iloc[22]) if len(row) > 22 and not pd.isna(row.iloc[22]) else 0
            fvv = str(row.iloc[30]).strip() if len(row) > 30 and not pd.isna(row.iloc[30]) else ""
            ssv = str(row.iloc[32]).strip() if len(row) > 32 and not pd.isna(row.iloc[32]) else ""
            ins_days = _days_until(ins_exp)
            lic_days = _days_until(lic_exp)
            risk = "green"
            if (ins_days is not None and ins_days < 0) or (lic_days is not None and lic_days < 0):
                risk = "red"
            elif (ins_days is not None and ins_days < 90) or (lic_days is not None and lic_days < 90):
                risk = "orange"
            director = str(row.iloc[4]).strip() if len(row) > 4 and not pd.isna(row.iloc[4]) else ""
            phone = str(row.iloc[5]).strip() if len(row) > 5 and not pd.isna(row.iloc[5]) else ""
            lic_num = str(row.iloc[8]).strip() if len(row) > 8 and not pd.isna(row.iloc[8]) else ""
            lic_date = _parse_date(row.iloc[9] if len(row) > 9 else None)
            rows.append({
                "name": str(row.iloc[1]).strip() if len(row) > 1 else "",
                "lat": float(lat),
                "lon": float(lon),
                "type": tur,
                "type_raw": str(row.iloc[7]).strip() if len(row) > 7 else "",
                "area_open": area_open,
                "area_closed": area_closed,
                "ins_sum": ins_sum,
                "ins_exp": ins_exp or "",
                "ins_days": ins_days,
                "lic_num": lic_num,
                "lic_date": lic_date or "",
                "lic_exp": lic_exp or "",
                "lic_days": lic_days,
                "director": director,
                "phone": phone,
                "fvv": fvv,
                "ssv": ssv,
                "risk": risk,
            })
        WAREHOUSE_REGISTRY_CACHE = rows
        return {"warehouses": rows, "file_date": file_date, "loaded": True}
    except Exception as exc:
        return {"warehouses": [], "file_date": "", "loaded": False, "error": str(exc)}


AVIA_AWB_PATH: Path | None = None
AVIA_AWB_CACHE: dict | None = None

YAROQLILIK_CACHE: dict | None = None

_AVIA_COUNTRY_LAT: dict[str, str] = {
    "НИДЕРЛАНДЫ": "Niderlandiya", "АКШ": "AQSH", "США": "AQSH",
    "СЯНГАН": "Gonkong", "ГОНКОНГ": "Gonkong", "ТУРКИЯ": "Turkiya",
    "ТУРЦИЯ": "Turkiya", "ХИТОЙ": "Xitoy", "КИТАЙ": "Xitoy",
    "ФРАНЦИЯ": "Frantsiya", "ЧЕХИЯ": "Chexiya", "ИЗРАИЛЬ": "Isroil",
    "БРАЗИЛИЯ": "Braziliya", "ИТАЛИЯ": "Italiya", "АВСТРИЯ": "Avstriya",
    "КАНАДА": "Kanada", "БОСНИЯ И ГЕРЦЕГОВИНА": "Bosniya va Gersegovina",
    "ИНДИЯ": "Hindiston", "ГЕРМАНИЯ": "Germaniya", "ВЬЕТНАМ": "Vyetnam",
    "ИСПАНИЯ": "Ispaniya", "ШВЕЦИЯ": "Shvetsiya", "БЕЛЬГИЯ": "Belgiya",
    "ПОКИСТОН": "Pokiston", "ПАКИСТАН": "Pokiston", "ЭКВАДОР": "Ekvador",
    "ВЕЛИКОБРИТАНИЯ": "Buyuk Britaniya", "РОССИЯ": "Rossiya",
    "ЯПОНИЯ": "Yaponiya", "АЗЕРБАЙДЖАН": "Ozarbayjon", "ПОЛЬША": "Polsha",
    "БИРЛАШГАН АРАБ АМИР.": "BAA", "ОАЭ": "BAA",
    "ОБЪЕДИНЕННЫЕ АРАБСКИЕ ЭМИРАТЫ": "BAA",
    "КОРЕЯ РЕСП.": "Janubiy Koreya", "РЕСПУБЛИКА КОРЕЯ": "Janubiy Koreya",
    "ТАЙВАНЬ": "Tayvan", "ЛАТВИЯ": "Latviya",
    "КИРГИЗИСТОН": "Qirgʻiziston", "КЫРГЫЗСТАН": "Qirgʻiziston",
    "КОЛУМБИЯ": "Kolumbiya", "ТОЖИКИСТОН": "Tojikiston",
    "ТАДЖИКИСТАН": "Tojikiston", "КОЗОГИСТОН": "Qozogʻiston",
    "КАЗАХСТАН": "Qozogʻiston", "ПОРТУГАЛИЯ": "Portugaliya",
    "СИНГАПУР": "Singapur", "БЕЛАРУСЬ": "Belarus",
    "ШВЕЙЦАРИЯ": "Shveytsariya", "САУДОВСКАЯ АРАВИЯ": "Saudiya Arabistoni",
    "УЗБЕКИСТОН": "Oʻzbekiston", "УЗБЕКИСТАН": "Oʻzbekiston",
    "КЕНИЯ": "Keniya", "ЛИТВА": "Litva", "НОВАЯ ЗЕЛАНДИЯ": "Yangi Zelandiya",
    "МАЛАЙЗИЯ": "Malayziya", "СЛОВАКИЯ РЕСПУБЛИКАСИ": "Slovakiya",
    "СЛОВАКИЯ": "Slovakiya", "ИРЛАНДИЯ": "Irlandiya", "ДАНИЯ": "Daniya",
    "ГРУЗИЯ": "Gruziya", "АВСТРАЛИЯ": "Avstraliya",
    "ФИНЛЯНДИЯ": "Finlandiya", "ВЕНГРИЯ": "Vengriya",
    "УКРАИНА": "Ukraina", "ЭСТОНИЯ": "Estoniya", "УГАНДА": "Uganda",
    "РУМЫНИЯ": "Ruminiya", "ИНДОНЕЗИЯ": "Indoneziya",
    "ТАИЛАНД": "Tailand", "КАТАР": "Qatar", "КУВЕЙТ": "Quvayt",
    "КОНГО": "Kongo", "АНГИЛЬЯ": "Angilya",
    "ИРАН": "Eron", "ИРОН": "Eron", "ЭРОН": "Eron",
}


def _norm_ru(s: str) -> str:
    return str(s).upper().strip().replace("H", "Н")


def _avia_country_latin(raw: str) -> str:
    normed = _norm_ru(raw)
    if normed in _AVIA_COUNTRY_LAT:
        return _AVIA_COUNTRY_LAT[normed]
    for key, val in _AVIA_COUNTRY_LAT.items():
        if key in normed:
            return val
    return raw.strip()


_LEGAL_FORM_RE = re.compile(
    r'\b(LLC|OOO|МЧЖ|MCHJ|MChJ|AJ|АЖ|AO|АО|LTD|CO|INC|CORP|GmbH|JSC|ЗАО|ОАО|ООО|ЧП|ИП|SP|SA|NV|PLC|SRL|BV|AG|PJSC|OJSC|XK|Xk)\b',
    re.IGNORECASE,
)
# Indicators that the rest of the string is an address / phone
_ADDR_SPLIT_RE = re.compile(
    r'\b(MFY|MFY\.|KUCHASI|KO[\'`]CHASI|STREET|STR\b|AVE\b|AVENUE|BLVD|PROSPEKT|MAHALLASI|'
    r'TASHKENT\s+TE|TE\s+00|00998|\+998)\b',
    re.IGNORECASE,
)

def _normalize_company(name: str) -> str:
    """Strip address/phone suffix; keep only the legal entity name."""
    name = str(name).strip()
    # Truncate at first address indicator
    addr_m = _ADDR_SPLIT_RE.search(name)
    if addr_m:
        name = name[:addr_m.start()].strip()
    matches = list(_LEGAL_FORM_RE.finditer(name))
    if not matches:
        words = name.split()
        return ' '.join(words[:5]) if len(words) > 5 else name
    last = matches[-1]
    # Legal form is at the START (Russian/CIS style: "LLC Company Name")
    if last.start() == 0:
        suffix = name[:last.end()]  # e.g. "LLC"
        rest = name[last.end():].strip()
        # Take up to 4 words of the rest as the actual name
        rest_words = rest.split()
        return (suffix + ' ' + ' '.join(rest_words[:4])).strip() if rest_words else suffix
    # Legal form is in the middle or end (English style: "Company Name LLC")
    return name[:last.end()].strip()


def load_avia_awb(path: Path | None = None) -> dict:
    global AVIA_AWB_PATH, AVIA_AWB_CACHE
    today = datetime.today().date()
    try:
        target = path or AVIA_AWB_PATH
        if target is None:
            found = sorted(
                DASHBOARD_DIR.glob("Yuklarni qabul qilish*.xlsx"),
                key=lambda p: p.stat().st_mtime, reverse=True,
            )
            if not found:
                return {"loaded": False, "error": "AWB Excel fayl topilmadi"}
            target = found[0]
        AVIA_AWB_PATH = target
        # Auto-detect header row: try header=0,1,2 and pick the one with AWB column
        df = None
        _awb_col = _arr_col = _country_col = _joylar_col = _vazn_col = _company_col = _goods_col = None
        _AWB_NAMES = ["awb raqami", "awb", "airwaybill", "air waybill", "хавойи юк хати"]
        _ARR_NAMES = ["kelish sanasi", "kelishi", "prибытие", "arrival", "sana"]
        _COUNTRY_NAMES = ["davlat", "mamlakat", "country", "страна", "jo'natuvchi davlat"]
        _JOYLAR_NAMES = ["joylar", "joyl", "places", "место", "qoldiqdagi joylar"]
        _VAZN_NAMES = ["vazn", "kg", "weight", "brutto", "gross"]
        _COMPANY_NAMES = ["korxona", "company", "yuklama egasi", "oluvchi", "консигнатор"]
        _GOODS_NAMES = ["tovar", "goods", "yuk nomi", "товар", "наименование"]

        def _find_col(cols, names):
            for n in names:
                for c in cols:
                    if n in str(c).lower():
                        return c
            return None

        for _hdr in [0, 1, 2]:
            _df = pd.read_excel(target, header=_hdr)
            _cols = [str(c).lower() for c in _df.columns]
            _awb_try = _find_col(_df.columns, _AWB_NAMES)
            if _awb_try is not None:
                df = _df
                _awb_col     = _awb_try
                _arr_col     = _find_col(_df.columns, _ARR_NAMES)
                _country_col = _find_col(_df.columns, _COUNTRY_NAMES)
                _joylar_col  = _find_col(_df.columns, _JOYLAR_NAMES)
                _vazn_col    = _find_col(_df.columns, _VAZN_NAMES)
                _company_col = _find_col(_df.columns, _COMPANY_NAMES)
                _goods_col   = _find_col(_df.columns, _GOODS_NAMES)
                break

        # Fallback: use positional columns (original logic)
        if df is None:
            df = pd.read_excel(target, header=1)
            _awb_col = df.columns[10] if len(df.columns) > 10 else None
            _arr_col = df.columns[6]  if len(df.columns) > 6  else None
            _country_col = df.columns[8]  if len(df.columns) > 8  else None
            _joylar_col  = df.columns[11] if len(df.columns) > 11 else None
            _vazn_col    = df.columns[12] if len(df.columns) > 12 else None
            _company_col = df.columns[14] if len(df.columns) > 14 else None
            _goods_col   = df.columns[18] if len(df.columns) > 18 else None

        def _cell(row, col):
            if col is None: return None
            v = row.get(col)
            return v if pd.notna(v) else None

        def _clean_awb(v):
            if v is None: return ""
            s = str(v).strip()
            # Remove trailing ".0" from numeric AWBs
            if s.endswith(".0") and s[:-2].isdigit():
                s = s[:-2]
            return s

        rows_raw = []
        skipped = 0
        for _, row in df.iterrows():
            try:
                awb = _clean_awb(_cell(row, _awb_col))
                if not awb or awb.lower() in ("nan", "none", "awb raqami", "awb", "-", ""):
                    skipped += 1
                    continue
                arr_raw = str(_cell(row, _arr_col) or "").strip()
                try:
                    arr_date = datetime.strptime(arr_raw, "%d.%m.%Y").date()
                except Exception:
                    arr_date = None
                joylar_v = _cell(row, _joylar_col)
                joylar = int(float(joylar_v)) if joylar_v is not None else 0
                vazn_v = _cell(row, _vazn_col)
                vazn_kg = float(vazn_v) if vazn_v is not None else 0.0
                country_v = _cell(row, _country_col)
                country_raw = str(country_v).strip() if country_v is not None else ""
                company_v = _cell(row, _company_col)
                company = _normalize_company(company_v) if company_v is not None else ""
                goods_v = _cell(row, _goods_col)
                goods = str(goods_v).strip() if goods_v is not None else ""
                if goods in ("nan", "None"):
                    goods = ""
                rows_raw.append({"awb": awb, "arr_date": arr_date, "joylar": joylar,
                                  "vazn_kg": vazn_kg, "country_raw": country_raw,
                                  "company": company, "goods": goods})
            except Exception:
                skipped += 1
                continue
        awb_groups: dict[str, list] = {}
        for r in rows_raw:
            awb_groups.setdefault(r["awb"], []).append(r)
        awb_list = []
        for awb, recs in awb_groups.items():
            total_joylar = sum(r["joylar"] for r in recs)
            total_vazn_kg = sum(r["vazn_kg"] for r in recs)
            countries = [r["country_raw"] for r in recs if r["country_raw"]]
            country_raw = max(set(countries), key=countries.count) if countries else ""
            company = next((r["company"] for r in recs if r["company"] and r["company"] != "nan"), "")
            goods = next((r["goods"] for r in recs if r["goods"]), "")
            dates = [r["arr_date"] for r in recs if r["arr_date"]]
            arr_date = max(dates) if dates else None
            days_ago = (today - arr_date).days if arr_date else None
            is_overdue = (days_ago is not None and days_ago > 15)
            awb_list.append({
                "awb": awb, "flights": len(recs), "company": company,
                "country": country_raw,
                "country_latin": _avia_country_latin(country_raw) if country_raw else "",
                "goods": goods, "joylar": total_joylar,
                "vazn": round(total_vazn_kg / 1000, 3),
                "arrival_date": arr_date.strftime("%d.%m.%Y") if arr_date else "",
                "days_ago": days_ago, "is_overdue": is_overdue,
            })
        awb_list.sort(key=lambda x: x.get("arrival_date", ""), reverse=True)
        total_vazn = sum(r["vazn"] for r in awb_list)
        total_joylar_all = sum(r["joylar"] for r in awb_list)
        active = [r for r in awb_list if not r["is_overdue"]]
        overdue = [r for r in awb_list if r["is_overdue"]]
        country_stats: dict[str, dict] = {}
        for r in awb_list:
            c = r["country_latin"] or r["country"]
            if not c:
                continue
            if c not in country_stats:
                country_stats[c] = {"name": c, "awb": 0, "vazn": 0.0, "joylar": 0}
            country_stats[c]["awb"] += 1
            country_stats[c]["vazn"] = round(country_stats[c]["vazn"] + r["vazn"], 3)
            country_stats[c]["joylar"] += r["joylar"]
        countries_list = sorted(country_stats.values(), key=lambda x: x["vazn"], reverse=True)
        company_stats: dict[str, dict] = {}
        for r in awb_list:
            c = r["company"]
            if not c or c == "nan":
                continue
            if c not in company_stats:
                company_stats[c] = {"company": c, "awb": 0, "vazn": 0.0, "joylar": 0, "overdue": 0}
            company_stats[c]["awb"] += 1
            company_stats[c]["vazn"] = round(company_stats[c]["vazn"] + r["vazn"], 3)
            company_stats[c]["joylar"] += r["joylar"]
            if r["is_overdue"]:
                company_stats[c]["overdue"] += 1
        companies_list = sorted(company_stats.values(), key=lambda x: x["vazn"], reverse=True)
        result = {
            "loaded": True, "file_name": target.name,
            "total_rows": len(rows_raw), "skipped_rows": skipped,
            "unique_awb": len(awb_list),
            "multi_flight_awb": sum(1 for r in awb_list if r["flights"] > 1),
            "total_joylar": total_joylar_all,
            "total_vazn": round(total_vazn, 3),
            "active_count": len(active),
            "active_vazn": round(sum(r["vazn"] for r in active), 3),
            "overdue_count": len(overdue),
            "overdue_vazn": round(sum(r["vazn"] for r in overdue), 3),
            "awb_list": awb_list,
            "companies": companies_list[:50],
            "countries": countries_list,
        }
        AVIA_AWB_CACHE = result
        return result
    except Exception as exc:
        import traceback as _tb
        return {"loaded": False, "error": f"{type(exc).__name__}: {exc}"}


# ── Yaroqlilik muddati (expired / expiring goods) ─────────────────────────────

_YAR_COL_MAP: dict[str, str] = {
    "Holat": "holat",
    "Deklaratsiya raqami": "decl_raqami",
    "Rejim": "rejim",
    "Rasmiylashtirilgan sana": "reg_sana",
    "STIR": "stir",
    "Yuk qabul qiluvchi": "korxona",
    "Tovar raqami": "tovar_raqami",
    "TIF TN kodi": "hs_kod",
    "Tovar nomi": "tovar_nomi",
    "Yaroqlilik muddati": "yaroqlilik",
    "Qolgan kun": "qolgan_kun",
    "O’tgan kun": "otgan_kun",
    "Topilgan sana matni": "topilgan_sana",
    "Izoh": "izoh",
    "Excel qatori": "excel_qatori",
    "Vazni, kg": "vazn",
    "Qiymati, ming doll.": "qiymat",
    "Vazni kg": "vazn",
    "Qiymati ming doll.": "qiymat",
}

_YAR_SKIP_SHEETS = {"xulosa", "summary", "natija"}


def _yar_detect_holat(sname: str) -> str:
    sl = sname.lower()
    if "1_oy" in sl or "1oy" in sl or "bir_oy" in sl:
        return "1 oy ichida muddati tugaydi"
    if "180" in sl:
        return "180 kundan kam qolgan"
    return "Muddati o'tgan"


def _yar_norm(s: str) -> str:
    """Normalize curly/typographic apostrophes to straight apostrophe."""
    return str(s).replace("‘", "'").replace("’", "'").replace("ʼ", "'")


def parse_yaroqlilik_excel(file_bytes: bytes) -> dict:
    """Parse yaroqlilik muddat Excel file, auto-detect sheets."""
    import io as _io
    buf = _io.BytesIO(file_bytes)
    report_date = datetime.today().strftime("%Y-%m-%d")

    try:
        all_sheets = pd.read_excel(buf, sheet_name=None, header=None)
    except Exception as exc:
        return {"items": [], "loaded": True, "error": str(exc), "report_date": report_date}

    all_items: list[dict] = []

    for sname, df_raw in all_sheets.items():
        if sname.lower() in _YAR_SKIP_SHEETS:
            continue

        hdr_row: int | None = None
        for i in range(min(10, len(df_raw))):
            row_vals = [str(v) for v in df_raw.iloc[i].tolist()]
            if any("Holat" in v or "Deklaratsiya" in v for v in row_vals):
                hdr_row = i
                break
        if hdr_row is None:
            continue

        buf.seek(0)
        try:
            df = pd.read_excel(buf, sheet_name=sname, header=hdr_row)
        except Exception:
            continue

        default_holat = _yar_detect_holat(sname)

        for _, row in df.iterrows():
            item: dict = {}
            for col in df.columns:
                mapped = _YAR_COL_MAP.get(_yar_norm(str(col)), "")
                if not mapped:
                    continue
                val = row[col]
                if pd.isna(val) or val is None:
                    val = None
                elif isinstance(val, pd.Timestamp):
                    val = val.strftime("%Y-%m-%d")
                elif isinstance(val, float):
                    val = int(val) if val == int(val) and abs(val) < 1e15 else round(val, 4)
                elif isinstance(val, str):
                    val = _yar_norm(val)
                item[mapped] = val

            if not item.get("decl_raqami"):
                continue
            if not item.get("holat"):
                item["holat"] = default_holat
            all_items.append(item)

    # Deduplicate by (decl + tovar_no + holat)
    seen: set[tuple] = set()
    unique: list[dict] = []
    for item in all_items:
        key = (item.get("decl_raqami", ""), item.get("tovar_raqami", ""), item.get("holat", ""))
        if key not in seen:
            seen.add(key)
            unique.append(item)

    expired_count = sum(1 for i in unique if i.get("holat") == "Muddati o'tgan")
    warn180_count = sum(1 for i in unique if i.get("holat") == "180 kundan kam qolgan")
    warn30_count = sum(1 for i in unique if i.get("holat") == "1 oy ichida muddati tugaydi")
    total_qiymat = sum(float(i.get("qiymat") or 0) for i in unique)
    total_vazn = sum(float(i.get("vazn") or 0) for i in unique)

    return {
        "items": unique,
        "loaded": True,
        "report_date": report_date,
        "count": len(unique),
        "expired_count": expired_count,
        "warn180_count": warn180_count,
        "warn30_count": warn30_count,
        "total_qiymat": round(total_qiymat, 3),
        "total_vazn": round(total_vazn, 3),
    }


def load_yaroqlilik_cached() -> dict:
    global YAROQLILIK_CACHE
    if YAROQLILIK_CACHE is not None:
        return YAROQLILIK_CACHE
    yar_path = DATA_DIR / "yaroqlilik.json"
    if yar_path.exists():
        try:
            YAROQLILIK_CACHE = json.loads(yar_path.read_text(encoding="utf-8"))
            return YAROQLILIK_CACHE
        except Exception:
            pass
    return {"items": [], "loaded": False, "report_date": ""}


def save_yaroqlilik(data: dict) -> None:
    global YAROQLILIK_CACHE
    YAROQLILIK_CACHE = data
    yar_path = DATA_DIR / "yaroqlilik.json"
    yar_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_cbu_rates() -> dict:
    try:
        from urllib.request import Request as _Req, urlopen as _urlopen
        req = _Req("https://cbu.uz/common/json/",
                   headers={"User-Agent": "IBK-Dashboard/1.0"})
        with _urlopen(req, timeout=6) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        rates = {}
        for item in data:
            ccy = item.get("Ccy", "")
            if ccy:
                rates[ccy] = {
                    "rate": float(item.get("Rate", 0)),
                    "nominal": int(item.get("Nominal", 1)),
                    "name": item.get("CcyNm_UZ") or item.get("CcyNm_EN") or ccy,
                    "diff": float(item.get("Diff", 0)),
                }
        date_str = data[0].get("Date", "") if data else ""
        return {"success": True, "rates": rates, "date": date_str}
    except Exception as exc:
        return {"success": False, "error": str(exc), "rates": {}, "date": ""}


_FR24_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
}


def _fr24_parse_flight(f: dict, typ: str) -> dict:
    fl = f.get("flight", {})
    ident = fl.get("identification", {})
    status = fl.get("status", {})
    generic = status.get("generic", {})
    airport = fl.get("airport", {})
    aircraft = fl.get("aircraft", {})
    times = fl.get("time", {})
    src = airport.get("origin" if typ == "arr" else "destination", {}) or {}
    pos = (src.get("position") or {})
    info = src.get("info") or {}
    event_ts = (generic.get("eventTime") or {}).get("utc")
    return {
        "callsign": ident.get("callsign", ""),
        "flight": (ident.get("number") or {}).get("default", ""),
        "from": (src.get("code") or {}).get("iata", "") if typ == "arr" else "TAS",
        "to": "TAS" if typ == "arr" else (src.get("code") or {}).get("iata", ""),
        "airport_name": src.get("name", ""),
        "city": (pos.get("region") or {}).get("city", ""),
        "country": (pos.get("country") or {}).get("name", ""),
        "lat": pos.get("latitude"),
        "lon": pos.get("longitude"),
        "first_seen": event_ts if typ == "dep" else None,
        "last_seen": event_ts if typ == "arr" else None,
        "status_text": status.get("text", ""),
        "status_type": (generic.get("status") or {}).get("text", ""),
        "status_color": (generic.get("status") or {}).get("color", "gray"),
        "live": bool(status.get("live", False)),
        "aircraft": (aircraft.get("model") or {}).get("text", ""),
        "reg": aircraft.get("registration", ""),
        "terminal": info.get("terminal", ""),
        "gate": info.get("gate", ""),
        "dep_ts": (times.get("real") or {}).get("departure") or (times.get("scheduled") or {}).get("departure"),
        "arr_ts": (times.get("estimated") or {}).get("arrival") or (times.get("other") or {}).get("eta") or (times.get("scheduled") or {}).get("arrival"),
    }


def fetch_tas_flights() -> dict:
    url = "https://api.flightradar24.com/common/v1/airport.json?code=TAS&plugin[]=schedule&limit=50"
    now = int(time.time())
    result = {"updated": now, "arrivals": [], "departures": [], "error": ""}
    try:
        req = Request(url, headers=_FR24_HEADERS)
        with urlopen(req, timeout=12) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        sched = raw["result"]["response"]["airport"]["pluginData"]["schedule"]
        result["arrivals"] = [_fr24_parse_flight(f, "arr") for f in (sched.get("arrivals") or {}).get("data", [])]
        result["departures"] = [_fr24_parse_flight(f, "dep") for f in (sched.get("departures") or {}).get("data", [])]
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
    return result


def get_tas_flights() -> dict:
    now = time.time()
    if FLIGHTS_CACHE["data"] is None or now - FLIGHTS_CACHE["ts"] > FLIGHTS_CACHE_TTL:
        FLIGHTS_CACHE["data"] = fetch_tas_flights()
        FLIGHTS_CACHE["ts"] = now
    return FLIGHTS_CACHE["data"]


# Toshkent atrofidagi hudud (Markaziy Osiyo) - jonli parvozlar xaritasi uchun
TAS_BBOX = {"lamin": 34.0, "lomin": 56.0, "lamax": 46.0, "lomax": 78.0}
LIVE_FLIGHTS_CACHE: dict = {"ts": 0.0, "data": None}
LIVE_FLIGHTS_CACHE_TTL = 25  # OpenSky anonim limitlariga ehtiyot


def fetch_live_states() -> dict:
    b = TAS_BBOX
    url = (
        "https://opensky-network.org/api/states/all"
        f"?lamin={b['lamin']}&lomin={b['lomin']}&lamax={b['lamax']}&lomax={b['lomax']}"
    )
    result = {"updated": int(time.time()), "planes": [], "error": ""}
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0 (IBK-Dashboard)"})
        with urlopen(req, timeout=10) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        for s in raw.get("states") or []:
            lon, lat = s[5], s[6]
            if lon is None or lat is None:
                continue
            result["planes"].append({
                "icao24": s[0],
                "callsign": (s[1] or "").strip(),
                "country": s[2],
                "lon": lon,
                "lat": lat,
                "alt": s[7] or s[13] or 0,
                "on_ground": bool(s[8]),
                "speed": s[9] or 0,
                "heading": s[10] or 0,
            })
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
    return result


def get_live_states() -> dict:
    now = time.time()
    if LIVE_FLIGHTS_CACHE["data"] is None or now - LIVE_FLIGHTS_CACHE["ts"] > LIVE_FLIGHTS_CACHE_TTL:
        LIVE_FLIGHTS_CACHE["data"] = fetch_live_states()
        LIVE_FLIGHTS_CACHE["ts"] = now
    return LIVE_FLIGHTS_CACHE["data"]


def ensure_job(job_id: str) -> dict:
    """JOBS va JOBS[job_id] hech qachon None bo'lib qolmasligi uchun himoya."""
    global JOBS
    if not isinstance(JOBS, dict):
        JOBS = {}
    if not job_id:
        job_id = "unknown_" + str(int(time.time() * 1000))
    job = JOBS.get(job_id)
    if not isinstance(job, dict):
        job = {"status": "navbatda", "data": {}}
    if not isinstance(job.get("data"), dict):
        job["data"] = {}
    JOBS[job_id] = job
    return job


def safe_files_dict(files) -> dict:
    """files har doim dict bo'lishini kafolatlaydi."""
    if isinstance(files, dict):
        return files
    return {"status": "tayyorlanmoqda", "excel": "", "pdf": "", "pngs": []}

DEFAULT_PERMS = ["view", "upload", "export", "release"]
ADMIN_PERMS = ["view", "upload", "export", "release", "admin", "settings"]
ROLES = {"admin", "rahbar", "inspektor", "foydalanuvchi", "user"}
ROLE_LABELS = {
    "admin": "Admin",
    "rahbar": "Rahbar",
    "inspektor": "Inspektor",
    "foydalanuvchi": "Foydalanuvchi",
    "user": "Foydalanuvchi",
}


def ensure_dirs():
    for path in [DATA_DIR, UPLOAD_DIR, REPORT_DIR, ASSET_DIR, TOLOV_OUTPUT_DIR]:
        path.mkdir(parents=True, exist_ok=True)
    if not USER_PATH.exists():
        save_json(USER_PATH, {
            "admin": {
                "salt": "ibk",
                "password": hash_password("admin123", "ibk"),
                "role": "admin",
                "enabled": True,
                "full_name": "Administrator",
                "position": "Admin",
                "phone": "",
                "lang": "uz",
                "post_code": "",
                "role_label": "Admin",
                "perms": ADMIN_PERMS,
            }
        })
    else:
        users = load_json(USER_PATH, {})
        changed = False
        for rec in users.values():
            if "enabled" not in rec:
                rec["enabled"] = True
                changed = True
            if "perms" not in rec:
                rec["perms"] = ADMIN_PERMS if rec.get("role") == "admin" else DEFAULT_PERMS
                changed = True
            for field, default in [("full_name", ""), ("position", ""), ("phone", ""), ("lang", "uz"), ("post_code", "")]:
                if field not in rec:
                    rec[field] = default
                    changed = True
            label = ROLE_LABELS.get(rec.get("role", "user"), "Foydalanuvchi")
            if rec.get("role_label") != label:
                rec["role_label"] = label
                changed = True
        if changed:
            save_json(USER_PATH, users)
    if not SESSION_PATH.exists():
        save_json(SESSION_PATH, {})
    if not INDEX_PATH.exists():
        save_json(INDEX_PATH, {"reports": []})


def load_json(path: Path, default):
    if not path.exists():
        return default
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default
    return default if data is None else data


def save_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def hash_password(password: str, salt: str) -> str:
    return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()


def safe_name(name: str) -> str:
    return re.sub(r"[^\w .,'`()+-]+", "_", name, flags=re.UNICODE).strip() or "file"


def parse_multipart(content_type: str, body: bytes) -> dict[str, dict]:
    match = re.search(r"boundary=(?P<b>[^;]+)", content_type or "")
    if not match:
        return {}
    boundary = match.group("b").strip().strip('"').encode("utf-8")
    result: dict[str, dict] = {}
    marker = b"--" + boundary
    for part in body.split(marker):
        part = part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        if part.endswith(b"--"):
            part = part[:-2].rstrip(b"\r\n")
        head, sep, content = part.partition(b"\r\n\r\n")
        if not sep:
            continue
        disposition = ""
        for line in head.decode("utf-8", errors="replace").split("\r\n"):
            if line.lower().startswith("content-disposition:"):
                disposition = line
                break
        name_match = re.search(r'name="([^"]+)"', disposition)
        if not name_match:
            continue
        filename_match = re.search(r'filename="([^"]*)"', disposition)
        field_name = name_match.group(1)
        item = {
            "filename": filename_match.group(1) if filename_match else "",
            "content": content.rstrip(b"\r\n"),
        }
        if field_name in result:
            if isinstance(result[field_name], list):
                result[field_name].append(item)
            else:
                result[field_name] = [result[field_name], item]
        else:
            result[field_name] = item
    return result


def report_date_from_name(name: str) -> datetime:
    return core.report_date_from_name(name) or datetime.now()


def fmt_date(dt: datetime) -> str:
    return dt.strftime("%d.%m.%Y")


def duplicate_report(filename: str, report_date: datetime) -> dict | None:
    needle = safe_name(filename)
    for item in load_json(INDEX_PATH, {"reports": []}).get("reports", []):
        if item.get("date") != fmt_date(report_date):
            continue
        source_name = Path(item.get("source", "")).name
        if source_name == needle or source_name == filename:
            return item
    return None


def clean_archive_records(reports: list[dict]) -> list[dict]:
    """Arxivda bir sana bitta qator bo'lib qolsin; eng foydali yozuv saqlanadi."""
    def score(rec: dict):
        source = str(rec.get("source", ""))
        deposit = str(rec.get("deposit", ""))
        return (
            1 if "IBK_Dashboard_SQLite" in source else 0,
            1 if deposit else 0,
            str(rec.get("id", "")),
        )
    best: dict[str, dict] = {}
    for rec in reports or []:
        if not isinstance(rec, dict):
            continue
        date = rec.get("date", "")
        if not date:
            continue
        if date not in best or score(rec) > score(best[date]):
            best[date] = rec
    def date_key(rec: dict):
        try:
            return datetime.strptime(rec.get("date", ""), "%d.%m.%Y")
        except Exception:
            return datetime.min
    return sorted(best.values(), key=date_key, reverse=True)


def digits(value) -> str:
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return re.sub(r"\D", "", str(value or ""))


def source_post_code(value) -> str:
    match = re.search(r"\d{5}", str(value or ""))
    return match.group(0) if match else ""


POST_CLASSIFIER = {
    # Qoraqalpog'iston Respublikasi
    "35001": "Nukus aeroporti chegara bojxona posti",
    "35002": "Nukus TIF bojxona posti",
    "35003": "Xo'jayli chegara bojxona posti",
    "35004": "Dovut-ota chegara bojxona posti",
    "35010": "Qoraqalpog'iston temir yo'l chegara bojxona posti",
    # Andijon viloyati
    "03002": "Do'stlik chegara bojxona posti",
    "03003": "Andijon aeroporti chegara bojxona posti",
    "03005": "Mingtepa chegara bojxona posti",
    "03006": "Qorasuv chegara bojxona posti",
    "03007": "Xonobod chegara bojxona posti",
    "03008": "Pushmon chegara bojxona posti",
    "03009": "Madaniyat chegara bojxona posti",
    "03011": "Andijon TIF bojxona posti",
    "03013": "Keskanyoʻr chegara bojxona posti",
    "03014": "Savay temir yo'l chegara bojxona posti",
    "03015": "Asaka TIF bojxona posti",
    # Buxoro viloyati
    "06001": "Buxoro aeroporti chegara bojxona posti",
    "06006": "Buxoro TIF bojxona posti",
    "06009": "Qorakoʻl TIF bojxona posti",
    "06010": "Olot chegara bojxona posti",
    "06011": "Xo'jadavlat temir yo'l chegara bojxona posti",
    # Jizzax viloyati
    "08003": "Uchturgon chegara bojxona posti",
    "08004": "Jizzax TIF bojxona posti",
    "08007": "Qo'shkent chegara bojxona posti",
    # Qashqadaryo viloyati
    "10002": "Nasaf TIF bojxona posti",
    "10007": "Qamashi-G'uzor TIF bojxona posti",
    "10008": "Qarshi-Kerki chegara bojxona posti",
    "10012": "Qarshi aeroporti chegara bojxona posti",
    # Navoiy viloyati
    "12002": "Navoiy aeroporti chegara bojxona posti",
    "12003": "Navoiy TIF bojxona posti",
    "12008": "Zarafshon TIF bojxona posti",
    # Namangan viloyati
    "14002": "Namangan aeroporti chegara bojxona posti",
    "14003": "Uchqo'rg'on chegara bojxona posti",
    "14004": "Kosonsoy chegara bojxona posti",
    "14005": "Pop chegara bojxona posti",
    "14010": "Namangan TIF bojxona posti",
    # Samarqand viloyati
    "18001": "Samarqand aeroporti chegara bojxona posti",
    "18002": "Jartepa chegara bojxona posti",
    "18005": "Samarqand TIF bojxona posti",
    "18007": "Ulug'bek TIF bojxona posti",
    # Surxondaryo viloyati
    "22002": "Termiz aeroporti chegara bojxona posti",
    "22003": "Sariosiyo chegara bojxona posti",
    "22004": "Sariosiyo temir yo'l chegara bojxona posti",
    "22005": "Termiz TIF bojxona posti",
    "22006": "Denov TIF bojxona posti",
    "22007": "Gulbahor chegara bojxona posti",
    "22011": "Daryo porti chegara bojxona posti",
    "22015": "Boldir temir yo'l chegara bojxona posti",
    "22017": "Ayritom chegara bojxona posti",
    "22022": "Termiz xalqaro savdo markazi TIF bojxona posti",
    # Sirdaryo viloyati
    "24002": "Xovosоbod chegara bojxona posti",
    "24004": "Sirdaryo chegara bojxona posti",
    "24006": "Oq oltin chegara bojxona posti",
    "24009": "Guliston TIF bojxona posti",
    "24014": "Malik chegara bojxona posti",
    # Toshkent viloyati
    "27001": "Yallama chegara bojxona posti",
    "27008": "Navoiy chegara bojxona posti",
    "27009": "S. Najimov chegara bojxona posti",
    "27011": "Oybek chegara bojxona posti",
    "27013": "Bekobod avto chegara bojxona posti",
    "27014": "Chirchiq TIF bojxona posti",
    "27015": "Olmaliq TIF bojxona posti",
    "27016": "Yangiyoʻl TIF bojxona posti",
    "27019": "Nazarbek TIF bojxona posti",
    "27020": "Keles TIF bojxona posti",
    "27021": "G'ishtkoprik chegara bojxona posti",
    "27023": "Farhod chegara bojxona posti",
    "27024": "Bekobod temir yo'l chegara bojxona posti",
    "27028": "Angren TIF bojxona posti",
    # Farg'ona viloyati
    "30001": "Farg'ona aeroporti chegara bojxona posti",
    "30002": "Qo'qon TIF bojxona posti",
    "30004": "Farg'ona chegara bojxona posti",
    "30005": "Andarxon chegara bojxona posti",
    "30006": "Rishton chegara bojxona posti",
    "30008": "Rovot chegara bojxona posti",
    "30009": "Vodiy TIF bojxona posti",
    "30010": "O'zbekiston chegara bojxona posti",
    "30012": "So'x chegara bojxona posti",
    # Xorazm viloyati
    "33001": "Shovot chegara bojxona posti",
    "33004": "Do'stlik chegara bojxona posti",
    "33007": "Urganch TIF bojxona posti",
    "33011": "Urganch aeroporti chegara bojxona posti",
    "33033": "Shovot chegaraoldi savdo zonasi chegara bojxona posti",
    # Toshkent shahri
    "26002": "Toshkent-tovar TIF bojxona posti",
    "26003": "Ark buloq TIF bojxona posti",
    "26004": "Chuqursoy TIF bojxona posti",
    "26009": "Keles temir yo'l chegara bojxona posti",
    "26010": "Sirg'ali TIF bojxona posti",
    "26013": "Chuqursoy texnik idora temir yo'l chegara bojxona posti",
    # Toshkent-AERO
    "00101": "Toshkent xalqaro aeroporti CHBP",
    "00102": "Avia yuklar TIF bojxona posti",
    "00107": "Elektron tijorat TIF bojxona posti",
    "00110": "Toshkent-Humo aeroporti CHBP",
}


def source_post_name(code: str) -> str:
    code = core.clean(code)
    return POST_CLASSIFIER.get(code, f"Post №{code}" if code else "-")


def source_transport(code: str, post_name: str = "") -> str:
    text = f"{code} {post_name}".lower()
    if code.startswith("001") or any(w in text for w in ["avia", "aero", "aeroport", "airport"]):
        return "Avia"
    if any(w in text for w in ["temir", "rail", "poezd", "vokzal", "yol", "yo'l"]):
        return "Temir yo'l"
    return "Avto"


def declaration_key(row) -> str:
    return core.clean(row[core.SRC["decl_no"]])


def read_report_source(source: Path) -> pd.DataFrame:
    df = core.read_source(source)
    df["_decl_key"] = df.apply(declaration_key, axis=1)
    df["_source_post_code"] = df[core.SRC["decl_no"]].map(source_post_code)
    df["_source_post_name"] = df["_source_post_code"].map(source_post_name)
    df["_source_transport"] = df.apply(lambda r: source_transport(r.get("_source_post_code", ""), r.get("_source_post_name", "")), axis=1)
    return df


def item_key(row) -> str:
    goods_no = core.clean(row[core.SRC.get("goods_no", core.SRC["decl_no"])])
    return f"{declaration_key(row)}::{goods_no}"


def snapshot_rows(df: pd.DataFrame) -> list[dict]:
    rows = []
    for _, r in df.iterrows():
        dt = r.get("_date", None)
        if pd.notna(dt):
            gtd_date = pd.Timestamp(dt).strftime("%Y-%m-%d")
        else:
            gtd_date = core.clean(r[core.SRC["decl_date"]])
        rows.append({
            "item_key": item_key(r),
            "decl": core.clean(r[core.SRC["decl_no"]]),
            "item_no": core.clean(r[core.SRC["goods_no"]]),
            "regime": core.clean(r.get("_regime_code", r[core.SRC["regime"]])),
            "post": core.to_latin(core.clean(r.get("_post_group", r[core.SRC["post"]]))),
            "source_post": core.clean(r.get("_source_post_code", "")),
            "source_post_name": core.clean(r.get("_source_post_name", "")),
            "transport": core.clean(r.get("_source_transport", "")),
            "stir": digits(r[core.SRC["stir"]]),
            "company": core.clean(r[core.SRC["company"]]),
            "gtd_date": gtd_date,
            "hs_code": core.clean(r[core.SRC["hs"]]),
            "goods": core.clean(r[core.SRC["goods"]]),
            "weight": float(r.get("_weight_tn", 0) or 0),
            "value": float(r.get("_value_usd_k", 0) or 0),
            "payment": float(r.get("_pay_total_mln", 0) or 0),
            "partiya": int(r.get("_partiya", 1) or 0),
            "country": core.clean(r[core.SRC["country"]]),
            "warehouse": core.clean(r[core.SRC["warehouse"]]),
            "reason": core.clean_reason_display(r[core.SRC["reason"]]),
        })
    return rows


def save_snapshot_to_store(report_id: str, source: Path, deposit: Path | None, report_date: datetime, df: pd.DataFrame | None = None):
    global STORE
    if STORE is None:
        STORE = IBKStore(DB_PATH)
    df = df if df is not None else read_report_source(source)
    snapshot_id = STORE.register_snapshot(report_id, source.name, str(source), str(deposit or ""), fmt_date(report_date))
    STORE.insert_active_items(snapshot_id, snapshot_rows(df))
    return snapshot_id


STORE_BACKFILLED = False


def ensure_store_backfilled():
    """Arxivdagi (2024 yildan boshlab) avval yuklangan, lekin SQLite snapshot
    bazasiga hali tushmagan hisobotlarni bir martalik tarzda bazaga qo'shadi.
    Shu orqali "Aylanma tendensiyasi" kabi tahlillar 2024 yildan boshlab
    mavjud bo'lgan barcha davrlarni qamrab oladi."""
    global STORE, STORE_BACKFILLED
    if STORE_BACKFILLED:
        return
    if STORE is None:
        STORE = IBKStore(DB_PATH)
    try:
        existing = {s.get("report_id") for s in STORE.list_snapshots()}
    except Exception:
        existing = set()
    archive = load_json(INDEX_PATH, {"reports": []})
    reports = archive.get("reports", []) if isinstance(archive, dict) else []
    for rec in reports:
        if not isinstance(rec, dict):
            continue
        rid = rec.get("id")
        if not rid or rid in existing:
            continue
        source = rec.get("source")
        if not source:
            continue
        src_path = Path(source)
        if not src_path.exists():
            continue
        try:
            report_date = report_date_from_name(src_path.name)
            df = read_report_source(src_path)
            deposit_raw = rec.get("deposit") or ""
            deposit = Path(deposit_raw) if deposit_raw else None
            save_snapshot_to_store(rid, src_path, deposit, report_date, df)
            existing.add(rid)
        except Exception:
            continue
    STORE_BACKFILLED = True


def is_own_warehouse_name(value: str) -> bool:
    text = core.clean(value)
    low = text.lower().replace("?", "'").replace("`", "'")
    if not text:
        return True
    address_words = [
        "toshkent shahar", "toshkent shahri", "tashkent", "toshkent viloyati",
        " r-n", "r-n ", "tumani", "tuman", "ko'chasi", "kochasi", "ko`chasi",
        "mfy", "mahalla", "-uy", " uy", "dom", "??????", "??????", "??????",
    ]
    return any(w in low for w in address_words)


def own_warehouse_label(value: str) -> str:
    return "O'z ombor" if is_own_warehouse_name(value) else core.clean(value)


def own_company_rows_web(df: pd.DataFrame, report_date: datetime, over_3_months: bool = False) -> list[dict]:
    own = df[df[core.SRC["warehouse"]].map(is_own_warehouse_name)].copy()
    if over_3_months:
        cutoff = pd.Timestamp(report_date) - pd.DateOffset(months=3)
        own = own[own["_date"].le(cutoff)].copy()
    if own.empty:
        return []
    g = own.groupby(core.SRC["stir"], dropna=False).agg(
        company=(core.SRC["company"], "first"),
        partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum"),
        first_date=("_date", "min"), goods=(core.SRC["goods"], core.goods_category),
    ).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    rows = []
    for r in g.itertuples(index=False):
        kun_hisobi = int((pd.Timestamp(report_date) - pd.Timestamp(r.first_date)).days) if pd.notna(r.first_date) else 0
        rows.append({"key": {"stir": digits(r[0])}, "korxona": core.clean(r.company), "stir": digits(r[0]), "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov), "muddat": r.first_date.strftime("%d.%m.%Y") if pd.notna(r.first_date) else "", "kun_hisobi": kun_hisobi, "tovar": core.to_latin(core.clean(r.goods))})
    return rows

def detail_rows(df: pd.DataFrame, filters: dict) -> list[dict]:
    d = df.copy()
    if filters.get("decl"):
        d = d[d[core.SRC["decl_no"]].map(core.clean).eq(filters["decl"])]
    if filters.get("stir"):
        d = d[d[core.SRC["stir"]].map(digits).eq(filters["stir"])]
    if filters.get("tnved"):
        d = d[d["_tnved_name"].map(core.to_latin).eq(filters["tnved"])]
    if filters.get("regime"):
        d = d[d["_regime_code"].eq(filters["regime"])]
    if filters.get("post"):
        d = d[d["_post_group"].map(core.to_latin).eq(filters["post"])]
    if filters.get("source_post"):
        d = d[d["_source_post_code"].map(core.clean).eq(filters["source_post"])]
    if filters.get("transport"):
        d = d[d["_source_transport"].map(core.clean).eq(filters["transport"])]
    if filters.get("food_name"):
        _food_cats = [
            ("Ichimliklar (alkogolsiz)", r"ичимлик|напит|drink|сок|water|вода"),
            ("Shakar va qandolat mahsulotlari", r"шакар|сахар|sugar|qand|конфет|шоколад"),
            ("Sut mahsulotlari, tuxum va asal", r"сут|молок|milk|tuxum|яйц|egg|asal|мед"),
            ("Go'sht va go'sht mahsulotlari", r"go'sht|мяс|meat|колбас|tovuq|куриц"),
            ("Yog' va moy mahsulotlari", r"yog'|yog\b|мой|масло|\boil\b"),
            ("Don, un va yorma mahsulotlari", r"\bdon\b|\bun\b|мука|круп|guruch|рис|bug'doy|пшениц|jo'xori|arpa"),
            ("Meva-sabzavot mahsulotlari", r"мева|сабзавот|овощ|фрукт|картоф|томат|пиёз|лук|pomidor|bodring"),
            ("Boshqa oziq-ovqatlar", r"oziq|овқат|озиқ|confection|кондитер|озуқ|пищ"),
        ]
        fname = filters["food_name"]
        hay = (d["_tnved_name"].map(core.clean) + " " + d[core.SRC["goods"]].map(core.clean))
        used = pd.Series(False, index=d.index)
        matched = False
        for label, pat in _food_cats:
            mask = hay.str.contains(pat, case=False, na=False, regex=True)
            if core.to_latin(core.clean(label)) == fname:
                d = d[mask & ~used]
                matched = True
                break
            used |= mask
        if not matched:
            d = d[~used & d["_tnved_name"].map(core.clean).str.contains(
                "озиқ|овқат|пищ|food|озуқ", case=False, na=False, regex=True)]
    out = []
    for _, r in d.head(2000).iterrows():
        out.append({
            "decl": core.clean(r[core.SRC["decl_no"]]),
            "date": r["_date_text"],
            "regime": r["_regime_code"],
            "post": core.to_latin(core.clean(r["_post_group"])),
            "source_post": core.clean(r.get("_source_post_code", "")),
            "source_post_name": core.clean(r.get("_source_post_name", "")),
            "transport": core.clean(r.get("_source_transport", "")),
            "stir": digits(r[core.SRC["stir"]]),
            "company": core.clean(r[core.SRC["company"]]),
            "hs": core.clean(r[core.SRC["hs"]]),
            "goods": core.clean(r[core.SRC["goods"]]),
            "partiya": int(r["_partiya"]),
            "vazn": float(r["_weight_tn"]),
            "qiymat": float(r["_value_usd_k"]),
            "tolov": float(r["_pay_total_mln"]),
            "reason": core.clean_reason_display(r[core.SRC["reason"]]),
        })
    return out


def released_between(old_source: Path, new_source: Path) -> dict:
    old = read_report_source(old_source)
    new = read_report_source(new_source)
    old_g = old.groupby("_decl_key", dropna=False).agg(
        company=(core.SRC["company"], "first"),
        stir=(core.SRC["stir"], "first"),
        regime=("_regime_code", "first"),
        post=("_post_group", "first"),
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    )
    new_g = new.groupby("_decl_key", dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    )
    rows = []
    for key, r in old_g.iterrows():
        n = new_g.loc[key] if key in new_g.index else None
        if n is None:
            dp, dv, dq, dt = int(r.partiya), float(r.vazn), float(r.qiymat), float(r.tolov)
        else:
            dp = max(0, int(r.partiya) - int(n.partiya))
            dv = max(0.0, float(r.vazn) - float(n.vazn))
            dq = max(0.0, float(r.qiymat) - float(n.qiymat))
            dt = max(0.0, float(r.tolov) - float(n.tolov))
        if dp > 0 or dv > 0.005 or dq > 0.005 or dt > 0.005:
            rows.append({
                "key": {"decl": key},
                "decl": key,
                "company": core.clean(r.company),
                "stir": digits(r.stir),
                "regime": r.regime,
                "post": core.to_latin(core.clean(r.post)),
                "partiya": dp,
                "vazn": dv,
                "qiymat": dq,
                "tolov": dt,
            })
    rows.sort(key=lambda x: (x["qiymat"], x["tolov"], x["partiya"]), reverse=True)
    return {
        "partiya": sum(x["partiya"] for x in rows),
        "vazn": sum(x["vazn"] for x in rows),
        "qiymat": sum(x["qiymat"] for x in rows),
        "tolov": sum(x["tolov"] for x in rows),
        "rows": rows[:300],
    }


def release_company_table(base_source: Path, final_source: Path) -> dict:
    base = read_report_source(base_source)
    final = read_report_source(final_source)

    def by_company(df: pd.DataFrame):
        g = df.groupby(core.SRC["stir"], dropna=False).agg(
            company=(core.SRC["company"], "first"),
            vazn=("_weight_tn", "sum"),
            qiymat=("_value_usd_k", "sum"),
            partiya=("_partiya", "sum"),
            tolov=("_pay_total_mln", "sum"),
        ).reset_index()
        g["stir"] = g[core.SRC["stir"]].map(digits)
        return g.set_index("stir")

    base_g = by_company(base)
    final_g = by_company(final)
    rows = []
    unreleased = []
    for stir, b in base_g.iterrows():
        f = final_g.loc[stir] if stir in final_g.index else None
        remain_vazn = float(f.vazn) if f is not None else 0.0
        remain_qiymat = float(f.qiymat) if f is not None else 0.0
        remain_partiya = int(f.partiya) if f is not None else 0
        remain_tolov = float(f.tolov) if f is not None else 0.0
        released_vazn = max(0.0, float(b.vazn) - remain_vazn)
        released_qiymat = max(0.0, float(b.qiymat) - remain_qiymat)
        released_partiya = max(0, int(b.partiya) - remain_partiya)
        released_tolov = max(0.0, float(b.tolov) - remain_tolov)
        row = {
            "key": {"stir": stir},
            "korxona": core.clean(b.company),
            "stir": stir,
            "base_vazn": float(b.vazn),
            "base_qiymat": float(b.qiymat),
            "base_partiya": int(b.partiya),
            "remain_vazn": remain_vazn,
            "remain_qiymat": remain_qiymat,
            "remain_partiya": remain_partiya,
            "released_vazn": released_vazn,
            "released_qiymat": released_qiymat,
            "released_pct": (released_qiymat / float(b.qiymat) * 100) if float(b.qiymat) else 0.0,
            "released_partiya": released_partiya,
            "released_tolov": released_tolov,
            "current_vazn": remain_vazn,
            "current_qiymat": remain_qiymat,
            "current_partiya": remain_partiya,
        }
        if released_vazn <= 0.005 and released_qiymat <= 0.005 and released_partiya <= 0 and released_tolov <= 0.005:
            unreleased.append(row)
            continue
        rows.append(row)
    rows.sort(key=lambda x: (x["released_qiymat"], x["released_vazn"], x["released_partiya"]), reverse=True)
    total = {
        "korxona": "Jami",
        "stir": "",
        "base_vazn": sum(r["base_vazn"] for r in rows),
        "base_qiymat": sum(r["base_qiymat"] for r in rows),
        "base_partiya": sum(r["base_partiya"] for r in rows),
        "remain_vazn": sum(r["remain_vazn"] for r in rows),
        "remain_qiymat": sum(r["remain_qiymat"] for r in rows),
        "remain_partiya": sum(r["remain_partiya"] for r in rows),
        "released_vazn": sum(r["released_vazn"] for r in rows),
        "released_qiymat": sum(r["released_qiymat"] for r in rows),
        "released_pct": 0.0,
        "released_partiya": sum(r["released_partiya"] for r in rows),
        "released_tolov": sum(r["released_tolov"] for r in rows),
        "current_vazn": sum(r["current_vazn"] for r in rows),
        "current_qiymat": sum(r["current_qiymat"] for r in rows),
        "current_partiya": sum(r["current_partiya"] for r in rows),
    }
    total["released_pct"] = (total["released_qiymat"] / total["base_qiymat"] * 100) if total["base_qiymat"] else 0.0
    partial = [r for r in rows if r["released_partiya"] == 0 and (r["released_qiymat"] > 0.005 or r["released_vazn"] > 0.005)]
    unreleased.sort(key=lambda x: (x["base_qiymat"], x["base_vazn"], x["base_partiya"]), reverse=True)
    return {"total": total, "rows": rows, "partial": partial[:100], "unreleased": unreleased[:100], "top_released": rows[:20]}


def find_report_by_date(date_text: str):
    for r in load_json(INDEX_PATH, {"reports": []})["reports"]:
        if r.get("date") == date_text:
            return r
    return None


def company_headers():
    return [
        {"k": "korxona", "t": "Korxona", "width": 42},
        {"k": "stir", "t": "STIR", "width": 14},
        {"k": "partiya", "t": "Partiya", "i": True, "width": 10},
        {"k": "vazn", "t": "Vazn (tn)", "n": True, "width": 14},
        {"k": "qiymat", "t": "Qiymat (ming $)", "n": True, "width": 16},
        {"k": "tolov", "t": "To'lov (mln so'm)", "n": True, "width": 18},
        {"k": "depozit", "t": "Depozit (mln so'm)", "n": True, "width": 18},
    ]


def expired_headers():
    return [
        {"k": "korxona", "t": "Korxona", "width": 42},
        {"k": "stir", "t": "STIR", "width": 14},
        {"k": "rejim", "t": "Rejim", "width": 10},
        {"k": "post", "t": "Post", "width": 22},
        {"k": "partiya", "t": "Partiya", "i": True, "width": 10},
        {"k": "kun", "t": "Kun hisobi", "i": True, "width": 12},
        {"k": "qiymat", "t": "Qiymat (ming $)", "n": True, "width": 16},
        {"k": "tolov", "t": "To'lov (mln so'm)", "n": True, "width": 18},
    ]


def goods_headers():
    return [
        {"k": "name", "t": "Tovar", "width": 46},
        {"k": "partiya", "t": "Partiya", "i": True, "width": 10},
        {"k": "qiymat", "t": "Qiymat (ming $)", "n": True, "width": 16},
        {"k": "tolov", "t": "To'lov (mln so'm)", "n": True, "width": 18},
    ]


def food_headers():
    return [
        {"k": "name", "t": "Oziq-ovqat turi", "width": 42},
        {"k": "vazn", "t": "Vazn (tn)", "n": True, "width": 14},
        {"k": "qiymat", "t": "Qiymat (ming $)", "n": True, "width": 16},
        {"k": "over_vazn", "t": "3 oy+ vazn (tn)", "n": True, "width": 16},
        {"k": "over_qiymat", "t": "3 oy+ qiymat (ming $)", "n": True, "width": 18},
        {"k": "ulush", "t": "Qiymatdagi ulushi (%)", "n": True, "width": 18},
    ]


def basic_headers(name_title: str = "Ko'rsatkich"):
    return [
        {"k": "name", "t": name_title, "width": 38},
        {"k": "partiya", "t": "Partiya", "i": True, "width": 10},
        {"k": "vazn", "t": "Vazn (tn)", "n": True, "width": 14},
        {"k": "qiymat", "t": "Qiymat (ming $)", "n": True, "width": 16},
        {"k": "tolov", "t": "To'lov (mln so'm)", "n": True, "width": 18},
    ]


def release_headers():
    return [
        {"k": "korxona", "t": "Korxona", "width": 42},
        {"k": "stir", "t": "STIR", "width": 14},
        {"k": "base_vazn", "t": "Boshlang'ich vazn (tn)", "n": True, "width": 18},
        {"k": "base_qiymat", "t": "Boshlang'ich qiymat (ming $)", "n": True, "width": 22},
        {"k": "base_partiya", "t": "Boshlang'ich partiya", "i": True, "width": 18},
        {"k": "remain_vazn", "t": "Yakuniy sanada boshlang'ichdan qoldiq vazn (tn)", "n": True, "width": 24},
        {"k": "remain_qiymat", "t": "Yakuniy sanada boshlang'ichdan qoldiq qiymat (ming $)", "n": True, "width": 28},
        {"k": "remain_partiya", "t": "Yakuniy sanada boshlang'ichdan qoldiq partiya", "i": True, "width": 24},
        {"k": "released_vazn", "t": "Yechilishi vazn (tn)", "n": True, "width": 18},
        {"k": "released_qiymat", "t": "Yechilishi qiymat (ming $)", "n": True, "width": 20},
        {"k": "released_pct", "t": "Yechilishi (%)", "n": True, "width": 14},
        {"k": "released_partiya", "t": "Yechilishi partiya", "i": True, "width": 16},
        {"k": "released_tolov", "t": "Yechilgan to'lov (mln so'm)", "n": True, "width": 20},
        {"k": "current_vazn", "t": "Yakuniy qoldiq vazn (tn)", "n": True, "width": 18},
        {"k": "current_qiymat", "t": "Yakuniy qoldiq qiymat (ming $)", "n": True, "width": 22},
        {"k": "current_partiya", "t": "Yakuniy qoldiq partiya", "i": True, "width": 18},
    ]


def make_xlsx(headers: list[dict], rows: list[dict], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadval"
    ws.cell(1, 1, title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h["t"])
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, h in enumerate(headers, 1):
            value = row.get(h["k"], "")
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if h.get("n") else "left", vertical="center")
            if h.get("i"):
                cell.number_format = "# ##0"
            elif h.get("n"):
                cell.number_format = "# ##0.00"
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(2, c).column_letter].width = h.get("width", 14)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_dashboard(report_id: str, source: Path, deposit: Path | None, report_date: datetime) -> dict:
    df = read_report_source(source)
    deposits, deposit_date, deposit_total = core.read_deposit(deposit)
    expired = core.with_expiry(df, report_date)
    expired = expired[expired["_expired"]].copy()

    companies = df.groupby(core.SRC["stir"], dropna=False).agg(
        company=(core.SRC["company"], "first"),
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).reset_index()
    companies["stir"] = companies[core.SRC["stir"]].map(digits)
    companies["depozit"] = companies["stir"].map(lambda s: float(deposits.get(s, 0.0)))

    def company_records(frame):
        rows = []
        for _, r in frame.iterrows():
            rows.append({
                "key": {"stir": r["stir"]},
                "korxona": core.clean(r.company),
                "stir": r["stir"],
                "partiya": int(r.partiya),
                "vazn": float(r.vazn),
                "qiymat": float(r.qiymat),
                "tolov": float(r.tolov),
                "depozit": float(r.get("depozit", 0.0)),
            })
        return rows

    goods_rows_web = []
    for r in core.goods_rows(df):
        goods_rows_web.append({"key": {"tnved": core.to_latin(core.clean(r[1]))}, "name": core.to_latin(core.clean(r[1])), "partiya": int(r[2]), "korxona": int(r[3]), "vazn": float(r[4]), "qiymat": float(r[5]), "tolov": float(r[6])})

    regimes = df.groupby("_regime_code", dropna=False).agg(
        partiya=("_partiya", "sum"),
        vazn=("_weight_tn", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
    ).sort_values("qiymat", ascending=False).reset_index()

    exp = expired.groupby([core.SRC["stir"], "_regime_code", "_post_group"], dropna=False).agg(
        company=(core.SRC["company"], "first"),
        partiya=("_partiya", "sum"),
        qiymat=("_value_usd_k", "sum"),
        tolov=("_pay_total_mln", "sum"),
        expiry=("_expiry", "min"),
    ).sort_values(["partiya", "qiymat"], ascending=False).reset_index()
    exp_rows = []
    for _, r in exp.iterrows():
        exp_rows.append({
            "key": {"stir": digits(r[core.SRC["stir"]]), "regime": r["_regime_code"], "post": core.to_latin(core.clean(r["_post_group"]))},
            "korxona": core.clean(r.company),
            "stir": digits(r[core.SRC["stir"]]),
            "rejim": r["_regime_code"],
            "post": core.to_latin(core.clean(r["_post_group"])),
            "partiya": int(r.partiya),
            "kun": max(0, int((report_date - r.expiry).days)) if pd.notna(r.expiry) else 0,
            "qiymat": float(r.qiymat),
            "tolov": float(r.tolov),
        })

    age_rows = []
    for label, cutoff in [("3 oy+", report_date - pd.DateOffset(months=3)), ("6 oy+", report_date - pd.DateOffset(months=6)), ("1 yil+", report_date - pd.DateOffset(years=1))]:
        d = df[df["_date"].le(cutoff)]
        age_rows.append({"muddat": label, "partiya": int(d["_partiya"].sum()), "vazn": float(d["_weight_tn"].sum()), "qiymat": float(d["_value_usd_k"].sum()), "tolov": float(d["_pay_total_mln"].sum())})

    summary_rows = [{
        "name": "Jami",
        "partiya": int(df["_partiya"].sum()),
        "vazn": float(df["_weight_tn"].sum()),
        "qiymat": float(df["_value_usd_k"].sum()),
        "tolov": float(df["_pay_total_mln"].sum()),
    }]
    for _, r in regimes.iterrows():
        summary_rows.append({"key": {"view": "regime_posts", "regime": r["_regime_code"]}, "name": r["_regime_code"], "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov)})
    expired_summary = [{
        "name": "Jami muddati o'tgan",
        "partiya": int(expired["_partiya"].sum()),
        "vazn": float(expired["_weight_tn"].sum()) if len(expired) else 0.0,
        "qiymat": float(expired["_value_usd_k"].sum()) if len(expired) else 0.0,
        "tolov": float(expired["_pay_total_mln"].sum()) if len(expired) else 0.0,
    }]

    archive = load_json(INDEX_PATH, {"reports": []})["reports"]
    by_date = {}
    for r in archive:
        by_date.setdefault(r.get("date", ""), r)
    release = {}
    for days in [1, 3, 5, 10, 30]:
        base_dt = report_date - timedelta(days=days)
        base_key = fmt_date(base_dt)
        item = by_date.get(base_key)
        old_path = Path(item["source"]) if item and item.get("source") else None
        if old_path and old_path.exists():
            try:
                release[str(days)] = released_between(old_path, source)
                release[str(days)]["base_date"] = base_key
            except Exception:
                release[str(days)] = {"partiya": 0, "vazn": 0, "qiymat": 0, "tolov": 0, "rows": [], "base_date": base_key, "missing_date": base_key}
        else:
            release[str(days)] = {"partiya": 0, "vazn": 0, "qiymat": 0, "tolov": 0, "rows": [], "base_date": "", "missing_date": base_key}

    food_total_value = 0.0
    food_data = []
    for row in core.food_rows(df):
        item = {"key": {"food_name": core.to_latin(core.clean(row[1]))}, "name": core.to_latin(core.clean(row[1])), "vazn": float(row[2]), "qiymat": float(row[3]), "over_vazn": float(row[4]), "over_qiymat": float(row[5]), "ulush": 0.0}
        food_data.append(item)
        food_total_value += item["qiymat"]
    for item in food_data:
        item["ulush"] = (item["qiymat"] / food_total_value * 100) if food_total_value else 0.0
    food_total = {"key": {}, "name": "IBK bo'yicha Jami", "vazn": sum(x["vazn"] for x in food_data), "qiymat": food_total_value, "over_vazn": sum(x["over_vazn"] for x in food_data), "over_qiymat": sum(x["over_qiymat"] for x in food_data), "ulush": 100.0 if food_total_value else 0.0}

    def list_to_basic(rows, name_idx=1, partiya_idx=2, vazn_idx=3, qiymat_idx=4, tolov_idx=5):
        out = []
        for row in rows:
            out.append({"key": {}, "name": core.to_latin(core.clean(row[name_idx])), "partiya": int(row[partiya_idx] or 0), "vazn": float(row[vazn_idx] or 0), "qiymat": float(row[qiymat_idx] or 0), "tolov": float(row[tolov_idx] or 0)})
        return out

    wh = df.copy()
    wh["_warehouse_web"] = wh[core.SRC["warehouse"]].map(own_warehouse_label)
    warehouse_data = []
    wh_g = wh.groupby("_warehouse_web", dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum")).reset_index().sort_values(["qiymat", "tolov"], ascending=False)
    for _, r in wh_g.iterrows():
        name = core.to_latin(core.clean(r["_warehouse_web"]))
        warehouse_data.append({"key": {"warehouse": name}, "name": name, "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov), "_class": "own-row" if name.lower().startswith("o'z ombor") else ""})
    reason_data = list_to_basic(core.reason_rows(df))
    own_all_data = own_company_rows_web(df, report_date, False)
    own_3m_data = own_company_rows_web(df, report_date, True)
    country_data = []
    for _, r in df.groupby(core.SRC["country"], dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum"), korxona=(core.SRC["stir"], "nunique")).reset_index().sort_values("qiymat", ascending=False).iterrows():
        country_data.append({"key": {"country": core.to_latin(core.clean(r[core.SRC["country"]]))}, "name": core.to_latin(core.clean(r[core.SRC["country"]])), "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov), "korxona": int(r.korxona)})
    all_post_summary = [{"key": {"post": r["_post_group"]}, "post": core.to_latin(core.clean(r["_post_group"])), "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov)} for _, r in df.groupby("_post_group", dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum")).reset_index().sort_values("qiymat", ascending=False).iterrows()]
    source_post_data = [{"key": {"source_post": r["_source_post_code"]}, "post_kodi": core.clean(r["_source_post_code"]) or "-", "post_nomi": core.clean(r["_source_post_name"]) or "-", "transport": core.clean(r["_source_transport"]) or "-", "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov), "korxona": int(r.korxona)} for _, r in df.groupby(["_source_post_code", "_source_post_name", "_source_transport"], dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum"), korxona=(core.SRC["stir"], "nunique")).reset_index().sort_values("qiymat", ascending=False).iterrows()]
    transport_data = [{"key": {"transport": r["_source_transport"]}, "name": core.clean(r["_source_transport"]) or "-", "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov)} for _, r in df.groupby("_source_transport", dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum")).reset_index().sort_values("qiymat", ascending=False).iterrows()]
    post_summary = [{"key": {"post": r["_post_group"]}, "post": core.to_latin(core.clean(r["_post_group"])), "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov)} for _, r in expired.groupby("_post_group", dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum")).reset_index().sort_values("qiymat", ascending=False).iterrows()]
    regime_posts = {}
    for regime, part in df.groupby("_regime_code", dropna=False):
        regime_posts[regime] = [{"key": {"regime": regime, "post": r["_post_group"]}, "post": core.to_latin(core.clean(r["_post_group"])), "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov)} for _, r in part.groupby("_post_group", dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum")).reset_index().sort_values("qiymat", ascending=False).iterrows()]
    year_series = df["_date_text"].astype(str).str.extract(r"(\d{4})", expand=False).fillna("")
    df["_gtd_year"] = year_series
    regime_year_post = []
    for _, r in df.groupby(["_post_group", "_regime_code", "_gtd_year"], dropna=False).agg(partiya=("_partiya", "sum"), vazn=("_weight_tn", "sum"), qiymat=("_value_usd_k", "sum"), tolov=("_pay_total_mln", "sum")).reset_index().sort_values(["_post_group", "_regime_code", "_gtd_year"]).iterrows():
        regime_year_post.append({"key": {"post": core.to_latin(core.clean(r["_post_group"])), "regime": r["_regime_code"]}, "post": core.to_latin(core.clean(r["_post_group"])), "rejim": core.clean(r["_regime_code"]), "yil": core.clean(r["_gtd_year"]) or "-", "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov)})
    expired_post_regime = []
    for row in core.expired_summary_rows(df, report_date):
        expired_post_regime.append({"key": {"post": core.to_latin(core.clean(row[0]))}, "post": core.to_latin(core.clean(row[0])), "jami_partiya": int(row[1]), "jami_qiymat": float(row[2]), "expired_partiya": int(row[3]), "expired_qiymat": float(row[4]), "ulush": float(row[5]) * 100, "im70_partiya": int(row[6]), "im70_qiymat": float(row[7]), "im74_partiya": int(row[8]), "im74_qiymat": float(row[9]), "tr80_partiya": int(row[10]), "tr80_qiymat": float(row[11])})
    expired_block = []
    for row in core.expired_block_rows(df, report_date):
        expired_block.append({"key": {}, "name": core.to_latin(core.clean(row[0])), "korxona": core.clean(row[1]), "stir": core.clean(row[2]), "partiya": int(row[3] or 0), "qiymat": float(row[4] or 0), "vazn": float(row[5] or 0), "tolov": float(row[6] or 0), "reason": core.to_latin(core.clean(row[7]))})

    return {
        "schema": 5,
        "id": report_id,
        "meta": {"date": fmt_date(report_date), "source": source.name, "deposit": deposit.name if deposit else "", "deposit_date": fmt_date(deposit_date) if deposit_date else ""},
        "kpis": {"partiya": int(df["_partiya"].sum()), "vazn": float(df["_weight_tn"].sum()), "qiymat": float(df["_value_usd_k"].sum()), "tolov": float(df["_pay_total_mln"].sum()), "depozit": float(deposit_total), "depozit_matched": float(companies[companies["depozit"].gt(0)]["depozit"].sum()), "expired": int(expired["_partiya"].sum()), "expired_value": float(expired["_value_usd_k"].sum()) if len(expired) else 0.0},
        "top_value": company_records(companies.sort_values("qiymat", ascending=False).head(30)),
        "top_deposit": company_records(companies[companies["depozit"].gt(0)].sort_values("depozit", ascending=False).head(20)),
        "all_companies": company_records(companies.sort_values("qiymat", ascending=False)),
        "regimes": [{"key": {"regime": r["_regime_code"]}, "rejim": r["_regime_code"], "partiya": int(r.partiya), "vazn": float(r.vazn), "qiymat": float(r.qiymat), "tolov": float(r.tolov)} for _, r in regimes.iterrows()],
        "summary": summary_rows,
        "expired_summary": expired_summary,
        "goods": goods_rows_web,
        "expired": exp_rows,
        "ages": age_rows,
        "all_post_summary": all_post_summary,
        "source_posts": source_post_data,
        "transport": transport_data,
        "post_summary": post_summary,
        "regime_posts": regime_posts,
        "regime_year_post": regime_year_post,
        "expired_post_regime": expired_post_regime,
        "expired_block": expired_block,
        "food": food_data,
        "food_total": food_total,
        "warehouse": warehouse_data,
        "reason": reason_data,
        "countries": country_data,
        "own_all": own_all_data,
        "own_3m": own_3m_data,
        "released": release,
    }


def update_report_files(job_id: str, report_dir: Path, report_date: datetime, error: str | None = None) -> dict:
    """Har doim dict qaytaradi. report_dir yoki report_date None bo'lsa ham xavfsiz ishlaydi."""
    # None himoyasi
    if report_dir is None or report_date is None:
        files: dict = {
            "status": "xatolik",
            "excel": "",
            "pdf": "",
            "pngs": [],
            "error": error or ("report_dir=None" if report_dir is None else "report_date=None"),
        }
        job = ensure_job(job_id)
        job["data"]["files"] = files
        job["files"] = files
        return files

    try:
        excel_path = report_dir / f"Ombor_malumot_{report_date:%Y_%m_%d}.xlsx"
        if not excel_path.exists():
            old_excel = report_dir / f"IBK_jamlanma_{report_date:%Y_%m_%d}.xlsx"
            if old_excel.exists():
                excel_path = old_excel
        png_path = report_dir / f"IBK_tahlil_{report_date:%Y_%m_%d}.png"
        pdf_path = report_dir / f"IBK_tahlil_{report_date:%Y_%m_%d}.pdf"
        pngs = sorted(report_dir.glob(f"{png_path.stem}*.png"))
        files = {
            "status": "tayyor" if excel_path.exists() and (pdf_path.exists() or pngs) else ("excel_tayyor" if excel_path.exists() else "tayyorlanmoqda"),
            "excel": excel_path.name if excel_path.exists() else "",
            "pdf": pdf_path.name if pdf_path.exists() else "",
            "pngs": [p.name for p in pngs if p.exists()],
        }
    except Exception as exc:
        files = {"status": "xatolik", "excel": "", "pdf": "", "pngs": [], "error": f"files_build: {exc}"}

    if error:
        files["error"] = error

    try:
        data_path = report_dir / "dashboard.json"
        data = load_json(data_path, {})
        if not isinstance(data, dict):
            data = {}
        data["files"] = files
        save_json(data_path, data)
    except Exception:
        pass

    job = ensure_job(job_id)
    if not isinstance(job.get("data"), dict):
        job["data"] = {}
    job["data"]["files"] = files
    job["files"] = files
    return files

def generate_artifacts(job_id: str, stored_source: Path, stored_deposit: Path | None, report_date: datetime, report_dir: Path):
    excel_path = report_dir / f"Ombor_malumot_{report_date:%Y_%m_%d}.xlsx"
    png_path = report_dir / f"IBK_tahlil_{report_date:%Y_%m_%d}.png"
    pdf_path = report_dir / f"IBK_tahlil_{report_date:%Y_%m_%d}.pdf"
    try:
        core.build(stored_source, TEMPLATE_PATH, excel_path, report_date, stored_deposit)
        update_report_files(job_id, report_dir, report_date)
    except Exception as exc:
        update_report_files(job_id, report_dir, report_date, f"Excel: {type(exc).__name__}: {exc}")
        return
    try:
        build_png_report(stored_source, png_path, pdf_path, report_date, stored_deposit)
        update_report_files(job_id, report_dir, report_date)
    except Exception as exc:
        update_report_files(job_id, report_dir, report_date, f"PNG/PDF: {type(exc).__name__}: {exc}")


def run_artifact_job(job_id: str, report_id: str):
    job = ensure_job(job_id)
    try:
        job["status"] = "Excel/PNG/PDF tayyorlanmoqda"
        archive = load_json(INDEX_PATH, {"reports": []})
        if not isinstance(archive, dict):
            archive = {"reports": []}
        reports = archive.get("reports") or []
        item = next((r for r in reports if isinstance(r, dict) and r.get("id") == report_id), None)
        if not item:
            raise FileNotFoundError("Hisobot topilmadi")
        report_dir = Path(item["dir"]) if item.get("dir") else None
        if report_dir is None:
            raise ValueError("run_artifact_job: report_dir=None")
        source = Path(item["source"]) if item.get("source") else None
        if source is None:
            raise ValueError("run_artifact_job: source=None")
        deposit = Path(item["deposit"]) if item.get("deposit") else None
        date_str = item.get("date", "")
        if not date_str:
            raise ValueError("run_artifact_job: date bo'sh")
        report_date = datetime.strptime(date_str, "%d.%m.%Y")
        generate_artifacts(report_id, source, deposit, report_date, report_dir)
        data = load_json(report_dir / "dashboard.json", {})
        if not isinstance(data, dict):
            data = {}
        job = ensure_job(job_id)
        job.update({"status": "tayyor", "data": data})
    except Exception as exc:
        job = ensure_job(job_id)
        job["status"] = "xatolik"
        job["error"] = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()}"



def tolov_summary_rows() -> list[dict]:
    summary = load_json(TOLOV_OUTPUT_DIR / "summary.json", {})
    rows = []
    for filename, name in TOLOV_FILES.items():
        rec = summary.get(filename, {})
        rows.append({
            "name": name,
            "rows": int(float(rec.get("rows", 0) or 0)),
            "sum": float(rec.get("sum", 0) or 0),
            "file": filename,
        })
    return rows


def build_tolov_from_source(source: Path) -> list[dict]:
    if not TOLOV_GENERATOR_PATH.exists():
        raise FileNotFoundError(f"To'lov generatori topilmadi: {TOLOV_GENERATOR_PATH}")
    spec = importlib.util.spec_from_file_location("ibk_tolov_generator", TOLOV_GENERATOR_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError("To'lov generatorini yuklab bo'lmadi")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    summary = module.build(source, TOLOV_OUTPUT_DIR)
    save_json(TOLOV_OUTPUT_DIR / "summary.json", summary)
    return tolov_summary_rows()


def run_job(job_id: str, source: Path, deposit: Path | None, report_date: datetime):
    job = ensure_job(job_id)
    try:
        # Kiruvchi parametrlarni tekshirish
        if report_date is None:
            raise ValueError("run_job: report_date=None, fayl nomidan sana aniqlanmadi")
        if source is None:
            raise ValueError("run_job: source=None")

        job["status"] = "Hisob-kitob qilinyapti"
        report_dir = REPORT_DIR / job_id
        report_dir.mkdir(parents=True, exist_ok=True)
        stored_source = report_dir / source.name
        shutil.copy2(source, stored_source)
        stored_deposit = None
        if deposit:
            stored_deposit = report_dir / deposit.name
            shutil.copy2(deposit, stored_deposit)

        df_for_store = read_report_source(stored_source)
        save_snapshot_to_store(job_id, stored_source, stored_deposit, report_date, df_for_store)
        data = build_dashboard(job_id, stored_source, stored_deposit, report_date)
        if not isinstance(data, dict):
            data = {}
        files = update_report_files(job_id, report_dir, report_date)
        if not isinstance(files, dict):
            files = {"status": "kutilmoqda", "excel": "", "pdf": "", "pngs": []}
        data["files"] = files
        if not data["files"].get("excel"):
            data["files"]["status"] = "kutilmoqda"
        save_json(report_dir / "dashboard.json", data)

        archive = load_json(INDEX_PATH, {"reports": []})
        if not isinstance(archive, dict):
            archive = {"reports": []}
        reports = archive.get("reports")
        if not isinstance(reports, list):
            reports = []
        reports = [r for r in reports if isinstance(r, dict) and r.get("id") != job_id]
        reports.append({"id": job_id, "date": fmt_date(report_date), "source": str(stored_source), "deposit": str(stored_deposit or ""), "dir": str(report_dir)})
        archive["reports"] = clean_archive_records(reports)
        save_json(INDEX_PATH, archive)
        job.update({"status": "tayyor", "data": data})
    except Exception as exc:
        job = ensure_job(job_id)
        job["status"] = "xatolik"
        job["error"] = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()}"


HTML = r"""<!doctype html><html lang="uz"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>IBK Dashboard</title>
<style>
:root{--ink:#172033;--muted:#64748b;--line:#dae2ec;--blue:#174a7c;--green:#1b9e77;--orange:#d95f02;--bg:#f6f8fb;--panel:#fff}*{box-sizing:border-box}body{margin:0;color:var(--ink);font:14px/1.45 "Segoe UI",Arial,sans-serif;background:linear-gradient(120deg,#f6f8fb,#eef6fb,#f7fbf5);background-size:280% 280%;animation:bgshift 18s ease-in-out infinite}header{position:sticky;top:0;background:#fff;border-bottom:1px solid var(--line);z-index:3;padding:12px 18px;display:flex;align-items:center;justify-content:space-between;gap:12px}h1{font-size:22px;margin:0}.muted{color:var(--muted)}main{max-width:1500px;margin:auto;padding:14px}.login,.upload,.panel{background:#fff;border:1px solid var(--line);border-radius:8px;padding:14px}.panel{overflow-x:auto}.upload{display:grid;grid-template-columns:1fr 1fr 170px auto;gap:10px;align-items:end}input,select{width:100%;padding:9px;border:1px solid var(--line);border-radius:6px;background:#fff}label{display:block;margin-bottom:5px;font-weight:700;color:var(--muted)}button,.btn{background:var(--blue);color:#fff;border:0;border-radius:6px;padding:10px 13px;font-weight:700;text-decoration:none;cursor:pointer}.btn.light,button.light{background:#e8eef6;color:var(--ink)}.kpis{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin:12px 0}.kpi{background:#fff;border:1px solid var(--line);border-radius:8px;padding:12px;display:flex;flex-direction:column}.kpi b{display:block;font-size:24px;margin-top:5px}.kpi-note{font-size:9.5px;color:var(--muted);margin-top:4px;font-style:italic}.tabs{display:flex;gap:8px;flex-wrap:wrap;margin:12px 0}.tab.active{background:var(--blue);color:#fff}.tab{background:#e8eef6;color:#172033}.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;align-items:start}h2{font-size:17px;margin:0 0 10px}table{width:100%;min-width:760px;border-collapse:collapse;table-layout:auto}.release-table{min-width:1120px;font-size:12px;table-layout:fixed}.release-table th,.release-table td{padding:4px 5px}.release-table th{white-space:normal!important;word-break:normal;line-height:1.12;vertical-align:middle;text-align:center}.release-table th:nth-child(1),.release-table td:nth-child(1){width:34px;text-align:center}.release-table th:nth-child(2),.release-table td:nth-child(2){width:245px}.release-table th:nth-child(3),.release-table td:nth-child(3){width:82px}.release-table th:nth-child(n+4),.release-table td:nth-child(n+4){width:82px}.release-table th:nth-child(6),.release-table td:nth-child(6),.release-table th:nth-child(9),.release-table td:nth-child(9),.release-table th:nth-child(12),.release-table td:nth-child(12){width:96px}th{background:var(--blue);color:#fff;text-align:left;padding:7px;font-size:13px}td{border:1px solid #d7e0ea;padding:5px 6px;vertical-align:middle}td.text,th.text{white-space:normal;overflow:visible;text-overflow:clip;word-break:normal}td.num,th.num{text-align:right;white-space:nowrap}tbody tr:first-child td{font-weight:800;background:#dce8f5}tbody tr.own-row td{font-weight:700;color:#155e75}th{white-space:normal;line-height:1.2}tr:nth-child(even) td{background:#f8fafc}tbody tr{cursor:pointer}.merged-row td{background:#dfeaf6!important;font-weight:800}.merged-row td:first-child{text-align:left}.bars{display:grid;gap:8px}.barrow{display:grid;grid-template-columns:220px 1fr 120px;gap:8px;align-items:center}.bar{height:24px;background:#e8eef6;position:relative}.bar span{display:block;height:100%;min-width:34px;background:var(--green);color:#fff;font-size:12px;font-weight:700;text-align:right;padding-right:5px;line-height:24px}dialog{width:min(1100px,96vw);border:0;border-radius:8px;padding:0}dialog .head{padding:12px 14px;border-bottom:1px solid var(--line);display:flex;justify-content:space-between}dialog .body{padding:14px;max-height:72vh;overflow:auto}.hidden{display:none}@media(max-width:1000px){.upload,.grid{grid-template-columns:1fr}.kpis{grid-template-columns:repeat(2,1fr)}header{align-items:flex-start;flex-direction:column}}
 .workspace{display:grid;grid-template-columns:210px 1fr;gap:14px;align-items:start}.tabs{position:sticky;top:78px;display:flex;flex-direction:column;gap:8px;margin:12px 0}.tab{width:100%;text-align:left;background:#e8eef6;color:#172033;border-left:5px solid transparent;transition:.18s transform,.18s background,.18s box-shadow}.tab:active,button:active,.btn:active{transform:translateY(1px) scale(.99)}.tab.active{background:#174a7c;color:#fff;border-left-color:#1b9e77;box-shadow:0 2px 8px rgba(23,74,124,.18)}.chart{margin-top:12px;padding-top:10px;border-top:1px solid var(--line)}.barrow{grid-template-columns:minmax(145px,220px) 1fr 110px}.barrow div:first-child{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.bar span{animation:growbar .55s ease-out}.wide{grid-column:1/-1}.partial{background:#fff4cc!important;color:#8a5a00;font-weight:800;border:1px solid #e6b84d}.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center}.danger{background:#b42318!important}.dark{--ink:#e5eefb;--muted:#9fb0c6;--line:#2f3f55;--blue:#2c7fb8;--bg:#101826;--panel:#172033}.dark header,.dark .login,.dark .upload,.dark .panel,.dark .kpi{background:#172033}.dark input,.dark select{background:#101826;color:var(--ink);border-color:var(--line)}.dark tr:nth-child(even) td{background:#1d2a3d}@keyframes bgshift{0%,100%{background-position:0% 50%}50%{background-position:100% 50%}}@keyframes growbar{from{width:0}to{}}@media(max-width:1000px){.workspace{grid-template-columns:1fr}.tabs{position:static;flex-direction:row;overflow:auto}.tab{min-width:max-content}.barrow{grid-template-columns:120px 1fr 82px}}

/* Customs emblem + airport movement background */
body::before{content:none;display:none}
body::after{content:none;display:none}
header::after{content:"";position:fixed;left:-120px;top:92px;width:110px;height:2px;background:linear-gradient(90deg,transparent,rgba(23,74,124,.45),transparent);box-shadow:24px -5px 0 -1px rgba(23,74,124,.22);transform:rotate(-12deg);animation:flightPath 18s linear infinite;pointer-events:none}
@keyframes runwayMove{from{background-position:0 0,0 0}to{background-position:360px 0,420px 0}}
@keyframes flightPath{0%{transform:translateX(-8vw) translateY(0) rotate(-12deg);opacity:0}12%{opacity:.5}70%{opacity:.35}100%{transform:translateX(115vw) translateY(34vh) rotate(-12deg);opacity:0}}
.dark body::before{opacity:.38;filter:brightness(.7)}



/* Premium background system */
body{background:#eef4f7;color:var(--ink)}
body::before{content:"";position:fixed;inset:0;z-index:-4;pointer-events:none;background:
  linear-gradient(120deg,rgba(245,248,250,.92),rgba(229,239,239,.88) 48%,rgba(247,250,247,.93)),
  url('/assets/gerb-bojxona.jpg') right 6vw top 96px / min(330px,28vw) auto no-repeat;
  filter:saturate(.92);opacity:1}
body::after{content:"";position:fixed;inset:0;z-index:-3;pointer-events:none;background:
  linear-gradient(118deg,transparent 0 36%,rgba(17,70,84,.12) 36.4%,transparent 37.1% 100%),
  linear-gradient(118deg,transparent 0 49%,rgba(28,113,89,.10) 49.2%,transparent 49.8% 100%),
  repeating-linear-gradient(118deg,transparent 0 118px,rgba(17,70,84,.055) 120px,transparent 124px),
  linear-gradient(180deg,rgba(255,255,255,.36),rgba(255,255,255,.72));
  animation:premiumDrift 38s linear infinite}
main::before{content:none;display:none}
header::before{content:"";position:fixed;left:-160px;top:118px;width:155px;height:3px;z-index:-1;pointer-events:none;background:linear-gradient(90deg,transparent,rgba(18,74,91,.58),rgba(27,158,119,.32),transparent);box-shadow:38px -8px 0 -1px rgba(18,74,91,.20),76px 8px 0 -2px rgba(27,158,119,.18);transform:rotate(-10deg);animation:premiumFlight 24s ease-in-out infinite}
.panel,.kpi,.login,.upload{box-shadow:0 14px 36px rgba(23,43,77,.07);background:rgba(255,255,255,.92);backdrop-filter:blur(8px)}
header{background:rgba(255,255,255,.88);backdrop-filter:blur(12px)}
body.bg-aero::before{background:
  linear-gradient(120deg,rgba(244,248,250,.95),rgba(235,244,243,.88)),
  url('/assets/gerb-bojxona.jpg') center 58% / min(360px,34vw) auto no-repeat;opacity:1}
body.bg-aero::after{background:
  linear-gradient(62deg,transparent 0 44%,rgba(23,74,124,.15) 44.25%,transparent 44.9%),
  linear-gradient(152deg,transparent 0 53%,rgba(27,158,119,.12) 53.25%,transparent 53.8%),
  repeating-linear-gradient(62deg,transparent 0 86px,rgba(23,74,124,.075) 88px,transparent 93px),
  repeating-linear-gradient(152deg,transparent 0 136px,rgba(217,95,2,.065) 138px,transparent 143px),
  linear-gradient(180deg,rgba(255,255,255,.34),rgba(255,255,255,.76));animation:aeroSweep 34s linear infinite}
body.bg-classic::before{background:linear-gradient(120deg,rgba(246,248,251,.97),rgba(238,246,251,.94),rgba(247,251,245,.97)),url('/assets/gerb-bojxona.jpg') center 46% / min(380px,36vw) auto no-repeat}
body.bg-classic::after{background:repeating-linear-gradient(115deg,transparent 0 92px,rgba(23,74,124,.055) 94px,transparent 99px);animation:premiumDrift 42s linear infinite}
@keyframes premiumDrift{from{background-position:0 0,0 0,0 0,0 0}to{background-position:520px 0,360px 0,480px 0,0 0}}
@keyframes aeroSweep{from{background-position:0 0,0 0,0 0,0 0,0 0}to{background-position:420px 0,-380px 0,520px 0,-460px 0,0 0}}
@keyframes premiumFlight{0%{transform:translateX(-8vw) translateY(0) rotate(-10deg);opacity:0}14%{opacity:.7}62%{opacity:.38}100%{transform:translateX(118vw) translateY(24vh) rotate(-10deg);opacity:0}}
.dark body::before,.dark.bg-aero::before,.dark.bg-classic::before{filter:brightness(.55) saturate(.85);opacity:.52}
.dark .panel,.dark .kpi,.dark .login,.dark .upload{background:rgba(23,32,51,.92)}

/* Animated Toshkent-AERO premium scene */
body::before,body::after,header::before{content:none!important;display:none!important}
.sky-scene{position:fixed;inset:0;z-index:0;overflow:hidden;pointer-events:none;background:linear-gradient(180deg,#dff4ff 0%,#f4fbff 42%,#dbeff3 68%,#c6e6ef 100%)}
.sky-scene:before{content:"";position:absolute;inset:0;background:radial-gradient(circle at 18% 18%,rgba(255,255,255,.88),rgba(255,255,255,0) 22%),linear-gradient(110deg,rgba(255,255,255,.38),rgba(255,255,255,0) 38%);opacity:.9}
.sky-scene:after{content:"";position:absolute;inset:0;background-image:linear-gradient(rgba(18,78,116,.045) 1px,transparent 1px),linear-gradient(90deg,rgba(18,78,116,.04) 1px,transparent 1px);background-size:52px 52px;mask-image:linear-gradient(180deg,transparent 0%,#000 36%,transparent 100%)}
.sky-layer{position:absolute;left:0;right:0}
.mountains{bottom:30vh;height:26vh;background:linear-gradient(135deg,transparent 0 22%,rgba(74,112,126,.28) 22% 34%,transparent 34% 42%,rgba(57,103,118,.36) 42% 58%,transparent 58% 66%,rgba(85,128,139,.24) 66% 78%,transparent 78%),linear-gradient(180deg,rgba(255,255,255,.45),rgba(255,255,255,0));clip-path:polygon(0 100%,0 52%,10% 66%,18% 32%,28% 72%,41% 25%,55% 70%,68% 22%,80% 62%,91% 36%,100% 58%,100% 100%);filter:blur(.2px)}
.water{bottom:0;height:24vh;background:linear-gradient(180deg,rgba(113,180,200,.34),rgba(77,145,172,.55)),repeating-linear-gradient(165deg,rgba(255,255,255,.48) 0 2px,transparent 2px 22px);animation:waterMove 16s linear infinite}
.city{bottom:18vh;height:24vh;width:220%;background:linear-gradient(180deg,transparent 0 20%,rgba(18,45,64,.08) 20% 100%),repeating-linear-gradient(90deg,rgba(24,57,78,.45) 0 30px,transparent 30px 44px,rgba(36,84,105,.35) 44px 76px,transparent 76px 88px,rgba(21,70,96,.42) 88px 120px,transparent 120px 140px);clip-path:polygon(0 100%,0 62%,2% 62%,2% 44%,5% 44%,5% 66%,7% 66%,7% 36%,10% 36%,10% 58%,13% 58%,13% 28%,16% 28%,16% 70%,19% 70%,19% 42%,23% 42%,23% 62%,27% 62%,27% 35%,31% 35%,31% 68%,35% 68%,35% 24%,39% 24%,39% 61%,43% 61%,43% 40%,47% 40%,47% 66%,51% 66%,51% 31%,55% 31%,55% 58%,60% 58%,60% 37%,64% 37%,64% 70%,70% 70%,70% 30%,75% 30%,75% 62%,82% 62%,82% 44%,88% 44%,88% 66%,94% 66%,94% 38%,100% 38%,100% 100%);animation:cityDrift 90s linear infinite;opacity:.78}
.city.front{bottom:15vh;height:18vh;opacity:.62;filter:blur(.3px);animation-duration:62s;background:repeating-linear-gradient(90deg,rgba(15,65,87,.58) 0 44px,transparent 44px 58px,rgba(33,91,112,.5) 58px 92px,transparent 92px 116px)}
.tower{position:absolute;right:18vw;bottom:23vh;width:44px;height:150px;background:linear-gradient(180deg,rgba(32,81,101,.52),rgba(22,56,74,.3));clip-path:polygon(22% 100%,38% 28%,16% 28%,16% 14%,84% 14%,84% 28%,62% 28%,78% 100%);box-shadow:0 0 34px rgba(54,118,145,.16)}.tower:before{content:"";position:absolute;left:-22px;top:0;width:88px;height:28px;border-radius:8px;background:linear-gradient(90deg,rgba(255,255,255,.62),rgba(77,132,151,.44));border:1px solid rgba(255,255,255,.5)}.tower:after{content:"";position:absolute;left:19px;top:-42px;width:6px;height:46px;background:rgba(37,93,115,.42);box-shadow:0 -12px 22px rgba(235,247,255,.9)}.runway{position:absolute;right:5vw;bottom:12vh;width:50vw;height:19vh;transform:perspective(620px) rotateX(62deg) skewX(-18deg);transform-origin:bottom right;background:linear-gradient(90deg,rgba(35,49,57,.3),rgba(29,39,47,.72)),repeating-linear-gradient(90deg,transparent 0 52px,rgba(255,255,255,.82) 52px 66px,transparent 66px 118px);border-left:2px solid rgba(255,255,255,.55);border-right:2px solid rgba(255,255,255,.38);box-shadow:0 -18px 80px rgba(20,65,90,.18)}
.plane-wrap{position:absolute;top:11vh;left:-48vw;width:min(760px,58vw);animation:planeCruise 30s linear infinite;animation-delay:-10s;filter:drop-shadow(0 20px 22px rgba(24,55,72,.22));will-change:transform}
.plane-svg{width:100%;height:auto;display:block}.plane-body{fill:url(#planeSkin);stroke:rgba(23,54,72,.38);stroke-width:2}.plane-wing{fill:#d8e9ef;stroke:rgba(23,54,72,.34);stroke-width:2}.plane-tail{fill:#c8dde6}.plane-window{fill:#183f55}.plane-text{font:700 29px Georgia,serif;fill:#155b49;letter-spacing:.4px}.seal-ring{fill:rgba(255,255,255,.78);stroke:#2b8a57;stroke-width:5}.engine{fill:#eef7fa;stroke:#35596a;stroke-width:2}.engine-dark{fill:#264b5c}.bird{position:absolute;width:38px;height:18px;opacity:.58;animation:birdFly 26s linear infinite}.bird:before,.bird:after{content:"";position:absolute;top:8px;width:18px;height:8px;border-top:2px solid rgba(38,78,94,.72);border-radius:50%}.bird:before{left:2px;transform:rotate(18deg)}.bird:after{right:2px;transform:rotate(-18deg)}.b1{top:14vh;left:-5vw;animation-delay:-4s}.b2{top:25vh;left:-12vw;animation-delay:-14s;transform:scale(.72)}.b3{top:18vh;left:-20vw;animation-delay:-22s;transform:scale(.55)}
header,main,dialog{position:relative;z-index:2}.panel,.kpi,.login,.upload{background:rgba(255,255,255,.9);backdrop-filter:blur(18px) saturate(1.15)}.dark .sky-scene{filter:brightness(.72) saturate(.82)}.dark .panel,.dark .kpi,.dark .login,.dark .upload{background:rgba(23,32,51,.9)}
.exec-summary{grid-column:1/-1}.summary-grid{display:grid;grid-template-columns:repeat(5,minmax(145px,1fr));gap:10px}.summary-item{border:1.5px solid rgba(25,100,180,.28);border-radius:16px;padding:12px;background:linear-gradient(180deg,rgba(255,255,255,.88),rgba(240,248,255,.72));box-shadow:0 2px 10px rgba(18,72,160,.1);color:#12304e!important}.summary-item b{display:block;color:var(--green);margin-bottom:5px}.summary-item span{color:#445566!important}.kpi{cursor:pointer;transition:transform .18s ease,box-shadow .18s ease}.kpi:hover{transform:translateY(-2px);box-shadow:0 14px 30px rgba(19,67,91,.14)}
@keyframes planeCruise{0%{transform:translate3d(-18vw,8vh,0) scale(.76) rotate(-4deg)}45%{transform:translate3d(58vw,-1vh,0) scale(.95) rotate(-1deg)}100%{transform:translate3d(155vw,-8vh,0) scale(1.08) rotate(-3deg)}}
@keyframes cityDrift{from{transform:translateX(0)}to{transform:translateX(-50%)}}@keyframes waterMove{from{background-position:0 0,0 0}to{background-position:0 0,260px 80px}}@keyframes birdFly{from{transform:translateX(0) translateY(0) scale(.8)}50%{transform:translateX(58vw) translateY(-4vh) scale(1)}to{transform:translateX(112vw) translateY(2vh) scale(.7)}}
@media(max-width:900px){.plane-wrap{width:86vw;animation-duration:24s}.summary-grid{grid-template-columns:1fr 1fr}.runway{width:76vw}.city{height:18vh}}
.tab-group-title{margin:8px 8px 6px;padding:10px 12px;border-radius:14px;background:linear-gradient(135deg,#173d67,#246f78);color:#fff;font-weight:800;letter-spacing:.08em;text-align:center;box-shadow:0 10px 24px rgba(25,70,100,.18)}
.tab.payments{background:linear-gradient(135deg,#e9fff6,#fff7df);border-color:#d5eadf;font-weight:800}.dark .tab.payments{background:linear-gradient(135deg,#163c34,#3a321b)}
.overview-note{font-size:12px;color:var(--muted);margin-top:8px}.mini-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}.module-intro{display:flex;justify-content:space-between;gap:12px;align-items:center;margin-bottom:10px}.module-intro .muted{max-width:760px}
@media(max-width:900px){.mini-grid{grid-template-columns:1fr}.tab-group-title{text-align:left}}


/* Cinematic aviation background inspired by provided reference videos */
.sky-scene{background:radial-gradient(circle at 78% 18%,rgba(255,177,79,.78),rgba(255,177,79,.12) 18%,transparent 37%),radial-gradient(circle at 18% 20%,rgba(80,137,190,.42),transparent 30%),linear-gradient(180deg,#07111f 0%,#12233b 38%,#2d3141 58%,#070b14 100%)!important}
.sky-scene:before{content:""!important;display:block!important;position:absolute;inset:-8%;background:radial-gradient(ellipse at 76% 29%,rgba(255,213,139,.48),transparent 25%),radial-gradient(ellipse at 46% 25%,rgba(255,255,255,.22),transparent 28%),radial-gradient(ellipse at 20% 45%,rgba(168,194,217,.18),transparent 31%),radial-gradient(ellipse at 70% 70%,rgba(248,164,79,.2),transparent 34%);filter:blur(18px);opacity:.95;animation:cloudCinema 34s ease-in-out infinite alternate}
.sky-scene:after{content:""!important;display:block!important;position:absolute;inset:0;background-image:radial-gradient(circle,rgba(255,218,149,.42) 0 1px,transparent 1.6px),radial-gradient(circle,rgba(134,194,232,.18) 0 1px,transparent 1.8px),linear-gradient(180deg,transparent 0%,rgba(0,0,0,.34) 100%);background-size:86px 86px,132px 132px,100% 100%;background-position:0 0,28px 48px,0 0;opacity:.72;mask-image:linear-gradient(180deg,transparent 0%,#000 46%,#000 100%);animation:citySpark 18s linear infinite}
.mountains{bottom:29vh;height:30vh;opacity:.26;background:linear-gradient(135deg,transparent 0 19%,rgba(101,137,151,.42) 19% 34%,transparent 34% 43%,rgba(72,105,126,.52) 43% 59%,transparent 59% 66%,rgba(107,139,145,.32) 66% 80%,transparent 80%)!important;filter:blur(1.5px)}
.water{bottom:0;height:23vh;background:linear-gradient(180deg,rgba(52,93,111,.2),rgba(8,19,32,.72)),repeating-linear-gradient(170deg,rgba(255,190,112,.12) 0 2px,transparent 2px 34px)!important;animation:waterMove 24s linear infinite!important;opacity:.7}
.city{bottom:8vh;height:44vh;width:260%;opacity:.9;transform-origin:bottom center;transform:perspective(820px) rotateX(58deg);clip-path:none!important;background:radial-gradient(circle,rgba(255,204,111,.92) 0 1.6px,transparent 2.8px),radial-gradient(circle,rgba(122,191,231,.62) 0 1px,transparent 2.4px),linear-gradient(90deg,transparent 0 48%,rgba(255,205,119,.38) 48% 50%,transparent 50% 100%),linear-gradient(0deg,transparent 0 47%,rgba(255,255,255,.08) 47% 50%,transparent 50% 100%)!important;background-size:42px 42px,76px 76px,180px 180px,150px 150px!important;background-position:0 0,22px 18px,0 0,0 0!important;animation:cityDrift 72s linear infinite!important;filter:drop-shadow(0 0 16px rgba(255,178,94,.22))}
.city.front{bottom:2vh;height:34vh;opacity:.82;animation-duration:48s!important;background:radial-gradient(circle,rgba(255,223,146,.95) 0 1.8px,transparent 3px),radial-gradient(circle,rgba(100,176,224,.42) 0 1px,transparent 2.4px),linear-gradient(90deg,transparent 0 46%,rgba(255,255,255,.13) 46% 50%,transparent 50% 100%)!important;background-size:34px 34px,66px 66px,128px 128px!important;filter:blur(.15px) drop-shadow(0 0 18px rgba(255,188,94,.18))}
.runway{left:31vw;right:auto;bottom:2vh;width:38vw;height:46vh;transform:perspective(780px) rotateX(68deg);transform-origin:bottom center;background:linear-gradient(90deg,transparent 0 43%,rgba(19,27,35,.86) 43% 57%,transparent 57% 100%),repeating-linear-gradient(0deg,transparent 0 34px,rgba(255,255,255,.7) 34px 48px,transparent 48px 90px),repeating-linear-gradient(90deg,transparent 0 32px,rgba(94,185,255,.5) 32px 38px,transparent 38px 72px)!important;border:0;box-shadow:0 0 90px rgba(62,159,223,.22),inset 0 0 60px rgba(255,207,128,.12);opacity:.74}
.tower{right:14vw;bottom:25vh;opacity:.62;filter:drop-shadow(0 0 18px rgba(255,199,110,.26))}.tower:after{box-shadow:0 -12px 30px rgba(255,214,139,.9)}
.bird{display:none}.plane-wrap{top:8vh;left:50%;width:min(900px,58vw);animation:planeCinematic 28s cubic-bezier(.45,0,.25,1) infinite!important;animation-delay:-8s!important;filter:drop-shadow(0 30px 30px rgba(0,0,0,.42)) drop-shadow(0 0 20px rgba(255,210,132,.18));transform-origin:center center}.plane-text{font:800 24px Georgia,serif!important;fill:#f8f3df!important;letter-spacing:.6px}.plane-body{fill:url(#planeSkin)!important;stroke:rgba(255,255,255,.38)!important}.plane-wing{fill:#cddde8!important;stroke:rgba(255,255,255,.26)!important}.plane-tail{fill:#b8cedb!important}.engine{fill:#dceaf1!important;stroke:#516675!important}.engine-dark{fill:#07131d!important}.seal-ring{fill:rgba(255,255,255,.82)!important;stroke:#2e9b62!important}.plane-window{fill:#0b1d2d!important}
.panel,.kpi,.login,.upload{background:rgba(255,255,255,.88)!important;box-shadow:0 18px 44px rgba(2,12,24,.12)}.dark .panel,.dark .kpi,.dark .login,.dark .upload{background:rgba(18,27,43,.9)!important}.tab-group-title{box-shadow:0 12px 30px rgba(0,0,0,.22)}
@keyframes planeCinematic{0%{transform:translate3d(-78vw,-6vh,0) scale(.34) rotate(-9deg);opacity:0}8%{opacity:.96}42%{transform:translate3d(-12vw,3vh,0) scale(.78) rotate(-2deg);opacity:1}68%{transform:translate3d(10vw,7vh,0) scale(1.08) rotate(1deg);opacity:1}100%{transform:translate3d(92vw,-1vh,0) scale(.48) rotate(7deg);opacity:0}}
@keyframes cloudCinema{from{transform:translate3d(-2%,1%,0) scale(1)}to{transform:translate3d(3%,-2%,0) scale(1.08)}}@keyframes citySpark{from{background-position:0 0,28px 48px,0 0}to{background-position:86px 86px,160px 180px,0 0}}


/* Cinematic polish: glass header and better plane path */
header{background:linear-gradient(180deg,rgba(255,255,255,.82),rgba(255,255,255,.54))!important;backdrop-filter:blur(18px) saturate(1.1);border-bottom:1px solid rgba(255,255,255,.42);box-shadow:0 12px 35px rgba(4,15,31,.08)}
.dark header{background:linear-gradient(180deg,rgba(15,24,40,.84),rgba(15,24,40,.56))!important}
@keyframes planeCinematic{0%{transform:translate3d(-84vw,-7vh,0) scale(.36) rotate(-8deg);opacity:0}10%{opacity:.95}36%{transform:translate3d(-22vw,0vh,0) scale(.72) rotate(-3deg);opacity:1}58%{transform:translate3d(-4vw,6vh,0) scale(1.08) rotate(0deg);opacity:1}78%{transform:translate3d(18vw,3vh,0) scale(.9) rotate(3deg);opacity:.95}100%{transform:translate3d(82vw,-6vh,0) scale(.42) rotate(8deg);opacity:0}}
@media(min-width:1000px){.plane-wrap{width:min(960px,62vw)}}


/* Final professional video-like background asset */
.sky-scene{background-image:linear-gradient(90deg,rgba(4,10,20,.58),rgba(4,10,20,.22) 38%,rgba(4,10,20,.48)),linear-gradient(180deg,rgba(5,11,22,.18),rgba(5,11,22,.34)),url('/assets/ibk-cinematic-airport.png')!important;background-size:cover,cover,cover!important;background-position:center center!important;animation:cinemaBgDrift 38s ease-in-out infinite alternate!important;transform-origin:center center}
.sky-scene:before{content:""!important;display:block!important;position:absolute;inset:0;background:radial-gradient(circle at 70% 34%,rgba(255,190,95,.16),transparent 23%),linear-gradient(90deg,rgba(0,0,0,.25),transparent 44%,rgba(0,0,0,.2));filter:none!important;opacity:1!important;animation:cinemaLightBreath 9s ease-in-out infinite alternate!important}
.sky-scene:after{content:""!important;display:block!important;position:absolute;inset:0;background:linear-gradient(180deg,rgba(255,255,255,.04),rgba(0,0,0,.1)),radial-gradient(circle at 50% 72%,rgba(255,211,131,.18),transparent 18%);background-size:100% 100%!important;opacity:1!important;mask-image:none!important;animation:none!important}
.sky-scene .mountains,.sky-scene .water,.sky-scene .city,.sky-scene .runway,.sky-scene .tower,.sky-scene .bird,.sky-scene .plane-wrap{display:none!important}
header{background:linear-gradient(180deg,rgba(255,255,255,.78),rgba(255,255,255,.48))!important}.panel,.kpi,.login,.upload{background:rgba(255,255,255,.86)!important;backdrop-filter:blur(20px) saturate(1.1)!important}.tab{background:rgba(245,248,253,.88)!important}.tab.active{background:#254f86!important}.tab-group-title{background:linear-gradient(135deg,rgba(18,54,88,.9),rgba(28,105,119,.86))!important}
@keyframes cinemaBgDrift{0%{transform:scale(1.02) translate3d(-.8%,.4%,0)}100%{transform:scale(1.075) translate3d(.9%,-.6%,0)}}@keyframes cinemaLightBreath{0%{opacity:.72}100%{opacity:1}}
@media(max-width:900px){.sky-scene{background-position:54% center!important}.panel,.kpi,.login,.upload{background:rgba(255,255,255,.9)!important}}


/* Visible premium background motion layers */
.sky-scene{animation:cinemaBgDrift 22s ease-in-out infinite alternate!important;will-change:transform,background-position}
.cinema-clouds,.cinema-runway,.cinema-glow,.cinema-vignette{position:absolute;inset:0;pointer-events:none}
.cinema-clouds{inset:-12%;background:radial-gradient(ellipse at 12% 28%,rgba(203,219,234,.18),transparent 24%),radial-gradient(ellipse at 78% 22%,rgba(255,191,105,.14),transparent 28%),radial-gradient(ellipse at 48% 48%,rgba(255,255,255,.08),transparent 32%);filter:blur(20px);mix-blend-mode:screen;animation:cloudLayerMove 18s ease-in-out infinite alternate;opacity:.95}
.cinema-runway{background:repeating-linear-gradient(180deg,transparent 0 34px,rgba(255,224,151,.72) 35px 37px,transparent 38px 78px),repeating-linear-gradient(90deg,transparent 0 42px,rgba(90,190,255,.42) 43px 45px,transparent 46px 92px);clip-path:polygon(42% 58%,58% 58%,82% 100%,18% 100%);filter:blur(.35px) drop-shadow(0 0 12px rgba(255,201,111,.42));opacity:.46;animation:runwayLightFlow 2.8s linear infinite;mix-blend-mode:screen}
.cinema-glow{background:radial-gradient(circle at 50% 38%,rgba(255,205,129,.22),transparent 18%),radial-gradient(circle at 52% 62%,rgba(255,238,174,.18),transparent 16%);mix-blend-mode:screen;animation:planeAuraPulse 4.8s ease-in-out infinite alternate;opacity:.72}
.cinema-vignette{background:linear-gradient(90deg,rgba(2,7,16,.44),transparent 35%,rgba(2,7,16,.38)),linear-gradient(180deg,rgba(0,0,0,.08),rgba(0,0,0,.46));opacity:.82}
@keyframes cloudLayerMove{0%{transform:translate3d(-2.5%,1.2%,0) scale(1.02)}100%{transform:translate3d(2.2%,-1.6%,0) scale(1.08)}}
@keyframes runwayLightFlow{0%{background-position:0 0,0 0;opacity:.34}50%{opacity:.58}100%{background-position:0 78px,92px 0;opacity:.34}}
@keyframes planeAuraPulse{0%{transform:scale(.96);opacity:.46}100%{transform:scale(1.08);opacity:.9}}
@keyframes cinemaBgDrift{0%{transform:scale(1.03) translate3d(-1.4%,.6%,0)}50%{transform:scale(1.065) translate3d(.2%,-.3%,0)}100%{transform:scale(1.095) translate3d(1.1%,-.9%,0)}}


/* Real video background layer */
.bg-video{position:absolute;inset:0;width:100%;height:100%;object-fit:cover;z-index:0;filter:saturate(1.08) contrast(1.04) brightness(.9)}
#bgCanvas{display:none}.cinema-clouds,.cinema-runway,.cinema-glow,.cinema-vignette{z-index:1}.sky-scene:before,.sky-scene:after{z-index:2}.sky-scene{background-image:linear-gradient(90deg,rgba(4,10,20,.58),rgba(4,10,20,.22) 38%,rgba(4,10,20,.48)),url('/assets/ibk-cinematic-airport.png')!important}


/* Lighter video background and expandable module navigation */
.bg-video{filter:saturate(.96) contrast(.98) brightness(1.18)!important;opacity:.92}.cinema-vignette{opacity:.28!important;background:linear-gradient(90deg,rgba(255,255,255,.10),transparent 45%,rgba(255,255,255,.04)),linear-gradient(180deg,rgba(255,255,255,.12),rgba(255,255,255,.02))!important}.cinema-clouds{opacity:.62!important}.cinema-runway{opacity:.28!important}.cinema-glow{opacity:.5!important}.sky-scene:before{background:radial-gradient(circle at 70% 34%,rgba(255,230,180,.24),transparent 24%),linear-gradient(90deg,rgba(255,255,255,.10),transparent 44%,rgba(255,255,255,.08))!important}.panel,.kpi,.login,.upload{background:rgba(255,255,255,.91)!important}.summary-item{background:linear-gradient(180deg,rgba(255,255,255,.86),rgba(255,255,255,.62))!important}.module-parent{margin:6px 8px;padding:12px 14px;border:0;border-radius:16px;background:linear-gradient(135deg,rgba(26,73,121,.94),rgba(40,129,139,.9));color:white;font-weight:850;text-align:left;box-shadow:0 12px 28px rgba(27,78,118,.2);cursor:pointer}.module-parent.pay{background:linear-gradient(135deg,rgba(36,124,96,.94),rgba(210,151,46,.86))}.module-parent.active{outline:2px solid rgba(255,255,255,.85)}.subtabs{padding:4px 0 10px}.tab.sub{margin-left:18px;width:calc(100% - 28px);font-size:13px}.pay-kpis{display:grid;grid-template-columns:repeat(4,minmax(160px,1fr));gap:12px;margin-bottom:14px}.pay-card{padding:14px;border:1px solid var(--border);border-radius:16px;background:rgba(255,255,255,.78)}.pay-card b{display:block;font-size:20px;margin-top:6px}.module-grid{display:grid;grid-template-columns:1.2fr .8fr;gap:14px}@media(max-width:900px){.pay-kpis,.module-grid{grid-template-columns:1fr}.tab.sub{margin-left:10px;width:calc(100% - 14px)}}
.excel-wrap{overflow:auto;border:1px solid #9fb9ce;border-radius:10px;background:white}.excel-table{width:100%;border-collapse:collapse;font-size:13px;background:white}.excel-table th{background:#1f4e78!important;color:white!important;border:1px solid #6f8da8;padding:7px 8px;text-align:center;vertical-align:middle;white-space:normal}.excel-table td{border:1px solid #b7c8d8;padding:6px 8px;vertical-align:middle;background:#fff}.excel-table td.num{text-align:center;font-variant-numeric:tabular-nums}.excel-table td.text{text-align:left}.excel-table tr.total td{background:#e2f0d9!important;font-weight:800}.excel-table tr:nth-child(even):not(.total) td{background:#f7fbff}.excel-table .download-cell{text-align:center}.excel-download{display:inline-flex;align-items:center;justify-content:center;min-width:74px;padding:6px 10px;border-radius:8px;background:#eef6ff;border:1px solid #9fc2e0;color:#164a73;font-weight:800;text-decoration:none}.excel-download:hover{background:#dff0ff}.excel-actions{display:flex;gap:8px;flex-wrap:wrap;margin:8px 0 10px}.excel-title{display:flex;justify-content:space-between;align-items:center;gap:10px}.dark .excel-table td{background:#182538;color:#eaf1f7;border-color:#36546d}.dark .excel-table tr:nth-child(even):not(.total) td{background:#1f3046}.dark .excel-table tr.total td{background:#253f32!important}


/* AeroInfo-style light aviation portal theme */
.sky-scene{background:linear-gradient(180deg,#eef8ff 0%,#f8fcff 44%,#eaf5fb 100%)!important;animation:none!important;transform:none!important}.bg-video{filter:none!important;opacity:1!important}.cinema-clouds,.cinema-runway,.cinema-glow,.cinema-vignette,.sky-scene:before,.sky-scene:after{display:none!important}header{background:rgba(255,255,255,.86)!important;border-bottom:1px solid rgba(17,90,142,.12)!important;box-shadow:0 8px 28px rgba(20,82,128,.08)!important}.panel,.kpi,.login,.upload{background:rgba(255,255,255,.88)!important;border:1px solid rgba(27,98,157,.13)!important;box-shadow:0 14px 34px rgba(38,98,148,.10)!important;backdrop-filter:blur(14px) saturate(1.04)!important}.summary-item{background:linear-gradient(180deg,rgba(255,255,255,.92),rgba(244,249,253,.78))!important;border-color:rgba(29,112,170,.13)!important}.tab{background:rgba(255,255,255,.84)!important;border:1px solid rgba(30,102,165,.13)!important;color:#14304c}.tab.active{background:linear-gradient(135deg,#2275b8,#2aa6b4)!important;color:white!important}.module-parent{background:linear-gradient(135deg,#1e76b7,#35b7c4)!important;box-shadow:0 12px 26px rgba(37,119,177,.18)!important}.module-parent.pay{background:linear-gradient(135deg,#2a9d8f,#7ccfbd)!important}.bar span{background:linear-gradient(90deg,#2f86c8,#4ecdc4)!important}.kpi b,h1,h2{color:#102a43}.muted{color:#5d7082!important}body{background:#eef8ff!important}
header{display:flex!important;align-items:center!important;justify-content:space-between!important;gap:12px!important}header>div:first-child{flex:0 0 auto!important;min-width:165px!important}h1{white-space:nowrap!important}#meta{max-width:280px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}#actions{margin-left:auto!important;display:flex!important;align-items:center!important;justify-content:flex-end!important;gap:8px!important;flex-wrap:wrap!important}.fixed-table{table-layout:fixed!important;min-width:1080px}.fixed-table th,.fixed-table td{overflow:hidden;text-overflow:ellipsis}.expired-total-table{table-layout:fixed!important;min-width:1180px;font-size:12px}.expired-total-table th,.expired-total-table td{text-align:center;padding:5px 6px}.expired-total-table th{white-space:normal!important;word-break:break-word!important;overflow-wrap:anywhere!important;line-height:1.12!important;height:48px!important;vertical-align:middle!important}.expired-total-table td.text{text-align:left}.admin-layout{display:flex;flex-direction:column;gap:14px}.admin-card{border:1px solid var(--line);border-radius:12px;padding:16px;background:rgba(255,255,255,.72);box-shadow:0 10px 24px rgba(23,43,77,.06)}.admin-form{display:grid;grid-template-columns:110px 1fr;gap:7px 12px;align-items:center}.admin-form input[type=hidden]{display:none}.admin-form .perm-grid,.admin-form .excel-actions{grid-column:1/-1}.admin-form>label{font-weight:600;color:#3a5a7a;font-size:13px;text-align:right;padding-right:4px}.admin-form input,.admin-form select{padding:7px 10px;border:1px solid var(--line);border-radius:8px;font-size:13px;background:var(--panel);color:var(--ink)}.perm-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;margin-top:2px}.perm-grid label{font-weight:600;color:var(--ink);background:#eef6ff;border:1px solid #c8dded;border-radius:8px;padding:8px}.role-pill{display:inline-block;border-radius:999px;padding:3px 9px;background:#e8f3ff;color:#174a7c;font-weight:800}.live-chart{position:relative;min-height:210px;padding:10px}.sparkline{height:150px;border-left:1px solid var(--line);border-bottom:1px solid var(--line);background:linear-gradient(180deg,rgba(34,117,184,.08),rgba(42,166,180,.02));display:flex;align-items:flex-end;gap:8px;padding:12px}.sparkline i{display:block;flex:1;border-radius:6px 6px 0 0;background:linear-gradient(180deg,#2275b8,#4ecdc4);animation:barRise .8s ease both}.sparkline i:nth-child(2n){background:linear-gradient(180deg,#1b9e77,#8bd3c7)}@keyframes barRise{from{height:0;opacity:.3}to{opacity:1}}@media(max-width:1000px){.admin-layout{grid-template-columns:1fr}}
.dark{--ink:#eaf2ff!important;--muted:#b7c6d8!important;--line:#3b5168!important;--blue:#3f8fd5!important;--panel:#17283c!important}.dark body,.dark{background:#0d1b2a!important;color:#eaf2ff!important}.dark header,.dark .panel,.dark .kpi,.dark .login,.dark .upload,.dark .admin-card,.dark .summary-item,.dark .pay-card{background:rgba(20,35,54,.94)!important;color:#eaf2ff!important;border-color:#39536d!important}.dark td{background:#142337!important;color:#edf6ff!important;border-color:#34506b!important}.dark tr:nth-child(even) td{background:#192b42!important}.dark th{background:#255d92!important;color:white!important}.dark input,.dark select{background:#0f1d2d!important;color:#fff!important;border-color:#4d6b86!important}.dark .muted{color:#bed1e5!important}.dark .tab{background:#16283d!important;color:#eaf2ff!important}.dark .tab.active,.dark .module-parent.active{box-shadow:0 0 0 2px #84d4ff inset!important}.dark .bar{background:#203650!important}.dark .excel-table td{background:#142337!important;color:#edf6ff!important}.dark .excel-download{background:#1d3a58;color:#fff;border-color:#4d80aa}.login{max-width:430px;margin:9vh auto!important;text-align:center;border-radius:24px!important;padding:24px!important;box-shadow:0 22px 60px rgba(22,83,132,.18)!important}.login:before{display:none!important;content:none!important}@keyframes sealSpin{from{rotate:0deg}to{rotate:0deg}}@keyframes sealFloat{0%,100%{translate:0 0}50%{translate:0 -7px}}.sky-scene{background:linear-gradient(180deg,rgba(229,246,255,.92),rgba(247,252,255,.86)),radial-gradient(circle at 72% 22%,rgba(77,184,225,.24),transparent 26%),linear-gradient(135deg,#eaf8ff,#f8fdff 50%,#dff2fb)!important}.sky-scene:after{display:block!important;content:"";position:absolute;inset:0;background:linear-gradient(10deg,transparent 60%,rgba(40,144,198,.08) 61%,transparent 62%),repeating-linear-gradient(115deg,transparent 0 90px,rgba(40,144,198,.055) 92px,transparent 96px);animation:bgPlaneDrift 14s linear infinite}.bg-video{opacity:.28!important;filter:saturate(.75) brightness(1.25)!important}.plane-wrap{opacity:.12!important}.module-parent{font-size:15px!important;letter-spacing:.2px}.tab.sub{border-radius:10px!important;background:rgba(255,255,255,.72)!important;color:#12304e!important}.tab.sub.active{background:linear-gradient(135deg,#1e76b7,#35b7c4)!important;color:white!important;font-weight:900!important}.dark .tab.sub{background:#1a314a!important;color:#edf7ff!important;border-color:#4c6c88!important}.dark .tab.sub.active{background:linear-gradient(135deg,#2d84d0,#38bac8)!important;color:#fff!important}.dark .module-parent,.dark .module-parent.pay{color:#fff!important}.dark h1,.dark h2,.dark h3,.dark b,.dark label,.dark .summary-item span{color:#f4fbff!important}.module-parent:hover,.tab:hover,button:hover,.btn:hover,tr:hover td,.archive-card:hover,.summary-item:hover{transform:translateY(-2px)!important;transition:transform .18s ease,box-shadow .18s ease,background .18s ease;box-shadow:0 12px 28px rgba(35,114,178,.16)!important}tr:hover td{background:#eaf6ff!important}.dark tr:hover td{background:#24435f!important}.viz-card{display:grid;grid-template-columns:170px 1fr;gap:16px;align-items:center}.donut{width:150px;height:150px;border-radius:50%;background:conic-gradient(#2477bd var(--p),#d8edf8 0);display:grid;place-items:center;animation:donutPulse 1.5s ease-in-out infinite}.donut:after{content:attr(data-label);display:grid;place-items:center;width:92px;height:92px;border-radius:50%;background:rgba(255,255,255,.94);font-weight:900;color:#17507b;text-align:center}.trend-svg{width:100%;height:160px}.trend-line{fill:none;stroke:#1f8ec1;stroke-width:4;stroke-linecap:round;stroke-dasharray:900;animation:drawLine 1.8s ease-in-out infinite alternate}.trend-area{fill:url(#trendGrad);opacity:.36}.trend-dot{fill:#1b9e77;animation:dotPulse 1.1s ease-in-out infinite}.dark .donut:after{background:#142337;color:#dff6ff}@keyframes donutPulse{0%,100%{filter:saturate(1)}50%{filter:saturate(1.5) brightness(1.05)}}@keyframes drawLine{from{stroke-dashoffset:900}to{stroke-dashoffset:0}}@keyframes dotPulse{0%,100%{r:4}50%{r:7}}@keyframes bgPlaneDrift{from{background-position:0 0,0 0}to{background-position:240px 0,320px 0}}

.flow-map{min-height:310px;border:1px solid rgba(42,116,176,.16);border-radius:18px;background:linear-gradient(180deg,rgba(239,249,255,.92),rgba(255,255,255,.78));padding:10px;overflow:hidden}.flow-map svg{width:100%;height:310px}.flow-map rect{fill:#f4fbff}.flow-map .land{fill:#dff1f8;stroke:#97c6dd;stroke-width:1}.flow-map .land.small{fill:#edf8fb}.flow-map .grid-map path{stroke:rgba(56,131,184,.13);stroke-width:1}.flow-map .route path{fill:none;stroke:url(#routeG);stroke-linecap:round;opacity:.58;stroke-dasharray:520;animation:routeDraw 2.2s ease-in-out infinite alternate}.flow-map .route circle{fill:#2bb7c5;stroke:white;stroke-width:2;filter:drop-shadow(0 3px 7px rgba(20,101,151,.25))}.flow-map text{font-size:9px;fill:#12304e;font-weight:800}.flow-map .uz circle{fill:#1f72b8;stroke:white;stroke-width:3;animation:dotPulse 1.2s ease-in-out infinite}.flow-map .uz text{font-size:12px;fill:#0d335b}.dark .flow-map{background:#142337;border-color:#3b5975}.dark .flow-map rect{fill:#142337}.dark .flow-map .land{fill:#1e3853;stroke:#517da2}.dark .flow-map text{fill:#eaf6ff}.dark .flow-map .grid-map path{stroke:rgba(172,214,245,.16)}@keyframes routeDraw{from{stroke-dashoffset:520;opacity:.34}to{stroke-dashoffset:0;opacity:.82}}
.globe-map{min-height:390px!important;padding:16px!important;background:radial-gradient(circle at 50% 28%,rgba(255,255,255,.98),rgba(226,247,255,.86) 46%,rgba(214,237,249,.72))!important}.globe-map svg{height:380px!important;display:block;margin:auto;max-width:960px}.globe-map rect{fill:rgba(247,253,255,.72)!important}.globe-shadow{fill:rgba(55,133,184,.18);filter:blur(1px)}.globe-sea{fill:url(#globeSea);stroke:rgba(255,255,255,.88);stroke-width:2.2;filter:drop-shadow(0 24px 34px rgba(43,116,170,.22))}.globe-lines ellipse{fill:none;stroke:rgba(255,255,255,.46);stroke-width:1.15}.continent{fill:rgba(255,255,255,.54);stroke:rgba(79,143,184,.42);stroke-width:1}.globe-map .route path{stroke:url(#routeG);stroke-linecap:round;opacity:.68;stroke-dasharray:520;animation:routeDraw 1.7s ease-in-out infinite alternate}.globe-map .route circle{fill:#37c9d3;stroke:#fff;stroke-width:2.2}.globe-map .route text{font-size:8.3px;fill:#153b5f;font-weight:850;text-shadow:0 1px 0 rgba(255,255,255,.75)}.globe-map .uz circle{fill:#1769aa;stroke:#fff;stroke-width:3.2;animation:dotPulse 1.05s ease-in-out infinite}.globe-map .uz text{font-size:12px;fill:#0c3154;font-weight:900;text-shadow:0 1px 0 rgba(255,255,255,.8)}.chart-under-globe{margin-top:14px;padding-top:14px;border-top:1px solid rgba(42,116,176,.16)}.dark .globe-map{background:radial-gradient(circle at 50% 26%,rgba(42,77,105,.92),rgba(18,39,62,.88) 50%,rgba(12,28,45,.86))!important}.dark .globe-map rect{fill:rgba(15,32,51,.72)!important}.dark .globe-shadow{fill:rgba(54,176,223,.14)}.dark .globe-sea{stroke:rgba(194,237,255,.45);filter:drop-shadow(0 24px 36px rgba(0,0,0,.32))}.dark .globe-lines ellipse{stroke:rgba(230,249,255,.28)}.dark .continent{fill:rgba(224,247,255,.26);stroke:rgba(180,229,252,.28)}.dark .globe-map .route text,.dark .globe-map .uz text{fill:#e9f8ff;text-shadow:0 1px 4px rgba(0,0,0,.55)}.dark .chart-under-globe{border-top-color:#375977}
.sample-release-table{min-width:1620px!important}.sample-release-table th{white-space:normal!important;word-break:break-word!important;overflow-wrap:anywhere!important;line-height:1.08!important;height:58px!important;padding:5px 4px!important;vertical-align:middle!important}.sample-release-table td{vertical-align:middle!important}.sample-release-table td:nth-child(2){text-align:left!important}.sample-release-table tr:first-child td{background:#e2f0d9!important;font-weight:850!important}.dark .sample-release-table tr:first-child td{background:#253f32!important}
.compact-archive{min-width:720px!important;table-layout:fixed!important}.compact-archive th,.compact-archive td{padding:6px 8px!important}.compact-archive td:nth-child(2),.compact-archive td:nth-child(3){white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.compact-archive button{padding:6px 10px!important}
/* Final UX polish */
body{background:#eef8ff!important}.sky-scene{background:linear-gradient(180deg,rgba(236,248,255,.95),rgba(248,253,255,.9)),radial-gradient(circle at 78% 18%,rgba(65,170,220,.18),transparent 28%),linear-gradient(120deg,#e7f6ff,#ffffff 48%,#dff1fb)!important}.sky-scene:before{display:block!important;content:""!important;position:absolute;inset:0;background:repeating-linear-gradient(110deg,transparent 0 78px,rgba(34,126,190,.055) 80px,transparent 84px),linear-gradient(15deg,transparent 58%,rgba(33,128,194,.08) 59%,transparent 61%)!important;animation:bgPlaneDrift 18s linear infinite!important}.sky-scene:after{display:block!important;content:""!important;position:absolute;inset:0;background:radial-gradient(ellipse at 50% 115%,rgba(31,91,145,.16),transparent 42%),linear-gradient(180deg,transparent 0 62%,rgba(65,150,201,.08) 63%,transparent 64%)!important;animation:none!important}
.login{position:relative!important;max-width:min(760px,94vw)!important;min-height:72vh!important;margin:5vh auto!important;display:grid!important;place-items:center!important;background:transparent!important;border:0!important;box-shadow:none!important;overflow:visible!important}.login:before{display:none!important}.login-seal-wrap{position:absolute;inset:0;display:grid;place-items:center;cursor:pointer;transition:opacity .9s ease,transform .9s ease}.login-seal{width:min(560px,82vw);height:auto;animation:sealFloatCalm 4.8s ease-in-out infinite;transform-origin:center center;mix-blend-mode:multiply}.login.active .login-seal-wrap{opacity:0;transform:scale(1.1);pointer-events:none}.login-box{position:relative;z-index:2;width:min(420px,92vw);padding:26px;border-radius:22px;background:rgba(255,255,255,.88);border:1px solid rgba(76,145,196,.18);box-shadow:0 24px 70px rgba(28,87,141,.18);backdrop-filter:blur(18px) saturate(1.08);opacity:0;transform:translateY(18px) scale(.96);pointer-events:none;transition:opacity .55s ease,transform .55s ease}.login.active .login-box{opacity:1;transform:translateY(0) scale(1);pointer-events:auto}.login-form-stack{display:grid!important;grid-template-columns:1fr!important;gap:12px!important}.pass-wrap{display:flex;gap:6px;align-items:center}.pass-wrap input{flex:1}.eye-btn{width:46px;padding:9px!important;border-radius:9px!important;background:#eaf3fb!important;color:#143b61!important}.designer-line{font-weight:800;color:#1f5f93}.header-clock{font-weight:800;color:#1d6a9f}.busy-spinner{display:inline-block;width:14px;height:14px;margin-left:8px;border:2px solid rgba(255,255,255,.45);border-top-color:#fff;border-radius:50%;vertical-align:-2px;animation:spinBusy .75s linear infinite}.btn.light .busy-spinner,button.light .busy-spinner{border-color:rgba(31,80,120,.25);border-top-color:#1f5f93}.is-busy{opacity:.82;pointer-events:none}#actions .btn,#actions button{min-height:36px}.dark #actions button.light,.dark #actions .btn.light{background:#284968!important;color:#f6fbff!important;border:1px solid #5e86a8!important}.dark #actions .btn,.dark #actions button{color:#fff!important}.dark .login-box{background:rgba(16,31,49,.92)!important;color:#eef8ff!important;border-color:#416684!important}.dark .eye-btn{background:#203b57!important;color:#fff!important}
table{table-layout:fixed!important}th{white-space:normal!important;word-break:break-word!important;overflow-wrap:anywhere!important;text-align:center!important;vertical-align:middle!important;line-height:1.12!important;min-height:42px!important}td{height:34px!important;max-height:34px!important;vertical-align:middle!important}td.text{white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;word-break:normal!important}td.num{text-align:center!important}.fixed-table th,.fixed-table td{overflow:hidden!important}.excel-table th{white-space:normal!important;word-break:break-word!important;overflow-wrap:anywhere!important}.excel-table td.text{white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.merged-row td,.expired-total-table td:first-child{white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}
.transport-table{min-width:1010px!important}.transport-table th{height:52px!important;padding:5px 4px!important}.transport-table td{height:34px!important;padding:4px 6px!important}.transport-table td:nth-child(2){text-align:left!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.transport-viz{display:grid;grid-template-columns:360px 1fr;gap:18px}.ring-grid{display:grid;grid-template-columns:repeat(2,minmax(140px,1fr));gap:12px}.transport-ring{position:relative;min-height:150px;border-radius:18px;background:linear-gradient(180deg,rgba(255,255,255,.92),rgba(236,248,255,.86));border:1px solid rgba(50,123,180,.16);display:grid;place-items:center;text-align:center;cursor:pointer;overflow:hidden;box-shadow:0 10px 24px rgba(30,96,150,.09);transition:.2s}.transport-ring:before{content:"";position:absolute;width:104px;height:104px;border-radius:50%;background:conic-gradient(#1e79bd calc(var(--p)*1%),#d9eef8 0);animation:ringLoad .9s ease both;animation-delay:var(--delay)}.transport-ring:after{content:"";position:absolute;width:68px;height:68px;border-radius:50%;background:rgba(255,255,255,.95)}.transport-ring b,.transport-ring span,.transport-ring small{position:relative;z-index:1}.transport-ring b{align-self:end;font-size:13px;color:#143b61}.transport-ring span{font-weight:900;font-size:22px;color:#0d6da6}.transport-ring small{align-self:start;color:#64748b;font-size:11px;max-width:130px}.transport-ring:hover{transform:translateY(-4px) scale(1.015);box-shadow:0 16px 36px rgba(31,111,178,.18)}.flow-list{display:grid;gap:8px}.flow-row{display:grid;grid-template-columns:255px 1fr 92px;gap:10px;align-items:center;padding:8px 10px;border:1px solid rgba(37,108,166,.13);border-radius:12px;background:rgba(255,255,255,.72);cursor:pointer;transition:.18s}.flow-row:hover{transform:translateX(4px);box-shadow:0 10px 24px rgba(32,103,163,.13)}.flow-name{min-width:0;display:grid;grid-template-columns:58px 1fr;gap:8px;align-items:center}.flow-name span{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.flow-track{height:24px;border-radius:999px;background:#e7f1f8;position:relative;overflow:hidden}.flow-track i{display:block;height:100%;border-radius:999px;background:linear-gradient(90deg,#32c6cf,#1d72b8);animation:flowGrow .75s ease both}.flow-track em{position:absolute;inset:0;display:grid;place-items:center;color:#fff;font-style:normal;font-weight:900;font-size:12px;text-shadow:0 1px 2px rgba(0,0,0,.32)}.flow-num{text-align:right;font-weight:800;color:#174a7c;white-space:nowrap}.flow-unit{font-size:10px;font-weight:500;color:#6b8ba8;margin-left:2px}.dark .transport-ring,.dark .flow-row{background:rgba(20,35,54,.88);border-color:#39536d}.dark .transport-ring:after{background:#142337}.dark .transport-ring b,.dark .flow-name b,.dark .flow-num{color:#eaf6ff}.dark .transport-ring small{color:#b7c6d8}.dark .flow-track{background:#263f59}.dark .flow-name span{color:#eaf6ff}@keyframes ringLoad{from{transform:scale(.84);opacity:.2}to{transform:scale(1);opacity:1}}@keyframes flowGrow{from{width:0}to{}}@media(max-width:1000px){.transport-viz{grid-template-columns:1fr}.flow-row{grid-template-columns:210px 1fr 82px}.ring-grid{grid-template-columns:repeat(2,1fr)}}
.sample-like{border-collapse:collapse!important;table-layout:fixed!important;font-family:"Segoe UI",Arial,sans-serif}.sample-like th{background:#b6dde8!important;color:#111!important;border:1px solid #333!important;font-size:12px!important;font-weight:800!important;white-space:normal!important;line-height:1.08!important;text-align:center!important;vertical-align:middle!important;padding:4px!important}.sample-like td{border:1px solid #333!important;height:24px!important;max-height:24px!important;font-size:12px!important;padding:3px 5px!important;vertical-align:middle!important}.sample-like tbody tr:first-child td,.sample-like .grand-total td{background:#eaf1dd!important;font-weight:800!important}.sample-like .merged-row td{background:#d9eaf7!important;font-weight:800!important;text-align:center!important}.sample-like .sub-total td{background:#f2f7fb!important;font-weight:800!important}.sample-like td.text,.sample-like .col-korxona,.sample-like .col-company,.sample-like .col-goods,.sample-like .col-tovar,.sample-like .col-reason{text-align:left!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.sample-like .col-reason,td.col-reason{white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.expired-sample{min-width:1120px!important}.expired-sample col:nth-child(1),.expired-sample td:nth-child(1),.expired-sample th:nth-child(1){width:330px!important}.food-sample{min-width:820px!important}.food-sample col:nth-child(1){width:56px!important}.food-sample col:nth-child(2){width:340px!important}.regime-year-table{max-width:920px!important;min-width:820px!important}.col-korxona,.col-company{min-width:260px;text-align:left!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.col-goods,.col-tovar{white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.col-reason{white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}body.login-screen header{display:flex!important}body.login-screen #actions{display:none!important}body.login-screen #app,body.login-screen #dash,body.login-screen #tabs,body.login-screen #view,body.login-screen #kpis,.login-screen .workspace,.login-screen .tabs{display:none!important}body.login-screen main{max-width:100%!important}.login-screen .plane-wrap,.login-screen .bird,.login-screen .city,.login-screen .mountains,.login-screen .water,.login-screen .runway,.login-screen .tower,.login-screen .cinema-clouds,.login-screen .cinema-runway,.login-screen .cinema-glow,.login-screen .cinema-vignette{display:none!important}.login-screen .bg-video{display:block!important;opacity:.56!important;filter:brightness(1.18) saturate(.82) contrast(.96)!important}.login-screen #bgCanvas{display:block!important}.login-screen .sky-scene{background:radial-gradient(circle at 50% 32%,rgba(255,255,255,.96),rgba(228,245,255,.84) 42%,rgba(207,232,248,.72))!important}.login-screen .sky-scene:before,.login-screen .sky-scene:after{display:none!important}.login-seal{mix-blend-mode:multiply!important;filter:none!important}.eye-btn{font-size:18px!important;line-height:1!important}
dialog{width:min(1760px,98.5vw)!important;max-width:98.5vw!important}dialog .body{max-height:82vh!important}.details-wide{min-width:1580px!important}.real-globe{min-height:430px!important;padding:12px!important}
.flow-legend{display:flex;flex-wrap:wrap;gap:14px;justify-content:center;align-items:center;margin-top:10px;font-size:12px;color:#557086;font-weight:700}
.flow-legend .lg{font-size:16px;line-height:1}
.flow-legend .lg.hi{color:#16a34a}.flow-legend .lg.mid{color:#f59e0b}.flow-legend .lg.lo{color:#5b9bd1}
.flow-legend .lg-size{font-size:11px;color:#7aa9c9}
.dark .flow-legend{color:#b9d4ea}
.globe-caption{text-align:center;margin-top:8px;font-size:12px;color:#557086;font-weight:700}.designer-line{display:block!important;font-size:12px!important;font-weight:700!important;color:#557086!important;margin-top:2px}.dark .designer-line{color:#b9d4ea!important}
.flights-map{height:460px;border-radius:14px;overflow:hidden;border:1px solid var(--line);margin-bottom:8px}
.plane-marker div{font-size:20px;line-height:1;color:#174a7c;text-shadow:0 0 3px #fff,0 0 1px #fff;transform-origin:center center}
.dark .plane-marker div{color:#9fd4ff;text-shadow:0 0 3px #0a1726,0 0 1px #0a1726}
.trend-chart-wrap{width:100%}
.trend-chart{width:100%;height:230px;display:block}
.trend-grid{stroke:rgba(120,150,180,.18);stroke-width:1}
.trend-axis{font-size:10px;fill:#7a93ab;font-weight:700}
.trend-line{fill:none;stroke-width:2.4;stroke-linejoin:round}
.trend-line.c0{stroke:#1f72b8}.trend-line.c1{stroke:#1b9e77}
.trend-dot.c0{fill:#1f72b8}.trend-dot.c1{fill:#1b9e77}
.trend-dot{stroke:#fff;stroke-width:1.4}
.trend-legend-row{display:flex;gap:16px;justify-content:center;margin-top:6px;font-size:12px;font-weight:700;color:#557086}
.trend-legend.c0{color:#1f72b8}.trend-legend.c1{color:#1b9e77}
.dark .trend-grid{stroke:rgba(180,229,252,.12)}
.rtab-wrap{background:#fff;border-radius:14px;overflow:hidden;border:1px solid #e8edf3}.rtab-head{padding:22px 28px 18px;border-bottom:1px solid #eef1f5;display:flex;flex-wrap:wrap;align-items:center;gap:16px}.rtab-title{display:flex;align-items:center;gap:14px;flex:none}.rtab-accent{width:4px;height:46px;border-radius:3px;background:linear-gradient(180deg,#1f6fb8,#35c1c9);box-shadow:0 2px 9px rgba(53,193,201,.4);flex-shrink:0}.rtab-sup{font-size:10px;font-weight:700;letter-spacing:.18em;text-transform:uppercase;color:#9aa6b8;margin-bottom:4px}.rtab-h1{font-size:22px;font-weight:800;letter-spacing:-.02em;color:#16243a;line-height:1}.rtab-gradient{background:linear-gradient(135deg,#1f6fb8,#35c1c9);-webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;color:transparent}.rtab-controls{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-left:auto}.rtab-search{position:relative;display:flex;align-items:center}.rtab-search input{width:240px;padding:10px 36px 10px 38px;border:1px solid #dde5ee;border-radius:12px;font-size:13px;color:#26384f;outline:none;background:#f5f8fc;transition:border-color .18s,box-shadow .18s}.rtab-search input:focus{border-color:#35c1c9;background:#fff;box-shadow:0 0 0 3px rgba(53,193,201,.18)}.rtab-search svg{position:absolute;left:13px;pointer-events:none}.rtab-search .rtab-clear{position:absolute;right:9px;width:20px;height:20px;border:none;border-radius:50%;background:#e6ebf2;color:#7d8aa0;cursor:pointer;display:inline-flex;align-items:center;justify-content:center;padding:0;transition:background .15s}.rtab-search .rtab-clear:hover{background:#d0d9e5}.rtab-sortbtns{display:inline-flex;gap:4px;padding:4px;background:#eaeff5;border:1px solid #e0e7f0;border-radius:14px}.rtab-sortbtns button{display:inline-flex;align-items:center;gap:6px;padding:8px 14px;border-radius:10px;font-size:13px;font-weight:600;cursor:pointer;border:1px solid transparent;background:transparent;color:#5b6b82;transition:background .2s,color .15s,box-shadow .2s,transform .15s}.rtab-sortbtns button:hover{background:#fff;color:#1f6fb8;box-shadow:0 2px 7px rgba(31,45,61,.09)}.rtab-sortbtns button.on{background:linear-gradient(135deg,#2477c2,#35c1c9);color:#fff;border-color:rgba(255,255,255,.3);box-shadow:0 5px 14px rgba(31,111,184,.38);transform:translateY(-1px)}.rtab-wrap{--rtab-grid:28px 1fr 72px repeat(3,minmax(130px,1.1fr))}.rtab-colhead{display:grid;grid-template-columns:var(--rtab-grid);gap:20px;padding:14px 24px;background:#27374d;font-size:12px;font-weight:800;letter-spacing:.05em;text-transform:uppercase;color:#cdd9e6}.rtab-colhead .on{color:#7ee6ec;background:rgba(95,216,223,.16);border-radius:8px;padding:4px 8px;margin:-4px -8px;box-shadow:inset 0 0 0 1px rgba(126,230,236,.28)}.rtab-colhead .num{text-align:right}.rtab-body{padding:4px 0}@keyframes rtabRowIn{from{transform:translateY(8px);opacity:0}to{transform:translateY(0);opacity:1}}@keyframes rtabBarGrow{from{transform:scaleX(0)}to{transform:scaleX(1)}}.rtab-row{display:grid;grid-template-columns:var(--rtab-grid);gap:20px;align-items:center;padding:12px 24px;border-bottom:1px solid #f2f5f8;animation:rtabRowIn .5s cubic-bezier(.22,1,.36,1) both;transition:background .15s;cursor:default}.rtab-row:hover{background:#f7fafd}.rtab-row:last-child{border-bottom:none}.rtab-rank{display:inline-flex;align-items:center;justify-content:center;min-width:24px;height:24px;padding:0 4px;border-radius:7px;font-size:12px;font-weight:800;font-variant-numeric:tabular-nums;animation:rtabRowIn .45s cubic-bezier(.22,1,.36,1) both}.rtab-rank.top{background:linear-gradient(135deg,#1f6fb8,#35c1c9);color:#fff;box-shadow:0 3px 9px rgba(31,111,184,.32)}.rtab-rank.reg{background:#eef2f6;color:#8c98ad}.rtab-name{font-size:14px;font-weight:600;color:#26384f;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.rtab-stir{display:flex;align-items:center;gap:4px;margin-top:2px}.rtab-stir span{font-size:11.5px;color:#9aa6b8;font-variant-numeric:tabular-nums}.rtab-stir button,.rtab-stir-copy{width:20px;height:20px;border:none;border-radius:5px;background:transparent;color:#9aa6b8;cursor:pointer;display:inline-flex;align-items:center;justify-content:center;padding:0;transition:background .15s,color .15s}.rtab-stir button:hover,.rtab-stir-copy:hover{background:#e8f1f8;color:#1f6fb8}.rtab-p{text-align:right;font-size:15px;font-weight:700;color:#3a4a61;font-variant-numeric:tabular-nums}.rtab-p.on{background:#eaf6fb;border-radius:10px;padding:8px 10px;margin:-8px -6px;box-shadow:inset 0 0 0 1px rgba(53,193,201,.18)}.rtab-cell{display:flex;flex-direction:column;gap:6px}.rtab-cell.on{background:#eaf6fb;border-radius:10px;padding:8px 10px;margin:-8px -6px;box-shadow:inset 0 0 0 1px rgba(53,193,201,.18)}.rtab-nums{display:flex;align-items:baseline;justify-content:space-between;gap:6px}.rtab-pct{font-size:12.5px;font-weight:700;color:#1f8fbf;font-variant-numeric:tabular-nums}.rtab-val{font-size:14px;font-weight:600;color:#3a4a61;font-variant-numeric:tabular-nums}.rtab-bar-bg{height:9px;border-radius:5px;background:#eef2f6;overflow:hidden}.rtab-bar-fg{height:100%;border-radius:5px;transform-origin:left center;animation:rtabBarGrow .65s cubic-bezier(.22,1,.36,1) both;transition:width .55s cubic-bezier(.4,0,.2,1)}.rtab-empty{display:flex;flex-direction:column;align-items:center;justify-content:center;gap:10px;padding:60px 24px;text-align:center;color:#9aa6b8}.dark .rtab-wrap{background:#0f1d2e;border-color:#2c4560}.dark .rtab-colhead{background:#152130}.dark .rtab-row{border-color:#1e3249}.dark .rtab-row:hover{background:#172840}.dark .rtab-name{color:#deeeff}.dark .rtab-val{color:#b8d4ee}.dark .rtab-pct{color:#4ecdc4}.dark .rtab-p{color:#deeeff}.dark .rtab-bar-bg{background:#1e3249}.dark .rtab-search input{background:#0f1d2e;border-color:#2c4560;color:#deeeff}.dark .rtab-sortbtns{background:#152130;border-color:#2c4560}.dark .rtab-sortbtns button{color:#9ab8d4}.dark .rtab-rank.reg{background:#1e3249;color:#7a9abc}
.rtab-row-exp{background:#fff8f8!important;border-left:3px solid #dc2626!important}
.dark .rtab-row-exp{background:#2a1515!important;border-left:3px solid #ef4444!important}
.dark .trend-axis{fill:#9fc3df}
.dark .trend-legend-row{color:#b9d4ea}
@keyframes sealFloatCalm{0%,100%{transform:translateY(0) scale(1)}50%{transform:translateY(-14px) scale(1.015)}}@keyframes spinBusy{to{transform:rotate(360deg)}}
.login-error{min-height:20px;margin-top:8px;color:#b42318;font-weight:800}.dark .login-error{color:#ffb4a8}.danger,.logout-btn{background:#c0392b!important;color:#fff!important}.logout-btn:hover{background:#9f2f25!important}.lang-btn{min-width:44px;padding:8px 10px!important}.regime-year-table{font-size:15px!important}.regime-year-table th,.regime-year-table td{text-align:center!important;font-size:15px!important;padding:5px 6px!important}.regime-year-table .col-name,.regime-year-table td.col-name{text-align:center!important;vertical-align:middle!important}.expired-detail-sample{min-width:1190px!important}.expired-detail-sample th{height:42px!important}.expired-detail-sample td{height:28px!important}.expired-detail-sample .col-reason{white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}.sample-like .merged-row td{text-align:center!important}.sample-like .col-korxona{text-align:left!important}.forgot-link{margin-top:6px;background:transparent!important;color:#174a7c!important;padding:4px 0!important;text-align:left}.dark .forgot-link{color:#9bd3ff!important}body.login-screen header{display:flex!important}body.login-screen #actions{display:none!important}body.login-screen #app,body.login-screen #dash,body.login-screen #tabs,body.login-screen #view,body.login-screen #kpis,.login-screen .workspace,.login-screen .tabs{display:none!important}.ibk-map-panel{background:linear-gradient(180deg,rgba(8,27,48,.94),rgba(13,43,70,.90))!important;color:#eaf7ff!important;border-color:rgba(119,201,255,.24)!important}.ibk-map-panel .muted,.ibk-map-panel span{color:#b9dff4!important}body.logged-in main{max-width:1500px!important;margin:auto!important;padding:14px!important;display:block!important}
body.logged-in #app{display:block!important}
body.logged-in #dash{display:block!important}
body.logged-in .workspace{display:grid!important;grid-template-columns:230px minmax(0,1fr)!important;gap:16px!important;align-items:start!important}
body.logged-in #tabs{display:flex!important;position:sticky!important;top:86px!important;flex-direction:column!important;gap:8px!important;margin:12px 0!important}
body.logged-in #view{min-width:0!important;width:100%!important}
body.logged-in .module-parent{display:flex!important;align-items:center!important;gap:9px!important;border-radius:10px!important;padding:12px 13px!important;background:rgba(232,238,246,.92)!important;color:#172033!important;border-left:5px solid transparent!important;text-align:left!important;box-shadow:0 8px 18px rgba(15,50,78,.06)!important}
body.logged-in .module-parent.active{background:#174a7c!important;color:#fff!important;border-left-color:#1b9e77!important}




body.logged-in .subtabs{display:flex!important;flex-direction:column!important;gap:6px!important;margin:2px 0 10px 12px!important}
body.logged-in .tab.sub{font-size:13px!important;padding:9px 10px!important;border-radius:8px!important}
body.login-screen #app,body.login-screen #dash,body.login-screen #tabs,body.login-screen #view,body.login-screen #kpis,.login-screen .workspace,.login-screen .tabs{display:none!important}
/* === HARD LOGIN/APP LAYER FIX === */
body.logged-in #login,
body:not(.login-screen) #login{
  display:none!important;
  visibility:hidden!important;
  opacity:0!important;
  pointer-events:none!important;
  position:fixed!important;
  left:-99999px!important;
  top:-99999px!important;
  width:0!important;
  height:0!important;
  overflow:hidden!important;
}

body.login-screen #login{
  display:grid!important;
  visibility:visible!important;
  opacity:1!important;
  pointer-events:auto!important;
  position:relative!important;
  left:auto!important;
  top:auto!important;
  width:auto!important;
  height:auto!important;
  overflow:visible!important;
}

body.login-screen #app,
body.login-screen #dash{
  display:none!important;
  visibility:hidden!important;
  opacity:0!important;
  pointer-events:none!important;
}

body.logged-in #app,
body.logged-in #dash{
  display:block!important;
  visibility:visible!important;
  opacity:1!important;
  pointer-events:auto!important;
}

body.logged-in main{
  display:block!important;
}

/* hex-bg */body{background:#f8fbff!important}body::before,body::after,header::after,header::before,main::before,main::after{content:none!important;display:none!important;animation:none!important}.sky-scene{position:fixed!important;top:0!important;left:0!important;right:0!important;bottom:0!important;width:100%!important;height:100%!important;z-index:0!important;overflow:hidden!important;pointer-events:none!important;animation:none!important;transform:none!important;background:#f8fbff!important}.sky-scene *{display:none!important}#hexBg{display:block!important;position:absolute!important;top:0!important;left:0!important;width:100%!important;height:100%!important;z-index:1!important}body .sky-scene::before,body .sky-scene::after{display:none!important;content:none!important;animation:none!important;background:none!important;opacity:0!important;visibility:hidden!important}.sky-scene::before,.sky-scene::after{display:none!important;content:none!important;animation:none!important;background:none!important;opacity:0!important;visibility:hidden!important}.login-screen .sky-scene::before,.login-screen .sky-scene::after{display:none!important;content:none!important;background:none!important;opacity:0!important;visibility:hidden!important}.logged-in .sky-scene::before,.logged-in .sky-scene::after{display:none!important;content:none!important;background:none!important;opacity:0!important}#bgVideo,.bg-video{display:none!important;opacity:0!important}.sky-scene>#bgCanvas{display:none!important;opacity:0!important}.login-screen .sky-scene{background:#f8fbff!important}.dark .sky-scene{background:#080f1e!important}.dark.login-screen .sky-scene{background:#080f1e!important}.panel,.kpi,.upload{background:rgba(255,255,255,.97)!important;backdrop-filter:blur(12px)}.dark .panel,.dark .kpi,.dark .upload{background:rgba(15,28,50,.95)!important}#login,main.login-screen-main{background:transparent!important;border:none!important;box-shadow:none!important;backdrop-filter:none!important}header{background:rgba(255,255,255,.08)!important;backdrop-filter:none!important;border-bottom:1px solid rgba(100,160,220,.15)!important;box-shadow:none!important}.dark header{background:rgba(8,15,30,.12)!important}body.login-screen header{background:transparent!important;border-bottom:none!important;box-shadow:none!important}.kpi{display:flex!important;flex-direction:column!important;justify-content:center!important;text-align:center!important;border:1.5px solid rgba(18,72,160,.14)!important;border-top:3px solid #2a6eb8!important;background:linear-gradient(160deg,rgba(42,110,184,.06),rgba(255,255,255,.97))!important;box-shadow:0 3px 14px rgba(0,0,0,.08)!important;border-radius:10px!important}.kpis .kpi:nth-child(1){border-top-color:#2a6eb8!important;background:linear-gradient(160deg,rgba(42,110,184,.07),rgba(255,255,255,.97))!important}.kpis .kpi:nth-child(2){border-top-color:#1b9e77!important;background:linear-gradient(160deg,rgba(27,158,119,.07),rgba(255,255,255,.97))!important}.kpis .kpi:nth-child(3){border-top-color:#e67e22!important;background:linear-gradient(160deg,rgba(230,126,34,.07),rgba(255,255,255,.97))!important}.kpis .kpi:nth-child(4){border-top-color:#8e44ad!important;background:linear-gradient(160deg,rgba(142,68,173,.07),rgba(255,255,255,.97))!important}.kpis .kpi:nth-child(5){border-top-color:#e74c3c!important;background:linear-gradient(160deg,rgba(231,76,60,.07),rgba(255,255,255,.97))!important}.kpis .kpi:nth-child(6){border-top-color:#16a085!important;background:linear-gradient(160deg,rgba(22,160,133,.07),rgba(255,255,255,.97))!important}.summary-item{border:1.5px solid rgba(18,72,160,.18)!important;border-top:3px solid #2a6eb8!important;background:linear-gradient(160deg,rgba(42,110,184,.06),rgba(255,255,255,.97))!important;box-shadow:0 4px 16px rgba(0,0,0,.09)!important;border-radius:12px!important}.summary-grid button.summary-item:nth-child(1){border-top-color:#2a6eb8!important;background:linear-gradient(160deg,rgba(42,110,184,.07),rgba(255,255,255,.97))!important}.summary-grid button.summary-item:nth-child(2){border-top-color:#1b9e77!important;background:linear-gradient(160deg,rgba(27,158,119,.07),rgba(255,255,255,.97))!important;border-color:rgba(27,158,119,.18)!important}.summary-grid button.summary-item:nth-child(3){border-top-color:#e67e22!important;background:linear-gradient(160deg,rgba(230,126,34,.07),rgba(255,255,255,.97))!important;border-color:rgba(230,126,34,.18)!important}.summary-item *{color:#12304e!important}.summary-item b,.summary-item>b{color:#1b9e77!important}.summary-item span{color:#445566!important}.module-parent{color:#fff!important}body.logged-in .module-parent{background:linear-gradient(135deg,#d4eaff,#dff0ff)!important;color:#0a2440!important;border:1.5px solid rgba(18,72,160,.22)!important;border-left:4px solid #2060b8!important;box-shadow:0 2px 10px rgba(18,72,160,.1)!important;font-weight:700!important}.module-parent.pay,body.logged-in .module-parent.pay{background:linear-gradient(135deg,#1d7f73,#269490)!important;color:#fff!important;border:1.5px solid #1a7a6a!important;border-left:4px solid #0dd4b4!important}.tab{color:#12304e!important}.tab.active{color:#fff!important}.tab.sub{color:#12304e!important}.tab.sub.active{color:#fff!important}.dark .tab,.dark .tab.sub{color:#eaf2ff!important}.dark .btn.light,.dark button.light{background:#1e3254!important;color:#deeeff!important;border:1px solid #3a587a!important}.dark button:not(.light):not(.danger){background:#2a6eb8!important;color:#fff!important}.dark select,.dark input{color:#deeeff!important}.dark .bar span{background:linear-gradient(90deg,#2a6eb8,#1b9e77)!important}.release-date-card{background:rgba(255,255,255,.97)!important;border:1px solid rgba(100,160,220,.18);border-radius:8px;padding:14px;margin-bottom:12px}.dark .release-date-card{background:rgba(15,28,50,.95)!important;border-color:rgba(59,81,104,.5)}.release-section-result table td{background-color:#fff}.release-section-result tr:nth-child(even) td{background-color:#f8fafc!important}.release-section-result tbody tr:first-child td{background-color:#d7e8d2!important}.dark .release-section-result table td{background-color:#1a2a3e}.dark .release-section-result tr:nth-child(even) td{background-color:#1d2a3d!important}
@keyframes cfmDash{to{stroke-dashoffset:-28}}.cfm-flow-line{animation:cfmDash 2.2s linear infinite}.cfm-legend{background:rgba(255,255,255,.93);border:1px solid #c8d8ea;border-radius:8px;padding:8px 12px;font-size:11px;line-height:1.7;color:#1a3a5c;box-shadow:0 2px 8px rgba(0,0,0,.1)}.dark .cfm-legend{background:rgba(18,32,50,.93);border-color:#3b5168;color:#eaf2ff}
.rating-list{display:flex;flex-direction:column;gap:8px}.rating-header{background:linear-gradient(135deg,#1f6fb8,#35c1c9);border-radius:14px;padding:14px 20px;display:flex;justify-content:space-between;align-items:center;color:#fff;font-weight:700;font-size:14px}.rating-row{background:rgba(255,255,255,.92);border-radius:10px;padding:11px 16px;display:flex;align-items:center;gap:12px;cursor:pointer;transition:transform .18s,box-shadow .18s}.rating-row:hover{transform:translateY(-2px);box-shadow:0 10px 24px rgba(31,111,184,.14)}.rating-rank{font-weight:900;color:#1f6fb8;min-width:26px;font-size:15px;text-align:center}.rating-name{flex:1;color:#4a6380;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.rating-bar-wrap{display:flex;align-items:center;gap:8px;min-width:240px;max-width:46%}.rating-bar{height:14px;border-radius:7px;background:linear-gradient(90deg,#1f6fb8,#35c1c9);flex-shrink:0;min-width:4px}.rating-val{font-size:12px;color:#1f6fb8;font-weight:700;white-space:nowrap;min-width:80px;text-align:right}.dark .rating-row{background:rgba(20,35,54,.88)!important}.dark .rating-name{color:#9fb2c8!important}.dark .rating-val,.dark .rating-rank{color:#7ec8ff!important}
@keyframes shimBar{0%{transform:translateX(-200%)}100%{transform:translateX(550%)}}
.bar span{position:relative!important;overflow:hidden!important}
.bar span::after{content:'';position:absolute;top:0;left:0;width:22%;height:100%;background:linear-gradient(90deg,transparent,rgba(255,255,255,.55),transparent);animation:shimBar 2.4s ease-in-out infinite;pointer-events:none;border-radius:999px}
.dark .bar span::after{background:linear-gradient(90deg,transparent,rgba(255,255,255,.22),transparent)}
@keyframes kpiGlow{0%,100%{filter:drop-shadow(0 2px 8px rgba(42,110,184,.07))}50%{filter:drop-shadow(0 5px 22px rgba(42,110,184,.48))}}
.kpi{animation:kpiGlow 3.5s ease-in-out infinite}
.wlayout{display:flex;flex-direction:column;gap:14px}
.stale-badge{display:inline-block;background:#fef3c7;border:1px solid #f59e0b;border-radius:4px;padding:1px 8px;font-size:11px;color:#92400e;font-weight:600;margin-left:6px;vertical-align:middle}
.sm-body{border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-top:4px}
.sm-row{display:flex;align-items:center;gap:8px;padding:9px 12px;border-bottom:1px solid var(--border);background:var(--card);cursor:default;user-select:none}
.sm-row:last-child{border-bottom:none}
.sm-row.dragging{opacity:.4;background:var(--hover)}
.sm-drag{cursor:grab;color:var(--muted);font-size:16px;flex-shrink:0}
.sm-eye{background:none;border:1px solid var(--border);border-radius:4px;padding:2px 7px;cursor:pointer;font-size:14px;flex-shrink:0}
.sm-eye.hidden-eye{opacity:.4}
.le-hidden{opacity:.45;text-decoration:line-through}
@keyframes summaryCard{0%,100%{filter:drop-shadow(0 2px 10px rgba(42,110,184,.06))}50%{filter:drop-shadow(0 7px 28px rgba(42,110,184,.35))}}
.exec-summary .summary-item{animation:summaryCard 4s ease-in-out infinite}
@keyframes ringShine{0%,100%{filter:saturate(1) brightness(1)}50%{filter:saturate(1.55) brightness(1.07)}}
.transport-ring{animation:ringShine 3.2s ease-in-out infinite!important}
.flow-track{overflow:hidden;position:relative}
.flow-track::after{content:'';position:absolute;top:0;left:-60%;width:45%;height:100%;background:linear-gradient(90deg,transparent,rgba(255,255,255,.48),transparent);animation:shimTrack 2.4s ease-in-out 0.8s infinite;pointer-events:none}
@keyframes shimTrack{0%{left:-60%}100%{left:120%}}
.exec-header{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;margin-bottom:12px}.exec-header-left{display:flex;align-items:center;gap:8px}.exec-pulse-dot{display:inline-block;width:9px;height:9px;border-radius:50%;background:#1b9e77;position:relative;flex-shrink:0}.exec-pulse-dot::after{content:'';position:absolute;inset:-4px;border-radius:50%;background:rgba(27,158,119,.28);animation:execPulse 1.8s ease-out infinite}.exec-title{font-size:15px;font-weight:700;color:#0a2440}.exec-badges{display:flex;gap:6px;flex-wrap:wrap}.exec-badge{font-size:11px;font-weight:700;padding:3px 10px;border-radius:20px;white-space:nowrap}.exec-badge-blue{background:#e4effe;color:#1a4f9a}.exec-badge-red{background:#fdeaea;color:#9b2020}.exec-badge-green{background:#e5f6ee;color:#0d6640}@keyframes execPulse{0%{transform:scale(.7);opacity:1}100%{transform:scale(2.4);opacity:0}}.dark .exec-title{color:#dff0ff!important}.dark .exec-badge-blue{background:#1a3258;color:#8ec4f4}.dark .exec-badge-red{background:#3a1a1a;color:#f48c8c}.dark .exec-badge-green{background:#0d3324;color:#7ddcb0}
.wr-map-wrap{border-radius:12px;overflow:hidden;border:1px solid var(--line);margin-bottom:16px}
.wr-filter-bar{padding:10px 12px 0;background:#fff;border-bottom:1px solid var(--line)}.dark .wr-filter-bar{background:#172033}
.wr-filter-row{display:flex;gap:8px;flex-wrap:wrap;padding-bottom:10px}
.wr-filter-chip{display:inline-flex;align-items:center;gap:6px;padding:5px 12px;border-radius:20px;border:1.5px solid #d3dde8;background:#f4f8fc;cursor:pointer;font-size:12px;font-weight:600;color:#3d5a7a;transition:.15s;user-select:none}
.wr-filter-chip input{position:absolute;opacity:0;width:0;height:0}
.wr-filter-chip.active{background:var(--blue);color:#fff;border-color:var(--blue)}
.wr-filter-chip:hover{border-color:var(--blue)}
.wr-chip-dot{width:10px;height:10px;border-radius:50%;display:inline-block}
.wr-filter-chip.active .wr-chip-dot{border:2px solid rgba(255,255,255,.6)}
@keyframes wr-popup-in{from{opacity:0;transform:translateY(16px) scale(.96)}to{opacity:1;transform:translateY(0) scale(1)}}
.wr-popup-card{position:absolute;bottom:16px;left:50%;transform:translateX(-50%);z-index:1200;width:320px;background:#fff;border-radius:14px;box-shadow:0 8px 32px rgba(0,0,0,.18);overflow:hidden;opacity:0;transition:opacity .25s,transform .25s;pointer-events:none}
.wr-popup-card.visible{opacity:1;pointer-events:auto;animation:wr-popup-in .28s ease-out both}
.wr-popup-head{padding:12px 16px;color:#fff;display:flex;flex-direction:column;gap:3px}
.wr-popup-head b{font-size:13px;line-height:1.3}
.wr-popup-head span{font-size:11px;opacity:.85}
.wr-popup-body{padding:10px 16px 14px;display:flex;flex-direction:column;gap:6px}
.wr-popup-row{display:flex;justify-content:space-between;align-items:baseline;gap:8px;font-size:12px}
.wr-popup-row span{color:var(--muted);flex-shrink:0}
.wr-popup-row b{text-align:right;font-size:12px}
.wr-popup-sep{height:1px;background:var(--line);margin:4px 0}
.wr-popup-sec span{font-size:11px;font-weight:700;color:var(--blue);text-transform:uppercase;letter-spacing:.4px}
.dark .wr-popup-card{background:#172033;box-shadow:0 8px 32px rgba(0,0,0,.5)}
.wr-type-legend{display:flex;gap:12px;flex-wrap:wrap;font-size:11px;padding:8px 0;margin-bottom:10px}
.wr-type-dot{display:inline-block;width:12px;height:12px;border-radius:50%;margin-right:4px;vertical-align:middle}
.wr-ins-table{width:100%;border-collapse:collapse;font-size:12px}
.wr-ins-table th{background:#1d72b8;color:#fff;padding:6px 8px;text-align:left;white-space:nowrap}
.wr-ins-table td{padding:5px 8px;border-bottom:1px solid rgba(0,0,0,.07)}
.wr-ins-table tr:nth-child(even) td{background:#f4f8ff}
.wr-risk-red{color:#b91c1c;font-weight:700}
.wr-risk-orange{color:#d97706;font-weight:700}
.wr-risk-green{color:#16a34a}
.dark .wr-ins-table th{background:#1a3a5c}
.dark .wr-ins-table td{border-color:rgba(255,255,255,.07)}
.dark .wr-ins-table tr:nth-child(even) td{background:rgba(255,255,255,.04)}
.wr-upload-form{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:12px;padding:12px;background:rgba(42,110,184,.06);border-radius:10px;border:1px dashed rgba(42,110,184,.25)}
.icon-btn{display:inline-flex;align-items:center;justify-content:center;width:30px;height:30px;border-radius:7px;border:1px solid var(--line);background:#f0f4f8;color:#3d5a7a;cursor:pointer;transition:.15s}.icon-btn:hover{background:#dde8f3;color:var(--blue)}.icon-btn.del-btn{background:#fff0f0;color:#b42318;border-color:#ffd7d7}.icon-btn.del-btn:hover{background:#ffd7d7}.dark .icon-btn{background:#1e2d40;color:#9fb0c6;border-color:#2f3f55}.dark .icon-btn.del-btn{background:#2a1a1a;color:#f87171;border-color:#7f1d1d}
.wr-badge-fvv{display:inline-block;background:#ef4444;color:#fff;border-radius:4px;padding:1px 5px;font-size:10px;font-weight:700;margin-left:4px}
.wr-badge-ssv{display:inline-block;background:#16a34a;color:#fff;border-radius:4px;padding:1px 5px;font-size:10px;font-weight:700;margin-left:4px}
.currency-widget{display:flex;gap:10px;align-items:center;flex-wrap:wrap;padding:4px 12px;font-size:12px;color:var(--muted);border-bottom:1px solid rgba(42,110,184,.09);min-height:22px;background:rgba(248,251,255,.7)}
.currency-widget .cur-item{display:inline-flex;gap:4px;align-items:center;padding:2px 7px;background:rgba(42,110,184,.07);border-radius:10px}
.currency-widget .cur-item b{color:var(--blue);font-weight:700}
.currency-widget .cur-date{color:var(--muted);font-size:11px}
.dark .currency-widget{background:rgba(10,18,40,.5);border-color:rgba(96,165,250,.12)}
tr.expired-row td{background:rgba(220,38,38,.04)!important}
.dark tr.expired-row td{background:rgba(220,38,38,.10)!important}
.uc-card{border-left:4px solid var(--blue)}.uc-header{display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap;margin-bottom:14px}.uc-header h2{margin:0;flex:1}.uc-badge{display:inline-flex;align-items:center;gap:4px;padding:3px 10px;border-radius:20px;font-size:12px;font-weight:600;white-space:nowrap}.uc-badge.ok{background:#dcfce7;color:#15803d}.uc-badge.warn{background:#fef9c3;color:#854d0e}.uc-body{display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap}.uc-field{display:flex;flex-direction:column;gap:5px;min-width:200px;flex:1}.uc-field label{font-size:12px;font-weight:600;color:var(--muted)}.uc-field input[type=file]{width:100%;font-size:12px;padding:6px;border:1px solid var(--line);border-radius:6px;background:#fff}.uc-btns{display:flex;gap:8px;align-items:flex-end;flex-shrink:0}.uc-note{font-size:12px;margin:8px 0 0;padding:8px 12px;background:#f0f9ff;border-radius:6px;color:#0369a1;border-left:3px solid #0ea5e9}.dark .uc-note{background:rgba(14,165,233,.1);color:#7dd3fc}.dark .uc-field input[type=file]{background:#1a2a3e;color:#deeeff;border-color:#3a587a}
</style></head><body class="login-screen"><div class="sky-scene" aria-hidden="true"><video id="bgVideo" class="bg-video" autoplay muted playsinline></video><canvas id="bgCanvas" width="1280" height="720"></canvas><div class="cinema-clouds"></div><div class="cinema-runway"></div><div class="cinema-glow"></div><div class="cinema-vignette"></div><div class="sky-layer mountains"></div><div class="sky-layer city"></div><div class="sky-layer city front"></div><div class="runway"></div><div class="tower"></div><div class="sky-layer water"></div><div class="bird b1"></div><div class="bird b2"></div><div class="bird b3"></div><div class="plane-wrap"><svg class="plane-svg" viewBox="0 0 900 360"><defs><linearGradient id="planeSkin" x1="0" y1="0" x2="0" y2="1"><stop offset="0" stop-color="#ffffff"/><stop offset="0.48" stop-color="#d8e5ec"/><stop offset="1" stop-color="#8ea5b4"/></linearGradient><clipPath id="sealClip"><circle cx="450" cy="132" r="24"/></clipPath></defs><path class="plane-tail" d="M420 93 450 18 480 93 464 112 436 112Z"/><path class="plane-wing" d="M103 168 450 116 797 168 770 204 494 176 465 268 435 268 406 176 130 204Z"/><path class="plane-body" d="M382 105c16-54 120-54 136 0 18 61 9 164-24 211-20 29-68 29-88 0-33-47-42-150-24-211Z"/><path class="plane-body" d="M338 148c47-30 177-30 224 0l-28 35c-38-21-130-21-168 0Z" opacity=".85"/><ellipse class="engine" cx="294" cy="203" rx="48" ry="35"/><ellipse class="engine-dark" cx="294" cy="205" rx="25" ry="20"/><ellipse class="engine" cx="606" cy="203" rx="48" ry="35"/><ellipse class="engine-dark" cx="606" cy="205" rx="25" ry="20"/><circle class="seal-ring" cx="450" cy="132" r="31"/><image href="/assets/gerb-bojxona.jpg" x="426" y="108" width="48" height="48" clip-path="url(#sealClip)"/><text class="plane-text" x="450" y="198" text-anchor="middle">Toshkent-AERO IBK</text><g><rect class="plane-window" x="409" y="91" width="14" height="8" rx="4"/><rect class="plane-window" x="431" y="86" width="14" height="8" rx="4"/><rect class="plane-window" x="453" y="86" width="14" height="8" rx="4"/><rect class="plane-window" x="475" y="91" width="14" height="8" rx="4"/></g><path d="M154 175c210-61 382-61 592 0" fill="none" stroke="rgba(255,255,255,.45)" stroke-width="4"/><path d="M410 302c25 19 55 19 80 0" fill="none" stroke="rgba(255,255,255,.38)" stroke-width="5" stroke-linecap="round"/></svg></div></div><header><div><h1>IBK Dashboard</h1><div class="muted" id="meta">Kirish kerak</div><div class="muted"><span id="clock" class="header-clock"></span><span class="designer-line">by @aero004</span></div></div><div id="actions"></div></header><main>
<section id="login" class="login login-closed"><div class="login-seal-wrap" onclick="activateLogin()"><img class="login-seal" src="/assets/sticker.webp" alt="Bojxona gerbi"></div><div class="login-box"><h2>Kirish</h2><div class="login-form-stack"><div><label>Login</label><input id="user" autocomplete="username" placeholder="Login"></div><div><label>Parol</label><div class="pass-wrap"><input id="pass" type="password" autocomplete="new-password" placeholder="Parol"><button class="eye-btn" type="button" onclick="togglePassword()" title="Ko'rsatish/yashirish">&#128065;</button></div></div><button id="loginBtn" onclick="doLogin()">Kirish</button><div id="loginError" class="login-error"></div><button type="button" class="forgot-link" onclick="forgotPassword()">Parolni unutdingizmi?</button></div><div class="muted">Gerb ustiga bosilganda kirish oynasi ochiladi.</div></div></section>
<section id="app" class="hidden"><div id="status" class="muted"></div><div id="currencyWidget" class="currency-widget"></div><div id="dash" class="hidden"><div class="kpis" id="kpis"></div><div class="workspace"><aside class="tabs" id="tabs"></aside><section id="view"></section></div></div></section>
<dialog id="dlg"><div class="head"><b id="dlgTitle">Asos</b><button class="light" onclick="dlg.close()">Yopish</button></div><div class="body" id="dlgBody"></div></dialog>
<script>
let TOKEN=localStorage.ibk_token||"", DATA=null, TAB="home", GROUP="home", ARCHIVE=[], PAYMENTS=[], ME=null, LANG=localStorage.ibk_lang||"uz", COMPANY_TRENDS={periods:[],companies:[]}, GOODS_TRENDS={periods:[],goods:[]}, AVIA_DATA=null, AVIA_STATS=null, YAROQLILIK_DATA=null, YAROQLILIK_FILTER='expired';
const I18N={
  uz:{archive:"Arxiv",upload:"Fayl yuklash",general:"Umumiy",companies:"Korxonalar",expired:"Muddati o'tgan",released:"Nazoratdan yechish",goods:"Tovarlar",food:"Oziq-ovqat",profile:"Profil",settings:"Sozlamalar",admin:"Admin",dark:"Tungi rejim",logout:"Chiqish",regimes:"Rejimlar",warehouses:"Omborlar",deadlines:"Muddatlar",validity:"Yaroqlilik",pay_overview:"Umumiy",pay_lists:"Hosil bo'lgan jadvallar",pay_analysis:"Tahlil"},
  uzc:{archive:"Архив",upload:"Файл юклаш",general:"Умумий",companies:"Корхоналар",expired:"Муддати ўтган",released:"Назоратдан ечиш",goods:"Товарлар",food:"Озиқ-овқат",profile:"Профил",settings:"Созламалар",admin:"Админ",dark:"Тунги режим",logout:"Чиқиш",regimes:"Режимлар",warehouses:"Омборлар",deadlines:"Муддатлар",validity:"Яроқлилик",pay_overview:"Умумий",pay_lists:"Ҳосил бўлган жадваллар",pay_analysis:"Таҳлил"},
  ru:{archive:"Архив",upload:"Загрузка файла",general:"Общий",companies:"Компании",expired:"Просроченные",released:"Снятие с контроля",goods:"Товары",food:"Продукты",profile:"Профиль",settings:"Настройки",admin:"Админ",dark:"Тёмный режим",logout:"Выход",regimes:"Режимы",warehouses:"Склады",deadlines:"Сроки",validity:"Срок годности",pay_overview:"Общий",pay_lists:"Сформированные таблицы",pay_analysis:"Анализ"}
};
function tr(k){return (I18N[LANG]||I18N.uz)[k]||k} function setBg(v){}function setLang(v){LANG=v;localStorage.ibk_lang=v;render()} const $=id=>document.getElementById(id);
const fmtN=v=>Math.abs(+v||0)<.005?"0":(+v).toLocaleString("ru-RU",{minimumFractionDigits:2,maximumFractionDigits:2}).replace(/\u00a0/g," "), fmtI=v=>Math.round(+v||0).toLocaleString("ru-RU").replace(/\u00a0/g," "), esc=s=>String(s??"").replace(/[&<>"']/g,m=>({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}[m]));
function activateLogin(){let el=$("login");if(el)el.classList.add("active")}
function togglePassword(){let p=$("pass");if(p)p.type=p.type==="password"?"text":"password"}
["user","pass"].forEach(id=>{let el=document.getElementById(id);if(el)el.addEventListener("keydown",e=>{if(e.key==="Enter"){e.preventDefault();doLogin()}})});
function updateClock(){let c=$("clock");if(c)c.textContent=new Date().toLocaleString("uz-UZ",{hour:"2-digit",minute:"2-digit",second:"2-digit",day:"2-digit",month:"2-digit",year:"numeric"}).replace(",", "")}
setInterval(updateClock,1000);setTimeout(updateClock,50);
function setBusy(btn,on,text){if(!btn)return;if(on){btn.dataset.old=btn.innerHTML;btn.classList.add("is-busy");btn.disabled=true;btn.innerHTML=(text||btn.textContent)+` <span class="busy-spinner"></span>`}else{btn.classList.remove("is-busy");btn.disabled=false;if(btn.dataset.old)btn.innerHTML=btn.dataset.old}}
function clearBusy(){document.querySelectorAll(".is-busy").forEach(b=>setBusy(b,false))}
async function api(url,opt={}){opt.headers=Object.assign({"X-Token":TOKEN},opt.headers||{});let r=await fetch(url,opt);if(r.status===401){showLogin();throw Error("login")};return await r.json()} function showLogin(){$("login").classList.remove("hidden");$("app").classList.add("hidden");$("meta").textContent="Kirish kerak"}
async function doLogin(){let btn=$("loginBtn"),err=$("loginError");try{if(err)err.textContent="";setBusy(btn,true,"Kirish");let user=($("user")?.value||"").trim(),pass=$("pass")?.value||"";if(!user||!pass){if(err)err.textContent="Login va parolni kiriting";return;}let j=await api("/api/login",{method:"POST",body:JSON.stringify({user,pass})});TOKEN=j.token;localStorage.ibk_token=TOKEN;ME=j.user;await showApp()}catch(e){if(err)err.textContent=(e&&e.message&&e.message!=="login")?e.message:"Login yoki parol xato";}finally{setBusy(btn,false)}} function logout(){localStorage.removeItem("ibk_token");TOKEN="";DATA=null;showLogin()}
async function loadAviaStats(){try{AVIA_STATS=await api('/api/avia_stats');}catch(e){AVIA_STATS=null;}}
let ARCHIVE_CURRENT_ID=null;
let DATA_IS_STALE=false;
let DIRECT_UPLOAD_URL=null;
let _uploadActive=false;
async function detectDirectUpload(){try{const j=await api('/api/server_info');const u=j.lan_url;const r=await fetch(u+'/api/server_info',{signal:AbortSignal.timeout(1500)});if(r.ok){DIRECT_UPLOAD_URL=u;console.log('[Direct upload] LAN orqali yuklash faol:',u);}}catch(e){DIRECT_UPLOAD_URL=null;}}
function todayUzDate(){let d=new Date();return String(d.getDate()).padStart(2,'0')+'.'+String(d.getMonth()+1).padStart(2,'0')+'.'+d.getFullYear();}
function staleBanner(rd){return `<div style="background:linear-gradient(135deg,#fef9c3,#fef3c7);border:1px solid #f59e0b;border-radius:8px;padding:10px 16px;margin-bottom:12px;display:flex;align-items:center;gap:10px;font-size:13px"><span style="font-size:18px">⚠️</span><div><b style="color:#92400e">Bugungi (${todayUzDate()}) ma'lumot yuklanmagan</b><span style="color:#78350f"> &nbsp;·&nbsp; Ko'rsatilayotgan holat: </span><b style="color:#b45309;font-size:14px">${rd}</b></div></div>`;}
function staleBadge(){if(!DATA||!DATA_IS_STALE)return '';let d=(DATA.meta&&DATA.meta.date)||DATA.date||'?';return `<span class="stale-badge">📅 ${d} holatiga</span>`;}
const WIDGET_DEFS={
  umumiy:[{id:'exec',l:"Umumiy ko'rsatkichlar"},{id:'post',l:"70-74-80: postlar kesimida"},{id:'expired_sum',l:"Jami muddati o'tgan (jadval)"},{id:'top_bars',l:"TOP 30 korxona diagrammasi"}],
  rejim:[{id:'jami',l:"Jami va 70-74-80 jadvali"},{id:'kesim',l:"Rejimlar kesimida jadvali"},{id:'qiymat',l:"Qiymat ulushi diagrammasi"},{id:'tolov',l:"To'lov ulushi diagrammasi"}],
  korxona:[{id:'rating',l:"TOP reyting bloki"},{id:'top_qiymat',l:"TOP 20 (qiymat) jadvali"},{id:'top_depozit',l:"TOP 20 (depozit) jadvali"},{id:'ulush',l:"Qiymat ulushi diagrammasi"},{id:'trend',l:"Davriy tendensiya grafigi"},{id:'transport_k',l:"Transport: korxonalar oqimi"},{id:'transport_t',l:"Transport turi tendensiyasi"}],
  ombor:[{id:'rating',l:"Omborlar reytingi"},{id:'reestri',l:"Omborlar reestri"},{id:'kesim',l:"Omborlar kesimida jadvali"},{id:'bars_group',l:"Qiymat, o'z ombor, ulush diagrammalari"},{id:'oborot',l:"Omborlar oboroti"},{id:'trend',l:"Yuk oqimi tendensiyasi"}],
  muddat:[{id:'kesim',l:"Muddatlar kesimida jadvali"},{id:'sabab',l:"Saqlanish sabablari jadvali"},{id:'bars_group',l:"Muddat va sabab diagrammalari"}],
  expired:[{id:'rejim',l:"Postlar va rejimlar jadvali"},{id:'post',l:"Postlar kesimida jadvali"},{id:'jamlanma',l:"Jamlanma jadvali"},{id:'korxona',l:"Korxonalar jadvali"}],
  goods:[{id:'rating',l:"Tovar reytingi bloki"},{id:'trend',l:"Davriy tendensiya grafigi"},{id:'jadval',l:"Tovarlar guruhlari jadvali"},{id:'bars_group',l:"Partiya, qiymat, vazn diagrammalari"}],
  food:[{id:'jadval',l:"Oziq-ovqatlar jadvali"},{id:'bars',l:"Qiymat ulushi diagrammasi"}],
};
function getTabLayout(tab){let defs=(WIDGET_DEFS[tab]||[]).map(d=>d.id);let saved=((UI_CONFIG||{}).tab_layouts||{})[tab];if(!saved||!saved.length)return defs;let order=saved.map(x=>x.id||x).filter(id=>defs.includes(id));defs.forEach(id=>{if(!order.includes(id))order.push(id);});return order;}
function getTabHidden(tab){let saved=((UI_CONFIG||{}).tab_layouts||{})[tab];if(!saved)return new Set();return new Set(saved.filter(x=>x&&x.hidden).map(x=>x.id));}
async function saveTabLayout(tab,items){if(!UI_CONFIG)UI_CONFIG={};if(!UI_CONFIG.tab_layouts)UI_CONFIG.tab_layouts={};UI_CONFIG.tab_layouts[tab]=items;try{await api('/api/ui_config',{method:'POST',body:JSON.stringify(UI_CONFIG)});}catch(e){}}
function wlayout(tab,widgets){let order=getTabLayout(tab);let hidden=getTabHidden(tab);let parts=order.filter(id=>!hidden.has(id)).map(id=>(widgets[id]?widgets[id]():'')).filter(Boolean);return `<div class="wlayout">${parts.join('')}</div>`;}
function layoutEditorPanel(){let tabs=Object.keys(WIDGET_DEFS);let lbl={umumiy:'Umumiy',rejim:'Rejimlar',korxona:'Korxonalar',ombor:'Omborlar',muddat:'Muddatlar',expired:"Muddati o'tgan",goods:'Tovarlar',food:'Oziq-ovqat'};if(!window._leTab||!tabs.includes(window._leTab))window._leTab=tabs[0];let act=window._leTab;let defs=WIDGET_DEFS[act]||[];let saved=((UI_CONFIG||{}).tab_layouts||{})[act]||[];let order=saved.length?saved.map(s=>s.id||s).filter(id=>defs.some(d=>d.id===id)):defs.map(d=>d.id);defs.forEach(d=>{if(!order.includes(d.id))order.push(d.id);});let hidSet=new Set(saved.filter(s=>s&&s.hidden).map(s=>s.id));let items=order.map(id=>{let d=defs.find(x=>x.id===id);return d?{id,label:d.l,hidden:hidSet.has(id)}:null}).filter(Boolean);return `<div class=panel id="lePanel"><h2>Ko'rinish tartibi — Drag & Drop</h2><div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px">${tabs.map(t=>`<button class="tab sub${t===act?' active':''}" onclick="window._leTab='${t}';$('lePanel').outerHTML=layoutEditorPanel();leBindDrag()">${lbl[t]||t}</button>`).join('')}</div><p class=muted style="font-size:12px;margin:0 0 8px">Sudrab joylashtiring · Ko'z tugmasi bilan yashiring/ko'rsating · Saqlash bosing</p><div id="leList" class="sm-body">${items.map(w=>`<div class="sm-row${w.hidden?' le-hidden':''}" draggable="true" data-le="${esc(w.id)}"><span class="sm-drag">⣿</span><span style="flex:1;font-size:13px">${esc(w.label)}</span><button class="sm-eye${w.hidden?' hidden-eye':''}" onclick="leEye(this)">${w.hidden?'🙈':'👁'}</button></div>`).join('')}</div><div style="margin-top:12px;display:flex;gap:8px;align-items:center"><button onclick="leSave('${act}')">Saqlash</button><button class=light onclick="leReset('${act}')">Asliyga qaytarish</button><span id="leSt" class=muted style="font-size:12px"></span></div></div>`;}
let _leDrag=null;
function leBindDrag(){let el=$('leList');if(!el)return;el.querySelectorAll('[data-le]').forEach(r=>{r.ondragstart=e=>{_leDrag=r.dataset.le;e.dataTransfer.effectAllowed='move';r.classList.add('dragging');};r.ondragend=()=>{_leDrag=null;el.querySelectorAll('[data-le]').forEach(x=>x.classList.remove('dragging'));};r.ondragover=e=>{e.preventDefault();if(!_leDrag||_leDrag===r.dataset.le)return;let rows=[...el.querySelectorAll('[data-le]')];let from=rows.find(x=>x.dataset.le===_leDrag);if(!from)return;if(rows.indexOf(from)<rows.indexOf(r))r.after(from);else r.before(from);};});}
function leEye(btn){let row=btn.closest('[data-le]');let h=!btn.classList.contains('hidden-eye');btn.classList.toggle('hidden-eye',h);btn.textContent=h?'🙈':'👁';row.classList.toggle('le-hidden',h);}
async function leSave(tab){let el=$('leList');if(!el)return;let items=[...el.querySelectorAll('[data-le]')].map(r=>({id:r.dataset.le,hidden:r.classList.contains('le-hidden')}));await saveTabLayout(tab,items);let s=$('leSt');if(s){s.textContent='✓ Saqlandi';setTimeout(()=>{if(s)s.textContent='';},2000);}}
async function leReset(tab){if(UI_CONFIG&&UI_CONFIG.tab_layouts)delete UI_CONFIG.tab_layouts[tab];try{await api('/api/ui_config',{method:'POST',body:JSON.stringify(UI_CONFIG||{})});}catch(e){}let p=$('lePanel');if(p){p.outerHTML=layoutEditorPanel();leBindDrag();}}
async function showApp(){$("login").classList.add("hidden");$("app").classList.remove("hidden");ME=await api("/api/me");LANG=ME.lang||localStorage.ibk_lang||"uz";await loadUIConfig();await loadArchive();await loadPayments();loadAviaStats();detectDirectUpload();if(ARCHIVE.length){let startId=ARCHIVE_CURRENT_ID&&ARCHIVE.find(r=>r.id===ARCHIVE_CURRENT_ID)?ARCHIVE_CURRENT_ID:ARCHIVE[0].id;DATA=await api("/api/reports/"+startId);TAB="umumiy";GROUP="bnrte";render();}else{DATA=null;TAB="home";GROUP="home";render();}} async function loadArchive(){let j=await api("/api/archive");ARCHIVE=j.reports||[];ARCHIVE_CURRENT_ID=j.current_id||null;let _t=todayUzDate();DATA_IS_STALE=ARCHIVE.length>0&&!ARCHIVE.some(r=>r.date===_t);} async function loadPayments(){try{let j=await api("/api/tolov");PAYMENTS=j.payments||[]}catch(e){PAYMENTS=[]}} async function loadReport(id){DATA=await api("/api/reports/"+id);if(TAB==="upload")TAB="umumiy";render()}
async function poll(id){try{let j=await api("/api/jobs/"+id);if($("status"))$("status").textContent=j.status;if(j.status==="xatolik"){if($("status"))$("status").textContent=j.error;return}if(j.status!=="tayyor"){setTimeout(()=>poll(id),1800);return}DATA=j.data;TAB="umumiy";await loadArchive();render()}catch(e){setTimeout(()=>poll(id),3000)}}
let ARTIFACT_POLL_ID=null;
async function prepareArtifacts(){if(!DATA)return;if(ARTIFACT_POLL_ID)return;try{let j=await api("/api/artifacts",{method:"POST",body:JSON.stringify({report:DATA.id})});ARTIFACT_POLL_ID=j.job_id;render();pollArtifacts(j.job_id)}catch(e){if($("status"))$("status").textContent="Xatolik: "+String(e)}}
async function pollArtifacts(id){if(ARTIFACT_POLL_ID!==id)return;try{let j=await api("/api/jobs/"+id);if($("status"))$("status").textContent=j.status==="tayyor"?"✓ Excel/PNG/PDF tayyor":j.status;if(j.status==="xatolik"){ARTIFACT_POLL_ID=null;render();return}if(j.status!=="tayyor"){setTimeout(()=>pollArtifacts(id),2500);return}ARTIFACT_POLL_ID=null;DATA=j.data;render()}catch(e){setTimeout(()=>pollArtifacts(id),5000)}}
async function openGroup(g,tab){if(GROUP===g){GROUP="home";TAB="home";render();return}GROUP=g;TAB=tab;if(g==="bnrte"&&!DATA&&ARCHIVE.length){await loadReport(ARCHIVE[0].id);return}render()}
function landingPanel(){return `<div class="panel wide"><h2>IBK Dashboard</h2><p class="muted">Toshkent-AERO IBK bo'yicha BNRTE nazoratdagi tovarlar, to'lovlar, arxiv, nazoratdan yechilish va tahliliy ko'rsatkichlar yagona dashboardda jamlanadi.</p><div class="summary-grid"><button class="summary-item" onclick="openGroup('bnrte','umumiy')"><b>BNRTE</b><span>Nazoratdagi tovarlar jamlanmasi, muddatlar, omborlar va muddati o'tgan tahlillar.</span></button><button class="summary-item" onclick="openGroup('payments','payments')"><b>To'lovlar</b><span>Baza fayl asosida to'lov turlari bo'yicha Excel jadvallar va tahlil.</span></button><button class="summary-item" onclick="openGroup('common','upload')"><b>Fayl yuklash</b><span>BNRTE yoki To'lovlar uchun yangi asos fayllarni yuklash.</span></button></div></div>${flightsPanelShell()}`}
function renderKpis(){let k=DATA.kpis||{};
  let depHtml=`<div class=kpi onclick="showKpi('depozit')" style="grid-row:span 1">
    <span>Jami depozit</span><b>${fmtN(k.depozit)}</b>
    <div style="margin-top:6px;padding-top:6px;border-top:1px solid rgba(0,0,0,.08);font-size:12px;color:var(--muted)">
      Nazoratda yuki bor qismi: <b style="color:var(--blue)">${fmtN(k.depozit_matched||0)}</b>
    </div>
  </div>`;
  let rows=[["Partiya",fmtI(k.partiya),"partiya"],["Vazn (tn)",fmtN(k.vazn),"vazn"],["Qiymat (ming $)",fmtN(k.qiymat),"qiymat"],["Kutilayotgan to'lov (mln so'm)",fmtN(k.tolov),"tolov"],["Muddati o'tgan partiya",fmtI(k.expired),"expired"]];
  let aviaHtml=AVIA_DATA&&AVIA_DATA.loaded?`<div class=kpi onclick="TAB='avia';GROUP='bnrte';render()" style="cursor:pointer"><span>✈ AVIA AWB</span><b>${fmtI(AVIA_DATA.unique_awb)}</b><div style="margin-top:6px;padding-top:6px;border-top:1px solid rgba(0,0,0,.08);font-size:12px;color:var(--muted)">Muddati o'tgan: <b style="color:#dc2626">${fmtI(AVIA_DATA.overdue_count)}</b></div></div>`:"";
  let aviaQiymatHtml=AVIA_STATS&&AVIA_STATS.jami_qiymat_k?`<div class=kpi onclick="TAB='avia';GROUP='bnrte';render()" style="cursor:pointer;border-top:3px solid #0ea5e9"><span>✈ Avia qiymat</span><b style="color:#0ea5e9">${fmtN(AVIA_STATS.jami_qiymat_k)}</b><div style="margin-top:6px;padding-top:6px;border-top:1px solid rgba(0,0,0,.08);font-size:12px;color:var(--muted)">ming $ • <b>${fmtN(AVIA_STATS.jami_vazn_tn)}</b> tn</div></div>`:"";
  return rows.map(x=>`<div class=kpi onclick="showKpi('${x[2]}')"><span>${x[0]}</span><b>${x[1]}</b></div>`).join("")+depHtml+aviaQiymatHtml+aviaHtml;}
function table(h,rows,cls=""){rows=rows||[];return `<table class="${cls}"><colgroup>${h.map(x=>`<col style="${x.w?'width:'+x.w:''}">`).join("")}</colgroup><thead><tr>${h.map(x=>`<th class="${x.n?'num':'text'}">${x.t}</th>`).join("")}</tr></thead><tbody>${rows.map((r,ri)=>`<tr class="${r._class||''}" onclick='detail(${JSON.stringify(r.key||{}).replaceAll("'","&#39;")})'>${h.map(x=>{let cls=x.n?'num':'text', raw=r[x.k], val=x.n&&(raw===""||raw===null||raw===undefined)?"":(x.f?x.f(raw):esc(raw)), tip=esc(raw);if(ri===0&&!x.n&&(val==="Jami"||String(val).startsWith("Jami ")))val="IBK bo'yicha Jami";if(x.k==="released_partiya"&&(+raw||0)===0&&((+r.released_qiymat||0)>0.005||(+r.released_vazn||0)>0.005)){cls+=" partial";tip="Qisman yechilgan";val="0"}return `<td class="${cls}" title="${tip}">${val}</td>`}).join("")}</tr>`).join("")}</tbody></table>`}
function total(rows,map){rows=rows||[];let out={};for(let k in map)out[k]=rows.reduce((a,r)=>a+(+r[map[k]]||0),0);return out} function basicTotal(rows,label="IBK bo'yicha Jami",nameKey="name"){let t=total(rows,{partiya:"partiya",vazn:"vazn",qiymat:"qiymat",tolov:"tolov"});let r={key:{},partiya:t.partiya,vazn:t.vazn,qiymat:t.qiymat,tolov:t.tolov};r[nameKey]=label;return [r].concat(rows||[])} function companyTotal(rows){let k=DATA.kpis||{};return [{key:{},korxona:"IBK bo'yicha Jami",stir:"",partiya:k.partiya||0,vazn:k.vazn||0,qiymat:k.qiymat||0,tolov:k.tolov||0,depozit:k.depozit_matched||0}].concat(rows||[])}function expiredTotal(rows){let t=total(rows,{partiya:"partiya",vazn:"vazn",qiymat:"qiymat",tolov:"tolov"});return [{key:{},korxona:"IBK bo'yicha Jami",stir:"",rejim:"",post:"",kun:"",partiya:t.partiya,vazn:t.vazn,qiymat:t.qiymat,tolov:t.tolov}].concat(rows||[])} function foodRows(){let t=DATA.food_total||{name:"IBK bo'yicha Jami",vazn:0,qiymat:0,over_vazn:0,over_qiymat:0,ulush:100};return [t].concat(DATA.food||[])} function regimeSummaryRows(){let rows=(DATA.summary||basicTotal(DATA.regimes||[])).map(r=>{if(["IM70","IM74","TR80"].includes(r.name||r.rejim))return Object.assign({key:{view:"regime_posts",regime:r.name||r.rejim}},r);return r});return rows}
function expiredSummaryRows(){return (DATA.expired_summary||basicTotal(DATA.expired||[])).map((r,i)=>i===0?Object.assign({key:{view:"expired_inline"}},r):r)} function periodRows(r){let rows=(r&&r.rows)||[], t=total(rows,{partiya:"partiya",qiymat:"qiymat",tolov:"tolov"});return [{key:{},company:"IBK bo'yicha Jami",stir:"",decl:"",partiya:t.partiya,qiymat:t.qiymat,tolov:t.tolov}].concat(rows)} function expiredPostRegimeRows(){let rows=DATA.expired_post_regime||[], t=total(rows,{jami_partiya:"jami_partiya",jami_qiymat:"jami_qiymat",expired_partiya:"expired_partiya",expired_qiymat:"expired_qiymat",im70_partiya:"im70_partiya",im74_partiya:"im74_partiya",tr80_partiya:"tr80_partiya"});t.ulush=t.jami_qiymat? t.expired_qiymat/t.jami_qiymat*100:0;t.post="IBK bo'yicha Jami";t.key={};return [t].concat(rows)} function executiveSummary(){if(!DATA)return "";let k=DATA.kpis||{}, top=(DATA.top_value||[])[0]||{}, dep=(DATA.top_deposit||[])[0]||{}, own=(DATA.warehouse||[]).find(r=>(r.name||"")==="O'z ombor")||{}, exp=(DATA.expired_post_regime||[])[0]||{};let items=[['Umumiy nazorat',`${fmtI(k.partiya)} partiya, ${fmtN(k.qiymat)} ming $ qiymat.`],['Muddati o\'tgan',`${fmtI(k.expired)} partiya. Asosiy kesim: ${esc(exp.post||'postlar')}.`],['Eng yirik korxona',`${esc(top.korxona||'-')} - ${fmtN(top.qiymat||0)} ming $.`],['O\'z ombor',`${fmtI(own.partiya||0)} partiya, ${fmtN(own.qiymat||0)} ming $.`],['Depozit yetakchisi',`${esc(dep.korxona||'-')} - ${fmtN(dep.depozit||0)} mln so\'m.`]];return `<div class="panel exec-summary"><h2>Rahbar uchun qisqa xulosa</h2><div class="summary-grid">${items.map(x=>`<div class="summary-item"><b>${x[0]}</b><span>${x[1]}</span></div>`).join("")}</div></div>`}
function showKpi(kind){if(!DATA)return;let titles={partiya:"Partiya kelib chiqishi",vazn:"Vazn kelib chiqishi",qiymat:"Qiymat kelib chiqishi",tolov:"Kutilayotgan to'lov kelib chiqishi",depozit:"Depozit kelib chiqishi",expired:"Muddati o'tgan partiya kelib chiqishi"};dlgTitle.textContent=titles[kind]||"KPI asosi";if(kind==="depozit"){dlgBody.innerHTML=table(companyCols(),companyTotal(DATA.top_deposit||[]));dlg.showModal();return}if(kind==="expired"){dlgBody.innerHTML=table([{k:"post",t:"Post",w:"22%"},{k:"jami_partiya",t:"Jami partiya",n:1,f:fmtI},{k:"jami_qiymat",t:"Jami qiymat (ming $)",n:1,f:fmtN},{k:"expired_partiya",t:"Muddati o'tgan partiya",n:1,f:fmtI},{k:"expired_qiymat",t:"Muddati o'tgan qiymat (ming $)",n:1,f:fmtN},{k:"ulush",t:"Partiyadagi ulushi (%)",n:1,f:fmtN},{k:"im70_partiya",t:"IM70 partiya",n:1,f:fmtI},{k:"im74_partiya",t:"IM74 partiya",n:1,f:fmtI},{k:"tr80_partiya",t:"TR80 partiya",n:1,f:fmtI}],expiredPostRegimeRows());dlg.showModal();return}dlgBody.innerHTML=table([{k:"rejim",t:"Rejim",w:"24%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"Kutilayotgan to'lov (mln so'm)",n:1,f:fmtN}],basicTotal(DATA.regimes||[],"IBK bo'yicha Jami","rejim"));dlg.showModal()}
 function numbered(rows){return (rows||[]).map((r,i)=>Object.assign({rn:i?i:""},r))} function dateOptions(){let seen=new Set();return ARCHIVE.filter(r=>!seen.has(r.date)&&seen.add(r.date)).map(r=>`<option value="${r.date}">${r.date}</option>`).join("")} function xls(kind){return DATA?`<a class="btn light" href="/api/export?kind=${kind}&report=${DATA.id}&token=${TOKEN}">Excel</a>`:""}
function bars(rows,label,value,fmt){rows=(rows||[]).filter(r=>(+r[value]||0)>0);let max=rows.reduce((m,r)=>Math.max(m,+r[value]||0),0)||1,sum=rows.reduce((a,r)=>a+(+r[value]||0),0)||1;return `<div class=bars>${rows.map(r=>{let pct=(+r[value]||0)/sum*100,w=(+r[value]||0)/max*100;return `<div class=barrow><div title="${esc(r[label])}">${esc(r[label])}</div><div class=bar><span style="width:${Math.max(5,w)}%">${pct.toFixed(1)}%</span></div><b>${fmt(r[value])}</b></div>`}).join("")}</div>`} function by(rows,k){return (rows||[]).slice().sort((a,b)=>(+b[k]||0)-(+a[k]||0))} function expiredBlockTable(rows){rows=rows||[];let body=rows.map(r=>{let group=!r.korxona&&!r.stir&&r.name;if(group)return `<tr class="merged-row"><td colspan="3" title="${esc(r.name)}">${esc(r.name)}</td><td class="num">${fmtI(r.partiya)}</td><td class="num">${fmtN(r.qiymat)}</td><td class="num">${fmtN(r.vazn)}</td><td class="num">${fmtN(r.tolov)}</td><td></td></tr>`;return `<tr title="${esc((r.korxona||'')+' '+(r.reason||''))}"><td class="num">${esc(r.name)}</td><td class="text col-korxona">${esc(r.korxona)}</td><td class="num">${esc(r.stir)}</td><td class="num">${fmtI(r.partiya)}</td><td class="num">${fmtN(r.qiymat)}</td><td class="num">${fmtN(r.vazn)}</td><td class="num">${fmtN(r.tolov)}</td><td class="text col-reason">${esc(r.reason)}</td></tr>`}).join("");return `<table class="sample-like expired-detail-sample"><colgroup><col style="width:44px"><col style="width:330px"><col style="width:116px"><col style="width:74px"><col style="width:96px"><col style="width:96px"><col style="width:116px"><col style="width:320px"></colgroup><thead><tr><th rowspan=2>T/r</th><th colspan=2>Korxona ma'lumotlari</th><th rowspan=2>Partiya</th><th rowspan=2>Qiymati<br>(ming $)</th><th rowspan=2>Vazni<br>(tn)</th><th rowspan=2>Kutilayotgan<br>(mln so'm)</th><th rowspan=2>Saqlanish sababi</th></tr><tr><th>Korxona nomi</th><th>STIR</th></tr></thead><tbody>${body}</tbody></table>`}
function uploadPanel(){return `<div class=stack><div class=panel><h2>Fayl yuklash</h2><div class=muted>Bu modul umumiy: BNRTE va To'lovlar uchun fayllar alohida yuklanadi.</div></div><div class=panel><h2>BNRTE jamlanma</h2><form class="upload" id="upload"><div><label>Asos fayl</label><input name="source" type="file" accept=".xls,.xlsx,.html,.htm" required></div><div><label>Depozit fayl</label><input name="deposit" type="file" accept=".xlsx"></div><div></div><button>BNRTE yuklash</button></form><div class=muted>Hisobot sanasi asos fayl nomidan avtomatik aniqlanadi.</div></div><div class=panel><h2>To'lovlar jadvallari</h2><form class="upload" id="tolovUpload"><div><label>To'lov baza fayli</label><input name="source" type="file" accept=".xlsx,.xls" required></div><div class=muted>04.06+07.06.2026 kabi asos fayl yuklanadi.</div><div></div><button>To'lov jadvallarini shakllantirish</button></form><div id="tolovUploadResult" class=muted>Natijada 13 ta Excel fayl shakllanadi va To'lovlar modulidagi yuklash tugmalariga ulanadi.</div></div><div class=panel><h2>Omborlar reestri</h2><div class="wr-upload-form"><label style="font-size:13px;font-weight:600">Yangi reestri faylini yuklash:</label><input type="file" id="wrFile" accept=".xlsx,.xls" style="font-size:12px"><button onclick="uploadWrRegistry()">Yuklash</button><span class="muted" id="wrUploadStatus"></span></div><div class=muted style="font-size:12px">omborlarReestri*.xlsx formatidagi fayl yuklanadi. Yuklangandan so'ng Omborlar → Reestri bo'limida aks etadi.</div></div><div class=panel><h2>✈ AVIA AWB</h2><div class="wr-upload-form"><label style="font-size:13px;font-weight:600">AWB Excel yuklash:</label><input type="file" id="aviaFile" accept=".xlsx,.xls" style="font-size:12px" onchange="uploadAviaAwbDirect(this.files[0])"><span class="muted" id="aviaUploadStatus"></span></div><div class=muted style="font-size:12px">"Yuklarni qabul qilish*.xlsx" formatidagi fayl yuklanadi. Yuklangandan so'ng BNRTE → AVIA AWB bo'limida aks etadi.</div></div></div>`} function bindUpload(){let f=$("upload");if(f)f.onsubmit=async e=>{e.preventDefault();$("status").textContent="BNRTE fayllari yuklanyapti...";let j=await api("/api/reports",{method:"POST",body:new FormData(f)});poll(j.job_id)};let tf=$("tolovUpload");if(tf)tf.onsubmit=async e=>{e.preventDefault();$("status").textContent="To'lovlar shakllantirilyapti...";let j=await api("/api/tolov",{method:"POST",body:new FormData(tf)});PAYMENTS=j.payments||[];$("tolovUploadResult").innerHTML=`Tayyor: ${fmtI(PAYMENTS.reduce((a,r)=>a+(+r.rows||0),0))} qator, ${fmtN(PAYMENTS.reduce((a,r)=>a+(+r.sum||0),0))} so'm. <button class="light" onclick="GROUP='payments';TAB='pay_lists';render()">To'lovlar jadvaliga o'tish</button>`;$("status").textContent="To'lovlar tayyor"}}
const ownCols=[{k:"korxona",t:"Korxona nomi",w:"20%"},{k:"stir",t:"STIR",w:"9%"},{k:"muddat",t:"Muddat",w:"9%"},{k:"kun_hisobi",t:"Kun hisobi",w:"7%",n:1,f:fmtI},{k:"partiya",t:"Partiya",w:"7%",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",w:"9%",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",w:"10%",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",w:"10%",n:1,f:fmtN},{k:"tovar",t:"Tovar",w:"19%"}];
function companyCols(){let k=DATA&&DATA.kpis||{};let dh=k.depozit?`Depozit ${fmtN(k.depozit)} (mln so'm)`:"Depozit (mln so'm)";return [{k:"korxona",t:"Korxona nomi",w:"48%"},{k:"stir",t:"STIR",w:"10%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN},{k:"depozit",t:dh,n:1,f:fmtN}]}
function _rtFmt(n){return(+n||0).toLocaleString('ru-RU',{minimumFractionDigits:2,maximumFractionDigits:2})}
function _rtPct(v,t){return t>0?((v/t)*100).toFixed(1).replace('.',',')+' %':'0 %'}
function _rtBar(v,mx,on){let w=mx>0?(v/mx*100).toFixed(1):0;return `height:100%;border-radius:5px;width:${w}%;background:${on?'linear-gradient(90deg,#1f6fb8,#35c1c9)':'#ccd6e0'};transform-origin:left center;animation:rtabBarGrow .65s cubic-bezier(.22,1,.36,1) both;`}
function copyText(text,btn){navigator.clipboard.writeText(String(text||'')).catch(()=>{let t=document.createElement('textarea');t.value=text;document.body.appendChild(t);t.select();document.execCommand('copy');document.body.removeChild(t)});if(btn){const s=btn.innerHTML;btn.innerHTML='✓';btn.style.color='#16a34a';setTimeout(()=>{btn.innerHTML=s;btn.style.color=''},900)}}
function animatePlaceholder(id,phrases,ms){ms=ms||2800;const el=document.getElementById(id);if(!el)return;clearInterval(el._ph);let i=0;el._ph=setInterval(()=>{if(document.activeElement===el||el.value)return;el.setAttribute('placeholder',phrases[i++%phrases.length])},ms)}
const _COPY_SVG=`<svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><rect x="9" y="9" width="13" height="13" rx="2"/><path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"/></svg>`;
const _SEARCH_SVG=`<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="#9aa6b8" stroke-width="2" stroke-linecap="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>`;
function _cpBtn(text){return `<button class="rtab-stir-copy" onclick="event.stopPropagation();copyText(${JSON.stringify(String(text||''))},this)" title="Nusxalash">${_COPY_SVG}</button>`}
function _rtMerge(tv,td,all){
  const depByStir={};(td||[]).forEach(r=>{if(r.stir)depByStir[r.stir]=r.depozit||0});
  const allArr=(all||[]).map(r=>Object.assign({},r,{depozit:depByStir[r.stir]||r.depozit||0}));
  let m={};(tv||[]).forEach(r=>{let k=r.stir||r.korxona;if(!m[k])m[k]=Object.assign({},r);else Object.assign(m[k],r)});
  (td||[]).forEach(r=>{let k=r.stir||r.korxona;if(!m[k])m[k]=Object.assign({},r);else{m[k].depozit=m[k].depozit||r.depozit;m[k].vazn=m[k].vazn||r.vazn;m[k].partiya=m[k].partiya||r.partiya;m[k].qiymat=m[k].qiymat||r.qiymat}});
  return {top30:Object.values(m),allData:allArr.length?allArr:Object.values(m)};
}
window._RT={metric:'q',query:'',data:[],allData:[]};
function _rtRows(){
  const{metric:m,query:q,data:top30,allData}=window._RT;
  const fld={p:'partiya',q:'qiymat',d:'depozit',v:'vazn'}[m];
  const pool=allData.length?allData:top30;
  const sorted=[...pool].sort((a,b)=>(+b[fld]||0)-(+a[fld]||0));
  sorted.forEach((r,i)=>{r._rank=i+1});
  const qL=(q||'').trim().toLowerCase();
  const vis=qL?sorted.filter(r=>(r.korxona||'').toLowerCase().includes(qL)||(r.stir||'').includes(qL)):sorted.slice(0,30);
  const tot={q:pool.reduce((s,r)=>s+(+r.qiymat||0),0),d:pool.reduce((s,r)=>s+(+r.depozit||0),0),v:pool.reduce((s,r)=>s+(+r.vazn||0),0)};
  const mx={q:sorted[0]?+sorted[0].qiymat||1:1,d:[...pool].sort((a,b)=>b.depozit-a.depozit)[0]?.depozit||1,v:[...pool].sort((a,b)=>b.vazn-a.vazn)[0]?.vazn||1};
  const expSet=new Set((DATA&&DATA.expired||[]).map(x=>x.stir||'').filter(Boolean));
  const expKSet=new Set((DATA&&DATA.expired||[]).map(x=>(x.korxona||'').toLowerCase()).filter(Boolean));
  if(!vis.length)return `<div class="rtab-empty"><svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="#c2ccd9" stroke-width="1.7" stroke-linecap="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg><div style="font-size:14px;font-weight:600;color:#5b6b82;">Hech narsa topilmadi</div></div>`;
  return vis.map((r,idx)=>{
    const isTop=r._rank<=3,pOn=m==='p',qOn=m==='q',dOn=m==='d',vOn=m==='v';
    const hasExp=!!(r.stir&&expSet.has(r.stir))||expKSet.has((r.korxona||'').toLowerCase());
    const nameHtml=hasExp
      ?`<span style="color:#dc2626;font-weight:700">${esc(r.korxona||'')}</span><span title="Muddati o'tgan tovarlari mavjud" style="display:inline-block;background:#fef2f2;border:1px solid #fca5a5;border-radius:4px;padding:0 5px;font-size:10px;font-weight:700;color:#dc2626;margin-left:5px;vertical-align:middle">⚠ muddati o'tgan</span>`
      :esc(r.korxona||'');
    return `<div class="rtab-row${hasExp?' rtab-row-exp':''}" style="animation-delay:${(idx*0.028).toFixed(3)}s" onclick='detail(${JSON.stringify({stir:r.stir}).replaceAll("'","&#39;")})'>
      <div><span class="rtab-rank ${isTop?'top':'reg'}" style="animation-delay:${(idx*0.028+0.08).toFixed(3)}s">${r._rank}</span></div>
      <div style="overflow:hidden"><div class="rtab-name" title="${esc(r.korxona||'')}">${nameHtml}</div><div class="rtab-stir"><span>STIR: ${esc(r.stir||'')}</span>${r.stir?_cpBtn(r.stir):''}</div></div>
      <div class="rtab-p${pOn?' on':''}">${(+r.partiya||0).toLocaleString('ru-RU')}</div>
      <div class="rtab-cell${qOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.qiymat||0,tot.q)}</span><span class="rtab-val">${_rtFmt(r.qiymat)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.qiymat||0,mx.q,qOn)}"></div></div></div>
      <div class="rtab-cell${dOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.depozit||0,tot.d)}</span><span class="rtab-val">${_rtFmt(r.depozit)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.depozit||0,mx.d,dOn)}"></div></div></div>
      <div class="rtab-cell${vOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.vazn||0,tot.v)}</span><span class="rtab-val">${_rtFmt(r.vazn)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.vazn||0,mx.v,vOn)}"></div></div></div>
    </div>`
  }).join('')
}
function _rtColHead(){
  const m=window._RT.metric;
  const h=(k,lbl,unit)=>`<div class="num${m===k?' on':''}">${lbl}${unit?` <span style="font-size:10px;opacity:.72">${unit}</span>`:''}</div>`;
  return `<div></div><div>Korxona / STIR</div>${h('p','Partiya','')}`+
    h('q','Qiymat','(ming $)')+h('d','Depozit','(mln. so\'m)')+h('v','Vazn','(tn.)');
}
function _rtBtn(k,lbl){
  const on=window._RT.metric===k;
  const icon=window._RT.icons[k]||'';
  const arrow=on?`<svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round"><path d="M12 5v14"/><path d="m19 12-7 7-7-7"/></svg>`:'';
  return `<button class="${on?'on':''}" onclick="rtSort('${k}')">${icon}<span>${lbl}</span>${arrow}</button>`;
}
function _rtRender(){
  const el=document.getElementById('rtabRows');if(el)el.innerHTML=_rtRows();
  const ch=document.getElementById('rtabCol');if(ch)ch.innerHTML=_rtColHead();
  const sb=document.getElementById('rtabSortBtns');
  if(sb)sb.innerHTML=_rtBtn('p','Partiya')+_rtBtn('q','Qiymat')+_rtBtn('d','Depozit')+_rtBtn('v','Vazn');
}
function rtSort(k){window._RT.metric=k;_rtRender()}
function rtSearch(v){window._RT.query=v;_rtRender()}
function korxonaRatingPanel(tv,td,allC){
  const merged=_rtMerge(tv,td,allC);
  window._RT.data=merged.top30;window._RT.allData=merged.allData;
  window._RT.metric='q';window._RT.query='';
  window._RT.icons={
    p:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="4" y1="9" x2="20" y2="9"/><line x1="4" y1="15" x2="20" y2="15"/><line x1="10" y1="3" x2="8" y2="21"/><line x1="16" y1="3" x2="14" y2="21"/></svg>`,
    q:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/></svg>`,
    d:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><path d="M21 12V7H5a2 2 0 0 1 0-4h14v4"/><path d="M3 5v14a2 2 0 0 0 2 2h16v-5"/><path d="M18 12a2 2 0 0 0 0 4h4v-4Z"/></svg>`,
    v:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="5" r="2"/><path d="M6.5 7h11l2.4 12.2a1 1 0 0 1-1 1.2H5.1a1 1 0 0 1-1-1.2L6.5 7Z"/></svg>`
  };
  setTimeout(()=>animatePlaceholder('rtSearchInput',['Korxona nomi yozing…','STIR raqam kiriting…','Masalan: Tashkent Cargo…','Qidirish uchun yozing…']),50);
  return `<div class="rtab-wrap panel" style="padding:0;overflow:hidden;">
    <div class="rtab-head">
      <div class="rtab-title"><div class="rtab-accent"></div><div><div class="rtab-sup">Interaktiv hisobot</div><div class="rtab-h1"><span class="rtab-gradient">TOP 30</span> korxona reytingi</div></div></div>
      <div class="rtab-controls">
        <div class="rtab-search">${_SEARCH_SVG}<input oninput="rtSearch(this.value)" placeholder="Korxona nomi yozing…" id="rtSearchInput" autocomplete="off"></div>
        <div class="rtab-sortbtns" id="rtabSortBtns">
          ${_rtBtn('p','Partiya')}${_rtBtn('q','Qiymat')}${_rtBtn('d','Depozit')}${_rtBtn('v','Vazn')}
        </div>
      </div>
    </div>
    <div class="rtab-colhead" id="rtabCol">${_rtColHead()}</div>
    <div class="rtab-body" id="rtabRows">${_rtRows()}</div>
  </div>`
}

// ═══════════════════════════════════════════════════════════════
// AWB RATING PANEL
// ═══════════════════════════════════════════════════════════════
const _AW_ICONS={
  j:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><rect x="3" y="3" width="18" height="18" rx="2"/><path d="M3 9h18M9 21V9"/></svg>`,
  v:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="5" r="2"/><path d="M6.5 7h11l2.4 12.2a1 1 0 0 1-1 1.2H5.1a1 1 0 0 1-1-1.2L6.5 7Z"/></svg>`,
  r:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><path d="M17.8 19.2L16 11l3.5-3.5C21 6 21 4 19.5 2.5S18 2 16.5 3.5L13 7 4.8 5.2"/><path d="m2 2 20 20"/></svg>`
};
window._AW={metric:'j',query:'',data:[]};
function _awBtn(k,lbl){
  const on=window._AW.metric===k;
  const arrow=on?`<svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round"><path d="M12 5v14"/><path d="m19 12-7 7-7-7"/></svg>`:'';
  return `<button class="${on?'on':''}" onclick="awSort('${k}')">${_AW_ICONS[k]||''}<span>${lbl}</span>${arrow}</button>`;
}
function _awColHead(){
  const m=window._AW.metric;
  const h=(k,l,u)=>`<div class="num${m===k?' on':''}">${l}${u?` <span style="font-size:10px;opacity:.72">${u}</span>`:''}</div>`;
  return `<div></div><div>AWB / Qabul qiluvchi</div><div class="num">Reys</div>${h('j','Joylar','')}${h('v','Vazn','(tn.)')}${h('r','','Holat')}`;
}
function _awRows(){
  const{metric:m,query:q,data}=window._AW;
  const fld={j:'joylar',v:'vazn',r:'flights'}[m]||'joylar';
  const sorted=[...data].sort((a,b)=>(+b[fld]||0)-(+a[fld]||0));
  sorted.forEach((r,i)=>{r._rank=i+1});
  const qL=(q||'').trim().toLowerCase();
  const vis=qL?sorted.filter(r=>(r.awb||'').toLowerCase().includes(qL)||(r.company||'').toLowerCase().includes(qL)):sorted.slice(0,30);
  const totJ=data.reduce((s,r)=>s+(+r.joylar||0),0);
  const totV=data.reduce((s,r)=>s+(+r.vazn||0),0);
  const mxJ=sorted[0]?+sorted[0].joylar||1:1;
  const mxV=[...data].sort((a,b)=>b.vazn-a.vazn)[0]?.vazn||1;
  const emptyIcon=`<svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="#c2ccd9" stroke-width="1.7" stroke-linecap="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>`;
  if(!vis.length)return `<div class="rtab-empty">${emptyIcon}<div style="font-size:14px;font-weight:600;color:#5b6b82">Hech narsa topilmadi</div></div>`;
  return vis.map((r,idx)=>{
    const isTop=r._rank<=3,jOn=m==='j',vOn=m==='v',rOn=m==='r';
    const od=r.is_overdue;
    const holatHtml=`<div style="font-size:11.5px;font-weight:700;color:${od?'#dc2626':'#16a34a'}">${od?'Muddati o\'tgan':'Kuzatuvda'}</div>${r.days_ago!=null?`<div style="font-size:11px;color:#9aa6b8;margin-top:1px">${r.days_ago} kun</div>`:''}`;
    return `<div class="rtab-row" style="animation-delay:${(idx*0.028).toFixed(3)}s">
      <div><span class="rtab-rank ${isTop?'top':'reg'}" style="animation-delay:${(idx*0.028+0.08).toFixed(3)}s">${r._rank}</span></div>
      <div style="overflow:hidden"><div class="rtab-name" title="${esc(r.awb||'')}">${esc(r.awb||'')} ${_cpBtn(r.awb||'')}${r.arrival_date?`<span style="font-size:10.5px;color:#9aa6b8;margin-left:4px">${esc(r.arrival_date)}</span>`:''}</div><div class="rtab-stir"><span>${esc(r.company||'—')}</span>${r.country_latin?`<span style="margin-left:5px;padding:1px 5px;background:#eef2f6;border-radius:4px;font-size:10px">${esc(r.country_latin)}</span>`:''}</div></div>
      <div class="rtab-p${rOn?' on':''}">${r.flights||1}</div>
      <div class="rtab-cell${jOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.joylar||0,totJ)}</span><span class="rtab-val">${(+r.joylar||0).toLocaleString('ru-RU')}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.joylar||0,mxJ,jOn)}"></div></div></div>
      <div class="rtab-cell${vOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.vazn||0,totV)}</span><span class="rtab-val">${_rtFmt(r.vazn)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.vazn||0,mxV,vOn)}"></div></div></div>
      <div>${holatHtml}</div>
    </div>`;
  }).join('');
}
function _awRender(){
  const el=document.getElementById('awRows');if(el)el.innerHTML=_awRows();
  const ch=document.getElementById('awCol');if(ch)ch.innerHTML=_awColHead();
  const sb=document.getElementById('awSortBtns');
  if(sb)sb.innerHTML=_awBtn('j','Joylar')+_awBtn('v','Vazn')+_awBtn('r','Reys');
}
function awSort(k){window._AW.metric=k;_awRender()}
function awSearch(v){window._AW.query=v;_awRender()}
function awbRatingPanel(awbData){
  if(!awbData||!awbData.loaded)return'';
  const list=awbData.awb_list||[];
  window._AW.data=list;window._AW.metric='j';window._AW.query='';
  setTimeout(()=>animatePlaceholder('awSearchInput',['AWB raqami yozing…','Korxona nomi kiriting…','Masalan: 555-12345678…','Qidirish uchun yozing…']),50);
  const totJ=list.reduce((s,r)=>s+(+r.joylar||0),0);
  const totV=list.reduce((s,r)=>s+(+r.vazn||0),0);
  const act=list.filter(r=>!r.is_overdue).length, od=list.filter(r=>r.is_overdue).length;
  return `<div class="rtab-wrap panel" style="padding:0;overflow:hidden;--rtab-grid:28px 1fr 54px minmax(100px,1fr) minmax(110px,1fr) 86px">
    <div class="rtab-head">
      <div class="rtab-title"><div class="rtab-accent" style="background:linear-gradient(180deg,#0ea5e9,#06b6d4)"></div><div><div class="rtab-sup">Interaktiv hisobot</div><div class="rtab-h1" style="font-size:19px"><span class="rtab-gradient" style="background:linear-gradient(135deg,#0ea5e9,#06b6d4);-webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent">✈ ${list.length}</span> ta AWB reytingi</div></div></div>
      <div class="rtab-controls">
        <div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center">
          <span style="font-size:12px;padding:3px 8px;background:#dcfce7;color:#16a34a;border-radius:8px;font-weight:600">Kuzatuvda: ${act}</span>
          <span style="font-size:12px;padding:3px 8px;background:#fee2e2;color:#dc2626;border-radius:8px;font-weight:600">O'tgan: ${od}</span>
          <span style="font-size:12px;color:#9aa6b8">${totJ.toLocaleString('ru-RU')} joy · ${_rtFmt(totV)} tn</span>
        </div>
        <div class="rtab-search">${_SEARCH_SVG}<input oninput="awSearch(this.value)" placeholder="AWB raqami yozing…" id="awSearchInput" autocomplete="off"></div>
        <div class="rtab-sortbtns" id="awSortBtns">${_awBtn('j','Joylar')}${_awBtn('v','Vazn')}${_awBtn('r','Reys')}</div>
      </div>
    </div>
    <div class="rtab-colhead" id="awCol">${_awColHead()}</div>
    <div class="rtab-body" id="awRows">${_awRows()}</div>
  </div>`;
}

// ═══════════════════════════════════════════════════════════════
// OMBOR RATING PANEL
// ═══════════════════════════════════════════════════════════════
window._WH={metric:'q',query:'',data:[],wrMap:{}};
function _whBtn(k,lbl){
  const on=window._WH.metric===k;
  const ICONS={p:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="4" y1="9" x2="20" y2="9"/><line x1="4" y1="15" x2="20" y2="15"/><line x1="10" y1="3" x2="8" y2="21"/><line x1="16" y1="3" x2="14" y2="21"/></svg>`,q:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/></svg>`,v:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="5" r="2"/><path d="M6.5 7h11l2.4 12.2a1 1 0 0 1-1 1.2H5.1a1 1 0 0 1-1-1.2L6.5 7Z"/></svg>`,t:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><rect x="2" y="5" width="20" height="14" rx="2"/><line x1="2" y1="10" x2="22" y2="10"/></svg>`};
  const arrow=on?`<svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round"><path d="M12 5v14"/><path d="m19 12-7 7-7-7"/></svg>`:'';
  return `<button class="${on?'on':''}" onclick="whSort('${k}')">${ICONS[k]||''}<span>${lbl}</span>${arrow}</button>`;
}
function _whColHead(){
  const m=window._WH.metric;
  const h=(k,l,u)=>`<div class="num${m===k?' on':''}">${l}${u?` <span style="font-size:10px;opacity:.72">${u}</span>`:''}</div>`;
  return `<div></div><div>Ombor / Litsenziya</div>${h('p','Partiya','')}${h('q','Qiymat','(ming $)')}${h('v','Vazn','(tn.)')}${h('t','To\'lov','(mln so\'m)')}`;
}
function _whRows(){
  const{metric:m,query:q,data,wrMap}=window._WH;
  const fld={p:'partiya',q:'qiymat',v:'vazn',t:'tolov'}[m]||'qiymat';
  const sorted=[...data].sort((a,b)=>(+b[fld]||0)-(+a[fld]||0));
  sorted.forEach((r,i)=>{r._rank=i+1});
  const qL=(q||'').trim().toLowerCase();
  const vis=qL?sorted.filter(r=>{
    const lic=(wrMap[(r.name||'').toLowerCase()]||{}).lic_num||'';
    return (r.name||'').toLowerCase().includes(qL)||lic.toLowerCase().includes(qL);
  }):sorted.slice(0,30);
  const tot={p:data.reduce((s,r)=>s+(+r.partiya||0),0),q:data.reduce((s,r)=>s+(+r.qiymat||0),0),v:data.reduce((s,r)=>s+(+r.vazn||0),0),t:data.reduce((s,r)=>s+(+r.tolov||0),0)};
  const mx={p:sorted[0]?+sorted[0].partiya||1:1,q:[...data].sort((a,b)=>b.qiymat-a.qiymat)[0]?.qiymat||1,v:[...data].sort((a,b)=>b.vazn-a.vazn)[0]?.vazn||1,t:[...data].sort((a,b)=>b.tolov-a.tolov)[0]?.tolov||1};
  const emptyIcon=`<svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="#c2ccd9" stroke-width="1.7" stroke-linecap="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>`;
  if(!vis.length)return `<div class="rtab-empty">${emptyIcon}<div style="font-size:14px;font-weight:600;color:#5b6b82">Hech narsa topilmadi</div></div>`;
  const pOn=m==='p',qOn=m==='q',vOn=m==='v',tOn=m==='t';
  return vis.map((r,idx)=>{
    const isTop=r._rank<=3;
    const wr=(wrMap[(r.name||'').toLowerCase()])||{};
    const lic=wr.lic_num||'';
    const badges=(wr.fvv?`<span class="wr-badge-fvv" style="font-size:9px;padding:0 3px">FVV</span>`:'')+(wr.ssv?`<span class="wr-badge-ssv" style="font-size:9px;padding:0 3px">GSP</span>`:'');
    return `<div class="rtab-row" style="animation-delay:${(idx*0.028).toFixed(3)}s">
      <div><span class="rtab-rank ${isTop?'top':'reg'}" style="animation-delay:${(idx*0.028+0.08).toFixed(3)}s">${r._rank}</span></div>
      <div style="overflow:hidden"><div class="rtab-name" title="${esc(r.name||'')}">${esc(r.name||'')}${badges}</div><div class="rtab-stir">${lic?`<span>${esc(lic)}</span>${_cpBtn(lic)}`:`<span style="color:#c2ccd9">Litsenziya yo'q</span>`}</div></div>
      <div class="rtab-p${pOn?' on':''}">${(+r.partiya||0).toLocaleString('ru-RU')}</div>
      <div class="rtab-cell${qOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.qiymat||0,tot.q)}</span><span class="rtab-val">${_rtFmt(r.qiymat)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.qiymat||0,mx.q,qOn)}"></div></div></div>
      <div class="rtab-cell${vOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.vazn||0,tot.v)}</span><span class="rtab-val">${_rtFmt(r.vazn)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.vazn||0,mx.v,vOn)}"></div></div></div>
      <div class="rtab-cell${tOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.tolov||0,tot.t)}</span><span class="rtab-val">${_rtFmt(r.tolov)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.tolov||0,mx.t,tOn)}"></div></div></div>
    </div>`;
  }).join('');
}
function _whRender(){
  const el=document.getElementById('whRows');if(el)el.innerHTML=_whRows();
  const ch=document.getElementById('whCol');if(ch)ch.innerHTML=_whColHead();
  const sb=document.getElementById('whSortBtns');
  if(sb)sb.innerHTML=_whBtn('p','Partiya')+_whBtn('q','Qiymat')+_whBtn('v','Vazn')+_whBtn('t','To\'lov');
}
function whSort(k){window._WH.metric=k;_whRender()}
function whSearch(v){window._WH.query=v;_whRender()}
function omborRatingPanel(rows,wrArg){
  rows=(rows||[]).filter(r=>r.name&&r.name!=="IBK bo'yicha Jami");
  const wrRows=Array.isArray(wrArg)?wrArg:(wrArg&&wrArg.warehouses)||[];
  const wrMap={};(wrRows||[]).forEach(w=>{if(w.name)wrMap[w.name.toLowerCase()]={lic_num:w.lic_num||'',lic_exp:w.lic_exp||'',fvv:w.fvv||'',ssv:w.ssv||''}});
  window._WH.data=rows;window._WH.metric='q';window._WH.query='';window._WH.wrMap=wrMap;
  setTimeout(()=>animatePlaceholder('whSearchInput',['Ombor nomi yozing…','Litsenziya raqami kiriting…','Qidirish uchun yozing…']),50);
  return `<div class="rtab-wrap panel" style="padding:0;overflow:hidden;--rtab-grid:28px 1fr minmax(80px,0.7fr) minmax(120px,1fr) minmax(120px,1fr) minmax(120px,1fr)">
    <div class="rtab-head">
      <div class="rtab-title"><div class="rtab-accent" style="background:linear-gradient(180deg,#7c3aed,#a78bfa)"></div><div><div class="rtab-sup">Interaktiv hisobot</div><div class="rtab-h1"><span class="rtab-gradient" style="background:linear-gradient(135deg,#7c3aed,#a78bfa);-webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent">${rows.length}</span> ta ombor reytingi</div></div></div>
      <div class="rtab-controls">
        <div class="rtab-search">${_SEARCH_SVG}<input oninput="whSearch(this.value)" placeholder="Ombor nomi yozing…" id="whSearchInput" autocomplete="off"></div>
        <div class="rtab-sortbtns" id="whSortBtns">${_whBtn('p','Partiya')}${_whBtn('q','Qiymat')}${_whBtn('v','Vazn')}${_whBtn('t','To\'lov')}</div>
      </div>
    </div>
    <div class="rtab-colhead" id="whCol">${_whColHead()}</div>
    <div class="rtab-body" id="whRows">${_whRows()}</div>
  </div>`;
}
function omborRatingUpdateWR(wrArg){
  if(!wrArg)return;
  const wrRows=Array.isArray(wrArg)?wrArg:(wrArg.warehouses)||[];
  if(!wrRows.length)return;
  const wrMap={};wrRows.forEach(w=>{if(w.name)wrMap[w.name.toLowerCase()]={lic_num:w.lic_num||'',lic_exp:w.lic_exp||''}});
  window._WH.wrMap=wrMap;_whRender();
}

// ═══════════════════════════════════════════════════════════════
// TOVAR RATING PANEL
// ═══════════════════════════════════════════════════════════════
window._GD={metric:'q',query:'',data:[]};
function _gdBtn(k,lbl){
  const on=window._GD.metric===k;
  const ICONS={p:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="4" y1="9" x2="20" y2="9"/><line x1="4" y1="15" x2="20" y2="15"/><line x1="10" y1="3" x2="8" y2="21"/><line x1="16" y1="3" x2="14" y2="21"/></svg>`,q:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/></svg>`,v:`<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="5" r="2"/><path d="M6.5 7h11l2.4 12.2a1 1 0 0 1-1 1.2H5.1a1 1 0 0 1-1-1.2L6.5 7Z"/></svg>`};
  const arrow=on?`<svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round"><path d="M12 5v14"/><path d="m19 12-7 7-7-7"/></svg>`:'';
  return `<button class="${on?'on':''}" onclick="gdSort('${k}')">${ICONS[k]||''}<span>${lbl}</span>${arrow}</button>`;
}
function _gdColHead(){
  const m=window._GD.metric;
  const h=(k,l,u)=>`<div class="num${m===k?' on':''}">${l}${u?` <span style="font-size:10px;opacity:.72">${u}</span>`:''}</div>`;
  return `<div></div><div>Tovar / TIF TN</div><div class="num">Korxona</div>${h('p','Partiya','')}${h('q','Qiymat','(ming $)')}${h('v','Vazn','(tn.)')}`;
}
function _gdRows(){
  const{metric:m,query:q,data}=window._GD;
  const fld={p:'partiya',q:'qiymat',v:'vazn'}[m]||'qiymat';
  const sorted=[...data].sort((a,b)=>(+b[fld]||0)-(+a[fld]||0));
  sorted.forEach((r,i)=>{r._rank=i+1});
  const qL=(q||'').trim().toLowerCase();
  const tnShort=(tn)=>{if(!tn)return'';let s=String(tn).replace(/\D/g,'');return s.slice(0,6)||s};
  const vis=qL?sorted.filter(r=>{
    const tn=String((r.key&&r.key.tnved)||r.tnved||'');
    return (r.name||'').toLowerCase().includes(qL)||tn.startsWith(qL)||tn.includes(qL);
  }):sorted.slice(0,30);
  const tot={p:data.reduce((s,r)=>s+(+r.partiya||0),0),q:data.reduce((s,r)=>s+(+r.qiymat||0),0),v:data.reduce((s,r)=>s+(+r.vazn||0),0)};
  const mx={p:sorted[0]?+sorted[0].partiya||1:1,q:[...data].sort((a,b)=>b.qiymat-a.qiymat)[0]?.qiymat||1,v:[...data].sort((a,b)=>b.vazn-a.vazn)[0]?.vazn||1};
  const emptyIcon=`<svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="#c2ccd9" stroke-width="1.7" stroke-linecap="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>`;
  if(!vis.length)return `<div class="rtab-empty">${emptyIcon}<div style="font-size:14px;font-weight:600;color:#5b6b82">Hech narsa topilmadi</div></div>`;
  const pOn=m==='p',qOn=m==='q',vOn=m==='v';
  return vis.map((r,idx)=>{
    const isTop=r._rank<=3;
    const tn=String((r.key&&r.key.tnved)||r.tnved||'').replace(/\D/g,'').slice(0,6);
    return `<div class="rtab-row" style="animation-delay:${(idx*0.028).toFixed(3)}s">
      <div><span class="rtab-rank ${isTop?'top':'reg'}" style="animation-delay:${(idx*0.028+0.08).toFixed(3)}s">${r._rank}</span></div>
      <div style="overflow:hidden"><div class="rtab-name" title="${esc(r.name||'')}">${esc(r.name||'')}</div><div class="rtab-stir">${tn?`<span>TIF TN: ${tn}</span>`:''}</div></div>
      <div class="rtab-p">${(+r.korxona||0).toLocaleString('ru-RU')}</div>
      <div class="rtab-cell${pOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.partiya||0,tot.p)}</span><span class="rtab-val">${(+r.partiya||0).toLocaleString('ru-RU')}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.partiya||0,mx.p,pOn)}"></div></div></div>
      <div class="rtab-cell${qOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.qiymat||0,tot.q)}</span><span class="rtab-val">${_rtFmt(r.qiymat)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.qiymat||0,mx.q,qOn)}"></div></div></div>
      <div class="rtab-cell${vOn?' on':''}"><div class="rtab-nums"><span class="rtab-pct">${_rtPct(+r.vazn||0,tot.v)}</span><span class="rtab-val">${_rtFmt(r.vazn)}</span></div><div class="rtab-bar-bg"><div style="${_rtBar(+r.vazn||0,mx.v,vOn)}"></div></div></div>
    </div>`;
  }).join('');
}
function _gdRender(){
  const el=document.getElementById('gdRows');if(el)el.innerHTML=_gdRows();
  const ch=document.getElementById('gdCol');if(ch)ch.innerHTML=_gdColHead();
  const sb=document.getElementById('gdSortBtns');
  if(sb)sb.innerHTML=_gdBtn('p','Partiya')+_gdBtn('q','Qiymat')+_gdBtn('v','Vazn');
}
function gdSort(k){window._GD.metric=k;_gdRender()}
function gdSearch(v){window._GD.query=v;_gdRender()}
function tovarRatingPanel(goodsRows){
  const rows=(goodsRows||[]).filter(r=>r.name&&r.name!=="IBK bo'yicha Jami"&&r.name!=="Boshqa");
  window._GD.data=rows;window._GD.metric='q';window._GD.query='';
  setTimeout(()=>animatePlaceholder('gdSearchInput',['Tovar nomini yozing…','TIF TN kodi kiriting…','Masalan: 8471…','Qidirish uchun yozing…']),50);
  const tot={p:rows.reduce((s,r)=>s+(+r.partiya||0),0),q:rows.reduce((s,r)=>s+(+r.qiymat||0),0),v:rows.reduce((s,r)=>s+(+r.vazn||0),0)};
  return `<div class="rtab-wrap panel" style="padding:0;overflow:hidden;--rtab-grid:28px 1fr 64px minmax(90px,0.8fr) minmax(120px,1fr) minmax(110px,1fr)">
    <div class="rtab-head">
      <div class="rtab-title"><div class="rtab-accent" style="background:linear-gradient(180deg,#f59e0b,#fb923c)"></div><div><div class="rtab-sup">Interaktiv hisobot</div><div class="rtab-h1"><span class="rtab-gradient" style="background:linear-gradient(135deg,#f59e0b,#fb923c);-webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent">${rows.length}</span> ta tovar guruhi</div></div></div>
      <div class="rtab-controls">
        <div style="font-size:12px;color:#9aa6b8">${tot.p.toLocaleString('ru-RU')} partiya · ${_rtFmt(tot.q)} ming $</div>
        <div class="rtab-search">${_SEARCH_SVG}<input oninput="gdSearch(this.value)" placeholder="Tovar nomini yozing…" id="gdSearchInput" autocomplete="off"></div>
        <div class="rtab-sortbtns" id="gdSortBtns">${_gdBtn('p','Partiya')}${_gdBtn('q','Qiymat')}${_gdBtn('v','Vazn')}</div>
      </div>
    </div>
    <div class="rtab-colhead" id="gdCol">${_gdColHead()}</div>
    <div class="rtab-body" id="gdRows">${_gdRows()}</div>
  </div>`;
}
const sumCols=[{k:"name",t:"Ko'rsatkich",w:"38%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"Kutilayotgan to'lov (mln so'm)",n:1,f:fmtN}];
const regimeCols=[{k:"rejim",t:"Rejim",w:"24%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"Kutilayotgan to'lov (mln so'm)",n:1,f:fmtN}];

function paymentRows(){let rows=PAYMENTS&&PAYMENTS.length?PAYMENTS:(DATA&&DATA.payments?DATA.payments:null);if(!rows)rows=[{name:"Sbor11-12",rows:402,sum:481739660,file:"1. Sbor11-12.xlsx"},{name:"Sbor 44",rows:2,sum:2472000,file:"2. Sbor 44.xlsx"},{name:"Sbor IM40",rows:470,sum:142964000,file:"3. Sbor im40.xlsx"},{name:"Sbor EK10",rows:89,sum:83018000,file:"4. Sbor ek10.xlsx"},{name:"Sbor DR",rows:2053,sum:292969000,file:"5. Sbor dr.xlsx"},{name:"29",rows:418,sum:15559074509.5,file:"6. 29.xlsx"},{name:"20",rows:296,sum:3935764049.81,file:"7. 20.xlsx"},{name:"27",rows:5,sum:37269405.9,file:"8. 27.xlsx"},{name:"25",rows:2,sum:20941130.06,file:"9. 25.xlsx"},{name:"30",rows:1964,sum:2250053286.17,file:"10. 30.xlsx"},{name:"74",rows:60,sum:2485871.08,file:"11. 74.xlsx"},{name:"79",rows:8,sum:494400000,file:"11. 79.xlsx"},{name:"IM42",rows:0,sum:0,file:"12. im42.xlsx"}];return rows.map(r=>Object.assign({},r,{file:r.file||paymentFileByName(r.name)}))}
function paymentFileByName(name){let m={"Sbor 11-12: yig'imlar":"1. Sbor11-12.xlsx","Sbor 44: TR80 yig'imi":"2. Sbor 44.xlsx","Sbor 10: IM40/ND40":"3. Sbor im40.xlsx","Sbor 10: EK10":"4. Sbor ek10.xlsx","Sbor 10: boshqa rejimlar":"5. Sbor dr.xlsx","29-kod to'lovi":"6. 29.xlsx","20-kod to'lovi":"7. 20.xlsx","27-kod to'lovi":"8. 27.xlsx","25-kod to'lovi":"9. 25.xlsx","30-kod to'lovi":"10. 30.xlsx","74-kod to'lovi":"11. 74.xlsx","79-kod to'lovi":"11. 79.xlsx","IM42 bo'yicha ro'yxat":"12. im42.xlsx","Sbor11-12":"1. Sbor11-12.xlsx","Sbor 44":"2. Sbor 44.xlsx","Sbor IM40":"3. Sbor im40.xlsx","Sbor EK10":"4. Sbor ek10.xlsx","Sbor DR":"5. Sbor dr.xlsx","29":"6. 29.xlsx","20":"7. 20.xlsx","27":"8. 27.xlsx","25":"9. 25.xlsx","30":"10. 30.xlsx","74":"11. 74.xlsx","79":"11. 79.xlsx","IM42":"12. im42.xlsx"};return m[name]||""}
function paymentTotal(){let rows=paymentRows();return {rows:rows.reduce((a,r)=>a+(+r.rows||0),0),sum:rows.reduce((a,r)=>a+(+r.sum||0),0)}}
function paymentTable(rows){let body=[{name:"IBK bo'yicha Jami",rows:paymentTotal().rows,sum:paymentTotal().sum,file:""}].concat(rows||[]);return `<div class="excel-wrap"><table class="excel-table"><thead><tr><th style="width:46px">T/r</th><th>To'lov turi</th><th style="width:110px">Qator</th><th style="width:190px">Summa (so'm)</th><th style="width:120px">Excel</th></tr></thead><tbody>${body.map((r,i)=>`<tr class="${i===0?'total':''}"><td class="num">${i===0?'':i}</td><td class="text">${esc(r.name)}</td><td class="num">${fmtI(r.rows)}</td><td class="num">${fmtN(r.sum)}</td><td class="download-cell">${r.file?`<a class="excel-download" href="/download/tolov/${encodeURIComponent(r.file)}?token=${TOKEN}">Yuklash</a>`:''}</td></tr>`).join("")}</tbody></table></div>`}
function paymentModule(mode="overview"){let rows=paymentRows(), t=paymentTotal(), sorted=by(rows,"sum");if(mode==="lists")return `<div class=stack><div class=panel><div class="excel-title"><h2>Hosil bo'lgan to'lov jadvallari</h2><a class="btn light" href="/download/tolov/_all?token=${TOKEN}">Hammasini ZIP</a></div>${paymentTable(rows)}</div><div class=panel><h2>Summa ulushi</h2>${bars(rows,"name","sum",fmtN)}</div></div>`;if(mode==="analysis")return `<div class=grid2><div class=panel><h2>To'lov turlari bo'yicha tahlil</h2>${paymentTable(sorted)}</div><div class=panel><h2>Eng katta summalar</h2>${bars(sorted.slice(0,8),"name","sum",fmtN)}</div><div class=panel><h2>Qatorlar soni</h2>${bars(by(rows,"rows"),"name","rows",fmtI)}</div></div>`;return `<div class=stack><div class=panel><h2>To'lovlar</h2><div class=pay-kpis><div class=pay-card>Jami qator<b>${fmtI(t.rows)}</b></div><div class=pay-card>Jami summa<b>${fmtN(t.sum)} so'm</b></div><div class=pay-card>To'lov turlari<b>${fmtI(rows.length)}</b></div><div class=pay-card>Eng katta tur<b>${esc(sorted[0]?.name||"-")}</b></div></div><div class=module-grid><div>${paymentTable(rows)}</div><div>${bars(sorted.slice(0,8),"name","sum",fmtN)}</div></div></div></div>`}

function overviewPanels(){return `<div class="grid2">${executiveSummary()}<div class=panel><h2>Jami va 70-74-80</h2>${table(sumCols,regimeSummaryRows())}<div class=overview-note>Rejim ustiga bosilganda postlar kesimida asos ochiladi.</div></div><div class=panel><h2 onclick="detail({view:'expired_inline'})">Jami muddati o'tgan</h2>${table(sumCols,expiredSummaryRows())}<div id="expiredInline" class="chart"></div></div><div class="panel wide"><h2>TOP 20 korxona: qiymat ulushi</h2>${bars(DATA.top_value||[],"korxona","qiymat",fmtN)}</div><div class=panel><h2>Rejimlar qiymat ulushi</h2>${bars(DATA.regimes||[],"rejim","qiymat",fmtN)}</div><div class=panel><h2>Muddati o'tgan postlar</h2>${bars(DATA.post_summary||[],"post","partiya",fmtI)}</div><div class=panel><h2>Tovar guruhlari partiya ulushi</h2>${bars(by(DATA.goods||[],"partiya"),"name","partiya",fmtI)}</div></div>`}

let WR_MAP=null,WR_DATA=null;
function wrTypeColor(t){return {ochiq:'#1d72b8',yopiq:'#0f4c8a',dutyfree:'#d97706',erkin:'#7c3aed'}[t]||'#1d72b8'}
function wrTypeName(t){return {ochiq:"Ochiq bojxona ombori",yopiq:"Yopiq bojxona ombori",dutyfree:"Boj olinmaydigan savdo",erkin:"Erkin ombor"}[t]||t}
function wrRiskColor(r){return {red:'#dc2626',orange:'#d97706',green:'#16a34a'}[r]||'#16a34a'}
function wrMarkerHtml(w){
  let clr=wrTypeColor(w.type),rClr=wrRiskColor(w.risk);
  let area=Math.max(w.area_open||0,w.area_closed||0);
  let sz=Math.round(Math.max(22,Math.min(42,22+Math.sqrt(area/10000)*20)));
  let border=w.risk==='red'?'3px solid #dc2626':w.risk==='orange'?'3px solid #d97706':'2px solid rgba(255,255,255,.8)';
  let badges='';
  if(w.fvv)badges+=`<div style="position:absolute;top:-6px;right:-6px;background:#ef4444;color:#fff;border-radius:4px;padding:1px 4px;font-size:9px;font-weight:900;white-space:nowrap;z-index:2;box-shadow:0 1px 3px rgba(0,0,0,.3)">🔥 FVV</div>`;
  if(w.ssv)badges+=`<div style="position:absolute;bottom:-6px;right:-6px;background:#16a34a;color:#fff;border-radius:4px;padding:1px 4px;font-size:9px;font-weight:900;white-space:nowrap;z-index:2;box-shadow:0 1px 3px rgba(0,0,0,.3)">GSP</div>`;
  let icon={ochiq:'📦',yopiq:'🔒',dutyfree:'🛍',erkin:'⭐'}[w.type]||'📦';
  return `<div style="position:relative;width:${sz}px;height:${sz}px;transform:translate(-50%,-50%)">${badges}<div style="width:${sz}px;height:${sz}px;border-radius:50%;background:${clr};border:${border};display:flex;align-items:center;justify-content:center;font-size:${Math.round(sz*0.44)}px;box-shadow:0 2px 8px rgba(0,0,0,.3);cursor:pointer;box-sizing:border-box">${icon}</div></div>`
}
function wrInsTable(wrs,fileDate){
  let sorted=[...wrs].sort((a,b)=>(b.ins_sum||0)-(a.ins_sum||0));
  let totalSum=wrs.reduce((a,w)=>a+(+w.ins_sum||0),0);
  let totalRow=`<tr><td>IBK bo'yicha Jami</td><td>—</td><td>—</td><td>—</td><td>—</td><td class="num">${totalSum?fmtN(totalSum/1e6)+' mln':'—'}</td><td>—</td><td>—</td></tr>`;
  let rows=sorted.map(w=>{
    let dCls=w.ins_days===null?'':(w.ins_days<0?'wr-risk-red':w.ins_days<90?'wr-risk-orange':'wr-risk-green');
    let dTxt=w.ins_days===null?'—':(w.ins_days<0?`${Math.abs(w.ins_days)} kun o'tgan!`:`${w.ins_days} kun qoldi`);
    let fvvBadge=w.fvv?`<span class="wr-badge-fvv">FVV</span>`:'';
    let ssvBadge=w.ssv?`<span class="wr-badge-ssv">GSP</span>`:'';
    return `<tr><td>${esc(w.name)}${fvvBadge}${ssvBadge}</td><td>${esc(wrTypeName(w.type))}</td><td>${esc(w.lic_num||'—')}</td><td>${esc(w.director||'—')}</td><td>${esc(w.phone||'—')}</td><td class="num">${w.ins_sum?fmtN(w.ins_sum/1e6)+' mln':'—'}</td><td>${esc(w.ins_exp||'—')}</td><td class="${dCls}">${dTxt}</td></tr>`;
  }).join('');
  return `<table class="wr-ins-table"><thead><tr><th>Ombor nomi</th><th>Tur</th><th>Litsenziya №</th><th>Direktor</th><th>Telefon</th><th>Sug'urta summasi</th><th>Muddat</th><th>Holat</th></tr></thead><tbody>${totalRow}${rows}</tbody></table>`;
}
function wrDeadlineTable(wrs,fileDate){
  let at_risk=wrs.filter(w=>w.ins_days!==null&&w.ins_days<90).sort((a,b)=>(a.ins_days||0)-(b.ins_days||0));
  if(!at_risk.length)return '<div class="muted" style="padding:12px">Yaqin 90 kunda muddati tugaydigan sug\'urta yo\'q.</div>';
  let overdue=at_risk.filter(w=>w.ins_days<0).length;
  let soon=at_risk.filter(w=>w.ins_days>=0).length;
  let sumTxt=overdue>0?`${overdue} ta muddati o'tgan, ${soon} ta yaqinlashyapti`:`${soon} ta ombor muddati yaqinlashyapti`;
  let totalRow=`<tr style="font-weight:700;background:#fef9ec;border-bottom:2px solid #f59e0b"><td>IBK bo'yicha jami</td><td>—</td><td>—</td><td>—</td><td>—</td><td>—</td><td class="${overdue>0?'wr-risk-red':'wr-risk-orange'}">${sumTxt}</td></tr>`;
  let rows=at_risk.map(w=>{
    let dCls=w.ins_days<0?'wr-risk-red':'wr-risk-orange';
    let dTxt=w.ins_days<0?`${Math.abs(w.ins_days)} kun o'tgan!`:`${w.ins_days} kun qoldi`;
    let fvvBadge=w.fvv?`<span class="wr-badge-fvv">FVV</span>`:'';
    let ssvBadge=w.ssv?`<span class="wr-badge-ssv">GSP</span>`:'';
    return `<tr><td>${esc(w.name)}${fvvBadge}${ssvBadge}</td><td>${esc(wrTypeName(w.type))}</td><td>${esc(w.lic_num||'—')}</td><td>${esc(w.director||'—')}</td><td>${esc(w.phone||'—')}</td><td>${esc(w.ins_exp||'—')}</td><td class="${dCls}">${dTxt}</td></tr>`;
  }).join('');
  return `<table class="wr-ins-table"><thead><tr><th>Ombor nomi</th><th>Tur</th><th>Litsenziya №</th><th>Direktor</th><th>Telefon</th><th>Sug'urta muddati</th><th>Holat</th></tr></thead><tbody>${totalRow}${rows}</tbody></table>`;
}
let WR_ACTIVE_TYPES=new Set(['ochiq','yopiq','dutyfree','erkin']);
let WR_POPUP_EL=null,WR_HIDE_TIMER=null;
function wrSvgMarker(w){
  let clr=wrTypeColor(w.type);
  let ring=w.risk==='red'?'#dc2626':w.risk==='orange'?'#f59e0b':'#22c55e';
  let sz=32;
  let icons={ochiq:'M4 20h16M8 4h8v12H8z',yopiq:'M12 2L4 7v6c0 5.25 3.58 10.15 8 11.34C16.42 23.15 20 18.25 20 13V7L12 2z',dutyfree:'M6 2l.5 4.5H9l.5-4.5h5l.5 4.5h2.5L18 2M3 7h18l-1.5 13H4.5L3 7z',erkin:'M12 2a7 7 0 0 1 7 7c0 5.25-7 13-7 13S5 14.25 5 9a7 7 0 0 1 7-7z'};
  let path=icons[w.type]||icons.ochiq;
  return `<svg width="${sz}" height="${sz}" viewBox="0 0 ${sz} ${sz}" xmlns="http://www.w3.org/2000/svg" style="filter:drop-shadow(0 2px 4px rgba(0,0,0,.28));cursor:pointer"><circle cx="16" cy="16" r="15" fill="${clr}" stroke="${ring}" stroke-width="3"/><g transform="translate(4,4)" stroke="#fff" stroke-width="1.6" fill="none" stroke-linecap="round" stroke-linejoin="round"><path d="${path}"/></g></svg>`;
}
function buildWrMap(wrs){
  let el=document.getElementById('wrMap');if(!el)return;
  if(WR_MAP){try{WR_MAP.remove()}catch(e){}WR_MAP=null}
  el.innerHTML='';
  // Build filter bar
  let fb=document.getElementById('wrFilterBar');
  if(fb){
    let types=['ochiq','yopiq','dutyfree','erkin'];
    fb.innerHTML=`<div class="wr-filter-row">${types.map(t=>`<label class="wr-filter-chip ${WR_ACTIVE_TYPES.has(t)?'active':''}"><input type="checkbox" ${WR_ACTIVE_TYPES.has(t)?'checked':''} onchange="wrToggleType('${t}',this.checked)"><span class="wr-chip-dot" style="background:${wrTypeColor(t)}"></span>${wrTypeName(t)}</label>`).join('')}</div>`;
  }
  _loadLeaflet(function(){
    try{
      let map=L.map(el,{center:[41.27,69.27],zoom:11,zoomControl:true});
      L.tileLayer('https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png',{attribution:'© OpenStreetMap © CARTO',maxZoom:19,subdomains:'abcd'}).addTo(map);
      WR_MAP=map;
      WR_MAP._wrMarkers=[];
      wrs.forEach(w=>{
        if(!w.lat||!w.lon)return;
        let icon=L.divIcon({html:wrSvgMarker(w),className:'',iconSize:[32,32],iconAnchor:[16,16]});
        let m=L.marker([w.lat,w.lon],{icon,title:w.name});
        m._wrType=w.type;
        m._wrData=w;
        if(!WR_ACTIVE_TYPES.has(w.type))m.options.opacity=0;
        m.on('mouseover',function(e){clearTimeout(WR_HIDE_TIMER);showWrPopup(w,e.originalEvent)});
        m.on('mouseout',function(){WR_HIDE_TIMER=setTimeout(hideWrPopup,220)});
        m.on('click',function(e){L.DomEvent.stopPropagation(e);clearTimeout(WR_HIDE_TIMER);showWrPopup(w,e.originalEvent)});
        m.addTo(map);
        WR_MAP._wrMarkers.push(m);
      });
      map.on('click',function(){clearTimeout(WR_HIDE_TIMER);hideWrPopup()});
    }catch(err){console.error('WR map error:',err)}
  });
}
function wrToggleType(type,on){
  if(on)WR_ACTIVE_TYPES.add(type);else WR_ACTIVE_TYPES.delete(type);
  document.querySelectorAll('.wr-filter-chip').forEach(el=>{
    let inp=el.querySelector('input');
    if(inp)el.classList.toggle('active',inp.checked);
  });
  if(!WR_MAP||!WR_MAP._wrMarkers)return;
  WR_MAP._wrMarkers.forEach(m=>{
    if(WR_ACTIVE_TYPES.has(m._wrType))m.setOpacity(1);else m.setOpacity(0);
  });
}
function showWrPopup(w,evt){
  clearTimeout(WR_HIDE_TIMER);
  if(WR_POPUP_EL&&WR_POPUP_EL.parentNode){WR_POPUP_EL.parentNode.removeChild(WR_POPUP_EL);WR_POPUP_EL=null;}
  let insDays=w.ins_days===null?'—':w.ins_days<0?`<span style="color:#dc2626">${Math.abs(w.ins_days)} kun o'tgan!</span>`:`<span style="color:#d97706">${w.ins_days} kun qoldi</span>`;
  let fvvRow=w.fvv?`<div class="wr-popup-row"><span>FVV:</span><b>${esc(w.fvv)}</b></div>`:'';
  let ssvRow=w.ssv?`<div class="wr-popup-row"><span>SSV:</span><b>${esc(w.ssv)}</b></div>`:'';
  let bnrteRow='';
  if(typeof DATA!=='undefined'&&DATA&&DATA.warehouse){
    let nm=w.name||'';
    let bn=(DATA.warehouse||[]).find(d=>(d.name||'').toLowerCase()===nm.toLowerCase()||(d.name||'').toLowerCase().includes(nm.toLowerCase())||nm.toLowerCase().includes((d.name||'').toLowerCase()));
    if(bn)bnrteRow=`<div class="wr-popup-sep"></div><div class="wr-popup-row wr-popup-sec"><span>Saqlanayotgan tovarlar:</span></div><div class="wr-popup-row"><span>Qiymati:</span><b>${fmtN(bn.qiymat||0)} ming $</b></div><div class="wr-popup-row"><span>Vazni:</span><b>${fmtN(bn.vazn||0)} tn</b></div><div class="wr-popup-row"><span>Partiya:</span><b>${fmtI(bn.partiya||0)}</b></div>`;
  }
  let div=document.createElement('div');
  div.className='wr-popup-card';
  div.innerHTML=`<div class="wr-popup-head" style="background:${wrTypeColor(w.type)}"><b>${esc(w.name)}</b><span>${esc(wrTypeName(w.type))}</span></div><div class="wr-popup-body"><div class="wr-popup-row"><span>Litsenziya №:</span><b>${esc(w.lic_num||'—')}</b></div><div class="wr-popup-row"><span>Sug'urta muddati:</span><b>${esc(w.ins_exp||'—')}</b></div><div class="wr-popup-row"><span>Holat:</span><b>${insDays}</b></div>${bnrteRow}${fvvRow}${ssvRow}</div>`;
  // Append to body with fixed positioning to avoid overflow:hidden clipping from map container
  document.body.appendChild(div);
  if(evt){
    let left=evt.clientX+14, top=evt.clientY+14;
    if(left+328>window.innerWidth-4) left=evt.clientX-340;
    if(top+260>window.innerHeight-4) top=evt.clientY-268;
    div.style.cssText=`position:fixed;left:${Math.max(2,left)}px;top:${Math.max(2,top)}px;bottom:auto;transform:none;z-index:9999`;
  } else {
    div.style.cssText='position:fixed;bottom:16px;left:50%;transform:translateX(-50%);z-index:9999';
  }
  WR_POPUP_EL=div;
  setTimeout(()=>div.classList.add('visible'),10);
  div.addEventListener('mouseenter',()=>clearTimeout(WR_HIDE_TIMER));
  div.addEventListener('mouseleave',()=>{WR_HIDE_TIMER=setTimeout(hideWrPopup,220)});
}
function hideWrPopup(){
  clearTimeout(WR_HIDE_TIMER);WR_HIDE_TIMER=null;
  if(WR_POPUP_EL){let el=WR_POPUP_EL;WR_POPUP_EL=null;el.classList.remove('visible');setTimeout(()=>{if(el.parentNode)el.parentNode.removeChild(el)},300);}
}
async function loadWarehouseRegistry(){
  let panel=document.getElementById('wrRegistryPanel');if(!panel)return;
  try{
    let j=await api('/api/warehouses');
    WR_DATA=j;
    let wrs=j.warehouses||[];
    omborRatingUpdateWR(wrs);
    let fd=j.file_date||'';
    let dateLabel=fd?` (${fd} holatiga)`:'';
    let stats={total:wrs.length,red:wrs.filter(w=>w.risk==='red').length,orange:wrs.filter(w=>w.risk==='orange').length,ssv:wrs.filter(w=>w.ssv).length,fvv:wrs.filter(w=>w.fvv).length};
    let statHtml=`<div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:14px;font-size:13px">
      <span><b style="color:#1d72b8">${stats.total}</b> ombor</span>
      <span style="color:#dc2626"><b>${stats.red}</b> muddati o'tgan</span>
      <span style="color:#d97706"><b>${stats.orange}</b> diqqat talab</span>
      <span style="color:#16a34a"><b>${stats.ssv}</b> SSV</span>
      ${stats.fvv?`<span style="color:#ef4444"><b>${stats.fvv}</b> FVV</span>`:''}
    </div>`;
    panel.innerHTML=`<h2>Omborlar reestri${dateLabel}</h2>
      ${statHtml}
      <div class="wr-map-wrap">
        <div class="wr-filter-bar" id="wrFilterBar"></div>
        <div id="wrMap" style="height:480px;width:100%"></div>
      </div>
      <div style="margin-top:20px"><h3 style="font-size:14px;margin-bottom:8px">Kafolat muddati yaqinlashayotganlar${dateLabel}</h3>${wrDeadlineTable(wrs,fd)}</div>
      <div style="margin-top:20px"><h3 style="font-size:14px;margin-bottom:8px">Sug'urta summasi bo'yicha${dateLabel}</h3>${wrInsTable(wrs,fd)}</div>`;
    buildWrMap(wrs);
  }catch(e){let panel=document.getElementById('wrRegistryPanel');if(panel)panel.innerHTML=`<h2>Omborlar reestri</h2><div class="muted">Ma'lumot yuklanmadi: ${esc(String(e))}</div>`}
}
async function uploadWrRegistry(){
  let inp=document.getElementById('wrFile'),st=document.getElementById('wrUploadStatus');
  if(!inp||!inp.files[0])return;
  if(st)st.textContent='Yuklanmoqda...';
  let fd=new FormData();fd.append('file',inp.files[0]);
  try{
    let j=await api('/api/upload_warehouses',{method:'POST',body:fd});
    WR_DATA=j;
    if(st)st.textContent=`Tayyor: ${(j.warehouses||[]).length} ta ombor yuklandi`;
    await loadWarehouseRegistry();
  }catch(e){if(st)st.textContent='Xatolik: '+String(e)}
}

function render(){if(_uploadActive&&!confirm("Fayl yuklanmoqda! Boshqa tabga o'tsangiz yuklash to'xtab qoladi. Davom etasizmi?"))return;if(TAB.startsWith("pay"))GROUP="payments";if(["umumiy","rejim","korxona","ombor","expired","released","goods","food","muddat","archive","avia","yaroqlilik"].includes(TAB))GROUP="bnrte";if(["upload","profile","settings","admin"].includes(TAB))GROUP="common";if(TAB==="home")GROUP="home";$("dash").classList.remove("hidden");$("meta").textContent=DATA&&DATA.meta?DATA.meta.date+" holatiga":"Tizimga xush kelibsiz";$("kpis").innerHTML=(DATA&&GROUP!=="home")?renderKpis():"";let bnrte=[["umumiy",tr("general")],["rejim",tr("regimes")],["korxona",tr("companies")],["ombor",tr("warehouses")],["expired",tr("expired")],["released",tr("released")],["goods",tr("goods")],["food",tr("food")],["muddat",tr("deadlines")],["yaroqlilik",tr("validity")],["archive",tr("archive")],["avia","✈ AVIA AWB"]], payTabs=[["payments",tr("pay_overview")],["pay_lists",tr("pay_lists")],["pay_analysis",tr("pay_analysis")]], commonTabs=[["upload",tr("upload")],["profile",tr("profile")],["settings",tr("settings")],["admin",tr("admin")]];let bnrteHtml=GROUP==="bnrte"?`<div class="subtabs">${bnrte.map(t=>`<button class="tab sub ${TAB===t[0]?'active':''}" onclick="TAB='${t[0]}';GROUP='bnrte';render()">${t[1]}</button>`).join("")}</div>`:"";let payHtml=GROUP==="payments"?`<div class="subtabs">${payTabs.map(t=>`<button class="tab sub ${TAB===t[0]?'active':''}" onclick="TAB='${t[0]}';GROUP='payments';render()">${t[1]}</button>`).join("")}</div>`:"";let commonHtml=GROUP==="common"?`<div class="subtabs">${commonTabs.map(t=>`<button class="tab sub ${TAB===t[0]?'active':''}" onclick="TAB='${t[0]}';GROUP='common';render()">${t[1]}</button>`).join("")}</div>`:"";$("tabs").innerHTML=`<button class="module-parent ${GROUP==='bnrte'?'active':''}" onclick="openGroup('bnrte','umumiy')">BNRTE</button>${bnrteHtml}<button class="module-parent pay ${GROUP==='payments'?'active':''}" onclick="openGroup('payments','payments')">To'lovlar</button>${payHtml}<button class="module-parent ${GROUP==='common'?'active':''}" onclick="openGroup('common','upload')">Boshqaruv</button>${commonHtml}`;let f=DATA&&DATA.files||{};let fileParts=[];if(f.excel)fileParts.push(`<a class=btn href="/download/${DATA.id}/${f.excel}?token=${TOKEN}">Jamlanma Excel</a>`);if(f.pdf)fileParts.push(`<a class=btn href="/download/${DATA.id}/${f.pdf}?token=${TOKEN}">PDF</a>`);if((f.pngs||[]).length)fileParts.push(`<a class=btn href="/download/${DATA.id}/${f.pngs[0]}?token=${TOKEN}">PNG</a>`);let prepBtn=ARTIFACT_POLL_ID?`<button class="light" disabled style="opacity:.7">⏳ Tayyorlanmoqda...</button>`:`<button class="light" onclick="prepareArtifacts()">Excel/PNG/PDF tayyorlash</button>`;let fileBtns=fileParts.length?fileParts.join(" ")+" "+prepBtn:prepBtn;if(!ARTIFACT_POLL_ID&&f.status&&f.status!=="tayyor")fileBtns+=` <span class="muted">${esc(f.status)}</span>`;$("status").textContent=f.error?"Excel/PNG/PDF tayyorlashda xatolik bor. Qayta tayyorlash tugmasini bosing yoki logni tekshiramiz.":"";$("actions").innerHTML=DATA?`${fileBtns} <button class="light" onclick="TAB='settings';render()">${tr("settings")}</button> <button class="logout-btn" onclick="logout()">${tr("logout")}</button>`:`<button class="light" onclick="TAB='settings';render()">${tr("settings")}</button> <button class="logout-btn" onclick="logout()">${tr("logout")}</button>`;view();}
function view(){let v=$("view");if(TAB==="home"){v.innerHTML=landingPanel();return}if(TAB==="archive"){let seen=new Set(), rows=ARCHIVE.filter(r=>{let k=r.date+"|"+(r.source||"").split(/[\\/]/).pop();if(seen.has(k))return false;seen.add(k);return true});v.innerHTML=`<div class=panel><h2>Arxiv</h2><div class=cards>${rows.map(r=>`<button class="archive-card" onclick="loadReport('${r.id}')"><b>${r.date}</b><span>Asos: ${esc((r.source||'').split(/[\\/]/).pop())}</span><span>Depozit: ${esc((r.deposit||'Depozitsiz').split(/[\\/]/).pop()||'Depozitsiz')}</span></button>`).join("")}</div></div>`;return}if(TAB==="upload"){v.innerHTML=uploadPanel();bindUpload();return}if(TAB==="profile"){v.innerHTML=`<div class=stack><div class=panel><h2>Profil</h2><b style="font-size:18px">${esc(ME.full_name||ME.user)}</b><p class=muted>${esc(ME.position||ME.role)}</p><p>Vakolatlar: ${(ME.perms||[]).map(p=>permTitle(p)).join(", ")||'—'}</p></div><div class=panel><h2>Parolni o'zgartirish</h2><form id="changePwForm" class=admin-form style="max-width:340px"><label>Joriy parol</label><div class="pass-wrap"><input id="cpOld" type=password placeholder="Joriy parol"><button class="eye-btn" type="button" onclick="let p=$('cpOld');p.type=p.type==='password'?'text':'password'">&#128065;</button></div><label>Yangi parol</label><div class="pass-wrap"><input id="cpNew" type=password placeholder="Kamida 6 ta belgi"><button class="eye-btn" type="button" onclick="let p=$('cpNew');p.type=p.type==='password'?'text':'password'">&#128065;</button></div><label>Yangi parol (takror)</label><div class="pass-wrap"><input id="cpNew2" type=password placeholder="Yangi parolni takrorlang"><button class="eye-btn" type="button" onclick="let p=$('cpNew2');p.type=p.type==='password'?'text':'password'">&#128065;</button></div><button onclick="selfChangePassword(event)">Saqlash</button><span id="cpMsg" class=muted style="margin-left:10px"></span></form></div></div>`;return}if(TAB==="settings"){v.innerHTML=`<div class=stack><div class=panel><h2>Sozlamalar</h2><div class=settings><label>Til</label><select onchange="setLang(this.value)"><option value=uz ${LANG==='uz'?'selected':''}>O'zbek lotin</option><option value=uzc ${LANG==='uzc'?'selected':''}>O'zbek kirill</option><option value=ru ${LANG==='ru'?'selected':''}>Rus tili</option></select><button onclick="document.body.classList.toggle('dark')">${tr("dark")}</button></div></div>${layoutEditorPanel()}</div>`;setTimeout(leBindDrag,0);return}if(TAB==="admin"){v.innerHTML=adminPanel();bindUserForm();loadUsers();return}if(!DATA){v.innerHTML=landingPanel();return}
if(TAB==="umumiy"){let topC=by(DATA.top_value||[],"qiymat").slice(0,30);v.innerHTML=wlayout('umumiy',{exec:()=>`<div>${executiveSummary()}</div>`,post:()=>`<div class=panel><h2>70-74-80: postlar kesimida ${staleBadge()}</h2>${table([{k:"post",t:"Post",w:"42%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"Kutilayotgan to'lov (mln so'm)",n:1,f:fmtN}],basicTotal(DATA.all_post_summary||[],"IBK bo'yicha Jami","post"),"fixed-table")}<div class=overview-note>Post qatorlari ustiga bosilganda asos deklaratsiyalar ochiladi.</div></div>`,expired_sum:()=>`<div class="panel wide"><h2>Jami muddati o'tgan: postlar va rejimlar kesimida</h2>${expiredTotalExcelTable()}</div>`,top_bars:()=>`<div class="panel wide"><h2>Qiymat bo'yicha TOP 30 korxona</h2>${bars(topC,"korxona","qiymat",fmtN)}</div>`});$('_cfmSlot')||v.querySelector('.wlayout')?.insertAdjacentHTML('beforeend','<div id="_cfmSlot"></div>');return}
if(TAB==="avia"){v.innerHTML=aviaPanel();loadAviaAwb();return}
if(TAB==="payments"){v.innerHTML=paymentModule("overview");return}if(TAB==="pay_lists"){v.innerHTML=paymentModule("lists");return}if(TAB==="pay_analysis"){v.innerHTML=paymentModule("analysis");return}
if(TAB==="rejim"){v.innerHTML=wlayout('rejim',{jami:()=>`<div class=panel><h2>Jami va 70-74-80 ${staleBadge()}</h2>${table(sumCols,regimeSummaryRows())}</div>`,kesim:()=>`<div class=panel><h2>70-74-80 rejimlar kesimida</h2>${table(regimeCols,basicTotal(DATA.regimes||[],"IBK bo'yicha Jami","rejim"))}</div>`,qiymat:()=>`<div class=panel><h2>Rejimlar qiymat ulushi</h2>${bars(DATA.regimes||[],"rejim","qiymat",fmtN)}</div>`,tolov:()=>`<div class=panel><h2>Rejimlar to'lov ulushi</h2>${bars(DATA.regimes||[],"rejim","tolov",fmtN)}</div>`});return}
if(TAB==="ombor"){v.innerHTML=wlayout('ombor',{rating:()=>omborRatingPanel(DATA.warehouse||[],WR_DATA),reestri:()=>`<div class="panel wide" id="wrRegistryPanel"><h2>Omborlar reestri</h2><div class="muted">Yuklanmoqda...</div></div>`,kesim:()=>`<div class="panel wide"><h2>Omborlar kesimida ${staleBadge()}</h2>${table(sumCols,basicTotal(DATA.warehouse||[]))}</div>`,bars_group:()=>`<div class=stack><div class=panel><h2>Omborlar qiymat ulushi</h2>${bars(DATA.warehouse||[],"name","qiymat",fmtN)}</div><div class=panel><h2>O'z ombor jami</h2>${table(ownCols,expiredTotal(DATA.own_all||[]).map(r=>Object.assign({korxona:r.korxona||"IBK bo'yicha Jami"},r)))}</div><div class="panel wide"><h2>O'z ombor 3 oy+</h2>${table(ownCols,expiredTotal(DATA.own_3m||[]).map(r=>Object.assign({korxona:r.korxona||"IBK bo'yicha Jami"},r)))}</div><div class=panel><h2>O'z ombor partiya ulushi</h2>${bars(DATA.own_all||[],"korxona","partiya",fmtI)}</div></div>`,oborot:()=>`<div class="panel wide"><h2>Omborlar oboroti (nazoratdan yechilgan)</h2><div class="filters compact-filters" style="margin-bottom:8px"><label style="font-size:12px;color:#557086">Boshlang'ich sana:</label><select id="omborOborotBase">${dateOptions()}</select><label style="font-size:12px;color:#557086">Yakuniy sana:</label><select id="omborOborotFinal">${dateOptions()}</select><button onclick="buildOmborOborot()">Hisoblash</button></div><div id="omborOborotResult" class="muted">Boshlang'ich sana (eskiroq) va yakuniy sana (yangiroq) tanlang, so'ng Hisoblash tugmasini bosing.</div></div>`,trend:()=>`<div class="panel wide" id="warehouseTrendPanel"><h3>Omborlar bo'yicha yuk oqimi tendensiyasi</h3><div class="muted">Yuklanmoqda...</div></div>`});let bSel=$('omborOborotBase'),fSel=$('omborOborotFinal');if(bSel&&fSel&&bSel.options.length>1){bSel.selectedIndex=bSel.options.length-1;fSel.selectedIndex=0;buildOmborOborot();}loadWarehouseTrends();loadWarehouseRegistry();return}
if(TAB==="muddat"){v.innerHTML=wlayout('muddat',{kesim:()=>`<div class=panel><h2>Muddatlar kesimida ${staleBadge()}</h2>${table([{k:"muddat",t:"Muddat",w:"30%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"Kutilayotgan to'lov (mln so'm)",n:1,f:fmtN}],basicTotal(DATA.ages||[],"IBK bo'yicha Jami","muddat"))}</div>`,sabab:()=>`<div class=panel><h2>Saqlanish sabablari</h2>${table(sumCols,basicTotal(DATA.reason||[]))}</div>`,bars_group:()=>`<div class=stack><div class=panel><h2>Muddatlar partiya ulushi</h2>${bars(DATA.ages||[],"muddat","partiya",fmtI)}</div><div class=panel><h2>Saqlanish sabablari qiymat ulushi</h2>${bars(DATA.reason||[],"name","qiymat",fmtN)}</div></div>`});return}
if(TAB==="korxona"){v.innerHTML=wlayout('korxona',{rating:()=>korxonaRatingPanel(DATA.top_value||[],DATA.top_deposit||[],DATA.all_companies||[]),top_qiymat:()=>`<div class=panel><h2>TOP 20 korxona (qiymat) ${xls("top_value")} ${staleBadge()}</h2>${table(companyCols(),companyTotal(DATA.top_value||[]))}</div>`,top_depozit:()=>`<div class=panel><h2>TOP 20 korxona (depozit mablag'lari) ${xls("top_deposit")}</h2>${table(companyCols(),companyTotal(DATA.top_deposit||[]))}</div>`,ulush:()=>`<div class=panel><h2>Qiymat ulushi</h2>${bars(DATA.top_value||[],"korxona","qiymat",fmtN)}</div>`,trend:()=>`<div class="panel wide" id="trendPanel"><h3>Korxonalar bo'yicha davriy tendensiya</h3><div class="muted">Yuklanmoqda...</div></div>`,transport_k:()=>`<div class="panel wide" id="transportCompanyPanel"><h3>Transport kesimida korxonalar — yuk oqimi tendensiyasi</h3><div class="muted">Yuklanmoqda...</div></div>`,transport_t:()=>`<div class="panel wide" id="transportTrendPanel"><h3>Transport turi bo'yicha davriy tendensiya</h3><div class="muted">Yuklanmoqda...</div></div>`});loadCompanyTrends();loadTransportCompanyTrends();loadTransportTrends();return}
if(TAB==="expired"){v.innerHTML=wlayout('expired',{rejim:()=>`<div class=panel><h2>Jami muddati o'tgan: postlar va rejimlar kesimida ${staleBadge()}</h2>${expiredTotalExcelTable()}<div class=overview-note>Jadval ustunlari jamlanma Excel vkladkasidagi ko'rinishga yaqin qat'iy kenglikda berildi.</div></div>`,post:()=>`<div class=panel><h2>Muddati o'tgan postlar kesimida</h2>${table([{k:"post",t:"Post",w:"42%"},{k:"partiya",t:"Partiya",n:1,f:fmtI,w:"90px"},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN,w:"120px"},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN,w:"130px"},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN,w:"135px"}],basicTotal(DATA.post_summary||[],"IBK bo'yicha Jami","post"),"fixed-table")}${bars(DATA.post_summary||[],"post","qiymat",fmtN)}${miniChart(DATA.post_summary||[],"qiymat","post")}</div>`,jamlanma:()=>`<div class=panel><h2>Muddati o'tgan jamlanma ${xls("expired")}</h2>${expiredBlockTable(DATA.expired_block||[])}</div>`,korxona:()=>`<div class=panel><h2>Muddati o'tgan korxonalar</h2>${table([{k:"korxona",t:"Korxona nomi",w:"300px"},{k:"stir",t:"STIR",w:"92px"},{k:"rejim",t:"Rejim",w:"70px"},{k:"post",t:"Nazorat posti",w:"170px"},{k:"kun",t:"Kun hisobi",n:1,f:fmtI,w:"80px"},{k:"partiya",t:"Partiya",n:1,f:fmtI,w:"78px"},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN,w:"120px"},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN,w:"125px"}],expiredTotal(DATA.expired||[]),"fixed-table")}</div>`});return}
if(TAB==="released"){let rel=DATA.released||{};v.innerHTML=`<div class=panel><h2>Nazoratdan yechilishi</h2><div class=filters><select id=relBase>${dateOptions()}</select><select id=relFinal>${dateOptions()}</select><button onclick="buildRelease()">Shakllantirish</button></div><div id=releaseResult></div></div><div class=stack>${["1","3","5","10","30"].map(d=>{let r=rel[d]||{rows:[]};let title=d==="30"?"1 oy ichida yechilgan":`${d} kun ichida yechilgan`;let note=r.missing_date?`<div class=muted>${r.missing_date} sanasidagi asos fayl arxivda yo'q.</div>`:`<div class=muted>Asos sana: ${r.base_date||"-"}</div>`;return `<div class=panel><h2>${title}</h2>${note}${table([{k:"company",t:"Korxona nomi",w:"40%"},{k:"stir",t:"STIR",w:"10%"},{k:"decl",t:"Deklaratsiya",w:"16%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN}],periodRows(r))}</div>`}).join("")}</div>`;return}
if(TAB==="goods"){let goodsCols=[{k:"name",t:"Tovar guruhi",w:"36%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"korxona",t:"Korxona",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN}],gp=topNWithOther(DATA.goods||[],30,"qiymat","name"),gpart=topNWithOther(DATA.goods||[],30,"partiya","name"),gvazn=topNWithOther(DATA.goods||[],30,"vazn","name");v.innerHTML=wlayout('goods',{rating:()=>tovarRatingPanel(DATA.goods||[]),trend:()=>`<div class="panel" id="goodsTrendPanel"><h3>Tovarlar bo'yicha davriy tendensiya</h3><div class="muted">Yuklanmoqda...</div></div>`,jadval:()=>`<div class="panel wide"><h2>Tovarlar guruhlari ${xls("goods")} ${staleBadge()}</h2>${table(goodsCols,basicTotal(gp))}</div>`,bars_group:()=>`<div class=stack><div class=panel><h2>Partiya bo'yicha ulushi</h2>${bars(gpart,"name","partiya",fmtI)}</div><div class=panel><h2>Qiymat bo'yicha ulushi</h2>${bars(gp,"name","qiymat",fmtN)}</div><div class=panel><h2>Vazn bo'yicha ulushi</h2>${bars(gvazn,"name","vazn",fmtN)}</div></div>`});loadGoodsTrends();return}
if(TAB==="food"){v.innerHTML=wlayout('food',{jadval:()=>`<div class=panel><h2>Oziq-ovqatlar kesimida ${xls("food")} ${staleBadge()}</h2>${table([{k:"name",t:"Oziq-ovqat turi",w:"42%"},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"over_vazn",t:"3 oy+ vazn (tn)",n:1,f:fmtN},{k:"over_qiymat",t:"3 oy+ qiymat (ming $)",n:1,f:fmtN},{k:"ulush",t:"Qiymatdagi ulushi (%)",n:1,f:fmtN}],foodRows())}</div>`,bars:()=>`<div class=panel><h2>Qiymat ulushi</h2>${bars(DATA.food||[],"name","qiymat",fmtN)}</div>`});return}
if(TAB==="yaroqlilik"){
  loadYaroqlilik();
  let yd=YAROQLILIK_DATA||{items:[],loaded:false};
  let items=yd.items||[];
  if(!yd.loaded){v.innerHTML=`<div class=panel><h2>Iste'mol muddati tahlili</h2><div class=muted>Ma'lumotlar yuklanmagan yoki yuklanmoqda. "Boshqaruv → Fayl yuklash" bo'limidan yaroqlilik Excel faylini yuklang.</div><button onclick="GROUP='common';TAB='upload';render()">Fayl yuklash</button></div>`;return}
  let expired=items.filter(r=>r.holat==="Muddati o'tgan");
  let warn180=items.filter(r=>r.holat==="180 kundan kam qolgan");
  let warn30=items.filter(r=>r.holat==="1 oy ichida muddati tugaydi");
  let activeF=YAROQLILIK_FILTER||'expired';
  let filtItems=activeF==='warn180'?warn180:activeF==='warn30'?warn30:expired;
  let byComp={};expired.forEach(r=>{let k=r.korxona||r.stir||'-';if(!byComp[k])byComp[k]={korxona:k,stir:r.stir||'',cnt:0,qiymat:0,vazn:0};byComp[k].cnt++;byComp[k].qiymat+=(+r.qiymat||0);byComp[k].vazn+=(+r.vazn||0)});
  let compRows=Object.values(byComp).sort((a,b)=>b.qiymat-a.qiymat).slice(0,15);
  let eqiymat=expired.reduce((s,r)=>s+(+r.qiymat||0),0);
  let tvazn=items.reduce((s,r)=>s+(+r.vazn||0),0);
  let tblRows=filtItems.map(r=>({decl:r.decl_raqami||'-',korxona:(r.korxona||'-').slice(0,50),stir:r.stir||'',hs:r.hs_kod||'',tovar:(r.tovar_nomi||'-').slice(0,80),expiry:r.yaroqlilik||'-',kun:r.otgan_kun?('+'+r.otgan_kun+' kun o\'tgan'):(r.qolgan_kun||0)+' kun qoldi',vazn:+r.vazn||0,qiymat:+r.qiymat||0}));
  v.innerHTML=`<div class=stack><div class="panel wide"><h2>Iste'mol muddati tahlili · ${yd.report_date||''} <button class="btn light" onclick="GROUP='common';TAB='upload';render()">Yangi fayl yuklash</button></h2><div class="summary-grid" style="grid-template-columns:repeat(5,minmax(120px,1fr));margin-bottom:12px"><div class="summary-item" style="border-color:rgba(200,40,40,.4)"><b style="font-size:22px;color:#c82828">${expired.length}</b><span>Muddati o'tgan tovar</span></div><div class="summary-item" style="border-color:rgba(220,140,0,.4)"><b style="font-size:22px;color:#d48c00">${warn180.length}</b><span>180 kun ichida tugaydi</span></div><div class="summary-item"><b style="font-size:22px">${warn30.length}</b><span>1 oy ichida tugaydi</span></div><div class="summary-item"><b style="font-size:22px">${fmtN(eqiymat)}</b><span>O'tgan tovar qiymati (ming $)</span></div><div class="summary-item"><b style="font-size:22px">${fmtN(tvazn/1000)}</b><span>Umumiy vazn (tonna)</span></div></div><div class="filters compact-filters" style="margin-bottom:8px"><button class="${activeF==='expired'?'btn':'btn light'}" style="font-size:12px;padding:4px 12px" onclick="YAROQLILIK_FILTER='expired';render()">Muddati o'tgan (${expired.length})</button><button class="${activeF==='warn180'?'btn':'btn light'}" style="font-size:12px;padding:4px 12px" onclick="YAROQLILIK_FILTER='warn180';render()">180 kun ichida (${warn180.length})</button><button class="${activeF==='warn30'?'btn':'btn light'}" style="font-size:12px;padding:4px 12px" onclick="YAROQLILIK_FILTER='warn30';render()">1 oy ichida (${warn30.length})</button></div>${table([{k:"decl",t:"Deklaratsiya raqami",w:"170px"},{k:"korxona",t:"Yuk qabul qiluvchi",w:"200px"},{k:"stir",t:"STIR",w:"88px"},{k:"hs",t:"TIF TN kodi",w:"90px"},{k:"tovar",t:"Tovar nomi",w:"240px"},{k:"expiry",t:"Yaroqlilik muddati",w:"110px"},{k:"kun",t:"Muddat holati",w:"120px"},{k:"vazn",t:"Vazni (kg)",n:1,f:fmtN,w:"85px"},{k:"qiymat",t:"Qiymati (ming $)",n:1,f:fmtN,w:"100px"}],tblRows,"fixed-table")}<div class=overview-note>Hisob-kitob sanasi: ${yd.report_date||"noma'lum"}. Jami ${items.length} ta tovar pozitsiyasi tahlil qilindi.</div></div><div class=panel><h2>Muddati o'tgan tovarlar — korxonalar kesimida</h2>${table([{k:"korxona",t:"Korxona nomi",w:"280px"},{k:"stir",t:"STIR",w:"90px"},{k:"cnt",t:"Pozitsiyalar",n:1,f:fmtI,w:"85px"},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN,w:"110px"},{k:"vazn",t:"Vazn (kg)",n:1,f:fmtN,w:"95px"}],compRows)}</div></div>`;return}}
async function buildRelease(){let base=$("relBase").value, final=$("relFinal").value, j=await api(`/api/release?base=${encodeURIComponent(base)}&final=${encodeURIComponent(final)}`);if(j.missing){$("releaseResult").innerHTML=`<div class=muted>Arxivda yo'q sana: ${j.missing.join(", ")}. Shu davr uchun asos fayl yuklash kerak.</div>`;return}let raw=(j.rows||[]).map(r=>Object.assign({type_text:r.release_type==="qisman"?"Qisman":"To'liq",korxona:r.korxona||r.company||""},r));let rows=numbered([Object.assign({korxona:"IBK bo'yicha Jami",stir:"",decl:"",item_no:"",regime:"",post:"",type_text:""},j.total||{})].concat(raw));let cols=[{k:"rn",t:"T/r",n:1,f:v=>v,w:"34px"},{k:"korxona",t:"Korxona nomi",w:"240px"},{k:"stir",t:"STIR",w:"82px"},{k:"decl",t:"Deklaratsiya",w:"135px"},{k:"item_no",t:"Tovar tartib raqami",w:"74px"},{k:"regime",t:"Rejim",w:"62px"},{k:"post",t:"Post",w:"145px"},{k:"base_partiya",t:`${base} holatiga partiya`,n:1,f:fmtI,w:"76px"},{k:"base_qiymat",t:`${base} holatiga qiymat (ming $)`,n:1,f:fmtN,w:"104px"},{k:"remain_partiya",t:`${final} holatiga qoldiq partiya`,n:1,f:fmtI,w:"80px"},{k:"remain_qiymat",t:`${final} holatiga qoldiq qiymat (ming $)`,n:1,f:fmtN,w:"104px"},{k:"released_partiya",t:"Yechilgan partiya",n:1,f:fmtI,w:"78px"},{k:"released_vazn",t:"Yechilgan vazn (tn)",n:1,f:fmtN,w:"96px"},{k:"released_qiymat",t:"Yechilgan qiymat (ming $)",n:1,f:fmtN,w:"105px"},{k:"released_pct",t:"Qiymatdagi ulushi (%)",n:1,f:fmtN,w:"88px"},{k:"type_text",t:"Holati",w:"82px"}];$("releaseResult").innerHTML=`<h2>${base} - ${final} <a class="btn light" href="/api/export?kind=release&base=${base}&final=${final}&token=${TOKEN}">Excel</a></h2><div class=overview-note>Boshlang'ich davr: ${base}. Yakuniy davr: ${final}. Qisman yechilgan qatorlar qiymat/vazn kamaygan, lekin partiya soni o'zgarmagan holatlarni bildiradi.</div>${table(cols,rows,"release-table fixed-table")}${miniChart(raw,"released_qiymat","korxona")}`}
async function detail(key){if(!DATA||!key||Object.keys(key).length===0)return;if(key.view==="expired_inline"){let el=$("expiredInline");if(el)el.innerHTML=`<h2>Postlar va rejimlar kesimida</h2>${table([{k:"post",t:"Post",w:"22%"},{k:"jami_partiya",t:"Jami partiya",n:1,f:fmtI},{k:"jami_qiymat",t:"Jami qiymat (ming $)",n:1,f:fmtN},{k:"expired_partiya",t:"Muddati o'tgan partiya",n:1,f:fmtI},{k:"expired_qiymat",t:"Muddati o'tgan qiymat (ming $)",n:1,f:fmtN},{k:"ulush",t:"Partiyadagi ulushi (%)",n:1,f:fmtN},{k:"im70_partiya",t:"IM70 partiya",n:1,f:fmtI},{k:"im74_partiya",t:"IM74 partiya",n:1,f:fmtI},{k:"tr80_partiya",t:"TR80 partiya",n:1,f:fmtI}],expiredPostRegimeRows())}`;return}if(key.view==="regime_posts"){let rows=(DATA.regime_posts||{})[key.regime]||[];dlgTitle.textContent=`${key.regime} - postlar kesimida`;dlgBody.innerHTML=table([{k:"post",t:"Post",w:"42%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN}],basicTotal(rows,"IBK bo'yicha Jami","post"));dlg.showModal();return}let q=new URLSearchParams({report:DATA.id,filters:JSON.stringify(key)}), j=await api("/api/details?"+q);dlgTitle.textContent="Asos deklaratsiyalar";dlgBody.innerHTML=table([{k:"decl",t:"Deklaratsiya"},{k:"date",t:"Sana"},{k:"regime",t:"Rejim"},{k:"stir",t:"STIR"},{k:"company",t:"Korxona"},{k:"hs",t:"TIF TN"},{k:"goods",t:"Tovar"},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN}],j.rows);dlg.showModal()}
function roleTitle(r){return ({admin:"Admin",rahbar:"Rahbar",inspektor:"Inspektor",foydalanuvchi:"Foydalanuvchi",user:"Foydalanuvchi"}[r]||r||"Foydalanuvchi")}function permTitle(p){return ({view:"Ko'rish",upload:"Fayl yuklash",export:"Excel/PDF/PNG",release:"Nazoratdan yechish",admin:"Admin",settings:"Sozlamalar"}[p]||p)}function miniChart(rows,k,label){rows=by(rows||[],k).slice(0,10);let max=rows.reduce((m,r)=>Math.max(m,+r[k]||0),0)||1;return `<div class="live-chart"><div class="sparkline">${rows.map((r,i)=>`<i title="${esc(r[label]||r.name||r.korxona||'')}: ${fmtN(r[k])}" style="height:${Math.max(8,(+r[k]||0)/max*100)}%;animation-delay:${i*.04}s"></i>`).join("")}</div></div>`}function expiredTotalExcelTable(){let cols=[{k:"post",t:"Post",w:"210px"},{k:"jami_partiya",t:"Jami partiya",n:1,f:fmtI,w:"82px"},{k:"jami_vazn",t:"Jami vazn (tn)",n:1,f:fmtN,w:"96px"},{k:"jami_qiymat",t:"Jami qiymat (ming $)",n:1,f:fmtN,w:"110px"},{k:"expired_partiya",t:"Muddati o'tgan partiya",n:1,f:fmtI,w:"92px"},{k:"expired_vazn",t:"Muddati o'tgan vazn (tn)",n:1,f:fmtN,w:"106px"},{k:"expired_qiymat",t:"Muddati o'tgan qiymat (ming $)",n:1,f:fmtN,w:"116px"},{k:"ulush",t:"Partiyadagi ulushi (%)",n:1,f:fmtN,w:"95px"},{k:"im70_partiya",t:"IM70 partiya",n:1,f:fmtI,w:"78px"},{k:"im74_partiya",t:"IM74 partiya",n:1,f:fmtI,w:"78px"},{k:"tr80_partiya",t:"TR80 partiya",n:1,f:fmtI,w:"78px"},{k:"note",t:"Izoh",w:"120px"}];let rows=expiredPostRegimeRows().map(r=>Object.assign({jami_vazn:r.jami_vazn||0,expired_vazn:r.expired_vazn||0,note:""},r));return table(cols,rows,"expired-total-table")}function adminPanel(){return `<div class="admin-layout"><div class="admin-card"><h2>Yangi xodim</h2><form id=userForm class=admin-form><input name=user placeholder="Login" required><input name=password placeholder="Dastlabki parol" required><input name=full_name placeholder="F.I.Sh."><input name=position placeholder="Lavozim"><input name=phone placeholder="Telefon"><input name=post_code placeholder="Post kodi, masalan 00102"><select name=role><option value=foydalanuvchi>Foydalanuvchi</option><option value=inspektor>Inspektor</option><option value=rahbar>Rahbar</option><option value=admin>Admin</option></select><select name=lang><option value=uz>O'zbek lotin</option><option value=uzc>O'zbek kirill</option><option value=ru>Rus tili</option></select><div class=perm-grid><label><input type=checkbox name=perm_view checked> Ko'rish</label><label><input type=checkbox name=perm_upload> Yuklash</label><label><input type=checkbox name=perm_export> Eksport</label><label><input type=checkbox name=perm_release> Yechish</label></div><button>Saqlash</button></form></div><div class="admin-card"><h2>Hodimlar ro'yxati</h2><div id=users></div></div></div>`}function bindUserForm(){let f=$("userForm");if(!f)return;f.onsubmit=async e=>{e.preventDefault();let perms=[];["view","upload","export","release"].forEach(p=>{if(f[`perm_${p}`]?.checked)perms.push(p)});await api("/api/users",{method:"POST",body:JSON.stringify({user:f.user.value,pass:f.password.value,full_name:f.full_name.value,position:f.position.value,phone:f.phone.value,post_code:f.post_code.value,role:f.role.value,lang:f.lang.value,perms})});loadUsers();f.reset();f.perm_view.checked=true}} async function loadUsers(){let box=$("users");if(!box)return;try{let j=await api("/api/users");let rows=(j.users||[]).map(u=>Object.assign({},u,{role_label:`<span class="role-pill">${roleTitle(u.role)}</span>`,perms_text:(u.perms||[]).map(permTitle).join(", ")}));box.innerHTML=table([{k:"user",t:"Login",w:"90px"},{k:"full_name",t:"F.I.Sh.",w:"210px"},{k:"role_label",t:"Rol",w:"120px"},{k:"post_code",t:"Post",w:"75px"},{k:"perms_text",t:"Vakolatlar",w:"260px"}],rows,"fixed-table").replaceAll("&lt;span class=&quot;role-pill&quot;&gt;","<span class=\"role-pill\">").replaceAll("&lt;/span&gt;","</span>")}catch(e){box.innerHTML="Admin vakolati kerak"}}
function svgLineChart(periods,series){let w=900,h=230,padL=56,padR=16,padT=18,padB=34;let allVals=series.flatMap(s=>s.values);let maxV=Math.max(1,...allVals);let n=Math.max(periods.length,1);let x=i=>padL+(n>1?i/(n-1)*(w-padL-padR):(w-padL-padR)/2);let y=val=>h-padB-(val/maxV)*(h-padT-padB);let grid=[0,.25,.5,.75,1].map(f=>{let val=maxV*f,yy=y(val);return `<line x1="${padL}" y1="${yy.toFixed(1)}" x2="${w-padR}" y2="${yy.toFixed(1)}" class="trend-grid"/><text x="${padL-8}" y="${(yy+4).toFixed(1)}" text-anchor="end" class="trend-axis">${fmtN(val)}</text>`}).join("");let lines=series.map((s,si)=>{let pts=s.values.map((v,i)=>`${x(i).toFixed(1)},${y(v).toFixed(1)}`).join(" ");let dots=s.values.map((v,i)=>`<circle cx="${x(i).toFixed(1)}" cy="${y(v).toFixed(1)}" r="3.6" class="trend-dot c${si}"><title>${esc(periods[i]||"")}: ${fmtN(v)}</title></circle>`).join("");return `<polyline points="${pts}" class="trend-line c${si}"/>${dots}`}).join("");let xlabels=periods.map((p,i)=>{if(periods.length>10&&i%Math.ceil(periods.length/10)!==0&&i!==periods.length-1)return "";return `<text x="${x(i).toFixed(1)}" y="${h-8}" text-anchor="middle" class="trend-axis">${esc(p)}</text>`}).join("");let legend=series.map((s,si)=>`<span class="trend-legend c${si}">\u25CF ${esc(s.label)}</span>`).join(" ");return `<div class="trend-chart-wrap"><svg viewBox="0 0 ${w} ${h}" class="trend-chart" preserveAspectRatio="xMidYMid meet">${grid}${lines}${xlabels}</svg><div class="trend-legend-row">${legend}</div></div>`}
function periodRangeCaption(periods){if(!periods.length)return "";if(periods.length===1)return `Davr: ${periods[0]} (hozircha yagona hisobot davri)`;return `Tahlil qilingan davrlar: ${periods[0]} dan ${periods[periods.length-1]} gacha, jami ${periods.length} ta hisobot davri (har bir davr - shu sanada yuklangan hisobot bo'yicha faol nazoratdagi qoldiq)`}
function seriesNZ(p){return (p.value>0.005)||(p.weight>0.005)}
function firstActiveIdx(series){for(let i=0;i<series.length;i++){if(seriesNZ(series[i]))return i}return -1}
function lastActiveIdx(series){for(let i=series.length-1;i>=0;i--){if(seriesNZ(series[i]))return i}return -1}
function trendBuckets(periods,items){let lastIdx=periods.length-1;function hadAny(series,upTo){return series.slice(0,upTo).some(seriesNZ)}let last2idx=periods.length>=2?periods.length-2:0;let active=items.filter(c=>seriesNZ(c.series[lastIdx]));let stoppedAll=items.filter(c=>hadAny(c.series,lastIdx)&&!seriesNZ(c.series[lastIdx]));let newAll=items.filter(c=>!hadAny(c.series,lastIdx)&&seriesNZ(c.series[lastIdx]));let stopped2=periods.length>=2?items.filter(c=>hadAny(c.series,last2idx)&&c.series.slice(last2idx).every(p=>!seriesNZ(p))):[];let new2=periods.length>=2?items.filter(c=>!hadAny(c.series,last2idx)&&c.series.slice(last2idx).some(seriesNZ)):[];return {active,stoppedAll,newAll,stopped2,new2}}
let COMP_PERIOD_M=3,GOODS_PERIOD_M=3;
function filterPeriodItems(periods,items,months){if(!periods.length)return{newItems:[],stoppedItems:[]};let lastDate=new Date(periods[periods.length-1]);let cutoff=new Date(lastDate);cutoff.setMonth(cutoff.getMonth()-months);let cutoffStr=cutoff.toISOString().slice(0,10);let bi=[],wi=[];periods.forEach((p,i)=>{if(p<cutoffStr)bi.push(i);else wi.push(i)});if(!wi.length)return{newItems:[],stoppedItems:[]};let newItems=items.filter(item=>{let hb=bi.some(i=>seriesNZ(item.series[i]));let hw=wi.some(i=>seriesNZ(item.series[i]));return!hb&&hw});let stoppedItems=items.filter(item=>{let hb=bi.some(i=>seriesNZ(item.series[i]));let hw=wi.some(i=>seriesNZ(item.series[i]));return hb&&!hw});return{newItems,stoppedItems}}
function periodBtns(fn,active){return[[1,'🗓 1 oy'],[3,'📅 3 oy'],[6,'📆 6 oy'],[12,'📊 1 yil']].map(([n,lbl])=>`<button class="${n===active?'btn':'btn light'}" style="padding:4px 12px;font-size:12px" onclick="${fn}(${n})">${lbl}</button>`).join('')}
function filterPeriodArr(periods,months){if(!periods||!periods.length)return[];let lastDate=new Date(periods[periods.length-1]);let cutoff=new Date(lastDate);cutoff.setMonth(cutoff.getMonth()-months);let cutoffStr=cutoff.toISOString().slice(0,10);return periods.filter(p=>p>=cutoffStr)}
function renderCompChart(m){let el=$('compChartSect');if(!el||!COMPANY_TRENDS)return;let periods=COMPANY_TRENDS.periods||[],companies=COMPANY_TRENDS.companies||[];if(!periods.length)return;let fp=filterPeriodArr(periods,m);let wi=fp.map(p=>periods.indexOf(p));let totalValues=fp.map((p,j)=>companies.reduce((a,c)=>a+(c.series[wi[j]]?.value||0),0));let totalWeights=fp.map((p,j)=>companies.reduce((a,c)=>a+(c.series[wi[j]]?.weight||0),0));let range=fp.length>=2?`${fp[0]} — ${fp[fp.length-1]}`:fp[0]||"";el.innerHTML=`<div class="panel"><h3>Umumiy oborot tendensiyasi (${range})</h3>${svgLineChart(fp,[{label:"Qiymat (ming $)",values:totalValues},{label:"Vazn (tn)",values:totalWeights}])}<div class="overview-note">Har qaysi korxona qatoriga bosilganda shu korxonaning alohida qiymat/vazn tendensiyasi ochiladi.</div></div>`}
function renderGoodsChart(m){let el=$('goodsChartSect');if(!el||!GOODS_TRENDS)return;let periods=GOODS_TRENDS.periods||[],items=GOODS_TRENDS.goods||[];if(!periods.length)return;let fp=filterPeriodArr(periods,m);let wi=fp.map(p=>periods.indexOf(p));let totalValues=fp.map((p,j)=>items.reduce((a,g)=>a+(g.series[wi[j]]?.value||0),0));let totalWeights=fp.map((p,j)=>items.reduce((a,g)=>a+(g.series[wi[j]]?.weight||0),0));let range=fp.length>=2?`${fp[0]} — ${fp[fp.length-1]}`:fp[0]||"";el.innerHTML=`<div class="panel"><h3>Umumiy tovar oqimi tendensiyasi (${range})</h3>${svgLineChart(fp,[{label:"Qiymat (ming $)",values:totalValues},{label:"Vazn (tn)",values:totalWeights}])}<div class="overview-note">Har qaysi tovar qatoriga bosilganda shu tovarning alohida qiymat/vazn tendensiyasi ochiladi.</div></div>`}
function renderCompNewStopped(m){COMP_PERIOD_M=m;renderCompChart(m);let el=$('compNewStoppedSect');if(!el||!COMPANY_TRENDS)return;let periods=COMPANY_TRENDS.periods||[],companies=COMPANY_TRENDS.companies||[];let{newItems,stoppedItems}=filterPeriodItems(periods,companies,m);let mLabel=m===1?'1 oy':m+' oy';el.innerHTML=`<div style="display:flex;gap:8px;align-items:center;margin:4px 0 12px;flex-wrap:wrap"><span style="font-size:12px;color:#557086;font-weight:700">Hisobot davri:</span>${periodBtns('renderCompNewStopped',m)}</div><div class="grid2">${companyTrendTable(periods,stoppedItems,'🚫 So\'nggi '+mLabel+' ichida faoliyatini to\'xtatgan korxonalar','Barcha korxonalar faol','stopped')}${companyTrendTable(periods,newItems,'🆕 So\'nggi '+mLabel+' ichida birinchi marta yuk olib kelgan korxonalar','Yangi korxona aniqlanmadi','new')}</div>`}
function renderGoodsNewStopped(m){GOODS_PERIOD_M=m;renderGoodsChart(m);let el=$('goodsNewStoppedSect');if(!el||!GOODS_TRENDS)return;let periods=GOODS_TRENDS.periods||[],items=GOODS_TRENDS.goods||[];let{newItems,stoppedItems}=filterPeriodItems(periods,items,m);let mLabel=m===1?'1 oy':m+' oy';el.innerHTML=`<div style="display:flex;gap:8px;align-items:center;margin:4px 0 12px;flex-wrap:wrap"><span style="font-size:12px;color:#557086;font-weight:700">Hisobot davri:</span>${periodBtns('renderGoodsNewStopped',m)}</div><div class="grid2">${goodsTrendTable(periods,stoppedItems,'🚫 So\'nggi '+mLabel+' ichida IBKdan chiqib ketgan tovarlar','Barcha tovarlar faol','stopped')}${goodsTrendTable(periods,newItems,'🆕 So\'nggi '+mLabel+' ichida yangi qo\'shilgan tovarlar','Yangi tovar aniqlanmadi','new')}</div>`}
async function buildOmborOborot(){let base=$('omborOborotBase').value,final=$('omborOborotFinal').value,box=$('omborOborotResult');if(!box)return;box.innerHTML='<div class=muted>Hisoblanmoqda...</div>';let j=await loadReleaseData(base,final,box);if(!j)return;box.innerHTML=warehouseTurnoverPanel(j)}
function companyTrendTable(periods,rows,title,emptyMsg,mode){if(!rows.length)return `<div class="panel"><h3>${esc(title)} (0)</h3><div class="muted">${esc(emptyMsg)}</div></div>`;let lastIdx=periods.length-1;let dateHead=mode==="new"?"Birinchi faol davr":"So'nggi faol davr";let valHead=mode==="new"?"Joriy qiymat (ming $)":"So'nggi qiymat (ming $)";let wHead=mode==="new"?"Joriy vazn (tn)":"So'nggi vazn (tn)";let body=rows.slice(0,150).map(c=>{let refIdx,dateLabel;if(mode==="new"){refIdx=lastIdx;let fi=firstActiveIdx(c.series);dateLabel=fi>=0?periods[fi]:"-"}else{refIdx=lastActiveIdx(c.series);dateLabel=refIdx>=0?periods[refIdx]:"-"}let s=c.series[refIdx>=0?refIdx:lastIdx]||{value:0,weight:0};return `<tr style="cursor:pointer" onclick="showCompanyTrend('${c.stir}')"><td class=text>${esc(c.company||"-")}</td><td>${esc(c.stir)}</td><td>${esc(dateLabel)}</td><td class=num>${fmtN(s.value)}</td><td class=num>${fmtN(s.weight)}</td></tr>`}).join("");return `<div class="panel"><h3>${esc(title)} (${rows.length})</h3><table class="fixed-table"><thead><tr><th>Korxona</th><th>STIR</th><th>${dateHead}</th><th>${valHead}</th><th>${wHead}</th></tr></thead><tbody>${body}</tbody></table></div>`}
function showCompKpiList(type){let b=window.COMP_BUCKETS||{},companies=window.COMP_ALL_COMPANIES||[],periods=(COMPANY_TRENDS&&COMPANY_TRENDS.periods)||[];let lastIdx=periods.length-1;let rows,title;if(type==='all'){rows=companies;title='Jami korxonalar ('+companies.length+')';}else if(type==='active'){rows=b.active||[];title='So\'nggi davrda faol ('+rows.length+')';}else if(type==='stopped'){rows=b.stoppedAll||[];title='Butunlay to\'xtatgan korxonalar ('+rows.length+')';}else if(type==='new'){rows=b.newAll||[];title='Yangi qo\'shilgan korxonalar ('+rows.length+')';}else return;dlgTitle.textContent=title;dlgBody.innerHTML=table([{k:"company",t:"Korxona nomi",w:"50%"},{k:"stir",t:"STIR",w:"110px"},{k:"last_value",t:"So\'nggi qiymat (ming $)",n:1,f:fmtN},{k:"last_weight",t:"Vazn (tn)",n:1,f:fmtN}],rows.slice(0,300).map(c=>{let s=c.series[lastIdx]||{value:0,weight:0};return {key:{stir:c.stir},company:c.company||c.stir,stir:c.stir,last_value:s.value,last_weight:s.weight}}),"fixed-table");dlg.showModal()}
async function loadCompanyTrends(){let el=$("trendPanel");if(!el)return;try{let j=await api("/api/company_trends");COMPANY_TRENDS=j;if(j.error&&!j.periods?.length){el.innerHTML=`<h3>Korxonalar bo'yicha davriy tendensiya</h3><div class="muted">Server xatoligi: ${esc(j.error)}</div>`;return}let periods=j.periods||[],companies=j.companies||[];if(!periods.length){el.innerHTML=`<h3>Korxonalar bo'yicha davriy tendensiya</h3><div class="muted">Hali tarixiy ma'lumot yetarli emas. Har bir yangi yuklangan fayl bilan tendensiya shakllanadi.</div>`;return}let b=trendBuckets(periods,companies);window.COMP_BUCKETS=b;window.COMP_ALL_COMPANIES=companies;let lastPeriod=periods[periods.length-1]||'';el.innerHTML=`<h3>Korxonalar bo'yicha davriy tendensiya</h3><div class="overview-note">${periodRangeCaption(periods)}. Har bir davr — shu sanada yuklab olingan hisobot asosidagi nazoratdagi qoldiq. <b>Faollik mezoni:</b> so'nggi davr qiymati yoki vazni 0 dan katta (> 0.005). <b>To'xtagan mezoni:</b> avvalgi davrlarda nazoratda bo'lgan, ammo ${lastPeriod} da qiymati va vazni nolga tushgan.</div><div class="kpis"><div class="kpi" onclick="showCompKpiList('all')" style="cursor:pointer" title="Barcha nazoratdan o\'tgan korxonalar (barcha davrlar bo\'yicha)"><span>Jami korxonalar</span><b>${fmtI(companies.length)}</b><small class="kpi-note">Barcha davrlar</small></div><div class="kpi" onclick="showCompKpiList('active')" style="cursor:pointer" title="So\'nggi hisobot davri (${lastPeriod}) da qiymati yoki vazni >0 bo\'lgan korxonalar"><span>So'nggi davrda faol</span><b>${fmtI(b.active.length)}</b><small class="kpi-note">${lastPeriod} holati</small></div><div class="kpi" onclick="showCompKpiList('stopped')" style="cursor:pointer" title="Avvalgi davrlarda nazoratda bo\'lgan, ammo ${lastPeriod} da qiymati va vazni 0 ga tushgan korxonalar"><span>Butunlay to'xtagan</span><b>${fmtI(b.stoppedAll.length)}</b><small class="kpi-note">${lastPeriod} da sifr</small></div><div class="kpi" onclick="showCompKpiList('new')" style="cursor:pointer" title="Ilgari nazoratda ko\'rinmagan, ammo ${lastPeriod} da birinchi marta faollashgan korxonalar"><span>Yangi qo'shilgan</span><b>${fmtI(b.newAll.length)}</b><small class="kpi-note">${lastPeriod} da yangi</small></div></div><div id="compChartSect"></div><div id="compNewStoppedSect"></div>`;renderCompChart(COMP_PERIOD_M);renderCompNewStopped(COMP_PERIOD_M)}catch(e){el.innerHTML=`<h3>Korxonalar bo'yicha davriy tendensiya</h3><div class="muted">Ma'lumotni yuklab bo'lmadi: ${esc(e.message||e)}</div>`}}
function showCompanyTrend(stir){let c=(COMPANY_TRENDS.companies||[]).find(x=>x.stir===stir);renderCompNewStopped(COMP_PERIOD_M);if(!c)return;let periods=COMPANY_TRENDS.periods||[];let values=c.series.map(s=>s.value),weights=c.series.map(s=>s.weight);dlgTitle.textContent=`${c.company||c.stir} - davriy tendensiya (STIR: ${c.stir})`;let rows=c.series.map((s,i)=>({date:periods[i],value:s.value,weight:s.weight,partiya:s.partiya,key:{}}));dlgBody.innerHTML=`<div class="overview-note">${periodRangeCaption(periods)}.</div>`+svgLineChart(periods,[{label:"Qiymat (ming $)",values},{label:"Vazn (tn)",values:weights}])+table([{k:"date",t:"Davr",w:"110px"},{k:"value",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"weight",t:"Vazn (tn)",n:1,f:fmtN},{k:"partiya",t:"Partiya",n:1,f:fmtI}],rows,"fixed-table");dlg.showModal()}
function goodsTrendTable(periods,rows,title,emptyMsg,mode){if(!rows.length)return `<div class="panel"><h3>${esc(title)} (0)</h3><div class="muted">${esc(emptyMsg)}</div></div>`;let lastIdx=periods.length-1;let dateHead=mode==="new"?"Birinchi faol davr":"So'nggi faol davr";let valHead=mode==="new"?"Joriy qiymat (ming $)":"So'nggi qiymat (ming $)";let wHead=mode==="new"?"Joriy vazn (tn)":"So'nggi vazn (tn)";let body=rows.slice(0,150).map(g=>{let refIdx,dateLabel;if(mode==="new"){refIdx=lastIdx;let fi=firstActiveIdx(g.series);dateLabel=fi>=0?periods[fi]:"-"}else{refIdx=lastActiveIdx(g.series);dateLabel=refIdx>=0?periods[refIdx]:"-"}let s=g.series[refIdx>=0?refIdx:lastIdx]||{value:0,weight:0};return `<tr style="cursor:pointer" title="${esc(g.goods||"-")}" onclick="showGoodsTrend('${g.hs_code}')"><td class=text>${esc(g.goods_short||g.hs_code||"-")}</td><td>${esc(g.hs_code)}</td><td>${esc(dateLabel)}</td><td class=num>${fmtN(s.value)}</td><td class=num>${fmtN(s.weight)}</td></tr>`}).join("");return `<div class="panel"><h3>${esc(title)} (${rows.length})</h3><table class="fixed-table"><thead><tr><th>Tovar</th><th>TIF TN kodi</th><th>${dateHead}</th><th>${valHead}</th><th>${wHead}</th></tr></thead><tbody>${body}</tbody></table></div>`}
async function loadGoodsTrends(){let el=$("goodsTrendPanel");if(!el)return;try{let j=await api("/api/goods_trends");GOODS_TRENDS=j;if(j.error&&!j.periods?.length){el.innerHTML=`<h3>Tovarlar bo'yicha davriy tendensiya</h3><div class="muted">Server xatoligi: ${esc(j.error)}</div>`;return}let periods=j.periods||[],items=j.goods||[];if(!periods.length){el.innerHTML=`<h3>Tovarlar bo'yicha davriy tendensiya</h3><div class="muted">Hali tarixiy ma'lumot yetarli emas. Har bir yangi yuklangan fayl bilan tendensiya shakllanadi.</div>`;return}let b=trendBuckets(periods,items);el.innerHTML=`<h3>Tovarlar bo'yicha davriy tendensiya</h3><div class="overview-note">${periodRangeCaption(periods)}. Har bir tovar TIF TN (HS) kodi bo'yicha guruhlangan.</div><div class="kpis"><div class="kpi"><span>Jami tovar turlari</span><b>${fmtI(items.length)}</b></div><div class="kpi"><span>So'nggi davrda faol</span><b>${fmtI(b.active.length)}</b></div><div class="kpi"><span>Chiqib ketgan</span><b>${fmtI(b.stoppedAll.length)}</b></div><div class="kpi"><span>Yangi qo'shilgan</span><b>${fmtI(b.newAll.length)}</b></div></div><div id="goodsChartSect"></div><div id="goodsNewStoppedSect"></div>`;renderGoodsChart(GOODS_PERIOD_M);renderGoodsNewStopped(GOODS_PERIOD_M)}catch(e){el.innerHTML=`<h3>Tovarlar bo'yicha davriy tendensiya</h3><div class="muted">Ma'lumotni yuklab bo'lmadi: ${esc(e.message||e)}</div>`}}
function showGoodsTrend(hs){let g=(GOODS_TRENDS.goods||[]).find(x=>x.hs_code===hs);renderGoodsNewStopped(GOODS_PERIOD_M);if(!g)return;let periods=GOODS_TRENDS.periods||[];let values=g.series.map(s=>s.value),weights=g.series.map(s=>s.weight);dlgTitle.textContent=`${g.goods||g.hs_code} - davriy tendensiya (TIF TN: ${g.hs_code})`;let rows=g.series.map((s,i)=>({date:periods[i],value:s.value,weight:s.weight,partiya:s.partiya,key:{}}));dlgBody.innerHTML=`<div class="overview-note">${periodRangeCaption(periods)}.</div>`+svgLineChart(periods,[{label:"Qiymat (ming $)",values},{label:"Vazn (tn)",values:weights}])+table([{k:"date",t:"Davr",w:"110px"},{k:"value",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"weight",t:"Vazn (tn)",n:1,f:fmtN},{k:"partiya",t:"Partiya",n:1,f:fmtI}],rows,"fixed-table");dlg.showModal()}

let WAREHOUSE_TRENDS={},TRANSPORT_TRENDS={},TRANSPORT_CO_TRENDS={};
let WAREHOUSE_PERIOD_M=6,TRANSPORT_PERIOD_M=3;
function renderWarehouseChart(m){WAREHOUSE_PERIOD_M=m;let el=$('whChartSect');if(!el||!WAREHOUSE_TRENDS?.periods?.length)return;let periods=WAREHOUSE_TRENDS.periods,warehouses=WAREHOUSE_TRENDS.warehouses;let fp=filterPeriodArr(periods,m);let wi=fp.map(p=>periods.indexOf(p));let totalValues=wi.map(i=>warehouses.reduce((a,w)=>a+(w.series[i]?.value||0),0));let totalWeights=wi.map(i=>warehouses.reduce((a,w)=>a+(w.series[i]?.weight||0),0));el.innerHTML=`<div style="display:flex;gap:8px;margin:4px 0 8px;flex-wrap:wrap"><span style="font-size:12px;color:#557086;font-weight:700">Hisobot davri:</span>${periodBtns('renderWarehouseChart',m)}</div><div class="panel"><h3>Barcha omborlar bo'yicha yuk oqimi (${fp[0]||''} — ${fp[fp.length-1]||''})</h3>${svgLineChart(fp,[{label:"Qiymat (ming $)",values:totalValues},{label:"Vazn (tn)",values:totalWeights}])}</div>`}
function renderTransportChart(m){TRANSPORT_PERIOD_M=m;let el=$('trChartSect');if(!el||!TRANSPORT_TRENDS?.periods?.length)return;let periods=TRANSPORT_TRENDS.periods,transports=TRANSPORT_TRENDS.transports;let fp=filterPeriodArr(periods,m);let wi=fp.map(p=>periods.indexOf(p));let colors=["#1d72b8","#16a34a","#f59e0b","#dc2626","#7c3aed","#0891b2","#ea580c"];let allSeries=transports.map((t,i)=>({label:t.transport,values:wi.map(idx=>t.series[idx]?.value||0),color:colors[i%colors.length]}));el.innerHTML=`<div style="display:flex;gap:8px;margin:4px 0 8px;flex-wrap:wrap"><span style="font-size:12px;color:#557086;font-weight:700">Hisobot davri:</span>${periodBtns('renderTransportChart',m)}</div><div class="panel"><h3>Transport turlari bo'yicha qiymat tendensiyasi (${fp[0]||''} — ${fp[fp.length-1]||''})</h3>${svgLineChart(fp,allSeries)}</div>`}
async function loadWarehouseTrends(){let el=$("warehouseTrendPanel");if(!el)return;try{let j=await api("/api/warehouse_trends");WAREHOUSE_TRENDS=j;if(j.error&&!j.periods?.length){el.innerHTML=`<h3>Omborlar tendensiyasi</h3><div class="muted">Server xatoligi: ${esc(j.error)}</div>`;return}let periods=j.periods||[],warehouses=j.warehouses||[];if(!periods.length){el.innerHTML=`<h3>Omborlar bo'yicha yuk oqimi tendensiyasi</h3><div class="muted">Hali tarixiy ma'lumot yetarli emas. Har bir yangi yuklangan fayl bilan tendensiya shakllanadi.</div>`;return}let growing=warehouses.filter(w=>{let s=w.series;if(s.length<2)return false;return s[s.length-1].value>s[s.length-2].value&&s[s.length-2].value>0});let shrinking=warehouses.filter(w=>{let s=w.series;if(s.length<2)return false;return s[s.length-1].value<s[s.length-2].value&&s[s.length-1].value>0});let whRows=warehouses.map(w=>{let s=w.series;let last=s[s.length-1]||{},prev=s.length>1?s[s.length-2]:{};let delta=(last.value||0)-(prev.value||0);let trend=delta>0?"📈 o'sish":delta<0?"📉 pasayish":"➡ barqaror";return {warehouse:w.warehouse,last_value:last.value||0,last_weight:last.weight||0,last_partiya:last.partiya||0,trend}});el.innerHTML=`<h3>Omborlar bo'yicha yuk oqimi tendensiyasi</h3><div class="overview-note">${periodRangeCaption(periods)}. Har qaysi ombor qatoriga bosish orqali alohida tendensiya grafiği ochiladi.</div><div class="kpis"><div class="kpi"><span>Jami omborlar</span><b>${fmtI(warehouses.length)}</b></div><div class="kpi"><span>So'nggi davrda faol</span><b>${fmtI(warehouses.filter(w=>w.series[w.series.length-1]?.value>0).length)}</b></div><div class="kpi"><span>Yuk oqimi o'sgan</span><b style="color:#16a34a">${fmtI(growing.length)}</b></div><div class="kpi"><span>Yuk oqimi kamaygan</span><b style="color:#dc2626">${fmtI(shrinking.length)}</b></div></div><div id="whChartSect"></div><div class="panel wide"><h3>Omborlar kesimida so'nggi holat</h3><table class="fixed-table"><colgroup><col style="width:42%"><col style="width:17%"><col style="width:15%"><col style="width:12%"><col style="width:14%"></colgroup><thead><tr><th class="text">Ombor</th><th class="num">Qiymat (ming $)</th><th class="num">Vazn (tn)</th><th class="num">Partiya</th><th class="text">Tendensiya</th></tr></thead><tbody>${whRows.map(r=>`<tr onclick="showWarehouseTrend(${JSON.stringify(r.warehouse)})" style="cursor:pointer"><td class="text">${esc(r.warehouse)}</td><td class="num">${fmtN(r.last_value)}</td><td class="num">${fmtN(r.last_weight)}</td><td class="num">${fmtI(r.last_partiya)}</td><td class="text">${r.trend}</td></tr>`).join("")}</tbody></table></div>`;renderWarehouseChart(WAREHOUSE_PERIOD_M)}catch(e){let el2=$("warehouseTrendPanel");if(el2)el2.innerHTML=`<h3>Omborlar tendensiyasi</h3><div class="muted">Ma'lumotni yuklab bo'lmadi: ${esc(e.message||e)}</div>`}}
function showWarehouseTrend(whName){let w=(WAREHOUSE_TRENDS.warehouses||[]).find(x=>x.warehouse===whName);if(!w)return;let periods=WAREHOUSE_TRENDS.periods||[];let values=w.series.map(s=>s.value),weights=w.series.map(s=>s.weight);dlgTitle.textContent=`${whName} — yuk oqimi tendensiyasi`;let rows=w.series.map((s,i)=>({date:periods[i],value:s.value,weight:s.weight,partiya:s.partiya,key:{}}));dlgBody.innerHTML=`<div class="overview-note">${periodRangeCaption(periods)}.</div>`+svgLineChart(periods,[{label:"Qiymat (ming $)",values},{label:"Vazn (tn)",values:weights}])+table([{k:"date",t:"Davr",w:"110px"},{k:"value",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"weight",t:"Vazn (tn)",n:1,f:fmtN},{k:"partiya",t:"Partiya",n:1,f:fmtI}],rows,"fixed-table");dlg.showModal()}
async function loadTransportTrends(){let el=$("transportTrendPanel");if(!el)return;try{let j=await api("/api/transport_trends");TRANSPORT_TRENDS=j;if(j.error&&!j.periods?.length){el.innerHTML=`<h3>Transport turi tendensiyasi</h3><div class="muted">Server xatoligi: ${esc(j.error)}</div>`;return}let periods=j.periods||[],transports=j.transports||[];if(!periods.length){el.innerHTML=`<h3>Transport turi bo'yicha davriy tendensiya</h3><div class="muted">Hali tarixiy ma'lumot yetarli emas. Har bir yangi yuklangan fayl bilan tendensiya shakllanadi.</div>`;return}let colors=["#1d72b8","#16a34a","#f59e0b","#dc2626","#7c3aed","#0891b2","#ea580c"];el.innerHTML=`<h3>Transport turi bo'yicha davriy tendensiya</h3><div class="overview-note">${periodRangeCaption(periods)}. Har qaysi transport turi bo'yicha qiymat (ming $) tendensiyasi.</div><div class="kpis">${transports.slice(0,4).map(t=>{let last=t.series[t.series.length-1]||{};return `<div class="kpi"><span>${esc(t.transport)}</span><b>${fmtN(last.value||0)}</b></div>`}).join("")}</div><div id="trChartSect"></div><div class="grid2">${transports.map((t,i)=>{let values=t.series.map(s=>s.value),weights=t.series.map(s=>s.weight);return `<div class="panel"><h3 onclick="showTransportTrend(${JSON.stringify(t.transport)})" style="cursor:pointer;color:${colors[i%colors.length]}">${esc(t.transport)}</h3>${svgLineChart(periods,[{label:"Qiymat",values,color:colors[i%colors.length]},{label:"Vazn (tn)",values:weights}])}</div>`}).join("")}</div>`;renderTransportChart(TRANSPORT_PERIOD_M)}catch(e){let el2=$("transportTrendPanel");if(el2)el2.innerHTML=`<h3>Transport turi tendensiyasi</h3><div class="muted">Ma'lumotni yuklab bo'lmadi: ${esc(e.message||e)}</div>`}}
function showTransportTrend(trName){let t=(TRANSPORT_TRENDS.transports||[]).find(x=>x.transport===trName);if(!t)return;let periods=TRANSPORT_TRENDS.periods||[];let values=t.series.map(s=>s.value),weights=t.series.map(s=>s.weight);dlgTitle.textContent=`${trName} — davriy tendensiya`;let rows=t.series.map((s,i)=>({date:periods[i],value:s.value,weight:s.weight,partiya:s.partiya,key:{}}));dlgBody.innerHTML=`<div class="overview-note">${periodRangeCaption(periods)}.</div>`+svgLineChart(periods,[{label:"Qiymat (ming $)",values},{label:"Vazn (tn)",values:weights}])+table([{k:"date",t:"Davr",w:"110px"},{k:"value",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"weight",t:"Vazn (tn)",n:1,f:fmtN},{k:"partiya",t:"Partiya",n:1,f:fmtI}],rows,"fixed-table");dlg.showModal()}
let TRANSPORT_CO_PERIOD_M=3;
function renderTransportCoTable(m){TRANSPORT_CO_PERIOD_M=m;let el=$('trCoSect');if(!el||!TRANSPORT_CO_TRENDS?.periods?.length)return;let periods=TRANSPORT_CO_TRENDS.periods,transports=TRANSPORT_CO_TRENDS.transports;let fp=filterPeriodArr(periods,m);let wi=fp.map(p=>periods.indexOf(p));let colors=["#1d72b8","#16a34a","#f59e0b","#dc2626","#7c3aed","#0891b2","#ea580c"];el.innerHTML=`<div style="display:flex;gap:8px;margin:4px 0 8px;flex-wrap:wrap"><span style="font-size:12px;color:#557086;font-weight:700">Hisobot davri:</span>${periodBtns('renderTransportCoTable',m)}</div>${transports.map((t,ti)=>{let cols=[{k:"company",t:"Korxona",w:"45%"},{k:"stir",t:"STIR",w:"90px"},{k:"period_value",t:"Davr qiymat (ming $)",n:1,f:fmtN},{k:"last_value",t:"So'nggi davr (ming $)",n:1,f:fmtN},{k:"trend",t:"Tendensiya",w:"80px"}];let rows=t.companies.map(c=>{let pVal=wi.reduce((a,i)=>a+(c.series[i]?.value||0),0);let last=c.series[c.series.length-1]||{},prev=c.series.length>1?c.series[c.series.length-2]:{};let delta=(last.value||0)-(prev.value||0);return {key:{stir:c.stir},company:c.company||c.stir,stir:c.stir,period_value:pVal,last_value:last.value||0,trend:delta>0?"📈":delta<0?"📉":"➡"}}).filter(r=>r.period_value>0).sort((a,b)=>b.period_value-a.period_value).slice(0,15);let totalV=rows.reduce((a,r)=>a+r.period_value,0);return rows.length?`<div class="panel wide"><h3 style="color:${colors[ti%colors.length]}">${esc(t.transport)} — TOP korxonalar (tanlangan davr: ${fmtN(totalV)} ming $)</h3>${table(cols,rows,"fixed-table")}</div>`:''}).join('')}`}
async function loadTransportCompanyTrends(){let el=$("transportCompanyPanel");if(!el)return;try{let j=await api("/api/transport_company_trends");TRANSPORT_CO_TRENDS=j;if(j.error&&!j.periods?.length){el.innerHTML=`<h3>Transport kesimida korxonalar</h3><div class="muted">Server xatoligi: ${esc(j.error)}</div>`;return}let periods=j.periods||[],transports=j.transports||[];if(!periods.length){el.innerHTML=`<h3>Transport kesimida korxonalar</h3><div class="muted">Hali tarixiy ma'lumot yetarli emas.</div>`;return}el.innerHTML=`<h3>Transport turlari kesimida korxonalar</h3><div class="overview-note">${periodRangeCaption(periods)}. Tanlangan davrda har bir transport turi bo'yicha eng faol korxonalar.</div><div id="trCoSect"></div>`;renderTransportCoTable(TRANSPORT_CO_PERIOD_M)}catch(e){let el2=$("transportCompanyPanel");if(el2)el2.innerHTML=`<h3>Transport kesimida korxonalar</h3><div class="muted">Ma'lumotni yuklab bo'lmadi: ${esc(e.message||e)}</div>`}}
function topNWithOther(rows,n,sortKey,labelKey){
  rows=by(rows||[],sortKey);let head=rows.slice(0,n), tail=rows.slice(n);
  if(tail.length){let other={key:{},name:"Boshqa tovarlar",korxona:"Boshqa tovarlar"};other[labelKey||"name"]="Boshqa tovarlar";["partiya","korxona","vazn","qiymat","tolov"].forEach(k=>other[k]=tail.reduce((a,r)=>a+(+r[k]||0),0));head.push(other)}
  return head;
}
function countryRows(){let rows=DATA.countries||DATA.country||[];return rows}
function chartBlock(title,rows,label,value,fmt){return `<div class=panel><h2>${title}</h2>${bars(rows,label,value,fmt)}${miniChart(rows,value,label)}</div>`}
function overviewPanels(){let topCompanies=by(DATA.top_value||[],"qiymat").slice(0,30);return `<div class="grid2">${executiveSummary()}<div class=panel><h2>70-74-80: postlar kesimida</h2>${table([{k:"post",t:"Post",w:"42%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"Kutilayotgan to'lov (mln so'm)",n:1,f:fmtN}],basicTotal(DATA.all_post_summary||[],"IBK bo'yicha Jami","post"),"fixed-table")}<div class=overview-note>Post qatorlari ustiga bosilganda asos deklaratsiyalar ochiladi.</div></div><div class="panel wide"><h2>Jami muddati o'tgan: postlar va rejimlar kesimida</h2>${expiredTotalExcelTable()}</div><div class="panel wide"><h2>Qiymat bo'yicha TOP 30 korxona</h2>${bars(topCompanies,"korxona","qiymat",fmtN)}</div><div id="_cfmSlot"></div></div>`}
function releaseAnalytics(rows){rows=rows||[];let byCompany={};rows.forEach(r=>{let k=r.stir||r.korxona||r.company||"";if(!byCompany[k])byCompany[k]={korxona:r.korxona||r.company||"",stir:r.stir||"",partiya:0,qiymat:0,vazn:0,count:0};byCompany[k].partiya+=+r.released_partiya||0;byCompany[k].qiymat+=+r.released_qiymat||0;byCompany[k].vazn+=+r.released_vazn||0;byCompany[k].count++});let list=Object.values(byCompany);let cols=[{k:"korxona",t:"Korxona nomi",w:"42%"},{k:"stir",t:"STIR",w:"12%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN}];return `<div class=grid2><div class=panel><h2>Eng ko'p nazoratdan yechgan korxonalar</h2>${table(cols,basicTotal(by(list,"qiymat").slice(0,20),"IBK bo'yicha Jami","korxona"),"fixed-table")}</div><div class=panel><h2>Eng tez nazoratdan yechadigan korxonalar</h2><div class=muted>To'liq tezlik tahlili bir nechta sanalar arxivga yuklangandan keyin avtomatik shakllanadi.</div>${table(cols,by(list,"partiya").slice(0,10),"fixed-table")}</div><div class=panel><h2>Eng sekin nazoratdan yechadigan korxonalar</h2><div class=muted>Yakuniy davrda qoldiq ko'p qolgan korxonalar alohida hisoblanadi.</div>${table(cols,by(list,"vazn").slice(-10).reverse(),"fixed-table")}</div></div>`}
async function buildRelease(){let base=$("relBase").value, final=$("relFinal").value, j=await api(`/api/release?base=${encodeURIComponent(base)}&final=${encodeURIComponent(final)}`);if(j.missing){$("releaseResult").innerHTML=`<div class=muted>Arxivda yo'q sana: ${j.missing.join(", ")}. Shu davr uchun asos fayl yuklash kerak.</div>`;return}let raw=(j.rows||[]).map(r=>Object.assign({korxona:r.korxona||r.company||"",partiya:r.released_partiya,qiymat:r.released_qiymat,vazn:r.released_vazn},r));let rows=numbered([Object.assign({korxona:"IBK bo'yicha Jami",stir:"",decl:"",regime:"",partiya:(j.total||{}).released_partiya,qiymat:(j.total||{}).released_qiymat,vazn:(j.total||{}).released_vazn},j.total||{})].concat(raw));let cols=[{k:"rn",t:"T/r",n:1,f:v=>v,w:"38px"},{k:"korxona",t:"Korxona nomi",w:"280px"},{k:"stir",t:"STIR",w:"90px"},{k:"decl",t:"Deklaratsiya",w:"145px"},{k:"regime",t:"Rejim",w:"70px"},{k:"partiya",t:"Partiya",n:1,f:fmtI,w:"80px"},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN,w:"120px"},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN,w:"110px"}];$("releaseResult").innerHTML=`<h2>${base} - ${final} <a class="btn light" href="/api/export?kind=release&base=${base}&final=${final}&token=${TOKEN}">Excel</a></h2><div class=overview-note>Boshlang'ich davr: ${base}. Yakuniy davr: ${final}.</div>${table(cols,rows,"release-table fixed-table")}${releaseAnalytics(raw)}${miniChart(raw,"qiymat","korxona")}`}
function startIdleTimer(){let t;function reset(){clearTimeout(t);t=setTimeout(()=>{logout();alert("20 daqiqa foydalanilmagani uchun profil avtomatik yopildi.")},1200000)}["click","keydown","mousemove","scroll","touchstart"].forEach(e=>document.addEventListener(e,reset,{passive:true}));reset()}startIdleTimer();
const oldRender=render;render=function(){oldRender();let a=$("actions");if(a){a.innerHTML=a.innerHTML.replaceAll("Jamlanma Excel","Ombor ma'lumot").replaceAll("Nazoratdan yechilgan","Nazoratdan yechish");a.insertAdjacentHTML("afterbegin",`<button class="light lang-btn" onclick="setLang('uz')">O'zb</button><button class="light lang-btn" onclick="setLang('uzc')">&#1038;&#1079;&#1073;</button><button class="light lang-btn" onclick="setLang('ru')">&#1056;&#1091;&#1089;</button><button class="light lang-btn" onclick="document.body.classList.toggle('dark')">&#9680;</button> `)}}
function parseUzDate(s){let m=String(s||"").match(/(\d{2})\.(\d{2})\.(\d{4})/);return m?new Date(+m[3],+m[2]-1,+m[1]).getTime():0}
topNWithOther=function(rows,n,sortKey,labelKey){rows=by(rows||[],sortKey);let head=rows.slice(0,n),tail=rows.slice(n);if(tail.length){let other={key:{},name:"Boshqalar",korxona:"Boshqalar"};other[labelKey||"name"]="Boshqalar";["partiya","vazn","qiymat","tolov"].forEach(k=>other[k]=tail.reduce((a,r)=>a+(+r[k]||0),0));head.push(other)}return head}
countryRows=function(){return topNWithOther(DATA.countries||DATA.country||[],30,"qiymat","name")}
bars=function(rows,label,value,fmt){rows=(rows||[]).filter(r=>(+r[value]||0)>0);let sum=rows.reduce((a,r)=>a+(+r[value]||0),0)||1;return `<div class=bars>${rows.map((r,i)=>{let pct=(+r[value]||0)/sum*100;return `<div class=barrow style="animation-delay:${i*.03}s"><div title="${esc(r[label])}">${esc(r[label])}</div><div class=bar><span style="width:${Math.max(6,pct)}%">${pct.toFixed(1)}%</span></div><b>${fmt(r[value])}</b></div>`}).join("")}</div>`}
miniChart=function(rows,k,label){rows=by(rows||[],k).slice(0,8);let sum=rows.reduce((a,r)=>a+(+r[k]||0),0)||1, pts=rows.map((r,i)=>[18+i*(260/Math.max(1,rows.length-1)),140-(+r[k]||0)/Math.max(...rows.map(x=>+x[k]||0),1)*104]);let path=pts.map((p,i)=>(i?'L':'M')+p[0].toFixed(1)+' '+p[1].toFixed(1)).join(' '), area=pts.length?`M18 150 ${path.slice(1)} L278 150 Z`:"";let p1=((+rows[0]?.[k]||0)/sum*100).toFixed(0)+"%";return `<div class="viz-card"><div class="donut" style="--p:${p1}" data-label="${p1}"></div><svg class="trend-svg" viewBox="0 0 300 170"><defs><linearGradient id="trendGrad" x1="0" y1="0" x2="0" y2="1"><stop offset="0" stop-color="#4ecdc4"/><stop offset="1" stop-color="#ffffff" stop-opacity="0"/></linearGradient></defs><path class="trend-area" d="${area}"></path><path class="trend-line" d="${path}"></path>${pts.map(p=>`<circle class="trend-dot" cx="${p[0]}" cy="${p[1]}" r="4"></circle>`).join("")}</svg></div>`}
chartBlock=function(title,rows,label,value,fmt){return `<div class=panel><h2>${title}</h2>${miniChart(rows,value,label)}${bars(rows,label,value,fmt)}</div>`}
executiveSummary=function(){if(!DATA)return "";let k=DATA.kpis||{},top=(DATA.top_value||[])[0]||{},dep=(DATA.top_deposit||[])[0]||{},own=(DATA.warehouse||[]).find(r=>(r.name||"")==="O'z ombor")||{},exp=by(DATA.post_summary||[],"partiya")[0]||{};let items=[["Umumiy nazorat",`${fmtI(k.partiya)} partiya, ${fmtN(k.qiymat)} ming $ qiymat.`],["Muddati o'tgan",`${fmtI(k.expired)} partiya. Eng ko'p partiya qayd etilgan post: ${esc(exp.post||'-')}.`],["Eng yirik korxona",`${esc(top.korxona||'-')} - ${fmtN(top.qiymat||0)} ming $.`],["O'z ombor",`${fmtI(own.partiya||0)} partiya, ${fmtN(own.qiymat||0)} ming $.`],["Depozit yetakchisi",`${esc(dep.korxona||'-')} - ${fmtN(dep.depozit||0)} mln so'm.`]];let dateStr=DATA.meta&&DATA.meta.date?DATA.meta.date:"bugun";return `<div class="panel exec-summary"><div class="exec-header"><div class="exec-header-left"><span class="exec-pulse-dot" aria-hidden="true"></span><span class="exec-title">Joriy holat</span></div><div class="exec-badges"><span class="exec-badge exec-badge-blue">${fmtI(k.partiya)} partiya</span><span class="exec-badge exec-badge-red">${fmtI(k.expired)} muddati o'tgan</span><span class="exec-badge exec-badge-green">Yangilangan: ${esc(dateStr)}</span></div></div><div class="summary-grid">${items.map(x=>`<div class="summary-item"><b>${x[0]}</b><span>${x[1]}</span></div>`).join("")}</div></div>`}
uploadPanel=function(){return `<div class=stack><div class=panel><h2>Fayl yuklash</h2><div class=muted>BNRTE, To'lovlar va yil davomida yig'ilgan asos fayllarni shu yerdan yuklaysiz.</div></div><div class=panel><h2>BNRTE jamlanma</h2><form class="upload" id="upload"><div><label>Asos fayl</label><input name="source" type="file" accept=".xls,.xlsx,.html,.htm" required></div><div><label>Depozit fayl</label><input name="deposit" type="file" accept=".xlsx"></div><div></div><button>BNRTE yuklash</button></form></div><div class=panel><h2>Yillik arxivni birdan yuklash</h2><form class="upload" id="bulkUpload"><div><label>Asos fayllar</label><input name="sources" type="file" accept=".xls,.xlsx,.html,.htm" multiple required></div><div><label>Depozit fayl ixtiyoriy</label><input name="deposit" type="file" accept=".xlsx"></div><div></div><button>Hammasini yuklash</button></form><div id=bulkResult class=muted>Fayllar sanasi nomidan olinadi va arxivga qo'shiladi.</div></div><div class=panel><h2>To'lovlar jadvallari</h2><form class="upload" id="tolovUpload"><div><label>To'lov baza fayli</label><input name="source" type="file" accept=".xlsx,.xls" required></div><div class=muted>04.06+07.06.2026 kabi asos fayl yuklanadi.</div><div></div><button>To'lov jadvallarini shakllantirish</button></form><div id="tolovUploadResult" class=muted>Natijada Excel fayllar shakllanadi.</div></div></div>`}
bindUpload=function(){let f=$("upload");if(f)f.onsubmit=async e=>{e.preventDefault();$("status").textContent="BNRTE fayllari yuklanyapti...";let j=await api("/api/reports",{method:"POST",body:new FormData(f)});poll(j.job_id)};let bf=$("bulkUpload");if(bf)bf.onsubmit=async e=>{e.preventDefault();$("status").textContent="Yillik asos fayllar yuklanyapti...";let j=await api("/api/reports_bulk",{method:"POST",body:new FormData(bf)});$("bulkResult").textContent=`${j.count||0} ta fayl navbatga qo'shildi. Arxiv sanalari yangilanadi.`;$("status").textContent="Bulk yuklash boshlandi";await loadArchive();};let tf=$("tolovUpload");if(tf)tf.onsubmit=async e=>{e.preventDefault();$("status").textContent="To'lovlar shakllantirilyapti...";let j=await api("/api/tolov",{method:"POST",body:new FormData(tf)});PAYMENTS=j.payments||[];$("tolovUploadResult").innerHTML=`Tayyor: ${fmtI(PAYMENTS.reduce((a,r)=>a+(+r.rows||0),0))} qator, ${fmtN(PAYMENTS.reduce((a,r)=>a+(+r.sum||0),0))} so'm.`;$("status").textContent="To'lovlar tayyor"}}
buildRelease=async function(){let base=$("relBase").value,final=$("relFinal").value;if(parseUzDate(base)>=parseUzDate(final)){$("releaseResult").innerHTML=`<div class=muted>Boshlang'ich sana yakuniy sanadan oldin bo'lishi kerak.</div>`;return}let j=await api(`/api/release?base=${encodeURIComponent(base)}&final=${encodeURIComponent(final)}`);if(j.missing){$("releaseResult").innerHTML=`<div class=muted>Arxivda yo'q sana: ${j.missing.join(", ")}. Shu davr uchun asos fayl yuklash kerak.</div>`;return}let raw=(j.rows||[]).map(r=>Object.assign({korxona:r.korxona||r.company||"",partiya:r.released_partiya,qiymat:r.released_qiymat,vazn:r.released_vazn},r));let rows=numbered([Object.assign({korxona:"IBK bo'yicha Jami",stir:"",decl:"",regime:"",partiya:(j.total||{}).released_partiya,qiymat:(j.total||{}).released_qiymat,vazn:(j.total||{}).released_vazn},j.total||{})].concat(raw));let cols=[{k:"rn",t:"T/r",n:1,f:v=>v,w:"38px"},{k:"korxona",t:"Korxona nomi",w:"280px"},{k:"stir",t:"STIR",w:"90px"},{k:"decl",t:"Deklaratsiya",w:"145px"},{k:"regime",t:"Rejim",w:"70px"},{k:"partiya",t:"Partiya",n:1,f:fmtI,w:"80px"},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN,w:"120px"},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN,w:"110px"}];$("releaseResult").innerHTML=`<h2>${base} - ${final} <a class="btn light" href="/api/export?kind=release&base=${base}&final=${final}&token=${TOKEN}">Excel</a></h2>${table(cols,rows,"release-table fixed-table")}${releaseAnalytics(raw)}${miniChart(raw,"qiymat","korxona")}`}
detail=async function(key){if(!DATA||!key||Object.keys(key).length===0)return;if(key.view==="expired_inline"){let el=$("expiredInline");if(el)el.innerHTML=expiredTotalExcelTable();return}if(key.view==="regime_posts"){let rows=(DATA.regime_posts||{})[key.regime]||[];dlgTitle.textContent=`${key.regime} - postlar kesimida`;dlgBody.innerHTML=table([{k:"post",t:"Post",w:"42%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN}],basicTotal(rows,"IBK bo'yicha Jami","post"));dlg.showModal();return}let filterText=JSON.stringify(key),q=new URLSearchParams({report:DATA.id,filters:filterText}),j=await api("/api/details?"+q);dlgTitle.textContent="Asos deklaratsiyalar";dlgBody.innerHTML=`<p><a class="btn light" href="/api/export_details?report=${DATA.id}&filters=${encodeURIComponent(filterText)}&token=${TOKEN}">Excelga yuklash</a></p>`+table([{k:"decl",t:"Deklaratsiya"},{k:"date",t:"Sana"},{k:"regime",t:"Rejim"},{k:"stir",t:"STIR"},{k:"company",t:"Korxona"},{k:"hs",t:"TIF TN"},{k:"goods",t:"Tovar"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN}],j.rows,"fixed-table details-wide");dlg.showModal()}
const renderFinal=render;render=function(){renderFinal();let tabs=$("tabs");if(tabs){let buttons=[...tabs.querySelectorAll("button")], rej=buttons.find(b=>b.textContent.trim()==="Rejimlar"), exp=buttons.find(b=>b.textContent.includes("Muddati"));if(rej&&exp&&exp.nextSibling!==rej)exp.after(rej)}let a=$("actions");if(a&&DATA){let f=DATA.files||{};if((f.pngs||[]).length>1&&!a.innerHTML.includes("_pngs"))a.insertAdjacentHTML("afterbegin",`<a class=btn href="/download/${DATA.id}/_pngs?token=${TOKEN}">Barcha PNG</a> `)}}
const apiRaw=api;api=async function(url,opt={}){opt.headers=Object.assign({"X-Token":TOKEN},opt.headers||{});let r=await fetch(url,opt);let j=await r.json().catch(()=>({error:`HTTP ${r.status}`}));if(r.status===401){showLogin();throw Error("login")}if(!r.ok||j.error)throw Error(j.error||`HTTP ${r.status}`);return j}
function countryPoint(name,i,total){let s=String(name||"").toLowerCase();let map=[["xitoy",70,126],["china",70,126],["rossiya",210,70],["russia",210,70],["turkiya",165,155],["turkey",165,155],["koreya",105,95],["germaniya",135,115],["germany",135,115],["italiya",125,148],["fransiya",118,130],["aqsh",48,88],["usa",48,88],["hindiston",98,182],["india",98,182],["qozog",238,116],["kazakh",238,116],["baa",146,205],["amirlik",146,205],["eron",170,185],["iran",170,185]];for(let m of map){if(s.includes(m[0]))return {x:m[1],y:m[2]}}let angle=(-120+(i/(Math.max(total-1,1))*240))*Math.PI/180;return {x:155+112*Math.cos(angle),y:150+82*Math.sin(angle)}}
const COUNTRY_COORDS=[["xitoy",35.86,104.19],["china",35.86,104.19],["rossiya",61.52,105.32],["russia",61.52,105.32],["turkiya",38.96,35.24],["turkey",38.96,35.24],["koreya",36.5,127.8],["germaniya",51.17,10.45],["germany",51.17,10.45],["italiya",41.87,12.57],["italy",41.87,12.57],["fransiya",46.23,2.21],["france",46.23,2.21],["aqsh",37.09,-95.71],["usa",37.09,-95.71],["amerika",37.09,-95.71],["hindiston",20.59,78.96],["india",20.59,78.96],["qozog",48.02,66.92],["kazakh",48.02,66.92],["baa",23.42,53.85],["amirlik",23.42,53.85],["eron",32.43,53.69],["iran",32.43,53.69],["polsha",51.92,19.15],["ispaniya",40.46,-3.75],["yaponiya",36.2,138.25],["japan",36.2,138.25],["belarus",53.7,27.95],["latviya",56.88,24.6],["litva",55.17,23.88],["ukraina",48.38,31.17],["misr",26.82,30.8],["egypt",26.82,30.8],["pokiston",30.38,69.35],["pakistan",30.38,69.35],["vietnam",14.06,108.28],["malayziya",4.21,101.98],["tailand",15.87,100.99],["indoneziya",-0.79,113.92],["braziliya",-14.24,-51.93],["meksika",23.63,-102.55],["saudiya",23.89,45.08],["gruziya",41.69,44.03],["georgia",41.69,44.03],["ozarbayjon",40.14,47.58],["azerbaijan",40.14,47.58],["armaniston",40.07,45.04],["armenia",40.07,45.04],["qirg'iziston",41.2,74.77],["kyrgyz",41.2,74.77],["tojikiston",38.86,71.28],["tajik",38.86,71.28],["turkmaniston",38.97,59.56],["turkmen",38.97,59.56],["afg'oniston",33.93,67.71],["afghan",33.93,67.71],["niderlandiya",52.13,5.29],["netherlands",52.13,5.29],["belgiya",50.50,4.47],["avstriya",47.52,14.55],["shveytsariya",46.82,8.23],["chexiya",49.82,15.47],["vengriya",47.16,19.50],["ruminiya",45.94,24.97],["bolgariya",42.73,25.49],["serbiya",44.02,21.01],["xorvatiya",45.10,15.20],["slovakiya",48.67,19.70],["sloveniya",46.15,14.99],["estoniya",58.60,25.01],["finlyandiya",61.92,25.75],["shvetsiya",60.13,18.64],["norvegiya",60.47,8.47],["daniya",56.26,9.50],["gollandiya",52.13,5.29],["avstraly",-25.27,133.78],["australia",-25.27,133.78],["kanada",56.13,-106.35],["canada",56.13,-106.35],["isroil",31.05,34.85],["israel",31.05,34.85],["iordaniya",30.59,36.24],["livan",33.85,35.86],["falastin",31.95,35.23],["suriya",34.80,38.99],["iroq",33.22,43.68],["iraq",33.22,43.68],["quvayt",29.31,47.48],["bahrayn",26.07,50.55],["qatar",25.35,51.18],["ummon",21.51,55.92],["yaman",15.55,48.52],["efiopiya",9.14,40.49],["marokash",31.79,-7.09],["tunis",33.89,9.54],["jazoir",28.03,1.66],["nigeria",9.08,8.68],["gana",7.95,-1.02],["keniya",-0.02,37.91],["tanzaniya",-6.37,34.89],["britaniya",55.38,-3.44],["buyuk brit",55.38,-3.44],["united king",55.38,-3.44],["england",51.50,-0.12],["scotland",56.49,-4.20],["wales",52.13,-3.78],["ireland",53.41,-8.24],["irlandiya",53.41,-8.24],["irlandia",53.41,-8.24],["portugal",39.40,-8.22],["portugaliya",39.40,-8.22],["gretsiya",37.98,23.73],["greece",37.98,23.73],["portugal",39.40,-8.22],["singapur",1.35,103.82],["singapore",1.35,103.82],["gonkong",22.32,114.17],["hong kong",22.32,114.17],["tayvan",23.70,121.00],["taiwan",23.70,121.00],["nz",40.90,174.89],["yangi zelend",40.90,174.89],["peru",-9.19,-75.02],["chili",-35.68,-71.54],["argentina",-38.42,-63.62],["kolumbiya",4.57,-74.30],["venetsuela",6.42,-66.59]];
function countryLatLon(name){let s=String(name||"").toLowerCase();for(let c of COUNTRY_COORDS){if(s.includes(c[0]))return {lat:c[1],lon:c[2]}}return null}
const COUNTRY_FLAGS={"xitoy":"🇨🇳","china":"🇨🇳","rossiya":"🇷🇺","russia":"🇷🇺","turkiya":"🇹🇷","turkey":"🇹🇷","koreya":"🇰🇷","germaniya":"🇩🇪","germany":"🇩🇪","italiya":"🇮🇹","fransiya":"🇫🇷","aqsh":"🇺🇸","usa":"🇺🇸","amerika":"🇺🇸","hindiston":"🇮🇳","india":"🇮🇳","qozog":"🇰🇿","kazakh":"🇰🇿","baa":"🇦🇪","amirlik":"🇦🇪","eron":"🇮🇷","iran":"🇮🇷","polsha":"🇵🇱","ispaniya":"🇪🇸","yaponiya":"🇯🇵","japan":"🇯🇵","belarus":"🇧🇾","latviya":"🇱🇻","litva":"🇱🇹","ukraina":"🇺🇦","misr":"🇪🇬","egypt":"🇪🇬","pokiston":"🇵🇰","pakistan":"🇵🇰","vietnam":"🇻🇳","malayziya":"🇲🇾","tailand":"🇹🇭","indoneziya":"🇮🇩","braziliya":"🇧🇷","meksika":"🇲🇽","saudiya":"🇸🇦","gretsiya":"🇬🇷","gruziya":"🇬🇪","britaniya":"🇬🇧"};
function countryFlag(name){let s=String(name||"").toLowerCase();for(let k in COUNTRY_FLAGS){if(s.includes(k))return COUNTRY_FLAGS[k]}return "🌐"}
const COUNTRY_LATIN={"ўзбекистон":"O'zbekiston","австралия":"Avstraliya","австрия":"Avstriya","ангилъя":"Angilya","аргентина":"Argentina","арманистон":"Armaniston","афғонистон":"Afg'oniston","ақш":"AQSh","баа":"BAA","бангладеш":"Bangladesh","бахрейн":"Bahreyn","белгия":"Belgiya","белоруссия":"Belorussiya","болгария":"Bolgariya","ботсвана":"Botsvana","бразилия":"Braziliya","британия территорияси":"Britaniya territoriyasi","буюк британия":"Buyuk Britaniya","вануату":"Vanuatu","венгрия":"Vengriya","въетнам":"Vyetnam","германия":"Germaniya","гонконг":"Gonkong","греция":"Gretsiya","грузия":"Gruziya","дания":"Daniya","жар":"JAR","индонезия":"Indoneziya","иордания":"Iordaniya","ирландия":"Irlandiya","исландия":"Islandiya","испания":"Ispaniya","исроил":"Isroil","италия":"Italiya","канада":"Kanada","кипр":"Kipr","корея (кхдр)":"Koreya (KXDR)","корея республикаси":"Koreya Respublikasi","куба":"Kuba","латвия":"Latviya","ливан":"Livan","литва":"Litva","лихтенштейн":"Lixtenshteyn","люксембург":"Lyuksemburg","мар":"MAR","маврикий":"Mavrikiy","мавритания":"Mavritaniya","македония":"Makedoniya","малайзия":"Malayziya","малдива":"Maldiva","малъта":"Malta","маршал ороллари":"Marshal orollari","мексика":"Meksika","миср":"Misr","молдова":"Moldova","монголия":"Mongoliya","нигерия":"Nigeriya","нидерландия":"Niderlandiya","озарбайжон":"Ozarbayjon","покистон":"Pokiston","полша":"Polsha","португалия":"Portugaliya","россия":"Rossiya","руминия":"Ruminiya","сан-марино":"San-Marino","саудия арабистони":"Saudiya Arabistoni","сейшел ороллари":"Seyshel orollari","сербия":"Serbiya","сингапур":"Singapur","сирия":"Siriya","словакия":"Slovakiya","словения":"Sloveniya","тайван":"Tayvan","тайланд":"Tayland","тожикистон":"Tojikiston","тунис":"Tunis","туркия":"Turkiya","туркманистон":"Turkmaniston","украина":"Ukraina","уругвай":"Urugvay","финландия":"Finlandiya","франция":"Fransiya","хитой":"Xitoy","хорватия":"Xorvatiya","черногория":"Chernogoriya","чеҳия":"Chexiya","чили":"Chili","швейцария":"Shveytsariya","швеция":"Shvetsiya","эквадор":"Ekvador","эрон":"Eron","эстония":"Estoniya","янги зеландия":"Yangi Zelandiya","япония":"Yaponiya","қатар":"Qatar","қирғизистон":"Qirg'iziston","қозоғистон":"Qozog'iston","ҳиндистон":"Hindiston","中国":"Xitoy","日本":"Yaponiya","한국":"Koreya","대한민국":"Koreya Respublikasi","الصين":"Xitoy","الهند":"Hindiston","روسيا":"Rossiya","ألمانيا":"Germaniya","فرنسا":"Fransiya","إيطاليا":"Italiya","البرازيل":"Braziliya","كندا":"Kanada","المملكة المتحدة":"Buyuk Britaniya","أمريكا":"AQSh","الولايات المتحدة":"AQSh","تركيا":"Turkiya","إيران":"Eron","باكستان":"Pokiston","بنغلاديش":"Bangladesh","إندونيسيا":"Indoneziya"};
function countryLatinName(name){let s=String(name||"").toLowerCase().trim();if(COUNTRY_LATIN[s])return COUNTRY_LATIN[s];for(let k in COUNTRY_LATIN){if(s===k.toLowerCase()||s.includes(k)||k.includes(s))return COUNTRY_LATIN[k]}return name}
let _CFM_ROWS=[];let COUNTRY_FLOW_MAP=null;let _LEAFLET_LOADED=false;let _LEAFLET_CBS=[];let COUNTRY_TRANSPORT_DATA={};let CFM_VEHICLE_ANIMS=[];
function stopCFMAnimations(){CFM_VEHICLE_ANIMS.forEach(a=>{a.cancelled=true;try{if(a.marker)a.marker.remove()}catch(e){}});CFM_VEHICLE_ANIMS=[]}
function cfmVehicleIcon(transport,size,bearing){let emoji=transport==='Avia'?'✈':transport==="Temir yo'l"?'🚂':'🚛';let s=Math.round(size);let rot=bearing!=null?`transform:rotate(${bearing.toFixed(1)}deg);transform-origin:center;`:'';return L.divIcon({html:`<div style="font-size:${s}px;line-height:1;user-select:none;pointer-events:none;display:inline-block;${rot}">${emoji}</div>`,className:'',iconSize:[s,s],iconAnchor:[s/2,s/2]})}
function animateCFMVehicle(map,fromLL,toLL,transport,size,idx,weight,maxWeight){let anim={cancelled:false,marker:null};let emojiNE=transport==='Avia'?45:0;let bearing=emojiNE-Math.atan2(toLL[0]-fromLL[0],toLL[1]-fromLL[1])*180/Math.PI;let icon=cfmVehicleIcon(transport,size,bearing);try{anim.marker=L.marker(fromLL,{icon,interactive:false}).addTo(map)}catch(e){return}let dlat=toLL[0]-fromLL[0],dlon=toLL[1]-fromLL[1];let dist=Math.sqrt(dlat*dlat+dlon*dlon*0.6);let wFactor=(weight&&maxWeight&&maxWeight>0)?1+(weight/maxWeight)*0.65:1;let duration=Math.max(7000,Math.min(30000,dist*700*wFactor));let offset=(idx*0.17+Math.random()*0.08)%1;let startT=null;function step(ts){if(anim.cancelled)return;if(!startT)startT=ts-(offset*duration);let t=((ts-startT)%duration)/duration;let lat=fromLL[0]+dlat*t;let lon=fromLL[1]+dlon*t;try{anim.marker.setLatLng([lat,lon])}catch(e){anim.cancelled=true;return}requestAnimationFrame(step)}requestAnimationFrame(step);CFM_VEHICLE_ANIMS.push(anim)}
function _loadLeaflet(cb){
  if(typeof L!=='undefined'){cb();return}
  _LEAFLET_CBS.push(cb);
  if(_LEAFLET_LOADED)return;
  _LEAFLET_LOADED=true;
  let lnk=document.createElement('link');lnk.rel='stylesheet';lnk.href='https://unpkg.com/leaflet@1.9.4/dist/leaflet.css';document.head.appendChild(lnk);
  let s=document.createElement('script');
  s.src='https://unpkg.com/leaflet@1.9.4/dist/leaflet.js';
  s.onload=function(){let cbs=_LEAFLET_CBS;_LEAFLET_CBS=[];cbs.forEach(fn=>fn())};
  s.onerror=function(){['cfmMap','flightsMap'].forEach(id=>{let el=document.getElementById(id);if(el)el.innerHTML='<div style="height:100%;display:flex;align-items:center;justify-content:center;color:#888;font-size:13px">Xarita kutubxonasi yuklanmadi</div>'})};
  document.head.appendChild(s);
}
function countryFlowMap(rows){
  return `<div style="border-radius:12px;overflow:hidden;border:1px solid #1e3a5f"><iframe src="/assets/tovarlar_oqimi.html?token=${TOKEN}" style="width:100%;height:580px;border:none;display:block" id="cfmFrame"></iframe></div>`
}
async function initCountryFlowMap(){
  let el=document.getElementById("cfmMap");if(!el)return;
  let rows=_CFM_ROWS;if(!rows.length)return;
  try{COUNTRY_TRANSPORT_DATA=await api("/api/country_transport")}catch(e){COUNTRY_TRANSPORT_DATA={}}
  _loadLeaflet(function(){_buildCFM(el,rows)});
}
function _buildCFM(el,rows){
  try{
    stopCFMAnimations();
    if(COUNTRY_FLOW_MAP){try{COUNTRY_FLOW_MAP.remove()}catch(e){}COUNTRY_FLOW_MAP=null}
    el.innerHTML='';
    let map=L.map(el,{center:[35,55],zoom:2});
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{attribution:'© OpenStreetMap',maxZoom:10}).addTo(map);
    COUNTRY_FLOW_MAP=map;
    let UZ=[41.3,69.3];
    let uzIcon=L.divIcon({html:'<div style="background:#1d72b8;color:#fff;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:900;white-space:nowrap;box-shadow:0 2px 6px rgba(0,0,0,.4);transform:translate(-50%,-50%)">🇺🇿 O\'zbekiston</div>',className:'',iconSize:[0,0]});
    L.marker(UZ,{icon:uzIcon}).addTo(map);
    let maxQ=Math.max(1,...rows.map(r=>+r.qiymat||0));
    let maxV=Math.max(1,...rows.map(r=>+r.vazn||0));
    let allLatLons=[];
    rows.forEach((r,ri)=>{
      let ll=countryLatLon(r.name);if(!ll)return;
      let q=+r.qiymat||0,v=+r.vazn||0,p=+r.partiya||0;
      let qS=q/maxQ;
      let clr=qS>=.6?'#16a34a':qS>=.3?'#d97706':'#1d72b8';
      let coord=[ll.lat,ll.lon];
      allLatLons.push(coord);
      L.polyline([coord,UZ],{color:clr,weight:1+qS*2.5,opacity:.5}).addTo(map);
      let sz=Math.round(Math.max(28,Math.min(50,28+Math.sqrt(v/maxV)*22)));
      let fs=sz>44?12:sz>36?11:10;
      let flag=countryFlag(r.name);
      let latinNm=countryLatinName(r.name);
      let nm=esc(latinNm).slice(0,18);
      let tip=`${flag} ${esc(latinNm)}\nPartiya: ${fmtI(p)}\nQiymat: ${fmtN(q)} ming $\nVazn: ${fmtN(v)} tn`;
      let html=`<div style="transform:translate(-50%,-50%);display:flex;flex-direction:column;align-items:center;gap:2px;cursor:pointer"><div style="width:${sz}px;height:${sz}px;border-radius:50%;background:${clr};border:2.5px solid #fff;display:flex;align-items:center;justify-content:center;font-size:${fs}px;font-weight:800;color:#fff;box-shadow:0 2px 8px rgba(0,0,0,.3);box-sizing:border-box">${fmtI(p)}</div><div style="white-space:nowrap;background:rgba(255,255,255,.93);padding:1px 5px;border-radius:3px;font-size:10px;font-weight:700;color:#1a3a5c;box-shadow:0 1px 3px rgba(0,0,0,.2)">${flag} ${nm}</div></div>`;
      let cIcon=L.divIcon({html,className:'',iconSize:[0,0]});
      L.marker(coord,{icon:cIcon,title:tip}).addTo(map);
      let ctData=COUNTRY_TRANSPORT_DATA[r.name]||null;
      let transport=ctData?ctData.dominant:'Avia';
      let vSize=Math.round(Math.max(14,Math.min(32,14+Math.sqrt(v/maxV)*18)));
      animateCFMVehicle(map,coord,UZ,transport,vSize,ri,v,maxV);
    });
    if(allLatLons.length){
      let lats=allLatLons.map(c=>c[0]).concat([UZ[0]]),lons=allLatLons.map(c=>c[1]).concat([UZ[1]]),pad=4;
      map.fitBounds([[Math.min(...lats)-pad,Math.min(...lons)-pad],[Math.max(...lats)+pad,Math.max(...lons)+pad]]);
    }
    let lgEl=document.createElement('div');
    lgEl.className='cfm-legend';
    lgEl.style.cssText='position:absolute;bottom:12px;right:12px;z-index:1000';
    lgEl.innerHTML='<b>Qiymat ulushi:</b><br><span style="color:#16a34a">● Yuqori (60%+)</span><br><span style="color:#d97706">● O\'rta (30-60%)</span><br><span style="color:#1d72b8">● Quyi (&lt;30%)</span><hr style="margin:4px 0"><b>Transport:</b> ✈ Avia · 🚛 Avto · 🚂 Temir yo\'l';
    el.style.position='relative';
    el.appendChild(lgEl);
  }catch(err){console.error('CFM map error:',err)}
}
const POST_NAMES={"35001":"Nukus aeroporti chegara bojxona posti","35002":"Nukus TIF bojxona posti","35003":"Xo'jayli chegara bojxona posti","35004":"Dovut-ota chegara bojxona posti","35010":"Qoraqalpog'iston temir yo'l chegara bojxona posti","03002":"Do'stlik chegara bojxona posti","03003":"Andijon aeroporti chegara bojxona posti","03005":"Mingtepa chegara bojxona posti","03006":"Qorasuv chegara bojxona posti","03007":"Xonobod chegara bojxona posti","03008":"Pushmon chegara bojxona posti","03009":"Madaniyat chegara bojxona posti","03011":"Andijon TIF bojxona posti","03013":"Keskanyol chegara bojxona posti","03014":"Savay temir yo'l chegara bojxona posti","03015":"Asaka TIF bojxona posti","06001":"Buxoro aeroporti chegara bojxona posti","06006":"Buxoro TIF bojxona posti","06009":"Qorakol TIF bojxona posti","06010":"Olot chegara bojxona posti","06011":"Xo'jadavlat temir yo'l chegara bojxona posti","08003":"Uchturgon chegara bojxona posti","08004":"Jizzax TIF bojxona posti","08007":"Qo'shkent chegara bojxona posti","10002":"Nasaf TIF bojxona posti","10007":"Qamashi-G'uzor TIF bojxona posti","10008":"Qarshi-Kerki chegara bojxona posti","10012":"Qarshi aeroporti chegara bojxona posti","12002":"Navoiy aeroporti chegara bojxona posti","12003":"Navoiy TIF bojxona posti","12008":"Zarafshon TIF bojxona posti","14002":"Namangan aeroporti chegara bojxona posti","14003":"Uchqo'rg'on chegara bojxona posti","14004":"Kosonsoy chegara bojxona posti","14005":"Pop chegara bojxona posti","14010":"Namangan TIF bojxona posti","18001":"Samarqand aeroporti chegara bojxona posti","18002":"Jartepa chegara bojxona posti","18005":"Samarqand TIF bojxona posti","18007":"Ulug'bek TIF bojxona posti","22002":"Termiz aeroporti chegara bojxona posti","22003":"Sariosiyo chegara bojxona posti","22004":"Sariosiyo temir yo'l chegara bojxona posti","22005":"Termiz TIF bojxona posti","22006":"Denov TIF bojxona posti","22007":"Gulbahor chegara bojxona posti","22011":"Daryo porti chegara bojxona posti","22015":"Boldir temir yo'l chegara bojxona posti","22017":"Ayritom chegara bojxona posti","22022":"Termiz xalqaro savdo markazi TIF bojxona posti","24002":"Xovosobod chegara bojxona posti","24004":"Sirdaryo chegara bojxona posti","24006":"Oq oltin chegara bojxona posti","24009":"Guliston TIF bojxona posti","24014":"Malik chegara bojxona posti","27001":"Yallama chegara bojxona posti","27008":"Navoiy chegara bojxona posti","27009":"S. Najimov chegara bojxona posti","27011":"Oybek chegara bojxona posti","27013":"Bekobod avto chegara bojxona posti","27014":"Chirchiq TIF bojxona posti","27015":"Olmaliq TIF bojxona posti","27016":"Yangiyol TIF bojxona posti","27019":"Nazarbek TIF bojxona posti","27020":"Keles TIF bojxona posti","27021":"G'ishtkoprik chegara bojxona posti","27023":"Farhod chegara bojxona posti","27024":"Bekobod temir yo'l chegara bojxona posti","27028":"Angren TIF bojxona posti","30001":"Farg'ona aeroporti chegara bojxona posti","30002":"Qo'qon TIF bojxona posti","30004":"Farg'ona chegara bojxona posti","30005":"Andarxon chegara bojxona posti","30006":"Rishton chegara bojxona posti","30008":"Rovot chegara bojxona posti","30009":"Vodiy TIF bojxona posti","30010":"O'zbekiston chegara bojxona posti","30012":"So'x chegara bojxona posti","33001":"Shovot chegara bojxona posti","33004":"Do'stlik chegara bojxona posti","33007":"Urganch TIF bojxona posti","33011":"Urganch aeroporti chegara bojxona posti","33033":"Shovot chegaraoldi savdo zonasi","26002":"Toshkent-tovar TIF bojxona posti","26003":"Ark buloq TIF bojxona posti","26004":"Chuqursoy TIF bojxona posti","26009":"Keles temir yo'l chegara bojxona posti","26010":"Sirg'ali TIF bojxona posti","26013":"Chuqursoy texnik idora temir yo'l chegara bojxona posti","00101":"Toshkent xalqaro aeroporti CHBP","00102":"Avia yuklar TIF bojxona posti","00107":"Elektron tijorat TIF bojxona posti","00110":"Toshkent-Humo aeroporti CHBP"};
function namedSourcePosts(){return (DATA.source_posts||[]).map(r=>{let nm=r.post_nomi;if(!nm||nm===r.post_kodi||/^\d{5}$/.test(String(nm||""))){nm=POST_NAMES[r.post_kodi]||(r.post_kodi?"Post №"+r.post_kodi:"-")}return Object.assign({},r,{post_nomi:nm})})}
function sourcePostInfographics(){let posts=by(namedSourcePosts(),"qiymat").slice(0,12),trans=DATA.transport||[],tp=trans.reduce((a,r)=>a+(+r.qiymat||0),0)||1,pp=posts.reduce((a,r)=>a+(+r.qiymat||0),0)||1;let rings=trans.map((r,i)=>{let pct=Math.round((+r.qiymat||0)/tp*100),dash=Math.max(0,Math.min(100,pct));return `<div class=transport-ring style="--p:${dash};--delay:${i*.12}s" onclick='detail(${JSON.stringify(r.key||{transport:r.name})})'><b>${esc(r.name||"-")}</b><span>${pct}%</span><small>${fmtI(r.partiya||0)} partiya В· ${fmtN(r.qiymat||0)} ming $</small></div>`}).join("");let barsHtml=posts.map((r,i)=>{let pct=Math.max(2,(+r.qiymat||0)/pp*100);return `<div class=flow-row onclick='detail(${JSON.stringify(r.key||{})})' title="${esc(r.post_nomi||r.post_kodi||"-")}"><div class=flow-name><b>${esc(r.post_kodi||"-")}</b><span>${esc(r.post_nomi||"-")}</span></div><div class=flow-track><i style="width:${pct.toFixed(1)}%;animation-delay:${i*.05}s"></i><em>${pct.toFixed(1)}%</em></div><div class=flow-num>${fmtN(r.qiymat||0)}<span class=flow-unit>ming $</span></div></div>`}).join("");return `<div class="panel wide"><h2>Nazoratga qo'yilgan postlar va transport turlari</h2><div class="overview-note" style="margin-bottom:6px">Barcha qiymatlar ming AQSh dollarida (ming $) berilgan. Post ulushi umumiy qiymat asosida hisoblanadi.</div><div class=transport-viz><div class=ring-grid>${rings}</div><div class=flow-list>${barsHtml}</div></div></div>`}
function transportPanel(){let cols=[{k:"post_kodi",t:"Post kodi",w:"78px"},{k:"post_nomi",t:"Post nomi",w:"260px"},{k:"transport",t:"Transport turi",w:"92px"},{k:"partiya",t:"Partiya",w:"78px",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",w:"92px",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",w:"112px",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",w:"112px",n:1,f:fmtN},{k:"korxona",t:"Korxona",w:"82px",n:1,f:fmtI}];let rows=namedSourcePosts();return `<div class=panel><h2>Deklaratsiya post kodi bo'yicha tahlil</h2>${table(cols,basicTotal(rows,"IBK bo'yicha Jami","post_nomi"),"fixed-table transport-table")}</div>${sourcePostInfographics()}${chartBlock("Transport turi bo'yicha ulushi · Qiymat, ming $",DATA.transport||[],"name","qiymat",fmtN)}`}
const overviewPanelsWithMap=overviewPanels;overviewPanels=function(){let html=overviewPanelsWithMap();let countries=countryRows();let countriesLatin=countries.map(r=>Object.assign({},r,{name:countryLatinName(r.name)}));let countryBlock=`<div class="panel wide"><h2>Davlatlar bo'yicha yo'nalishlar</h2>${countryFlowMap(countries)}<div class="chart-under-globe">${bars(countriesLatin,"name","qiymat",fmtN)}</div></div>${transportPanel()}${flightsPanelShell()}`;return html.replace('<div id="_cfmSlot"></div>',countryBlock)}
function flightsPanelShell(){
  return `<div class="panel wide" id="flightsPanelWrap">
<h2>Toshkent xalqaro aeroporti — jonli parvozlar</h2>
<div style="border-radius:14px;overflow:hidden;border:1px solid var(--line);margin-bottom:8px">
  <iframe src="/assets/tas_parvozlar.html?token=${TOKEN}" style="width:100%;height:580px;border:none;display:block" id="tasFrame"></iframe>
</div>
</div>`}
function initFlightsMap(){}
detail=async function(key){if(!DATA||!key||Object.keys(key).length===0)return;if(key.view==="expired_inline"){let el=$("expiredInline");if(el)el.innerHTML=expiredTotalExcelTable();return}if(key.view==="regime_posts"){let rows=(DATA.regime_posts||{})[key.regime]||[];dlgTitle.textContent=`${key.regime} - postlar kesimida`;dlgBody.innerHTML=table([{k:"post",t:"Post",w:"42%"},{k:"partiya",t:"Partiya",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",n:1,f:fmtN},{k:"tolov",t:"To'lov (mln so'm)",n:1,f:fmtN}],basicTotal(rows,"IBK bo'yicha Jami","post"));dlg.showModal();return}let filterText=JSON.stringify(key),q=new URLSearchParams({report:DATA.id,filters:filterText}),j=await api("/api/details?"+q);dlgTitle.textContent="Asos deklaratsiyalar";dlgBody.innerHTML=`<p><a class="btn light" href="/api/export_details?report=${DATA.id}&filters=${encodeURIComponent(filterText)}&token=${TOKEN}">Excelga yuklash</a></p>`+table([{k:"decl",t:"Deklaratsiya",w:"160px"},{k:"source_post",t:"Boshlang'ich post kodi",w:"90px"},{k:"source_post_name",t:"Boshlang'ich post nomi",w:"220px"},{k:"transport",t:"Transport",w:"92px"},{k:"date",t:"Sana",w:"86px"},{k:"regime",t:"Rejim",w:"64px"},{k:"post",t:"Nazorat posti",w:"170px"},{k:"stir",t:"STIR",w:"105px"},{k:"company",t:"Korxona",w:"220px"},{k:"hs",t:"TIF TN",w:"105px"},{k:"goods",t:"Tovar",w:"240px"},{k:"partiya",t:"Partiya",w:"70px",n:1,f:fmtI},{k:"vazn",t:"Vazn (tn)",w:"90px",n:1,f:fmtN},{k:"qiymat",t:"Qiymat (ming $)",w:"105px",n:1,f:fmtN}],j.rows,"fixed-table details-wide");dlg.showModal()}
function cleanTopActions(){let a=$("actions");if(!a)return;[...a.querySelectorAll("button")].forEach(b=>{if(["Sozlamalar","Nastroyki","Settings"].includes(b.textContent.trim()))b.remove()});[...a.querySelectorAll("a")].forEach(x=>{let t=x.textContent.trim();if(/^PNG\s+\d+$/i.test(t)||t==="Barcha PNG")x.remove()});if(DATA){let f=DATA.files||{};let has=a.querySelector("[data-pngzip]");if(!has&&(f.pngs||[]).length){let href=(f.pngs||[]).length>1?`/download/${DATA.id}/_pngs?token=${TOKEN}`:`/download/${DATA.id}/${f.pngs[0]}?token=${TOKEN}`;a.insertAdjacentHTML("afterbegin",`<a data-pngzip class=btn href="${href}">PNG</a> `)}}}
function translateRuPage(){if(LANG!=="ru")return;let pairs=[["Kirish","Р’С…РѕРґ"],["Chiqish","Р’С‹С…РѕРґ"],["Ombor ma'lumot","РЎРєР»Р°РґСЃРєР°СЏ СЃРІРѕРґРєР°"],["Fayl yuklash","Р—Р°РіСЂСѓР·РєР° С„Р°Р№Р»РѕРІ"],["Umumiy","РћР±С‰РµРµ"],["Korxonalar","РџСЂРµРґРїСЂРёСЏС‚РёСЏ"],["Muddati o'tgan","РџСЂРѕСЃСЂРѕС‡РµРЅРЅС‹Рµ"],["Nazoratdan yechish","РЎРЅСЏС‚РёРµ СЃ РєРѕРЅС‚СЂРѕР»СЏ"],["Tovarlar","РўРѕРІР°СЂС‹"],["Omborlar","РЎРєР»Р°РґС‹"],["Rejimlar","Р РµР¶РёРјС‹"],["Oziq-ovqat","РџСЂРѕРґРѕРІРѕР»СЊСЃС‚РІРёРµ"],["Muddatlar","РЎСЂРѕРєРё"],["Boshqaruv","РЈРїСЂР°РІР»РµРЅРёРµ"],["Arxiv","РђСЂС…РёРІ"],["To'lovlar","РџР»Р°С‚РµР¶Рё"],["Davlatlar bo'yicha yo'nalishlar","РњР°СЂС€СЂСѓС‚С‹ РїРѕ СЃС‚СЂР°РЅР°Рј"],["Deklaratsiya post kodi bo'yicha tahlil","РђРЅР°Р»РёР· РїРѕ РєРѕРґСѓ РїРѕСЃС‚Р° РґРµРєР»Р°СЂР°С†РёРё"],["Transport turi bo'yicha ulushi","Р”РѕР»СЏ РїРѕ РІРёРґСѓ С‚СЂР°РЅСЃРїРѕСЂС‚Р°"],["Qiymat bo'yicha TOP 30 korxona","РўРћРџ-30 РїСЂРµРґРїСЂРёСЏС‚РёР№ РїРѕ СЃС‚РѕРёРјРѕСЃС‚Рё"],["Rahbar uchun qisqa xulosa","РљСЂР°С‚РєР°СЏ СЃРІРѕРґРєР° РґР»СЏ СЂСѓРєРѕРІРѕРґРёС‚РµР»СЏ"]];document.querySelectorAll("button,h2,h1,b,label,.muted,.btn").forEach(el=>{let s=el.childNodes.length===1?el.textContent.trim():"";let p=pairs.find(x=>x[0]===s);if(p)el.textContent=p[1]})}
adminPanel=function(){return `<div class=stack><div class=admin-layout><div class="admin-card"><h2>👤 Hodim qo'shish yoki tahrirlash</h2><form id=userForm class=admin-form><input type=hidden name=edit_mode value=""><label>Login</label><input name=user required><label>Yangi parol</label><div class="pass-wrap"><input name=password type=password placeholder="Tahrirda bo'sh qoldirish mumkin"><button class="eye-btn" type="button" onclick="let p=this.previousElementSibling;p.type=p.type==='password'?'text':'password'" title="Ko'rsatish/yashirish">&#128065;</button></div><label>F.I.Sh.</label><input name=full_name><label>Lavozim</label><input name=position><label>Telefon</label><input name=phone><label>Post kodi</label><input name=post_code placeholder="00 = IBK, aks holda post kodi"><label>Rol</label><select name=role><option value=foydalanuvchi>Foydalanuvchi</option><option value=inspektor>Inspektor</option><option value=rahbar>Rahbar</option><option value=admin>Admin</option></select><label>Til</label><select name=lang><option value=uz>O'zbek lotin</option><option value=uzc>O'zbek kirill</option><option value=ru>Rus tili</option></select><div class=perm-grid><label><input type=checkbox name=perm_view checked> Ko'rish</label><label><input type=checkbox name=perm_upload> Yuklash</label><label><input type=checkbox name=perm_export> Eksport</label><label><input type=checkbox name=perm_release> Yechish</label></div><div class=excel-actions><button>Saqlash</button><button type=button class=light onclick="resetUserForm()">Tozalash</button></div></form></div><div class="admin-card"><h2>Hodimlar ro'yxati</h2><div id=users></div></div></div>${siteMapPanel()}</div>`}
function resetUserForm(){let f=$("userForm");if(!f)return;f.reset();f.edit_mode.value="";f.user.disabled=false;f.perm_view.checked=true}
bindUserForm=function(){let f=$("userForm");if(!f)return;f.onsubmit=async e=>{e.preventDefault();let perms=[];["view","upload","export","release"].forEach(p=>{if(f[`perm_${p}`]?.checked)perms.push(p)});let body={user:f.user.value,pass:f.password.value,full_name:f.full_name.value,position:f.position.value,phone:f.phone.value,post_code:f.post_code.value,role:f.role.value,lang:f.lang.value,perms,enabled:true};await api(f.edit_mode.value?"/api/users/update":"/api/users",{method:"POST",body:JSON.stringify(body)});resetUserForm();loadUsers()}}
window.editUser=function(login){api("/api/users").then(j=>{let u=(j.users||[]).find(x=>x.user===login),f=$("userForm");if(!u||!f)return;f.edit_mode.value="1";f.user.value=u.user;f.user.disabled=true;f.password.value="";f.full_name.value=u.full_name||"";f.position.value=u.position||"";f.phone.value=u.phone||"";f.post_code.value=u.post_code||"";f.role.value=u.role||"foydalanuvchi";f.lang.value=u.lang||"uz";["view","upload","export","release"].forEach(p=>f[`perm_${p}`].checked=(u.perms||[]).includes(p));window.scrollTo({top:0,behavior:"smooth"})})}
window.deleteUser=async function(login){if(!confirm(`${login} hodimini o'chirasizmi?`))return;await api("/api/users/delete",{method:"POST",body:JSON.stringify({user:login})});loadUsers()}
window.selfChangePassword=async function(e){
  e.preventDefault();
  let old=$('cpOld'),nw=$('cpNew'),nw2=$('cpNew2'),msg=$('cpMsg');
  if(!old?.value){msg.textContent='Joriy parolni kiriting';return}
  if(!nw?.value||nw.value.length<6){msg.textContent='Yangi parol kamida 6 ta belgi';return}
  if(nw.value!==nw2?.value){msg.textContent='Yangi parollar mos kelmadi';return}
  try{
    await api("/api/users/change_password",{method:"POST",body:JSON.stringify({old_pass:old.value,new_pass:nw.value})});
    msg.style.color='#166534';msg.textContent='✓ Parol muvaffaqiyatli o\'zgartirildi';
    old.value='';nw.value='';nw2.value='';
  }catch(er){msg.style.color='#b42318';msg.textContent='Xatolik: '+er.message}
}
window.changePassword=async function(login){
  let np=prompt(`"${login}" uchun yangi parol kiriting:`);
  if(!np||!np.trim())return;
  if(np.length<4){alert('Parol kamida 4 ta belgidan iborat bo\'lsin');return}
  try{
    await api("/api/users/update",{method:"POST",body:JSON.stringify({user:login,pass:np,_pwonly:true})});
    alert(`✓ "${login}" paroli muvaffaqiyatli o'zgartirildi`);
  }catch(e){alert('Xatolik: '+String(e))}
}
loadUsers=async function(){let box=$("users");if(!box)return;try{let j=await api("/api/users");let users=(j.users||[]).filter(u=>u.enabled!==false);
const SVG_EDIT=`<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M11 4H4a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2v-7"/><path d="M18.5 2.5a2.121 2.121 0 0 1 3 3L12 15l-4 1 1-4 9.5-9.5z"/></svg>`;
const SVG_KEY=`<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="7.5" cy="15.5" r="5.5"/><path d="M21 2l-9.6 9.6"/><path d="M15.5 7.5l3 3L22 7l-3-3"/></svg>`;
const SVG_DEL=`<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"/><path d="M19 6l-1 14a2 2 0 0 1-2 2H8a2 2 0 0 1-2-2L5 6"/><path d="M10 11v6"/><path d="M14 11v6"/><path d="M9 6V4a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v2"/></svg>`;
let rows=users.map(u=>{let su=esc(u.user||''),ju=(u.user||'').replaceAll("'","");
return `<tr style="border-bottom:1px solid var(--line)"><td style="padding:7px 9px;font-family:monospace;font-size:12px">${su}</td><td style="padding:7px 9px">${esc(u.full_name||'')}</td><td style="padding:7px 9px">${esc(roleTitle(u.role))}</td><td style="padding:7px 9px;text-align:center">${esc(u.post_code||'')}</td><td style="padding:7px 9px;font-size:11px;color:#5a7a9a">${(u.perms||[]).map(p=>esc(permTitle(p))).join(', ')}</td><td style="padding:4px 8px;text-align:center;white-space:nowrap"><div style="display:flex;gap:4px;justify-content:center;align-items:center"><button class="icon-btn edit-btn" onclick="editUser('${ju}')" title="Tahrirlash">${SVG_EDIT}</button><button class="icon-btn" style="background:#f0fdf4;color:#166534;border-color:#bbf7d0" onmouseover="this.style.background='#bbf7d0'" onmouseout="this.style.background='#f0fdf4'" onclick="changePassword('${ju}')" title="Parol o'zgartirish">${SVG_KEY}</button><button class="icon-btn del-btn" onclick="deleteUser('${ju}')" title="O'chirish">${SVG_DEL}</button></div></td></tr>`}).join('');
box.innerHTML=`<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:13px"><colgroup><col style="width:90px"><col style="width:200px"><col style="width:110px"><col style="width:70px"><col><col style="width:100px"></colgroup><thead><tr style="background:#f0f4fb;border-bottom:2px solid var(--line)"><th style="padding:8px 9px;text-align:left;font-weight:700">Login</th><th style="padding:8px 9px;text-align:left;font-weight:700">F.I.Sh.</th><th style="padding:8px 9px;text-align:left;font-weight:700">Rol</th><th style="padding:8px 9px;text-align:center;font-weight:700">Post</th><th style="padding:8px 9px;text-align:left;font-weight:700">Vakolatlar</th><th style="padding:8px 9px;text-align:center;font-weight:700">Amal</th></tr></thead><tbody>${rows}</tbody></table></div>`}catch(e){box.innerHTML="Admin vakolati kerak"}}
const bindUploadFinal=bindUpload;bindUpload=function(){bindUploadFinal();let f=$("upload");if(f)f.onsubmit=async e=>{e.preventDefault();try{$("status").textContent="BNRTE fayllari yuklanyapti...";let j=await api("/api/reports",{method:"POST",body:new FormData(f)});poll(j.job_id)}catch(err){$("status").textContent=err.message}};let bf=$("bulkUpload");if(bf)bf.onsubmit=async e=>{e.preventDefault();try{$("status").textContent="Yillik asos fayllar yuklanyapti...";let j=await api("/api/reports_bulk",{method:"POST",body:new FormData(bf)});let skip=(j.skipped||[]).length?` O'tkazib yuborildi: ${j.skipped.map(x=>x.filename).join(", ")}`:"";$("bulkResult").textContent=`${j.count||0} ta fayl navbatga qo'shildi.${skip}`;$("status").textContent="Bulk yuklash boshlandi";await loadArchive()}catch(err){$("status").textContent=err.message}}}
// ── UI Config (sayt xaritasi) ─────────────────────────────────────────────
let UI_CONFIG={};
async function loadUIConfig(){try{UI_CONFIG=await api('/api/ui_config');}catch(e){}}
const SM_ALL_TABS=[['umumiy','Umumiy'],['rejim','Rejimlar'],['korxona','Korxonalar'],['ombor','Omborlar'],['expired',"Muddati o'tgan"],['released',"Nazoratdan yechish"],['goods','Tovarlar'],['food','Oziq-ovqat'],['muddat','Muddatlar'],['yaroqlilik','Yaroqlilik'],['archive','Arxiv'],['avia','AVIA AWB'],['payments',"To'lovlar: Umumiy"],['pay_lists',"To'lovlar ro'yxati"],['pay_analysis',"To'lovlar tahlili"]];
const SM_ROLES=['admin','rahbar','inspektor','foydalanuvchi'];
const SM_RLBL={admin:'Admin',rahbar:'Rahbar',inspektor:'Inspektor',foydalanuvchi:"Foydalanuvchi"};

function applyUIConfig(){
  if(!ME)return;
  const myRole=ME.role||'foydalanuvchi';
  if(myRole==='admin')return; // admin sees everything
  const hidden=UI_CONFIG.tab_hidden||[];
  const tabRoles=UI_CONFIG.tab_roles||{};
  const tabOrder=UI_CONFIG.tab_order||[];
  // Hide forbidden tabs
  document.querySelectorAll('#tabs .tab.sub').forEach(btn=>{
    const m=btn.getAttribute('onclick')?.match(/TAB='([^']+)'/);
    if(!m)return;
    const id=m[1];
    const shouldHide=hidden.includes(id)||(tabRoles[id]&&!tabRoles[id].includes(myRole));
    btn.style.display=shouldHide?'none':'';
    if(shouldHide&&TAB===id){const v=[...document.querySelectorAll('#tabs .tab.sub:not([style*="none"])')];if(v.length){const tm=v[0].getAttribute('onclick')?.match(/TAB='([^']+)'/);if(tm){TAB=tm[1];view();}}}
  });
  // Reorder tabs by drag order
  if(tabOrder.length){
    document.querySelectorAll('#tabs .subtabs').forEach(container=>{
      const btns=[...container.querySelectorAll('.tab.sub')];
      tabOrder.forEach(id=>{const btn=btns.find(b=>{const m=b.getAttribute('onclick')?.match(/TAB='([^']+)'/);return m&&m[1]===id;});if(btn)container.appendChild(btn);});
    });
  }
}

const SM_MODULES=[
  {id:'bnrte',label:'BNRTE',tabs:['umumiy','rejim','korxona','ombor','expired','released','goods','food','muddat','yaroqlilik','archive','avia']},
  {id:'pay',label:"To'lovlar",tabs:['payments','pay_lists','pay_analysis']},
  {id:'common',label:'Boshqaruv',tabs:['upload','profile','settings','admin']}
];
function siteMapPanel(){
  const order=UI_CONFIG.tab_order||SM_ALL_TABS.map(t=>t[0]);
  const hidden=UI_CONFIG.tab_hidden||[];
  const tabRoles=UI_CONFIG.tab_roles||{};
  const tabMap=Object.fromEntries(SM_ALL_TABS);
  function sortedTabs(ids){return[...ids].sort((a,b)=>{let ai=order.indexOf(a),bi=order.indexOf(b);return(ai<0?999:ai)-(bi<0?999:bi);});}
  function modSection(mod){
    const rows=sortedTabs(mod.tabs).map(id=>{
      const label=tabMap[id]||id;
      const isH=hidden.includes(id);
      const roles=tabRoles[id]||SM_ROLES;
      const roleCells=SM_ROLES.map(r=>`<label style="cursor:pointer;display:flex;align-items:center;gap:3px;white-space:nowrap"><input type=checkbox data-tab="${id}" data-role="${r}" ${roles.includes(r)?'checked':''}><span style="font-size:11px">${SM_RLBL[r]}</span></label>`).join('');
      return `<tr draggable="true" data-sm-id="${id}" data-sm-mod="${mod.id}" ondragstart="smDs(event,this)" ondragover="smDo(event,this)" ondrop="smDp(event,this)" ondragend="smDe()" style="border-bottom:1px solid #e8edf3;background:${isH?'#fef2f2':'#fff'}">
<td style="padding:7px 5px;text-align:center;color:#94a3b8;font-size:17px;user-select:none;cursor:grab">⠿</td>
<td style="padding:7px 10px;font-weight:600;min-width:130px;font-size:13px">${label}</td>
<td style="padding:7px 10px;width:110px"><label style="cursor:pointer;display:flex;align-items:center;gap:4px;font-size:12px;color:${isH?'#dc2626':'#16a34a'}"><input type=checkbox data-tab="${id}" data-sm-hide="1" ${isH?'checked':''} onchange="smToggle(this)">${isH?'Yashirin':"Ko'rinadi"}</label></td>
<td style="padding:7px 10px"><div style="display:flex;gap:8px;flex-wrap:wrap">${roleCells}</div></td>
</tr>`;}).join('');
    return `<div style="margin-bottom:14px;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden">
<div style="background:#f1f5f9;padding:9px 14px;font-weight:700;font-size:13px;border-bottom:1px solid #e2e8f0;letter-spacing:.3px">${mod.label}</div>
<table style="width:100%;border-collapse:collapse">
<thead><tr style="background:#f8fafc;border-bottom:1px solid #e8edf3"><th style="width:36px"></th><th style="padding:6px 10px;text-align:left;font-size:11px;font-weight:700;color:#64748b">Tab</th><th style="padding:6px 10px;text-align:left;font-size:11px;font-weight:700;color:#64748b;width:110px">Holat</th><th style="padding:6px 10px;text-align:left;font-size:11px;font-weight:700;color:#64748b">Kimga ko'rinsin</th></tr></thead>
<tbody data-sm-mod="${mod.id}">${rows}</tbody></table></div>`;
  }
  return `<div class=panel style="margin-top:14px"><h2>Sayt xaritasi — ko'rinish va tartib</h2>
<p class=muted style="margin-bottom:14px">Har modul ichida tablarni sudrab tartibini o'zgartiring. Qaysi rol qaysi tabni ko'rishini belgilang. Admin har doim barcha tabni ko'radi.</p>
${SM_MODULES.map(m=>modSection(m)).join('')}
<div style="margin-top:12px;display:flex;align-items:center;gap:12px"><button onclick="saveSiteMap()">Saqlash</button><button class=light onclick="resetSiteMap()">Asl sozlamalar</button><span id="smMsg" class=muted></span></div></div>`;
}
let _smDrag=null,_smPrev=null;
function smDs(e,el){_smDrag=el;e.dataTransfer.effectAllowed='move';el.style.opacity='.45';}
function smDo(e,el){e.preventDefault();if(el===_smDrag||!_smDrag)return;if(el.dataset.smMod!==_smDrag.dataset.smMod)return;if(_smPrev)_smPrev.style.borderTop='';_smPrev=el;el.style.borderTop='2px solid #3b82f6';}
function smDe(){if(_smDrag)_smDrag.style.opacity='';if(_smPrev)_smPrev.style.borderTop='';_smDrag=_smPrev=null;}
function smDp(e,target){e.preventDefault();if(!_smDrag||_smDrag===target)return;if(_smDrag.dataset.smMod!==target.dataset.smMod)return;const tbody=target.closest('tbody');const rows=[...tbody.querySelectorAll('tr')];rows.indexOf(_smDrag)<rows.indexOf(target)?target.after(_smDrag):target.before(_smDrag);}
function smToggle(cb){const row=cb.closest('tr');row.style.background=cb.checked?'#fef2f2':'#fff';cb.nextElementSibling.textContent=cb.checked?'Yashirin':"Ko'rinadi";cb.nextElementSibling.style.color=cb.checked?'#dc2626':'#16a34a';}
async function saveSiteMap(){
  const tab_order=[];const tab_hidden=[];const tab_roles={};
  SM_MODULES.forEach(mod=>{
    const tbody=document.querySelector(`[data-sm-mod="${mod.id}"]`);if(!tbody)return;
    [...tbody.querySelectorAll('tr[data-sm-id]')].forEach(row=>{
      const id=row.dataset.smId;
      tab_order.push(id);
      if(row.querySelector('[data-sm-hide="1"]')?.checked)tab_hidden.push(id);
      const checked=[...row.querySelectorAll('[data-role]')].filter(el=>el.checked).map(el=>el.dataset.role);
      tab_roles[id]=checked.includes('admin')?checked:[...checked,'admin'];
    });
  });
  try{
    const cfg={...UI_CONFIG,tab_order,tab_hidden,tab_roles};
    await api('/api/ui_config',{method:'POST',body:JSON.stringify(cfg)});
    UI_CONFIG=cfg;
    const msg=$('smMsg');if(msg){msg.textContent='✓ Saqlandi';setTimeout(()=>msg.textContent='',3000);}
    render();
  }catch(e){const msg=$('smMsg');if(msg)msg.textContent='Xatolik: '+e.message;}
}
async function resetSiteMap(){
  if(!confirm('Barcha tab sozlamalari asl holatga qaytariladi?'))return;
  try{await api('/api/ui_config',{method:'POST',body:JSON.stringify({})});UI_CONFIG={};TAB='umumiy';render();}catch(e){}
}
// ─────────────────────────────────────────────────────────────────────────────
const renderClean=render;render=function(){renderClean();cleanTopActions();if(LANG==="ru")translateRuPage();applyUIConfig()}
function releaseCompanyRows(data){
  let all=[...(data.rows||[]),...(data.unreleased||[])], byc={};
  all.forEach(r=>{
    let k=r.stir||r.korxona||r.company||"-";
    if(!byc[k])byc[k]={key:{stir:r.stir||""},korxona:r.korxona||r.company||"",stir:r.stir||"",base_vazn:0,base_qiymat:0,base_partiya:0,remain_vazn:0,remain_qiymat:0,remain_partiya:0,released_vazn:0,released_qiymat:0,released_partiya:0};
    ["base_vazn","base_qiymat","base_partiya","remain_vazn","remain_qiymat","remain_partiya","released_vazn","released_qiymat","released_partiya"].forEach(x=>byc[k][x]+=+r[x]||0);
  });
  return by(Object.values(byc).map(r=>Object.assign(r,{released_pct:r.base_qiymat?r.released_qiymat/r.base_qiymat*100:0,remain_pct:r.base_qiymat?r.remain_qiymat/r.base_qiymat*100:0,current_vazn:r.remain_vazn,current_qiymat:r.remain_qiymat,current_partiya:r.remain_partiya})),"base_qiymat");
}
function releaseTotalRow(rows){
  let t={korxona:"IBK bo'yicha Jami",stir:"",base_vazn:0,base_qiymat:0,base_partiya:0,remain_vazn:0,remain_qiymat:0,remain_partiya:0,released_vazn:0,released_qiymat:0,released_partiya:0,current_vazn:0,current_qiymat:0,current_partiya:0};
  rows.forEach(r=>Object.keys(t).forEach(k=>{if(typeof t[k]==="number")t[k]+=+r[k]||0}));
  t.released_pct=t.base_qiymat?t.released_qiymat/t.base_qiymat*100:0;
  t.remain_pct=t.base_qiymat?t.remain_qiymat/t.base_qiymat*100:0;
  return t;
}
function releaseCols(base,final){
  return [
    {k:"rn",t:"№",n:1,f:v=>v,w:"34px"},{k:"korxona",t:"Korxona",w:"305px"},{k:"stir",t:"STIR",w:"88px"},
    {k:"base_vazn",t:`${base} holatiga qoldiq vazn (tn)`,n:1,f:fmtN,w:"96px"},{k:"base_qiymat",t:`${base} holatiga qoldiq qiymat (ming $)`,n:1,f:fmtN,w:"104px"},{k:"base_partiya",t:`${base} holatiga qoldiq partiya`,n:1,f:fmtI,w:"72px"},
    {k:"remain_vazn",t:`${final} holatiga ${base}dan qoldiq vazn (tn)`,n:1,f:fmtN,w:"96px"},{k:"remain_qiymat",t:`${final} holatiga ${base}dan qoldiq qiymat (ming $)`,n:1,f:fmtN,w:"104px"},{k:"remain_partiya",t:`${final} holatiga ${base}dan qoldiq partiya`,n:1,f:fmtI,w:"72px"},
    {k:"released_vazn",t:"Yechilishi vazn (tn)",n:1,f:fmtN,w:"88px"},{k:"released_qiymat",t:"Yechilishi qiymat (ming $)",n:1,f:fmtN,w:"96px"},{k:"released_pct",t:"Yechilishi %da",n:1,f:fmtN,w:"68px"},{k:"released_partiya",t:"Yechilishi partiya",n:1,f:fmtI,w:"72px"},
    {k:"current_vazn",t:`${final} holatiga qoldiq vazn (tn)`,n:1,f:fmtN,w:"96px"},{k:"current_qiymat",t:`${final} holatiga qoldiq qiymat (ming $)`,n:1,f:fmtN,w:"104px"},{k:"current_partiya",t:`${final} holatiga qoldiq partiya`,n:1,f:fmtI,w:"72px"}
  ];
}
function releaseSpeedPanels(rows,base,final){
  rows=rows||[];
  let days=Math.max(1,Math.round((parseUzDate(final)-parseUzDate(base))/86400000));
  let fast=rows.filter(r=>r.released_qiymat>0).map(r=>Object.assign({mezoni:`${fmtN(r.released_pct)}% / ${days} kun`},r)).sort((a,b)=>(b.released_pct-a.released_pct)||(b.released_qiymat-a.released_qiymat)).slice(0,15);
  let slow=rows.filter(r=>r.base_qiymat>0).map(r=>Object.assign({mezoni:`qoldiq ${fmtN(r.remain_pct)}%`},r)).sort((a,b)=>(b.remain_pct-a.remain_pct)||(b.remain_qiymat-a.remain_qiymat)).slice(0,15);
  let cols=[{k:"korxona",t:"Korxona nomi",w:"290px"},{k:"stir",t:"STIR",w:"86px"},{k:"released_qiymat",t:"Yechilgan qiymat (ming $)",n:1,f:fmtN,w:"105px"},{k:"released_vazn",t:"Yechilgan vazn (tn)",n:1,f:fmtN,w:"96px"},{k:"released_pct",t:"Yechilish ulushi (%)",n:1,f:fmtN,w:"86px"},{k:"remain_qiymat",t:"Qoldiq qiymat (ming $)",n:1,f:fmtN,w:"105px"},{k:"mezoni",t:"Mezon",w:"110px"}];
  return `<div class=grid2><div class=panel><h2>Eng tez nazoratdan yechadigan korxonalar</h2><div class=muted>Mezon: yechilgan qiymatning boshlang'ich qiymatga nisbati, teng holatda yechilgan qiymat. TOP 15.</div>${table(cols,basicTotal(fast,"IBK bo'yicha Jami","korxona"),"fixed-table")}</div><div class=panel><h2>Eng sekin nazoratdan yechadigan korxonalar</h2><div class=muted>Mezon: yakuniy sanada qolgan qiymat ulushi, teng holatda qoldiq qiymat. TOP 15.</div>${table(cols,basicTotal(slow,"IBK bo'yicha Jami","korxona"),"fixed-table")}</div>${chartBlock("Korxonalar yechilishi qiymat bo'yicha ulushi",fast,"korxona","released_qiymat",fmtN)}</div>`;
}
function warehouseTurnoverRows(data){
  let all=[...(data.rows||[]),...(data.unreleased||[])], map={};
  all.forEach(r=>{
    let name=r.warehouse||"O'z ombor";
    if(!map[name])map[name]={key:{warehouse:name},name,base_vazn:0,remain_vazn:0,released_vazn:0,base_qiymat:0,released_qiymat:0,remain_qiymat:0,partiya:0};
    ["base_vazn","remain_vazn","released_vazn","base_qiymat","released_qiymat","remain_qiymat"].forEach(k=>map[name][k]+=+r[k]||0);
    map[name].partiya+=+r.released_partiya||0;
  });
  return Object.values(map).map(r=>Object.assign(r,{released_pct:r.base_vazn?r.released_vazn/r.base_vazn*100:0,remain_pct:r.base_vazn?r.remain_vazn/r.base_vazn*100:0})).sort((a,b)=>b.released_vazn-a.released_vazn);
}
function warehouseTurnoverPanel(data){
  let rows=warehouseTurnoverRows(data);
  let fast=rows.filter(r=>r.released_vazn>0).sort((a,b)=>(b.released_pct-a.released_pct)||(b.released_vazn-a.released_vazn)).slice(0,15);
  let slow=rows.filter(r=>r.base_vazn>0).sort((a,b)=>(b.remain_pct-a.remain_pct)||(b.remain_vazn-a.remain_vazn)).slice(0,15);
  let cols=[{k:"name",t:"Ombor",w:"320px"},{k:"base_vazn",t:"Boshlang'ich vazn (tn)",n:1,f:fmtN,w:"105px"},{k:"released_vazn",t:"Yechilgan vazn (tn)",n:1,f:fmtN,w:"105px"},{k:"released_pct",t:"Yechilish ulushi (%)",n:1,f:fmtN,w:"90px"},{k:"remain_vazn",t:"Qoldiq vazn (tn)",n:1,f:fmtN,w:"105px"},{k:"partiya",t:"Partiya",n:1,f:fmtI,w:"75px"}];
  return `<div class="panel wide"><h2>Omborlar oboroti: nazoratdan yechilgan vazn bo'yicha</h2>${table(cols,basicTotal(rows,"IBK bo'yicha Jami","name"),"fixed-table")}</div><div class=grid2><div class=panel><h2>Eng tez oborot qiladigan omborlar</h2>${table(cols,basicTotal(fast,"IBK bo'yicha Jami","name"),"fixed-table")}</div><div class=panel><h2>Eng sekin oborot qiladigan omborlar</h2>${table(cols,basicTotal(slow,"IBK bo'yicha Jami","name"),"fixed-table")}</div>${chartBlock("Omborlar yechilgan vazn bo'yicha ulushi",fast,"name","released_vazn",fmtN)}</div>`;
}
buildRelease=async function(){
  let base=$("relBase").value, final=$("relFinal").value;
  if(parseUzDate(base)>=parseUzDate(final)){
    $("releaseResult").innerHTML=`<div class=muted>Boshlang'ich sana yakuniy sanadan oldin bo'lishi kerak.</div>`;
    return;
  }
  let j=await api(`/api/release?base=${encodeURIComponent(base)}&final=${encodeURIComponent(final)}`);
  if(j.missing){
    $("releaseResult").innerHTML=`<div class=muted>Arxivda yo'q sana: ${j.missing.join(", ")}. Shu davr uchun asos fayl yuklash kerak.</div>`;
    return;
  }
  window.LAST_RELEASE=j;
  let companies=releaseCompanyRows(j);
  let rows=numbered([releaseTotalRow(companies)].concat(companies));
  $("releaseResult").innerHTML=`<h2>${base} - ${final} <a class="btn light" href="/api/export?kind=release&base=${base}&final=${final}&token=${TOKEN}">Excel</a></h2><div class=overview-note>Boshlang'ich davr: ${base}. Yakuniy davr: ${final}. Jadval namuna fayldagi kabi korxonalar kesimida yig'iladi; ustun nomlari ixcham bo'lishi uchun keyingi qatorga ko'shiriladi.</div>${table(releaseCols(base,final),rows,"release-table sample-release-table fixed-table")}${releaseSpeedPanels(companies,base,final)}`;
  window.LAST_RELEASE=j;
}
function uniqueArchiveRows(){
  let best={};
  (ARCHIVE||[]).forEach(r=>{
    if(!r||!r.date)return;
    let score=(String(r.source||"").includes("IBK_Dashboard_SQLite")?2:0)+(r.deposit?1:0)+String(r.id||"").length/1000;
    if(!best[r.date]||score>best[r.date]._score)best[r.date]=Object.assign({_score:score},r);
  });
  return Object.values(best).sort((a,b)=>parseUzDate(b.date)-parseUzDate(a.date));
}
function compactArchivePanel(){
  let rows=uniqueArchiveRows();
  let isAdmin=ME&&ME.role==="admin";
  let body=rows.map(r=>{
    let source=(r.source||"").split(/[\\/]/).pop(), deposit=r.deposit?((r.deposit||"").split(/[\\/]/).pop()||"Bor"):"-";
    let isCurrent=r.id===ARCHIVE_CURRENT_ID||((!ARCHIVE_CURRENT_ID)&&r.id===ARCHIVE[0]?.id);
    let isLoaded=DATA&&DATA.id===r.id;
    let badge=isCurrent?`<span style="background:#166534;color:#fff;font-size:11px;padding:2px 7px;border-radius:10px;margin-left:6px">Joriy</span>`:"";
    let loadedBadge=isLoaded&&!isCurrent?`<span style="background:#1d4ed8;color:#fff;font-size:11px;padding:2px 7px;border-radius:10px;margin-left:6px">Ochiq</span>`:"";
    let adminBtns=isAdmin?`<button class="light" style="color:#b42318;border-color:#fca5a5" onclick="deleteArchiveEntry('${esc(r.id)}','${esc(r.date)}')" title="O'chirish">🗑</button>${!isCurrent?`<button class="light" style="font-size:12px" onclick="setCurrentArchive('${esc(r.id)}')" title="Joriy qilib belgilash">★ Joriy</button>`:""}`:""
    return `<tr><td class=num>${esc(r.date)}${badge}${loadedBadge}</td><td class=text title="${esc(source)}">${esc(source)}</td><td class=text title="${esc(deposit)}">${esc(deposit)}</td><td class=num style="white-space:nowrap"><div style="display:flex;gap:4px;justify-content:center"><button class="light" onclick="loadReport('${esc(r.id)}')">Ochish</button>${adminBtns}</div></td></tr>`;
  }).join("");
  let html=`<table class="fixed-table compact-archive"><colgroup><col style="width:130px"><col style="width:320px"><col style="width:130px"><col></colgroup><thead><tr><th>Sana</th><th>Asos fayl</th><th>Depozit</th><th></th></tr></thead><tbody>${body}</tbody></table>`;
  return `<div class=panel><h2>Arxiv</h2><div class=muted>${rows.length} ta sana. "Joriy" — dashar ochilganda yuklanadigan hisobot. Admin o'chira yoki joriy qilib belgilashi mumkin.</div>${html}</div>`;
}
async function deleteArchiveEntry(id,date){
  if(!confirm(`"${date}" arxiv yozuvini o'chirasizmi?`))return;
  try{
    await api("/api/archive/delete",{method:"POST",body:JSON.stringify({id})});
    await loadArchive();
    if(DATA&&DATA.id===id){DATA=ARCHIVE.length?await api("/api/reports/"+ARCHIVE[0].id):null;}
    render();
  }catch(e){alert("Xatolik: "+e.message)}
}
async function setCurrentArchive(id){
  try{
    await api("/api/archive/set_current",{method:"POST",body:JSON.stringify({id})});
    await loadArchive();
    render();
  }catch(e){alert("Xatolik: "+e.message)}
}
const renderArchiveCompact=render;render=function(){
  renderArchiveCompact();
  if(TAB==="archive"){
    let v=$("view");
    if(v)v.innerHTML=compactArchivePanel();
  }
}
const showLoginBase=showLogin;showLogin=function(){showLoginBase();let el=$("login");if(el)el.classList.remove("active");clearBusy()}
const pollBase=poll;poll=async function(id){try{await pollBase(id)}finally{if(DATA)clearBusy()}}
const prepareArtifactsBase=prepareArtifacts;prepareArtifacts=async function(){let btn=(typeof event!=="undefined"&&event)?event.currentTarget:null;try{setBusy(btn,true,"Tayyorlash");await prepareArtifactsBase()}catch(e){clearBusy();throw e}}
const bindUploadBusy=bindUpload;bindUpload=function(){
  bindUploadBusy();
  let f=$("upload");if(f)f.onsubmit=async e=>{e.preventDefault();let btn=e.submitter;try{setBusy(btn,true,"Yuklanmoqda");$("status").textContent="BNRTE fayllari yuklanyapti...";let j=await api("/api/reports",{method:"POST",body:new FormData(f)});poll(j.job_id)}catch(err){$("status").textContent=err.message;setBusy(btn,false)}};
  let bf=$("bulkUpload");if(bf)bf.onsubmit=async e=>{e.preventDefault();let btn=e.submitter;try{setBusy(btn,true,"Yuklanmoqda");$("status").textContent="Yillik asos fayllar yuklanyapti...";let j=await api("/api/reports_bulk",{method:"POST",body:new FormData(bf)});let skip=(j.skipped||[]).length?` O'tkazib yuborildi: ${j.skipped.map(x=>x.filename).join(", ")}`:"";$("bulkResult").textContent=`${j.count||0} ta fayl navbatga qo'shildi.${skip}`;$("status").textContent="Bulk yuklash boshlandi";await loadArchive()}catch(err){$("status").textContent=err.message}finally{setBusy(btn,false)}};
  let tf=$("tolovUpload");if(tf)tf.onsubmit=async e=>{e.preventDefault();let btn=e.submitter;try{setBusy(btn,true,"Shakllanmoqda");$("status").textContent="To'lovlar shakllantirilyapti...";let j=await api("/api/tolov",{method:"POST",body:new FormData(tf)});PAYMENTS=j.payments||[];$("tolovUploadResult").innerHTML=`Tayyor: ${fmtI(PAYMENTS.reduce((a,r)=>a+(+r.rows||0),0))} qator, ${fmtN(PAYMENTS.reduce((a,r)=>a+(+r.sum||0),0))} so'm.`;$("status").textContent="To'lovlar tayyor"}catch(err){$("status").textContent=err.message}finally{setBusy(btn,false)}};
}
const cleanTopActionsBase=cleanTopActions;cleanTopActions=function(){
  cleanTopActionsBase();
  let a=$("actions");if(!a)return;
  let pngLinks=[...a.querySelectorAll("a")].filter(x=>x.textContent.trim()==="PNG"||x.hasAttribute("data-pngzip"));
  pngLinks.slice(1).forEach(x=>x.remove());
  ["O'zb","\u040e\u0437\u0431","\u0420\u0443\u0441","\u25d0"].forEach(label=>{
    let buttons=[...a.querySelectorAll("button")].filter(b=>b.textContent.trim()===label);
    buttons.slice(1).forEach(b=>b.remove());
  });
}
const renderFinalPolish=render;render=function(){renderFinalPolish();cleanTopActions();updateClock();if(TOKEN){let l=document.getElementById("login");if(l){l.style.cssText="display:none!important";l.classList.add("hidden");}}}

// ===== UPLOAD YANGILANMASI =====
function showUploadProgress(label,pct){
  let el=$("uploadStatus");if(!el)return;
  if(!label){el.innerHTML='';return}
  let c=pct>=100?'#16a34a':pct===0?'#dc2626':'#1d72b8';
  el.innerHTML=`<div style="background:#f0f6ff;border:1px solid #c8dff5;border-radius:8px;padding:10px 14px"><div style="font-size:13px;font-weight:700;color:${c};margin-bottom:6px">${esc(label)}</div><div style="background:#dde8f5;border-radius:6px;height:8px;overflow:hidden"><div style="width:${Math.max(3,pct)}%;background:${c};height:100%;border-radius:6px;transition:width .4s ease"></div></div></div>`;
}
poll=async function(id){
  try{
    let j=await api("/api/jobs/"+id);
    let pMap={"navbatda":8,"Hisob-kitob qilinyapti":30,"SQLite bazaga saqlanmoqda":55,"Dashboard JSON tayyorlanmoqda":70,"Excel/PNG/PDF tayyorlanmoqda":82,"Depozit qayta ishlanmoqda":45};
    let pct=j.status==="tayyor"?100:j.status==="xatolik"?0:(pMap[j.status]||35);
    showUploadProgress(j.status==="tayyor"?"✓ Tayyor!":j.status==="xatolik"?"✗ Xatolik":j.status+" ...",pct);
    $("status").textContent=j.status;
    if(j.status==="xatolik"){showUploadProgress("✗ "+(j.error||"Noma'lum xato"),0);clearBusy();return}
    if(j.status!=="tayyor"){setTimeout(()=>poll(id),1800);return}
    DATA=j.data;TAB="umumiy";await loadArchive();clearBusy();setTimeout(()=>showUploadProgress("",0),4000);render();
  }catch(err){$("status").textContent=err.message;clearBusy()}
};
async function uploadDepositOnly(btn){
  let f=$("depositOnlyForm");if(!f)return;
  let sel=f.querySelector("[name=report_id]"),dep=f.querySelector("[name=deposit]");
  if(!sel||!dep||!dep.files.length){$("depositResult").textContent="Hisobot va depozit fayl tanlang";return}
  setBusy(btn,true,"Yuklanmoqda");$("depositResult").textContent="Depozit yuklanyapti...";
  showUploadProgress("Depozit fayl yuklanyapti...",10);
  try{
    let fd=new FormData();fd.append("report_id",sel.value);fd.append("deposit",dep.files[0]);
    let j=await api("/api/deposit",{method:"POST",body:fd});
    $("depositResult").textContent="Qayta hisoblanmoqda...";
    poll(j.job_id);
  }catch(err){$("depositResult").textContent="Xato: "+err.message;showUploadProgress("✗ "+err.message,0)}
  finally{setBusy(btn,false)}
}
async function refreshCurrentReport(btn){
  if(!DATA)return;
  setBusy(btn,true,"Yangilanmoqda");$("status").textContent="Ma'lumotlar yangilanmoqda...";
  try{DATA=await api("/api/reports/"+DATA.id);render();$("status").textContent="Yangilandi"}catch(err){$("status").textContent=err.message}finally{setBusy(btn,false)}
}
async function chunkedUpload(file,onProgress){
  _uploadActive=true;
  const CHUNK=1024*1024;
  const total=Math.max(1,Math.ceil(file.size/CHUNK));
  const uid=Date.now().toString(36)+Math.random().toString(36).slice(2,6);
  if(DIRECT_UPLOAD_URL===null){
    try{
      const j=await fetch('/api/server_info',{headers:{'X-Token':TOKEN}}).then(r=>r.json());
      const t=await fetch(j.lan_url+'/api/server_info',{signal:AbortSignal.timeout(2000)});
      DIRECT_UPLOAD_URL=t.ok?j.lan_url:'';
    }catch{DIRECT_UPLOAD_URL='';}
  }
  const base=DIRECT_UPLOAD_URL||'';
  const isLAN=!!DIRECT_UPLOAD_URL;
  const PARALLEL=isLAN?4:3;
  const MAX_RETRY=3;
  let done=0;
  async function uploadOne(i){
    const blob=file.slice(i*CHUNK,(i+1)*CHUNK);
    let lastErr;
    for(let attempt=0;attempt<MAX_RETRY;attempt++){
      if(attempt>0)await new Promise(r=>setTimeout(r,1500*attempt));
      try{
        const r=await fetch(base+'/api/chunk_upload',{method:'POST',headers:{
          'X-Token':TOKEN,'X-Upload-Id':uid,'X-Chunk-Index':String(i),
          'X-Total-Chunks':String(total),'X-Filename':encodeURIComponent(file.name),
          'Content-Type':'application/octet-stream'
        },body:blob});
        if(!r.ok){
          const j=await r.json().catch(()=>({}));
          lastErr=new Error(j.error||`Chunk ${i+1}/${total} xato: HTTP ${r.status}`);
          if(r.status===401||r.status===403)throw lastErr;
          continue;
        }
        done++;if(onProgress)onProgress(done/total);
        return;
      }catch(e){lastErr=e;if(e.message&&(e.message.includes('401')||e.message.includes('403')))throw e;}
    }
    throw lastErr;
  }
  try{
    for(let i=0;i<total;i+=PARALLEL)await Promise.all(Array.from({length:Math.min(PARALLEL,total-i)},(_,k)=>uploadOne(i+k)));
    return {upload_id:uid,filename:file.name,total_chunks:total,direct:isLAN};
  }finally{_uploadActive=false;}
}
async function ucUploadBnrte(btn){
  let src=$('ucBnrteSource'),dep=$('ucBnrteDeposit'),st=$('ucBnrteStatus');
  if(!src?.files?.length){if(st)st.textContent='Asos fayl tanlang';return}
  setBusy(btn,true,'Yuklanmoqda');
  try{
    const mode=DIRECT_UPLOAD_URL?'⚡ To\'g\'ridan-to\'g\'ri (LAN)':'☁️ Cloudflare orqali';
    if(st)st.textContent='Yuklanyapti... '+mode;
    showUploadProgress('Asos fayl yuklanyapti... '+mode,5);
    const srcInfo=await chunkedUpload(src.files[0],p=>{
      showUploadProgress(`Asos fayl: ${Math.round(p*100)}%`,5+p*70);
      if(st)st.textContent=`Asos fayl: ${Math.round(p*100)}% yuklandi`;
    });
    let depInfo={upload_id:'',filename:''};
    if(dep?.files?.length){
      if(st)st.textContent='Depozit fayl yuklanyapti...';
      showUploadProgress('Depozit fayl yuklanyapti...',76);
      const d=await chunkedUpload(dep.files[0],p=>{
        showUploadProgress(`Depozit: ${Math.round(p*100)}%`,76+p*15);
      });
      depInfo={upload_id:d.upload_id,filename:d.filename};
    }
    if(st)st.textContent='Hisoblanmoqda...';showUploadProgress('Hisoblanmoqda...',92);
    const j=await api('/api/chunk_finalize',{method:'POST',body:JSON.stringify({...srcInfo,deposit_upload_id:depInfo.upload_id,deposit_filename:depInfo.filename})});
    poll(j.job_id);
  }catch(e){if(st)st.textContent='Xatolik: '+e.message;showUploadProgress('✗ '+e.message,0);setBusy(btn,false)}
}
async function ucUploadDepozit(btn){
  if(!DATA){if($('ucDepozitStatus'))$('ucDepozitStatus').textContent='Avval BNRTE hisobotini yuklang';return}
  let src=$('ucDepozitFile'),st=$('ucDepozitStatus');
  if(!src?.files?.length){if(st)st.textContent='Fayl tanlang';return}
  setBusy(btn,true,'Yuklanmoqda');if(st)st.textContent='Yuklanyapti...';
  let fd=new FormData();fd.append('report_id',DATA.id);fd.append('deposit',src.files[0]);
  try{let j=await api('/api/deposit',{method:'POST',body:fd});if(st)st.textContent='Hisoblanmoqda...';poll(j.job_id);}
  catch(e){if(st)st.textContent='Xatolik: '+e.message;setBusy(btn,false)}
}
async function ucUploadTolov(btn){
  let src=$('ucTolovSource'),st=$('ucTolovStatus');
  if(!src?.files?.length){if(st)st.textContent='Fayl tanlang';return}
  setBusy(btn,true,'Yuklanmoqda');if(st)st.textContent='Yuklanyapti...';
  let fd=new FormData();fd.append('source',src.files[0]);
  try{
    let j=await api('/api/tolov',{method:'POST',body:fd});PAYMENTS=j.payments||[];
    if(st)st.textContent=`Tayyor: ${PAYMENTS.length} tur, ${fmtN(PAYMENTS.reduce((a,r)=>a+(+r.sum||0),0))} so'm`;
    if($('kpis'))$('kpis').innerHTML=renderKpis();
  }catch(e){if(st)st.textContent='Xatolik: '+e.message}finally{setBusy(btn,false)}
}
async function ucRecalcTolov(btn){
  setBusy(btn,true,'Yuklanmoqda');
  try{let j=await api('/api/tolov');PAYMENTS=j.payments||[];
    if($('ucTolovStatus'))$('ucTolovStatus').textContent=`Yangilandi: ${PAYMENTS.length} tur`;render();}
  catch(e){}finally{setBusy(btn,false)}
}
async function ucUploadWarehouse(btn){
  let src=$('ucWrFile'),st=$('ucWrStatus');
  if(!src?.files?.length){if(st)st.textContent='Fayl tanlang';return}
  setBusy(btn,true,'Yuklanmoqda');if(st)st.textContent='Yuklanyapti...';
  let fd=new FormData();fd.append('file',src.files[0]);
  try{
    let j=await api('/api/upload_wr_registry',{method:'POST',body:fd});
    if(j.warehouses)WR_DATA=j.warehouses;
    if(st)st.textContent=j.success?`Tayyor: ${(j.warehouses||[]).length} ombor`:'Xatolik';
  }catch(e){if(st)st.textContent='Xatolik: '+e.message}finally{setBusy(btn,false)}
}
async function ucRecalcWarehouse(btn){
  setBusy(btn,true,'Yuklanmoqda');
  try{let j=await api('/api/warehouses');WR_DATA=j.warehouses||[];
    if($('ucWrStatus'))$('ucWrStatus').textContent=`Yangilandi: ${WR_DATA.length} ombor`;}
  catch(e){}finally{setBusy(btn,false)}
}
async function ucUploadYaroqlilik(btn){
  let src=$('ucYarFile'),st=$('ucYarStatus');
  if(!src?.files?.length){if(st)st.textContent='Fayl tanlang';return}
  setBusy(btn,true,'Yuklanmoqda');if(st)st.textContent='Tahlil qilinmoqda...';
  let fd=new FormData();fd.append('file',src.files[0]);
  try{
    let j=await api('/api/upload_yaroqlilik',{method:'POST',body:fd});
    YAROQLILIK_DATA=j;
    if(st)st.textContent=j.loaded?`Tayyor: ${j.expired_count||0} muddati o'tgan, ${j.warn180_count||0} ogohlantirish (180 kun), ${j.warn30_count||0} (1 oy)`:`Xatolik: ${esc(j.error||'')}`;
    if(TAB==='yaroqlilik')render();
  }catch(e){if(st)st.textContent='Xatolik: '+e.message}finally{setBusy(btn,false)}
}
async function ucUploadAvia(btn){
  let src=$('ucAviaFile'),st=$('ucAviaStatus');
  if(!src?.files?.length){if(st)st.textContent='Fayl tanlang';return}
  setBusy(btn,true,'Yuklanmoqda');if(st)st.textContent='Yuklanyapti...';
  let fd=new FormData();fd.append('file',src.files[0]);
  try{
    let j=await api('/api/upload_avia_awb',{method:'POST',body:fd});AVIA_DATA=j;
    if(st)st.textContent=j.loaded?`Tayyor: ${fmtI(j.unique_awb)} AWB yuklandi`:`Xatolik: ${esc(j.error||'')}`;
    if($('kpis'))$('kpis').innerHTML=renderKpis();
    if(TAB==='avia')loadAviaAwb();
  }catch(e){if(st)st.textContent='Xatolik: '+e.message}finally{setBusy(btn,false)}
}
async function ucRecalcAvia(btn){
  let st=$('ucAviaStatus');
  setBusy(btn,true,'Yuklanmoqda');if(st)st.textContent='Qayta yuklanmoqda...';
  try{
    let [j,stats]=await Promise.all([api('/api/avia_awb?force=1'),api('/api/avia_stats')]);
    AVIA_DATA=j;AVIA_STATS=stats;
    if(st)st.textContent=j.loaded?`Yangilandi: ${fmtI(j.unique_awb)} AWB, ${fmtN(stats?.jami_qiymat_k||0)} ming $`:`Xatolik: ${esc(j.error||'')}`;
    if($('kpis'))$('kpis').innerHTML=renderKpis();
    if(TAB==='avia')loadAviaAwb();
  }catch(e){if(st)st.textContent='Xatolik: '+String(e)}finally{setBusy(btn,false)}
}
uploadPanel=function(){
  let repDate=DATA?DATA.meta&&DATA.meta.date||DATA.date||'?':null;
  let bnrteSt=repDate?`<span class="uc-badge ok">📅 ${repDate} holatiga — ${fmtI(DATA.kpis?.partiya||0)} partiya</span>`:`<span class="uc-badge warn">Yuklanmagan</span>`;
  let depSt=repDate?`<span class="uc-badge ok">Joriy: ${repDate}</span>`:`<span class="uc-badge warn">BNRTE yuklanmagan</span>`;
  let tolovSt=PAYMENTS&&PAYMENTS.length?`<span class="uc-badge ok">💳 ${PAYMENTS.length} tur</span>`:`<span class="uc-badge warn">Yuklanmagan</span>`;
  let wrSt=WR_DATA&&WR_DATA.length?`<span class="uc-badge ok">🏢 ${WR_DATA.length} ombor</span>`:`<span class="uc-badge warn">Yuklanmagan</span>`;
  let aviaSt=AVIA_DATA&&AVIA_DATA.loaded?`<span class="uc-badge ok">📦 ${fmtI(AVIA_DATA.unique_awb)} AWB</span>`:`<span class="uc-badge warn">Yuklanmagan</span>`;
  let aviaDbSt=AVIA_STATS?`<span class="uc-badge ok">💲 ${fmtN(AVIA_STATS.jami_qiymat_k)} ming $</span>`:'';
  let yarSt=YAROQLILIK_DATA&&YAROQLILIK_DATA.loaded?`<span class="uc-badge ok">⚠ ${YAROQLILIK_DATA.expired_count||0} muddati o'tgan</span>`:`<span class="uc-badge warn">Yuklanmagan</span>`;
  return `<div class=stack>
<div class=panel><h2>📂 Ma'lumotlar boshqaruvi</h2><p class=muted>Har bir modul uchun alohida yuklash va qayta hisoblash. <b>Qayta hisoblash</b> — yangi fayl yuklamasdan serverdan qayta oladi.</p><div id=uploadStatus></div></div>
<div class="panel uc-card"><div class=uc-header><h2>📋 BNRTE — Nazoratdagi tovarlar</h2>${bnrteSt}</div><div class=uc-body><div class=uc-field><label>Asos fayl (xls/xlsx)</label><input type=file id="ucBnrteSource" accept=".xls,.xlsx,.html,.htm"></div><div class=uc-btns><button onclick="ucUploadBnrte(this)">Yuklash</button><button class="btn light" onclick="refreshCurrentReport(this)">🔄 Qayta hisoblash</button></div></div><div class=muted id="ucBnrteStatus" style="margin-top:8px"></div></div>
<div class="panel uc-card"><div class=uc-header><h2>💾 Depozit fayl — alohida yuklash</h2>${depSt}</div><div class=uc-body><div class=uc-field><label>Depozit fayl (xlsx)</label><input type=file id="ucDepozitFile" accept=".xlsx"></div><div class=uc-btns><button onclick="ucUploadDepozit(this)"${DATA?'':' disabled title="Avval BNRTE yuklang"'}>Yuklash</button></div></div><div class=muted style="margin-top:6px;font-size:12px">Joriy hisobotga (${repDate||'—'}) depozit ma'lumotlarini biriktiradi va qayta hisoblab chiqadi.</div><div class=muted id="ucDepozitStatus" style="margin-top:4px"></div></div>
<div class="panel uc-card"><div class=uc-header><h2>💰 To'lovlar jadvallari</h2>${tolovSt}</div><div class=uc-body><div class=uc-field><label>To'lov baza fayli (xlsx)</label><input type=file id="ucTolovSource" accept=".xlsx,.xls"></div><div class=uc-btns><button onclick="ucUploadTolov(this)">Yuklash</button><button class="btn light" onclick="ucRecalcTolov(this)">🔄 Qayta hisoblash</button></div></div><div class=muted id="ucTolovStatus" style="margin-top:8px"></div></div>
<div class="panel uc-card"><div class=uc-header><h2>🏢 Omborlar reestri</h2>${wrSt}</div><div class=uc-body><div class=uc-field><label>Reestri fayli (omborlarReestri*.xlsx)</label><input type=file id="ucWrFile" accept=".xlsx,.xls"></div><div class=uc-btns><button onclick="ucUploadWarehouse(this)">Yuklash</button><button class="btn light" onclick="ucRecalcWarehouse(this)">🔄 Qayta hisoblash</button></div></div><div class=muted id="ucWrStatus" style="margin-top:8px"></div></div>
<div class="panel uc-card" style="border-left-color:#0ea5e9"><div class=uc-header><h2>✈ AVIA AWB</h2><div style="display:flex;gap:6px;flex-wrap:wrap">${aviaSt} ${aviaDbSt}</div></div><div class=uc-note>AWB Excel (Yuklarni qabul qilish.xlsx) → AWB ro'yxati, joylar, vazn. Qiymat (ming $) BNRTE asos faylidan olinadi — yangilash uchun BNRTE → Qayta hisoblash.</div><div class=uc-body style="margin-top:12px"><div class=uc-field><label>AWB Excel (Yuklarni qabul qilish*.xlsx)</label><input type=file id="ucAviaFile" accept=".xlsx,.xls"></div><div class=uc-btns><button onclick="ucUploadAvia(this)">Yuklash</button><button class="btn light" onclick="ucRecalcAvia(this)">🔄 Qayta hisoblash</button></div></div><div class=muted id="ucAviaStatus" style="margin-top:8px"></div></div>
<div class="panel uc-card" style="border-left-color:#e8560a"><div class=uc-header><h2>⚠ Iste'mol muddati tahlili</h2>${yarSt}</div><div class=uc-note>yaroqlilik_muddati_tahlili.xlsx yoki yaroqlilik_1_oy_ichida.xlsx — muddati o'tgan va yaqinda tugaydigan tovarlar.</div><div class=uc-body style="margin-top:12px"><div class=uc-field><label>Yaroqlilik Excel fayli</label><input type=file id="ucYarFile" accept=".xlsx,.xls"></div><div class=uc-btns><button onclick="ucUploadYaroqlilik(this)">Yuklash</button></div></div><div class=muted id="ucYarStatus" style="margin-top:8px"></div></div>
<div class="panel uc-card"><div class=uc-header><h2>📚 Yillik arxivni birdan yuklash</h2></div><div class=uc-note>Bir vaqtning o'zida bir nechta asos fayl tanlanadi. Fayllar navbatga qo'shiladi va server mustaqil qayta hisoblab chiqadi — admin kutib o'tirmasligi kerak. Fayl sanasi nom ichidan avtomatik aniqlanadi.</div><form id="bulkUpload"><div class=uc-body><div class=uc-field><label>Asos fayllar (bir nechta)</label><input name="sources" type="file" accept=".xls,.xlsx,.html,.htm" multiple required></div><div class=uc-btns><button>Hammasini yuklash</button></div></div></form><div id=bulkResult class=muted style="margin-top:8px">Fayllar tanlanmagan.</div></div>
</div>`;
}
const bindUploadFull=bindUpload;bindUpload=function(){
  let f=$("upload");
  if(f)f.onsubmit=async e=>{
    e.preventDefault();let btn=e.submitter;
    try{setBusy(btn,true,"Yuklanmoqda");showUploadProgress("Fayl yuklanyapti...",5);
      let j=await api("/api/reports",{method:"POST",body:new FormData(f)});poll(j.job_id);
    }catch(err){$("status").textContent=err.message;showUploadProgress("✗ "+err.message,0)}
    finally{setBusy(btn,false)};
  };
  let bf=$("bulkUpload");
  if(bf)bf.onsubmit=async e=>{
    e.preventDefault();let btn=e.submitter;
    try{
      setBusy(btn,true,"Yuklanmoqda");
      const files=[...(bf.querySelector('[name=sources]')?.files||[])];
      if(!files.length){$("bulkResult").textContent="Fayl tanlanmagan";return}
      let done=0,skipped=[];
      for(let i=0;i<files.length;i++){
        const f=files[i];
        showUploadProgress(`${i+1}/${files.length}: ${f.name} yuklanyapti...`,Math.round(i/files.length*90));
        try{
          const info=await chunkedUpload(f,p=>showUploadProgress(`${i+1}/${files.length}: ${f.name} — ${Math.round(p*100)}%`,Math.round((i+p)/files.length*90)));
          const j=await api('/api/chunk_finalize',{method:'POST',body:JSON.stringify(info)});
          if(j.error)skipped.push(f.name+': '+j.error); else done++;
        }catch(err){skipped.push(f.name+': '+err.message)}
      }
      let skipTxt=skipped.length?` O\'tkazildi: ${skipped.join(', ')}`:'';
      $("bulkResult").textContent=`${done} ta fayl navbatga qo\'shildi.${skipTxt}`;
      showUploadProgress(`${done} ta fayl navbatga qo\'shildi`,95);
      await loadArchive();setTimeout(()=>showUploadProgress("",0),3000);
    }catch(err){$("status").textContent=err.message;showUploadProgress("✗ "+err.message,0)}
    finally{setBusy(btn,false)};
  };
  let tf=$("tolovUpload");
  if(tf)tf.onsubmit=async e=>{
    e.preventDefault();let btn=e.submitter;
    try{setBusy(btn,true,"Shakllanmoqda");showUploadProgress("To\'lovlar shakllantirilyapti...",10);
      let j=await api("/api/tolov",{method:"POST",body:new FormData(tf)});
      PAYMENTS=j.payments||[];
      $("tolovUploadResult").innerHTML=`Tayyor: ${fmtI(PAYMENTS.reduce((a,r)=>a+(+r.rows||0),0))} qator, ${fmtN(PAYMENTS.reduce((a,r)=>a+(+r.sum||0),0))} so\'m.`;
      showUploadProgress("To\'lovlar tayyor!",100);setTimeout(()=>showUploadProgress("",0),3000);
    }catch(err){$("status").textContent=err.message;showUploadProgress("✗ "+err.message,0)}
    finally{setBusy(btn,false)};
  };
};
// ===== /UPLOAD YANGILANMASI =====

function startBackgroundVideo(){
  const video=document.getElementById('bgVideo'), canvas=document.getElementById('bgCanvas');
  if(!video||!canvas||!canvas.captureStream)return;
  const ctx=canvas.getContext('2d');
  function resize(){const d=Math.min(1.25,Math.max(1,window.devicePixelRatio||1));canvas.width=Math.round(1280*d);canvas.height=Math.round(720*d)}
  resize(); window.addEventListener('resize',resize,{passive:true});
  const routes=[{x1:.08,y1:.68,x2:.44,y2:.38,x3:.82,y3:.58,delay:0},{x1:.18,y1:.28,x2:.54,y2:.18,x3:.92,y3:.34,delay:.22},{x1:.05,y1:.48,x2:.38,y2:.60,x3:.76,y3:.26,delay:.48},{x1:.30,y1:.82,x2:.62,y2:.50,x3:.98,y3:.72,delay:.68}];
  function plane(ctx,x,y,a,scale,alpha){ctx.save();ctx.translate(x,y);ctx.rotate(a);ctx.scale(scale,scale);ctx.globalAlpha=alpha;ctx.fillStyle='#1d77b8';ctx.strokeStyle='rgba(255,255,255,.75)';ctx.lineWidth=1.2;ctx.beginPath();ctx.moveTo(18,0);ctx.lineTo(-14,-6);ctx.lineTo(-8,0);ctx.lineTo(-14,6);ctx.closePath();ctx.fill();ctx.stroke();ctx.fillStyle='#42b9c7';ctx.fillRect(-20,-2,8,4);ctx.restore()}
  function bez(p0,p1,p2,t){let u=1-t;return {x:u*u*p0.x+2*u*t*p1.x+t*t*p2.x,y:u*u*p0.y+2*u*t*p1.y+t*t*p2.y}}
  function draw(t){
    const w=canvas.width,h=canvas.height, time=t/1000;
    let g=ctx.createLinearGradient(0,0,0,h);g.addColorStop(0,'#eaf8ff');g.addColorStop(.46,'#ffffff');g.addColorStop(1,'#e9f5fb');ctx.fillStyle=g;ctx.fillRect(0,0,w,h);
    ctx.save();ctx.globalAlpha=.34;ctx.strokeStyle='rgba(48,128,188,.16)';ctx.lineWidth=1;for(let x=((time*10)%64)-64;x<w;x+=64){ctx.beginPath();ctx.moveTo(x,0);ctx.lineTo(x+h*.15,h);ctx.stroke()}for(let y=((time*7)%56)-56;y<h;y+=56){ctx.beginPath();ctx.moveTo(0,y);ctx.lineTo(w,y+h*.04);ctx.stroke()}ctx.restore();
    ctx.save();ctx.globalCompositeOperation='multiply';for(let i=0;i<7;i++){let x=((time*18+i*260)%(w+360))-180,y=h*(.12+.13*Math.sin(time*.21+i));let rg=ctx.createRadialGradient(x,y,0,x,y,w*.17);rg.addColorStop(0,'rgba(198,228,244,.42)');rg.addColorStop(1,'rgba(198,228,244,0)');ctx.fillStyle=rg;ctx.fillRect(0,0,w,h)}ctx.restore();
    ctx.save();ctx.strokeStyle='rgba(31,127,190,.24)';ctx.lineWidth=2;ctx.setLineDash([10,12]);for(const r of routes){ctx.beginPath();ctx.moveTo(r.x1*w,r.y1*h);ctx.quadraticCurveTo(r.x2*w,r.y2*h,r.x3*w,r.y3*h);ctx.stroke()}ctx.restore();
    ctx.save();ctx.globalCompositeOperation='screen';for(let i=0;i<3;i++){let cx=w*(.16+i*.31),cy=h*(.34+.12*Math.sin(time*.18+i));let rr=((time*38+i*70)%210)+30;ctx.strokeStyle='rgba(67,178,198,'+(0.24-rr/1100)+')';ctx.lineWidth=2;ctx.beginPath();ctx.arc(cx,cy,rr,0,Math.PI*2);ctx.stroke()}ctx.restore();
    for(const r of routes){let tt=(time*.065+r.delay)%1,p0={x:r.x1*w,y:r.y1*h},p1={x:r.x2*w,y:r.y2*h},p2={x:r.x3*w,y:r.y3*h},p=bez(p0,p1,p2,tt),q=bez(p0,p1,p2,Math.min(.995,tt+.01));plane(ctx,p.x,p.y,Math.atan2(q.y-p.y,q.x-p.x),1.0+tt*.25,.75)}
    ctx.save();ctx.globalAlpha=.7;ctx.fillStyle='rgba(255,255,255,.46)';ctx.fillRect(0,0,w,h);ctx.restore();
    requestAnimationFrame(draw);
  }
  function attach(){const stream=canvas.captureStream(30);window.BG_STREAM=stream;video.muted=true;video.playsInline=true;video.autoplay=true;video.srcObject=stream;video.play().catch(()=>{})}
  requestAnimationFrame(draw);setTimeout(attach,120);
}

/* Excel-like table polish and sample-table renderers */
const tableCore=table;
table=function(h,rows,cls=""){rows=rows||[];return `<table class="${cls}"><colgroup>${h.map(x=>`<col style="${x.w?'width:'+x.w:''}">`).join("")}</colgroup><thead><tr>${h.map(x=>`<th class="${x.n?'num':'text'} col-${x.k}">${x.t}</th>`).join("")}</tr></thead><tbody>${rows.map((r,ri)=>`<tr class="${r._class||''}" onclick='detail(${JSON.stringify(r.key||{}).replaceAll("'","&#39;")})'>${h.map(x=>{let cls=(x.n?'num':'text')+` col-${x.k}`,raw=r[x.k],val=x.n&&(raw===""||raw===null||raw===undefined)?"":(x.f?x.f(raw):esc(raw)),tip=esc(raw);if(ri===0&&!x.n&&(val==="Jami"||String(val).startsWith("Jami ")))val="IBK bo'yicha Jami";if(x.k==="released_partiya"&&(+raw||0)===0&&((+r.released_qiymat||0)>0.005||(+r.released_vazn||0)>0.005)){cls+=" partial";tip="Qisman yechilgan";val="0"}return `<td class="${cls}" title="${tip}">${val}</td>`}).join("")}</tr>`).join("")}</tbody></table>`}
function excelExpiredPostRegime(){let rows=expiredPostRegimeRows();let body=rows.map(r=>`<tr onclick='detail(${JSON.stringify(r.key||{}).replaceAll("'","&#39;")})'><td class=text title="${esc(r.post||'')}">${esc(r.post||'')}</td><td class=num>${fmtI(r.jami_partiya)}</td><td class=num>${fmtN(r.jami_qiymat)}</td><td class=num>${fmtI(r.expired_partiya)}</td><td class=num>${fmtN(r.expired_qiymat)}</td><td class=num>${fmtN(r.ulush)}</td><td class=num>${fmtI(r.im70_partiya)}</td><td class=num>${fmtN(r.im70_qiymat||0)}</td><td class=num>${fmtI(r.im74_partiya)}</td><td class=num>${fmtN(r.im74_qiymat||0)}</td><td class=num>${fmtI(r.tr80_partiya)}</td><td class=num>${fmtN(r.tr80_qiymat||0)}</td></tr>`).join("");return `<table class="sample-like expired-sample"><thead><tr><th rowspan=4>Bojxona postlari</th><th colspan=2 rowspan=3>Jami nazoratdagi</th><th colspan=9>Rejim muddati o'tgan tovarlar</th></tr><tr><th colspan=3 rowspan=2>Jami</th><th colspan=6>Shundan</th></tr><tr><th colspan=2>Vaqtincha saqlash IM70</th><th colspan=2>Bojxona ombori IM74</th><th colspan=2>Tranzit TR80</th></tr><tr><th>Partiya</th><th>Qiymati,<br>ming doll.</th><th>Partiya</th><th>Qiymati,<br>ming doll.</th><th>Partiyadagi<br>ulushi (%)</th><th>Partiya</th><th>Qiymati,<br>ming doll.</th><th>Partiya</th><th>Qiymati,<br>ming doll.</th><th>Partiya</th><th>Qiymati,<br>ming doll.</th></tr></thead><tbody>${body}</tbody></table>`}
expiredTotalExcelTable=function(){return excelExpiredPostRegime()}
function excelFoodTable(){let rows=foodRows().map((r,i)=>Object.assign({rn:i?i:"Jami:"},r));let body=rows.map(r=>`<tr onclick='detail(${JSON.stringify(r.key||{}).replaceAll("'","&#39;")})'><td class=num>${esc(r.rn)}</td><td class=text title="${esc(r.name)}">${esc(r.name)}</td><td class=num>${fmtN(r.vazn)}</td><td class=num>${fmtN(r.qiymat)}</td><td class=num>${fmtN(r.over_vazn)}</td><td class=num>${fmtN(r.over_qiymat)}</td><td class=num>${fmtN(r.ulush)}</td></tr>`).join("");return `<table class="sample-like food-sample"><colgroup><col style="width:36px"><col style="width:auto"><col style="width:86px"><col style="width:110px"><col style="width:86px"><col style="width:110px"><col style="width:90px"></colgroup><thead><tr><th rowspan=2>№</th><th rowspan=2>Tovarlar turi</th><th rowspan=2>Vazni,<br>tn</th><th rowspan=2>Qiymati,<br>ming AQSH doll.</th><th colspan=2>Shundan, 3 oydan oshganlari</th><th rowspan=2>Qiymatdagi<br>ulushi (%)</th></tr><tr><th>Vazni,<br>tn</th><th>Qiymati,<br>ming AQSH doll.</th></tr></thead><tbody>${body}</tbody></table>`}
function excelRegimeYearTable(){let rows=DATA.regime_year_post||[],posts=[...new Set(rows.map(r=>r.post))],out=[];posts.forEach(p=>{let pr=rows.filter(r=>r.post===p);out.push({name:p,_class:"merged-row"});["IM70","IM74","TR80"].forEach(reg=>{let rr=pr.filter(x=>x.rejim===reg);if(!rr.length)return;out.push({name:reg,partiya:rr.reduce((a,x)=>a+(+x.partiya||0),0),vazn:rr.reduce((a,x)=>a+(+x.vazn||0),0),qiymat:rr.reduce((a,x)=>a+(+x.qiymat||0),0),tolov:rr.reduce((a,x)=>a+(+x.tolov||0),0),_class:"sub-total"});rr.forEach(x=>out.push({name:(x.yil||'-')+' yil',partiya:x.partiya,vazn:x.vazn,qiymat:x.qiymat,tolov:x.tolov,key:x.key}))})});let total={name:"IBK bo'yicha Jami",partiya:rows.reduce((a,x)=>a+(+x.partiya||0),0),vazn:rows.reduce((a,x)=>a+(+x.vazn||0),0),qiymat:rows.reduce((a,x)=>a+(+x.qiymat||0),0),tolov:rows.reduce((a,x)=>a+(+x.tolov||0),0),_class:"grand-total"};return table([{k:"name",t:"Ko'rsatkich nomi",w:"230px"},{k:"partiya",t:"Partiya soni",w:"90px",n:1,f:fmtI},{k:"vazn",t:"Vazni<br>(tn.)",w:"115px",n:1,f:fmtN},{k:"qiymat",t:"Qiymati<br>(ming $)",w:"126px",n:1,f:fmtN},{k:"tolov",t:"To'lov<br>(mln.so'm)",w:"132px",n:1,f:fmtN}],[total].concat(out),"sample-like regime-year-table")}
const oldTransportPanel=transportPanel;transportPanel=function(){return oldTransportPanel().replace("Deklaratsiya post kodi bo'yicha tahlil","Transport turi bo'yicha tahlil").replace("Post kodi","Jo'natuvchi post")}
const renderSampleTables=render;render=function(){renderSampleTables();let v=$("view");if(!v||!DATA)return;if(TAB==="rejim")v.innerHTML=`<div class=stack><div class=panel><h2>70-74-80: post, rejim va yillar kesimida</h2>${excelRegimeYearTable()}</div><div class=panel><h2>Rejimlar qiymat bo'yicha ulushi</h2>${bars(DATA.regimes||[],"rejim","qiymat",fmtN)}</div></div>`;if(TAB==="food")v.innerHTML=`<div class=panel><h2>Oziq-ovqatlar kesimida ${xls("food")}</h2>${excelFoodTable()}${bars(DATA.food||[],"name","qiymat",fmtN)}</div>`}
const oldOverviewPanels3=overviewPanels;overviewPanels=function(){return oldOverviewPanels3()}
document.body.classList.add("login-screen");setTimeout(()=>{let eye=document.querySelector(".eye-btn");if(eye)eye.innerHTML="&#128065;";},0);
const showLoginScreenBase=showLogin;showLogin=function(){
  document.body.classList.add("login-screen");document.body.classList.remove("logged-in");
  let login=$("login"),app=$("app"),dash=$("dash"),tabs=$("tabs"),view=$("view"),kpis=$("kpis"),actions=$("actions"),err=$("loginError"),meta=$("meta");
  if(login){login.classList.remove("hidden");login.style.display="grid";}
  if(app){app.classList.add("hidden");app.style.display="none";}
  if(dash){dash.classList.add("hidden");dash.style.display="none";}
  if(tabs)tabs.innerHTML="";if(view)view.innerHTML="";if(kpis)kpis.innerHTML="";if(actions)actions.innerHTML="";
  if(err)err.textContent="";if(meta)meta.textContent="Kirish kerak";
}
doLogin=async function(){
  let btn=$("loginBtn"),err=$("loginError");
  if(err)err.textContent="";
  try{
    setBusy(btn,true,"Kirish");
    let user=($("user")?.value||"").trim(),pass=$("pass")?.value||"";
    if(!user||!pass)throw Error("Login va parolni kiriting");
    let j=await api("/api/login",{method:"POST",body:JSON.stringify({user,pass})});
    TOKEN=j.token;localStorage.ibk_token=TOKEN;ME=j.user;
    await showApp();
  }catch(e){
    forceLoginView();
    activateLogin();
    if(err)err.textContent=(e&&e.message&&e.message!=="login")?e.message:"Login yoki parol xato";
  }finally{setBusy(btn,false)}
}
function forgotPassword(){
  let err=$("loginError"),u=($("user")?.value||"").trim();
  if(err)err.textContent=u?`${u} uchun vaqtinchalik parol SMS orqali yuborish moduli tayyorlanmoqda. SMS provayder ulangandan keyin ishlaydi.`:"Avval loginni kiriting, keyin SMS orqali tiklash so'rovi yuboriladi.";
}
const showAppStableBase=showApp;
showApp=async function(){
  document.body.classList.remove("login-screen");
  document.body.classList.add("logged-in");
  let loginEl=$("login"),appEl=$("app"),dashEl=$("dash"),tabsEl=$("tabs"),viewEl=$("view"),kpisEl=$("kpis"),actionsEl=$("actions"),metaEl=$("meta");
  if(loginEl){loginEl.classList.add("hidden");loginEl.style.display="none";}
  if(appEl){appEl.classList.remove("hidden");appEl.style.display="block";}
  if(dashEl){dashEl.classList.remove("hidden");dashEl.style.display="block";}
  ME=await api("/api/me");
  LANG=ME.lang||localStorage.ibk_lang||"uz";
  await loadArchive();
  await loadPayments();
  DATA=null;TAB="home";GROUP="home";
  if(metaEl)metaEl.textContent="Tizimga xush kelibsiz";
  if(kpisEl)kpisEl.innerHTML="";
  if(tabsEl)tabsEl.innerHTML=`<button class="module-parent" onclick="openGroup('bnrte','umumiy')"><span>▦</span>BNRTE</button><button class="module-parent pay" onclick="openGroup('payments','payments')"><span>$</span>To'lovlar</button><button class="module-parent" onclick="openGroup('common','upload')"><span>⚙</span>Boshqaruv</button>`;
  if(actionsEl)actionsEl.innerHTML=`<button class="light lang-btn" onclick="setLang('uz')">O'zb</button><button class="light lang-btn" onclick="setLang('uzc')">Ўзб</button><button class="light lang-btn" onclick="setLang('ru')">Рус</button><button class="light lang-btn" onclick="document.body.classList.toggle('dark')">◐</button> <button class="logout-btn" onclick="logout()">Chiqish</button>`;
  if(viewEl) viewEl.innerHTML=landingPanel();
  setTimeout(()=>{initCountryFlowMap();},200);
}
let AUTO_LOGOUT_MS=20*60*1000,autoLogoutTimer=null;
function resetAutoLogout(){clearTimeout(autoLogoutTimer);if(TOKEN)autoLogoutTimer=setTimeout(()=>{logout();let e=$("loginError");if(e)e.textContent="20 daqiqa faollik bo'lmagani uchun qayta kirish kerak."},AUTO_LOGOUT_MS)}
["click","keydown","mousemove","touchstart","scroll"].forEach(ev=>document.addEventListener(ev,resetAutoLogout,{passive:true}));
const showAppAutoLogout=showApp;showApp=async function(){await showAppAutoLogout();resetAutoLogout()}
const logoutAutoBase=logout;logout=function(){clearTimeout(autoLogoutTimer);logoutAutoBase();}
const renderRealGlobe=render;render=function(){renderRealGlobe();setTimeout(()=>{initCountryFlowMap();},80)}
document.body.classList.remove("bg-aero","bg-classic");localStorage.removeItem("ibk_bg");
startBackgroundVideo();
function releaseDatePair(id,title){return `<div class="release-date-card"><h3>${title}</h3><div class="filters compact-filters"><select id="${id}Base">${dateOptions()}</select><select id="${id}Final">${dateOptions()}</select><button onclick="buildReleaseSection('${id}')">Shakllantirish</button></div><div id="${id}Result" class="release-section-result"></div></div>`}
function releaseDashboardPanel(){return `<div class="stack"><div class="panel"><h2>Nazoratdan yechish</h2><div class="overview-note">Tepadagi sanalar barcha jadvallar uchun umumiy ishlashi yoki har jadval alohida muddat bilan shakllanishi mumkin.</div><div class="filters compact-filters"><label class="inline-check"><input type="checkbox" id="relGlobalUse" checked onchange="syncReleaseDates()"> Barcha jadvallar uchun</label><select id="relGlobalBase" onchange="syncReleaseDates()">${dateOptions()}</select><select id="relGlobalFinal" onchange="syncReleaseDates()">${dateOptions()}</select><button onclick="buildAllReleaseSections()">Barchasini shakllantirish</button></div></div>${releaseDatePair('relMain','Nazoratdan yechilishi jadvali')}${releaseDatePair('relSpeed','Korxonalar: eng ko\'p, eng tez va eng sekin')}${releaseDatePair('relWh','Omborlar oboroti')}</div>`}
function syncReleaseDates(){let use=$('relGlobalUse')?.checked,base=$('relGlobalBase')?.value,final=$('relGlobalFinal')?.value;['relMain','relSpeed','relWh'].forEach(id=>{let b=$(`${id}Base`),f=$(`${id}Final`);if(!b||!f)return;if(use){b.value=base;f.value=final;b.disabled=true;f.disabled=true}else{b.disabled=false;f.disabled=false}})}
function releaseDatesFor(id){let use=$('relGlobalUse')?.checked;if(use)return {base:$('relGlobalBase').value,final:$('relGlobalFinal').value};return {base:$(`${id}Base`).value,final:$(`${id}Final`).value}}
async function loadReleaseData(base,final,target){if(parseUzDate(base)>=parseUzDate(final)){target.innerHTML=`<div class=muted>Boshlang'ich sana yakuniy sanadan oldin bo'lishi kerak.</div>`;return null}let j=await api(`/api/release?base=${encodeURIComponent(base)}&final=${encodeURIComponent(final)}`);if(j.missing){target.innerHTML=`<div class=muted>Arxivda yo'q sana: ${j.missing.join(', ')}. Shu davr uchun asos fayl yuklash kerak.</div>`;return null}window.LAST_RELEASE=j;return j}
async function buildReleaseSection(id){let d=releaseDatesFor(id),box=$(`${id}Result`);if(!box)return;box.innerHTML='<div class=muted>Hisoblanmoqda...</div>';let j=await loadReleaseData(d.base,d.final,box);if(!j)return;let companies=releaseCompanyRows(j);if(id==='relMain'){let rows=numbered([releaseTotalRow(companies)].concat(companies));box.innerHTML=`<h2>${d.base} - ${d.final} <a class="btn light" href="/api/export?kind=release&base=${d.base}&final=${d.final}&token=${TOKEN}">Excel</a></h2><div class=overview-note>Boshlang'ich davr: ${d.base}. Yakuniy davr: ${d.final}.</div>${table(releaseCols(d.base,d.final),rows,'release-table sample-release-table fixed-table')}`;return}if(id==='relSpeed'){box.innerHTML=releaseSpeedPanels(companies,d.base,d.final);return}if(id==='relWh'){box.innerHTML=warehouseTurnoverPanel(j);return}}
async function buildAllReleaseSections(){syncReleaseDates();for(const id of ['relMain','relSpeed','relWh']) await buildReleaseSection(id)}
const renderReleaseDateControls=render;render=function(){renderReleaseDateControls();if(TAB==='released'){let v=$('view');if(v){v.innerHTML=releaseDashboardPanel();setTimeout(syncReleaseDates,0)}}}
/* === BOOT FIX: tokenni o'chirib tashlamaymiz; holatni to'g'ri ochamiz === */
function forceLoginView(){
  document.body.classList.add("login-screen");
  document.body.classList.remove("logged-in");
  let login=$("login"), app=$("app"), dash=$("dash"), tabs=$("tabs"), view=$("view"), kpis=$("kpis"), actions=$("actions"), meta=$("meta");
  if(login){login.classList.remove("hidden");login.style.display="grid";}
  if(app){app.classList.add("hidden");app.style.display="none";}
  if(dash){dash.classList.add("hidden");dash.style.display="none";}
  if(tabs)tabs.innerHTML="";
  if(view)view.innerHTML="";
  if(kpis)kpis.innerHTML="";
  if(actions)actions.innerHTML="";
  if(meta)meta.textContent="Kirish kerak";
}

function forceAppView(){
  document.body.classList.remove("login-screen");
  document.body.classList.add("logged-in");
  let login=$("login"), app=$("app"), dash=$("dash");
  if(login){login.classList.add("hidden");login.style.display="none";}
  if(app){app.classList.remove("hidden");app.style.display="block";}
  if(dash){dash.classList.remove("hidden");dash.style.display="block";}
}

const showLoginFinalBase = showLogin;
showLogin = function(){
  let p=$("pass"); if(p) p.value='';
  forceLoginView();
}

const showAppFinalBase = showApp;
showApp = async function(){
  forceAppView();
  await showAppFinalBase();
  forceAppView();
  window.scrollTo({top:0, behavior:"auto"});
  loadExchangeRates();
}

/* doLogin final version is set above */

forceLoginView();
/* Hex monitoring background */
(function(){
  var sc=document.querySelector('.sky-scene');
  if(!sc)return;
  /* Kill all pseudo-element patterns that bleed through the canvas */
  var ks=document.createElement('style');
  ks.textContent=
    'body .sky-scene::before,body .sky-scene::after{display:none!important;content:none!important;animation:none!important;background:none!important;opacity:0!important;visibility:hidden!important}'+
    '.sky-scene::before,.sky-scene::after{display:none!important;content:none!important;animation:none!important;background:none!important;opacity:0!important;visibility:hidden!important}'+
    '.login-screen .sky-scene::before,.login-screen .sky-scene::after{display:none!important;content:none!important;background:none!important;opacity:0!important;visibility:hidden!important}'+
    '.logged-in .sky-scene::before,.logged-in .sky-scene::after{display:none!important;content:none!important;background:none!important;opacity:0!important}';
  document.head.appendChild(ks);
  /* Remove every old sky-scene child — eliminates all CSS specificity fights */
  while(sc.firstChild)sc.removeChild(sc.firstChild);
  /* Force full-screen via setProperty so !important beats any stylesheet rule */
  var sp=sc.style;
  ['position:fixed','top:0','left:0','right:0','bottom:0','width:100%','height:100%',
   'z-index:0','overflow:hidden','pointer-events:none','background:#f8fbff',
   'animation:none','transform:none'].forEach(function(r){
    var i=r.indexOf(':');sp.setProperty(r.slice(0,i),r.slice(i+1),'important');
  });
  /* Build hex canvas */
  var cv=document.createElement('canvas');
  cv.id='hexBg';
  sp=cv.style;
  ['position:absolute','top:0','left:0','width:100%','height:100%',
   'display:block','z-index:1','pointer-events:none'].forEach(function(r){
    var i=r.indexOf(':');sp.setProperty(r.slice(0,i),r.slice(i+1),'important');
  });
  sc.appendChild(cv);
  var ctx=cv.getContext('2d');
  var R=28,W=0,H=0,cells=[];
  function dk(){return document.body.classList.contains('dark');}
  function resize(){
    W=window.innerWidth;H=window.innerHeight;
    cv.width=W;cv.height=H;
    var cols=Math.ceil(W/(R*1.73))+2,rows=Math.ceil(H/(R*1.5))+2;
    cells=[];
    for(var row=0;row<rows;row++)
      for(var col=0;col<cols;col++)
        cells.push({x:col*R*1.73+(row%2)*R*.87,y:row*R*1.5,
                    p:Math.random()*Math.PI*2,s:.028+Math.random()*.032});
  }
  function hx(x,y){
    ctx.beginPath();
    for(var i=0;i<6;i++){var a=Math.PI/3*i-Math.PI/6;
      ctx.lineTo(x+R*Math.cos(a),y+R*Math.sin(a));}
    ctx.closePath();
  }
  function draw(){
    var d=dk();
    ctx.clearRect(0,0,W,H);
    ctx.fillStyle=d?'#080f1e':'#f8fbff';ctx.fillRect(0,0,W,H);
    for(var i=0;i<cells.length;i++){
      var c=cells[i];c.p+=c.s;
      var v=(Math.sin(c.p)+1)/2;
      hx(c.x,c.y);
      ctx.fillStyle=d?('rgba(59,130,246,'+(v*.32).toFixed(3)+')')
                     :('rgba(100,160,220,'+(v*.2).toFixed(3)+')');
      ctx.fill();
      ctx.strokeStyle=d?('rgba(96,165,250,'+(.1+v*.22).toFixed(3)+')')
                       :('rgba(100,160,220,'+(.05+v*.1).toFixed(3)+')');
      ctx.lineWidth=0.5;ctx.stroke();
    }
    requestAnimationFrame(draw);
  }
  resize();window.addEventListener('resize',resize,{passive:true});draw();
})();

/* ===== AVIA AWB MODULE ===== */
function aviaPanel(){
  return `<div class=stack><div class=panel><h2>✈ AVIA AWB — Nazoratdagi havo yuklar</h2><div class=muted>Havo yuki AWB (Air Waybill) ro'yxati. Har bir AWB bir necha reysda bo'linishi mumkin — ular yig'ilib ko'rsatiladi. 15 kun o'tgan AWBlar muddati o'tgan hisoblanadi.</div></div><div id=aviaStatsContainer><div class=muted style="padding:16px">Statistika yuklanmoqda...</div></div><div id=aviaContainer><div class=muted style="padding:16px">AWB ro'yxati yuklanmoqda...</div></div></div>`;
}

async function loadYaroqlilik(){
  if(YAROQLILIK_DATA&&YAROQLILIK_DATA.loaded)return;
  try{
    let j=await api('/api/yaroqlilik');
    YAROQLILIK_DATA=j;
    if(TAB==='yaroqlilik')render();
  }catch(e){}
}

async function loadAviaAwb(){
  let box=$('aviaContainer'), sbox=$('aviaStatsContainer');
  if(!box&&!sbox)return;
  try{
    let [j, st]=await Promise.all([api('/api/avia_awb'), api('/api/avia_stats').catch(()=>null)]);
    AVIA_DATA=j; if(st)AVIA_STATS=st;
    if(sbox)sbox.innerHTML=st?renderAviaStats(st):'';
    if($('kpis'))$('kpis').innerHTML=renderKpis();
    if(!j.loaded){
      if(box)box.innerHTML=`<div class="panel"><div class=muted>${esc(j.error||"Ma'lumot topilmadi")}<br><br><button class=light onclick="GROUP='common';TAB='upload';render()">Boshqaruv → Fayl yuklashga o'tish</button></div></div>`;
      return;
    }
    if(box)box.innerHTML=renderAviaContent(j);
  }catch(e){
    if(box)box.innerHTML=`<div class=panel><div class=muted>Xatolik: ${esc(e.message||String(e))}</div></div>`;
  }
}

function renderAviaStats(st){
  if(!st||!st.jami_qiymat_k)return'';
  let oyCols=[{k:'oy',t:'Oy',w:'18%'},{k:'qiymat_k',t:'Qiymat (ming $)',n:1,f:fmtN},{k:'vazn_tn',t:'Vazn (tn)',n:1,f:fmtN}];
  let compCols=[{k:'company',t:'Korxona',w:'50%'},{k:'stir',t:'STIR',w:'12%'},{k:'qiymat_k',t:'Qiymat (ming $)',n:1,f:fmtN},{k:'vazn_tn',t:'Vazn (tn)',n:1,f:fmtN}];
  let cntCols=[{k:'country',t:'Davlat',w:'30%'},{k:'qiymat_k',t:'Qiymat (ming $)',n:1,f:fmtN},{k:'vazn_tn',t:'Vazn (tn)',n:1,f:fmtN}];
  let oyRows=(st.by_month||[]).map(r=>Object.assign({key:{}},r));
  let compRows=(st.by_company||[]).map(r=>Object.assign({key:{}},r));
  let cntRows=(st.by_country||[]).map(r=>Object.assign({key:{}},r));
  let compTotal={key:{},company:'Jami (TOP 20)',stir:'',qiymat_k:compRows.reduce((a,r)=>a+(r.qiymat_k||0),0),vazn_tn:compRows.reduce((a,r)=>a+(r.vazn_tn||0),0)};
  let cntTotal={key:{},country:'Jami (TOP 15)',qiymat_k:cntRows.reduce((a,r)=>a+(r.qiymat_k||0),0),vazn_tn:cntRows.reduce((a,r)=>a+(r.vazn_tn||0),0)};
  return `<div class="panel" style="border-top:3px solid #0ea5e9"><h2>✈ BNRTE Avia statistika (ming $)</h2>
    <div class="kpis" style="margin-bottom:12px">
      <div class=kpi style="border-top:3px solid #0ea5e9"><span>Jami qiymat</span><b style="color:#0ea5e9">${fmtN(st.jami_qiymat_k)}</b><div style="font-size:11px;color:var(--muted);margin-top:4px">ming $</div></div>
      <div class=kpi><span>Jami vazn</span><b>${fmtN(st.jami_vazn_tn)}</b><div style="font-size:11px;color:var(--muted);margin-top:4px">tonna</div></div>
    </div>
    <div class=grid2>
      <div class=panel><h2>Oylar bo'yicha (ming $)</h2>${table(oyCols,oyRows)}</div>
      <div class=panel><h2>Davlatlar bo'yicha (ming $) — TOP 15</h2>${table(cntCols,[cntTotal].concat(cntRows))}</div>
      <div class="panel wide"><h2>Korxonalar bo'yicha (ming $) — TOP 20</h2>${table(compCols,[compTotal].concat(compRows))}</div>
    </div>
  </div>`;
}

async function uploadAviaAwb(file){
  if(!file)return;
  let box=$('aviaContainer');
  if(box)box.innerHTML='<div class=panel><div class=muted>Yuklanmoqda...</div></div>';
  let fd=new FormData();fd.append('file',file);
  try{
    let j=await api('/api/upload_avia_awb',{method:'POST',body:fd});
    AVIA_DATA=j;
    if(box)box.innerHTML=j.loaded?renderAviaContent(j):`<div class=panel><div class=muted>Xatolik: ${esc(j.error||'Yuklanmadi')}</div></div>`;
    if(j.loaded&&$('kpis'))$('kpis').innerHTML=renderKpis();
  }catch(e){
    if(box)box.innerHTML=`<div class=panel><div class=muted>Xatolik: ${esc(e.message||String(e))}</div></div>`;
  }
}
async function uploadAviaAwbDirect(file){
  if(!file)return;
  let st=$('aviaUploadStatus');
  if(st)st.textContent='Yuklanmoqda...';
  let fd=new FormData();fd.append('file',file);
  try{
    let j=await api('/api/upload_avia_awb',{method:'POST',body:fd});
    AVIA_DATA=j;
    if(st)st.textContent=j.loaded?`Tayyor: ${fmtI(j.unique_awb)} AWB yuklandi`:`Xatolik: ${esc(j.error||'Yuklanmadi')}`;
  }catch(e){
    if(st)st.textContent='Xatolik: '+String(e);
  }
}

function renderAviaContent(j){
  let kpiHtml=`<div class=kpis style="grid-column:1/-1">
    <div class=kpi><span>Jami unikal AWB</span><b>${fmtI(j.unique_awb)}</b></div>
    <div class=kpi><span>Jami joylar</span><b>${fmtI(j.total_joylar)}</b></div>
    <div class=kpi><span>Jami vazn (tn)</span><b>${fmtN(j.total_vazn)}</b></div>
    <div class=kpi><span style="color:#16a34a">Muddati o'tmagan</span><b style="color:#16a34a">${fmtI(j.active_count)}</b><div style="font-size:11px;color:var(--muted);margin-top:4px">${fmtN(j.active_vazn)} tn</div></div>
    <div class=kpi style="border-color:#dc2626"><span style="color:#dc2626">Muddati o'tgan (15 kun+)</span><b style="color:#dc2626">${fmtI(j.overdue_count)}</b><div style="font-size:11px;color:var(--muted);margin-top:4px">${fmtN(j.overdue_vazn)} tn</div></div>
    <div class=kpi><span>Ko'p reysli AWB</span><b>${fmtI(j.multi_flight_awb)}</b><div style="font-size:11px;color:var(--muted);margin-top:4px">2+ reys bilan yetkazilgan</div></div>
  </div>`;
  let compCols=[{k:'company',t:'Qabul qiluvchi korxona',w:'44%'},{k:'awb',t:'AWB',n:1,f:fmtI,w:'8%'},{k:'joylar',t:'Joylar',n:1,f:fmtI,w:'8%'},{k:'vazn',t:'Vazn (tn)',n:1,f:fmtN,w:'12%'},{k:'overdue',t:'Muddati o\'tgan',n:1,f:fmtI,w:'12%'}];
  let compRows=(j.companies||[]).slice(0,30).map(r=>Object.assign({_class:r.overdue>0?'expired-row':'',key:{}},r));
  let compTotal={key:{},company:'Jami (TOP 30)',awb:compRows.reduce((a,r)=>a+(r.awb||0),0),joylar:compRows.reduce((a,r)=>a+(r.joylar||0),0),vazn:compRows.reduce((a,r)=>a+(r.vazn||0),0),overdue:compRows.reduce((a,r)=>a+(r.overdue||0),0),_class:''};
  let awbCols=[{k:'awb',t:'AWB raqami',w:'16%'},{k:'flights',t:'Reys',n:1,f:fmtI,w:'6%'},{k:'company',t:'Qabul qiluvchi',w:'26%'},{k:'country_latin',t:'Davlat',w:'12%'},{k:'joylar',t:'Joylar',n:1,f:fmtI,w:'6%'},{k:'vazn',t:'Vazn (tn)',n:1,f:fmtN,w:'10%'},{k:'arrival_date',t:'Kelgan sana',w:'10%'},{k:'status_label',t:'Holat',w:'12%'}];
  let allAwbRows=(j.awb_list||[]).map(r=>Object.assign({status_label:r.is_overdue?'🔴 O\'tgan':'🟢 Kuzatuvda',_class:r.is_overdue?'expired-row':'',key:{}},r));
  let overdueAwbRows=allAwbRows.filter(r=>r.is_overdue);
  let countries15=topNWithOther(j.countries||[],15,'vazn','name');
  let countryHtml=`<div class=panel><h2>Davlatlar bo'yicha (vazn, tn) — TOP 15</h2>${bars(countries15,'name','vazn',fmtN)}</div>`;
  let overdueTableHtml=overdueAwbRows.length?table(awbCols,overdueAwbRows,'fixed-table'):'<div class=muted style="padding:12px">Muddati o\'tgan AWB yo\'q.</div>';
  return `${awbRatingPanel(j)}<div class=grid2>
    ${kpiHtml}
    ${countryHtml}
    <div class="panel wide"><h2>Korxonalar bo'yicha (TOP 30)</h2>${table(compCols,[compTotal].concat(compRows),'fixed-table')}</div>
    <div class="panel wide"><h2>Muddati o'tgan AWBlar — ${fmtI(overdueAwbRows.length)} ta</h2>${overdueTableHtml}</div>
  </div>`;
}

/* ===== VALYUTA KURSLARI ===== */
let EXCHANGE_RATES=null;
async function loadExchangeRates(force=false){
  if(EXCHANGE_RATES&&!force)return;
  try{
    let j=await api('/api/exchange_rates');
    EXCHANGE_RATES=j;
    renderCurrencyWidget(j);
  }catch(e){}
}

function renderCurrencyWidget(j){
  let el=$('currencyWidget');
  if(!el||!j||!j.rates)return;
  let keys=['USD','EUR','RUB','CNY','GBP'];
  let html=keys.map(k=>{let r=j.rates[k];return r?`<span class="cur-item"><b>${k}</b> ${fmtI(Math.round(r.rate))}</span>`:''}).filter(Boolean).join('');
  el.innerHTML=`<span class=cur-date>${j.date||''}</span> ${html}`;
}
</script></main></body></html>"""


class Handler(BaseHTTPRequestHandler):
    def _cors_origin(self) -> str:
        origin = self.headers.get("Origin", "")
        return origin if origin in ALLOWED_ORIGINS else ""

    def json(self, data, status=HTTPStatus.OK, cors=False):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        if cors:
            allowed = self._cors_origin()
            if allowed:
                self.send_header("Access-Control-Allow-Origin", allowed)
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        parsed = urlparse(self.path)
        if parsed.path in ("/api/chunk_upload", "/api/chunk_finalize", "/api/server_info"):
            allowed = self._cors_origin()
            self.send_response(HTTPStatus.OK)
            if allowed:
                self.send_header("Access-Control-Allow-Origin", allowed)
            self.send_header("Access-Control-Allow-Methods", "POST, GET, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "X-Token, X-Upload-Id, X-Chunk-Index, X-Total-Chunks, X-Filename, Content-Type")
            self.send_header("Content-Length", "0")
            self.end_headers()
        else:
            self.send_error(405)

    def body_json(self):
        return json.loads(self.rfile.read(int(self.headers.get("Content-Length", "0") or "0")).decode("utf-8") or "{}")

    def token(self):
        parsed = urlparse(self.path)
        q = parse_qs(parsed.query)
        return self.headers.get("X-Token") or q.get("token", [""])[0]

    def user(self):
        token = self.token()
        sess = SESSIONS.get(token)
        return sess["user"] if sess else None

    def require_user(self):
        user = self.user()
        if not user:
            self.json({"error": "login kerak"}, HTTPStatus.UNAUTHORIZED)
            return None
        return user

    def require_admin(self):
        user = self.require_user()
        if not user:
            return None
        users = load_json(USER_PATH, {})
        rec = users.get(user, {})
        if rec.get("role") != "admin" and "admin" not in rec.get("perms", []):
            self.json({"error": "admin huquqi kerak"}, HTTPStatus.FORBIDDEN)
            return None
        return user

    def require_perm(self, perm: str):
        user = self.require_user()
        if not user:
            return None
        rec = load_json(USER_PATH, {}).get(user, {})
        if rec.get("role") == "admin" or perm in rec.get("perms", []):
            return user
        self.json({"error": f"{perm} vakolati kerak"}, HTTPStatus.FORBIDDEN)
        return None

    def do_GET(self):
        global STORE
        parsed = urlparse(self.path)
        if parsed.path.startswith("/assets/"):
            name = Path(unquote(parsed.path.rsplit("/", 1)[-1])).name
            path = ASSET_DIR / name
            if not path.exists() or not path.is_file():
                self.send_error(404)
                return
            data = path.read_bytes()
            ext = path.suffix.lower()
            ctype = (
                "text/html; charset=utf-8" if ext in {".html", ".htm"} else
                "text/javascript; charset=utf-8" if ext in {".js"} else
                "text/css; charset=utf-8" if ext in {".css"} else
                "image/jpeg" if ext in {".jpg", ".jpeg"} else
                "image/png" if ext in {".png"} else
                "application/octet-stream"
            )
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if parsed.path == "/":
            body = HTML.encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if parsed.path == "/api/flights":
            if not self.require_user():
                return
            self.json(get_tas_flights())
            return
        if parsed.path == "/api/flights_live":
            if not self.require_user():
                return
            self.json(get_live_states())
            return
        if parsed.path == "/api/company_trends":
            if not self.require_user():
                return
            try:
                if STORE is None:
                    STORE = IBKStore(DB_PATH)
                ensure_store_backfilled()
                self.json(STORE.company_series_all())
            except Exception as exc:
                self.json({"periods": [], "companies": [], "error": str(exc)})
            return
        if parsed.path == "/api/goods_trends":
            if not self.require_user():
                return
            try:
                if STORE is None:
                    STORE = IBKStore(DB_PATH)
                ensure_store_backfilled()
                self.json(STORE.goods_series_all())
            except Exception as exc:
                self.json({"periods": [], "goods": [], "error": str(exc)})
            return
        if parsed.path == "/api/warehouse_trends":
            if not self.require_user():
                return
            try:
                if STORE is None:
                    STORE = IBKStore(DB_PATH)
                ensure_store_backfilled()
                self.json(STORE.warehouse_series_all())
            except Exception as exc:
                self.json({"periods": [], "warehouses": [], "error": str(exc)})
            return
        if parsed.path == "/api/transport_trends":
            if not self.require_user():
                return
            try:
                if STORE is None:
                    STORE = IBKStore(DB_PATH)
                ensure_store_backfilled()
                self.json(STORE.transport_series_all())
            except Exception as exc:
                self.json({"periods": [], "transports": [], "error": str(exc)})
            return
        if parsed.path == "/api/transport_company_trends":
            if not self.require_user():
                return
            try:
                if STORE is None:
                    STORE = IBKStore(DB_PATH)
                ensure_store_backfilled()
                self.json(STORE.transport_company_series_all())
            except Exception as exc:
                self.json({"periods": [], "transports": [], "error": str(exc)})
            return
        if parsed.path == "/api/country_flow":
            if not self.require_user():
                return
            try:
                if STORE is None:
                    STORE = IBKStore(DB_PATH)
                q = parse_qs(parsed.query)
                dates = STORE.available_dates()
                date_text = q.get("date", [""])[0].strip()
                if not date_text and dates:
                    date_text = dates[0]
                countries = STORE.country_flow_by_date(date_text) if date_text else []
                countries = [dict(c, name=core.to_latin(c['name'])) for c in countries]
                self.json({"countries": countries, "date": date_text, "available": dates})
            except Exception as exc:
                self.json({"countries": [], "date": "", "available": [], "error": str(exc)})
            return
        if parsed.path == "/api/country_transport":
            if not self.require_user():
                return
            try:
                if STORE is None:
                    STORE = IBKStore(DB_PATH)
                tr = STORE.country_transport_summary()
                self.json({core.to_latin(k): v for k, v in tr.items()})
            except Exception as exc:
                self.json({"error": str(exc)})
            return
        if parsed.path == "/api/warehouses":
            if not self.require_user():
                return
            try:
                self.json(load_warehouse_registry())
            except Exception as exc:
                self.json({"error": str(exc), "warehouses": []})
            return
        if parsed.path == "/api/yaroqlilik":
            if not self.require_user():
                return
            self.json(load_yaroqlilik_cached())
            return
        if parsed.path == "/api/avia_awb":
            if not self.require_user():
                return
            try:
                global AVIA_AWB_CACHE
                if "force=1" in (parsed.query or ""):
                    AVIA_AWB_CACHE = None
                self.json(AVIA_AWB_CACHE if AVIA_AWB_CACHE is not None else load_avia_awb())
            except Exception as exc:
                self.json({"error": str(exc), "loaded": False})
            return
        if parsed.path == "/api/avia_stats":
            if not self.require_user():
                return
            try:
                stats = STORE.avia_db_stats()
                stats["by_country"] = [
                    dict(r, country=core.to_latin(r["country"])) for r in stats["by_country"]
                ]
                self.json(stats)
            except Exception as exc:
                self.json({"error": str(exc), "decl_soni": 0, "jami_qiymat_k": 0})
            return
        if parsed.path == "/api/exchange_rates":
            try:
                self.json(fetch_cbu_rates())
            except Exception as exc:
                self.json({"success": False, "error": str(exc), "rates": {}, "date": ""})
            return
        if parsed.path == "/api/tolov":
            if not self.require_user():
                return
            self.json({"payments": tolov_summary_rows()})
            return
        if parsed.path == "/api/archive":
            if not self.require_user():
                return
            archive = load_json(INDEX_PATH, {"reports": []})
            archive["reports"] = clean_archive_records(archive.get("reports", []))
            save_json(INDEX_PATH, archive)
            self.json(archive)
            return
        if parsed.path == "/api/me":
            user = self.require_user()
            if not user:
                return
            rec = load_json(USER_PATH, {}).get(user, {})
            self.json({"user": user, "role": rec.get("role", "user"), "full_name": rec.get("full_name", ""), "position": rec.get("position", ""), "phone": rec.get("phone", ""), "lang": rec.get("lang", "uz"), "perms": rec.get("perms", [])})
            return
        if parsed.path == "/api/ui_config":
            self.json(load_json(UI_CONFIG_PATH, {}))
            return
        if parsed.path == "/api/server_info":
            self.json({"lan_url": f"http://{LAN_IP}:{PORT}", "port": PORT}, cors=True)
            return
        if parsed.path == "/api/users":
            if not self.require_admin():
                return
            users = load_json(USER_PATH, {})
            self.json({"users": [{"user": u, "role": r.get("role", "user"), "role_label": r.get("role_label", ROLE_LABELS.get(r.get("role", "user"), "Foydalanuvchi")), "post_code": r.get("post_code", ""), "enabled": r.get("enabled", True), "full_name": r.get("full_name", ""), "position": r.get("position", ""), "phone": r.get("phone", ""), "lang": r.get("lang", "uz"), "perms": r.get("perms", [])} for u, r in users.items()]})
            return
        if parsed.path.startswith("/api/jobs/"):
            if not self.require_user():
                return
            self.json(JOBS.get(parsed.path.rsplit("/", 1)[-1], {"status": "xatolik", "error": "Job topilmadi"}))
            return
        if parsed.path.startswith("/api/reports/"):
            if not self.require_user():
                return
            report_id = parsed.path.rsplit("/", 1)[-1]
            item = next((r for r in load_json(INDEX_PATH, {"reports": []})["reports"] if r["id"] == report_id), None)
            if not item:
                self.json({"error": "Topilmadi"}, HTTPStatus.NOT_FOUND)
                return
            data_path = Path(item["dir"]) / "dashboard.json"
            data = load_json(data_path, {})
            needs_rebuild = data and (
                data.get("schema") != 5 or
                "food" not in data or "post_summary" not in data or "regime_posts" not in data or "regime_year_post" not in data or "expired_post_regime" not in data or "expired_block" not in data or "warehouse" not in data or
                "vazn" not in data.get("kpis", {}) or "expired_value" not in data.get("kpis", {}) or
                "depozit_matched" not in data.get("kpis", {})
            )
            if needs_rebuild:
                try:
                    deposit_path = Path(item["deposit"]) if item.get("deposit") else None
                    rebuilt = build_dashboard(item["id"], Path(item["source"]), deposit_path, datetime.strptime(item["date"], "%d.%m.%Y"))
                    rebuilt["files"] = data.get("files") or {"status": "tayyorlanmoqda", "excel": "", "pdf": "", "pngs": []}
                    data = rebuilt
                    save_json(data_path, data)
                except Exception:
                    pass
            if data:
                files = data.setdefault("files", {})
                if files.get("status") == "tayyorlanmoqda" or not files.get("excel"):
                    try:
                        files = update_report_files(report_id, Path(item["dir"]), datetime.strptime(item["date"], "%d.%m.%Y"))
                        data["files"] = files
                    except Exception:
                        pass
                if files.get("excel") and (files.get("pdf") or files.get("pngs")) and not files.get("status"):
                    files["status"] = "tayyor"
                k = data.setdefault("kpis", {})
                exp = data.get("expired", [])
                k.setdefault("vazn", sum(x.get("vazn", 0) for x in data.get("regimes", [])))
                k.setdefault("expired_value", sum(x.get("qiymat", 0) for x in exp))
                if "summary" not in data:
                    data["summary"] = [{"name": "IBK bo'yicha Jami", "partiya": k.get("partiya", 0), "vazn": k.get("vazn", 0), "qiymat": k.get("qiymat", 0), "tolov": k.get("tolov", 0)}] + [
                        {"name": r.get("rejim", ""), "partiya": r.get("partiya", 0), "vazn": r.get("vazn", 0), "qiymat": r.get("qiymat", 0), "tolov": r.get("tolov", 0)}
                        for r in data.get("regimes", [])
                    ]
                if "expired_summary" not in data:
                    data["expired_summary"] = [{"name": "IBK bo'yicha Jami", "partiya": sum(x.get("partiya", 0) for x in exp), "vazn": sum(x.get("vazn", 0) for x in exp), "qiymat": sum(x.get("qiymat", 0) for x in exp), "tolov": sum(x.get("tolov", 0) for x in exp)}]
                data.setdefault("post_summary", [])
                data.setdefault("warehouse", [])
                data.setdefault("reason", [])
                data.setdefault("own_all", [])
                data.setdefault("own_3m", [])
                data.setdefault("food", [])
                data.setdefault("food_total", {"name": "IBK bo'yicha Jami", "vazn": 0, "qiymat": 0, "over_vazn": 0, "over_qiymat": 0, "ulush": 0})
                rel = data.setdefault("released", {})
                for d in ["1", "3", "5", "10", "30"]:
                    rel.setdefault(d, {"partiya": 0, "vazn": 0, "qiymat": 0, "tolov": 0, "rows": [], "base_date": "", "missing_date": ""})
                data.setdefault("all_companies", data.get("top_value", []))
            self.json(data)
            return
        if parsed.path == "/api/details":
            if not self.require_user():
                return
            q = parse_qs(parsed.query)
            report_id = q.get("report", [""])[0]
            filters = json.loads(q.get("filters", ["{}"])[0])
            item = next((r for r in load_json(INDEX_PATH, {"reports": []})["reports"] if r["id"] == report_id), None)
            if not item:
                self.json({"rows": []})
                return
            self.json({"rows": detail_rows(read_report_source(Path(item["source"])), filters)})
            return
        if parsed.path == "/api/export_details":
            if not self.require_perm("export"):
                return
            q = parse_qs(parsed.query)
            report_id = q.get("report", [""])[0]
            filters = json.loads(q.get("filters", ["{}"])[0])
            item = next((r for r in load_json(INDEX_PATH, {"reports": []})["reports"] if r["id"] == report_id), None)
            if not item:
                self.send_error(404)
                return
            rows = detail_rows(read_report_source(Path(item["source"])), filters)
            headers = [
                {"k": "decl", "t": "Deklaratsiya", "width": 20},
                {"k": "date", "t": "Sana", "width": 14},
                {"k": "regime", "t": "Rejim", "width": 10},
                {"k": "post", "t": "Post", "width": 24},
                {"k": "stir", "t": "STIR", "width": 14},
                {"k": "company", "t": "Korxona", "width": 42},
                {"k": "hs", "t": "TIF TN", "width": 14},
                {"k": "goods", "t": "Tovar", "width": 46},
                {"k": "partiya", "t": "Partiya", "i": True, "width": 10},
                {"k": "vazn", "t": "Vazn (tn)", "n": True, "width": 14},
                {"k": "qiymat", "t": "Qiymat (ming $)", "n": True, "width": 18},
                {"k": "tolov", "t": "To'lov (mln so'm)", "n": True, "width": 18},
                {"k": "reason", "t": "Saqlanish sababi", "width": 30},
            ]
            xlsx = make_xlsx(headers, rows, "Asos deklaratsiyalar")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename*=UTF-8''asos_deklaratsiyalar.xlsx")
            self.send_header("Content-Length", str(len(xlsx)))
            self.end_headers()
            self.wfile.write(xlsx)
            return
        if parsed.path == "/api/release":
            if not self.require_perm("release"):
                return
            q = parse_qs(parsed.query)
            base_date = q.get("base", [""])[0]
            final_date = q.get("final", [""])[0]
            base_item = find_report_by_date(base_date)
            final_item = find_report_by_date(final_date)
            missing = []
            if not base_item:
                missing.append(base_date or "boshlang'ich sana")
            if not final_item:
                missing.append(final_date or "yakuniy sana")
            if missing:
                self.json({"missing": missing, "rows": [], "total": {}})
                return
            base_snapshot = STORE.snapshot_id_by_date(base_date) if STORE else None
            final_snapshot = STORE.snapshot_id_by_date(final_date) if STORE else None
            if base_snapshot and final_snapshot:
                result = STORE.compute_released(base_snapshot, final_snapshot)
            else:
                result = release_company_table(Path(base_item["source"]), Path(final_item["source"]))
            result.update({"base": base_date, "final": final_date})
            self.json(result)
            return
        if parsed.path == "/api/export":
            if not self.require_perm("export"):
                return
            q = parse_qs(parsed.query)
            kind = q.get("kind", [""])[0]
            report_id = q.get("report", [""])[0]
            headers = []
            rows = []
            title = "IBK Dashboard jadval"
            if kind == "release":
                base_date = q.get("base", [""])[0]
                final_date = q.get("final", [""])[0]
                base_item = find_report_by_date(base_date)
                final_item = find_report_by_date(final_date)
                if not base_item or not final_item:
                    self.send_error(404)
                    return
                base_snapshot = STORE.snapshot_id_by_date(base_date) if STORE else None
                final_snapshot = STORE.snapshot_id_by_date(final_date) if STORE else None
                if base_snapshot and final_snapshot:
                    data = STORE.compute_released(base_snapshot, final_snapshot)
                else:
                    data = release_company_table(Path(base_item["source"]), Path(final_item["source"]))
                rows = [data["total"]] + data["rows"]
                title = f"Nazoratdan yechilishi {base_date} - {final_date}"
                headers = release_headers()
            else:
                item = next((r for r in load_json(INDEX_PATH, {"reports": []})["reports"] if r["id"] == report_id), None)
                if not item:
                    self.send_error(404)
                    return
                data = load_json(Path(item["dir"]) / "dashboard.json", {})
                if kind == "top_value":
                    rows, headers, title = data.get("top_value", []), company_headers(), "TOP 20 korxona qiymat"
                elif kind == "top_deposit":
                    rows, headers, title = data.get("top_deposit", []), company_headers(), "TOP 20 korxona depozit"
                elif kind == "expired":
                    rows, headers, title = data.get("expired", []), expired_headers(), "Muddati o'tgan"
                elif kind == "goods":
                    rows, headers, title = data.get("goods", []), goods_headers(), "Tovar guruhlari"
                elif kind == "food":
                    rows, headers, title = [data.get("food_total", {})] + data.get("food", []), food_headers(), "Oziq-ovqat"
                else:
                    rows, headers, title = data.get("top_value", []), company_headers(), "Jadval"
            xlsx = make_xlsx(headers, rows, title)
            filename = re.sub(r"[^\w.-]+", "_", title, flags=re.UNICODE) + ".xlsx"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
            self.send_header("Content-Length", str(len(xlsx)))
            self.end_headers()
            self.wfile.write(xlsx)
            return
        if parsed.path.startswith("/download/tolov/"):
            if not self.require_perm("export"):
                return
            name = Path(unquote(parsed.path.split("/download/tolov/", 1)[1])).name
            if name == "_all":
                import zipfile
                buf = BytesIO()
                with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for filename in TOLOV_FILES:
                        path = TOLOV_OUTPUT_DIR / filename
                        if path.exists() and path.is_file():
                            zf.write(path, arcname=filename)
                data = buf.getvalue()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/zip")
                self.send_header("Content-Disposition", "attachment; filename*=UTF-8''tolov_jadvallari.zip")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            if name not in TOLOV_FILES:
                self.send_error(404)
                return
            path = TOLOV_OUTPUT_DIR / name
            if not path.exists() or not path.is_file():
                self.send_error(404)
                return
            data = path.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(path.name)}")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if parsed.path.startswith("/download/"):
            if not self.require_perm("export"):
                return
            _, _, report_id, filename = parsed.path.split("/", 3)
            item = next((r for r in load_json(INDEX_PATH, {"reports": []})["reports"] if r["id"] == report_id), None)
            if not item:
                self.send_error(404)
                return
            if Path(unquote(filename)).name == "_pngs":
                import zipfile
                report_dir = Path(item["dir"])
                data_path = report_dir / "dashboard.json"
                dashboard = load_json(data_path, {})
                png_names = dashboard.get("files", {}).get("pngs", [])
                buf = BytesIO()
                with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for name in png_names:
                        p = report_dir / Path(name).name
                        if p.exists():
                            zf.write(p, arcname=p.name)
                data = buf.getvalue()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/zip")
                self.send_header("Content-Disposition", "attachment; filename*=UTF-8''IBK_tahlil_pnglar.zip")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            path = Path(item["dir"]) / Path(unquote(filename)).name
            if not path.exists():
                self.send_error(404)
                return
            data = path.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/octet-stream")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{path.name}")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        self.send_error(404)

    def do_POST(self):
        try:
            self._do_POST_inner()
        except Exception as exc:
            import traceback as _tb
            try:
                self.json({"error": f"Server xatosi: {type(exc).__name__}: {exc}", "trace": _tb.format_exc()[-800:]}, HTTPStatus.INTERNAL_SERVER_ERROR)
            except Exception:
                pass

    def _do_POST_inner(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/login":
            client_ip = self.client_address[0]
            now = time.time()
            fail = LOGIN_FAILS.get(client_ip, {"count": 0, "until": 0})
            if fail["until"] > now:
                mins = max(1, int((fail["until"] - now) / 60) + 1)
                self.json({"error": f"Juda ko'p noto'g'ri urinish. {mins} daqiqadan so'ng qayta urining."}, HTTPStatus.TOO_MANY_REQUESTS)
                return
            data = self.body_json()
            users = load_json(USER_PATH, {})
            rec = users.get(data.get("user", ""))
            if not rec or rec["password"] != hash_password(data.get("pass", ""), rec["salt"]):
                fail["count"] = fail.get("count", 0) + 1
                if fail["count"] >= LOGIN_BLOCK_AFTER:
                    fail["until"] = now + LOGIN_BLOCK_SECS
                    fail["count"] = 0
                LOGIN_FAILS[client_ip] = fail
                self.json({"error": "Login yoki parol xato"}, HTTPStatus.UNAUTHORIZED)
                return
            if not rec.get("enabled", True):
                self.json({"error": "Foydalanuvchiga ruxsat berilmagan"}, HTTPStatus.FORBIDDEN)
                return
            LOGIN_FAILS.pop(client_ip, None)
            token = secrets.token_urlsafe(32)
            SESSIONS[token] = {"user": data["user"], "created": time.time()}
            save_json(SESSION_PATH, SESSIONS)
            self.json({"token": token, "user": {"user": data["user"], "role": rec.get("role", "user"), "role_label": rec.get("role_label", ROLE_LABELS.get(rec.get("role", "user"), "Foydalanuvchi")), "post_code": rec.get("post_code", ""), "full_name": rec.get("full_name", ""), "position": rec.get("position", ""), "phone": rec.get("phone", ""), "lang": rec.get("lang", "uz"), "perms": rec.get("perms", [])}})
            return
        if parsed.path == "/api/ui_config":
            if not self.require_admin():
                return
            data = self.body_json()
            save_json(UI_CONFIG_PATH, data)
            self.json({"ok": True})
            return
        if parsed.path == "/api/users":
            if not self.require_admin():
                return
            data = self.body_json()
            user = re.sub(r"[^\w.@-]+", "", data.get("user", "")).strip()
            password = data.get("pass", "")
            role = data.get("role", "user") if data.get("role") in ROLES else "user"
            perms = [p for p in data.get("perms", DEFAULT_PERMS) if p in ADMIN_PERMS]
            if not user or not password:
                self.json({"error": "Login va parol kerak"}, HTTPStatus.BAD_REQUEST)
                return
            users = load_json(USER_PATH, {})
            salt = secrets.token_hex(8)
            users[user] = {"salt": salt, "password": hash_password(password, salt), "role": role, "role_label": ROLE_LABELS.get(role, "Foydalanuvchi"), "enabled": True, "full_name": data.get("full_name", ""), "position": data.get("position", ""), "phone": data.get("phone", ""), "lang": data.get("lang", "uz"), "post_code": data.get("post_code", ""), "perms": ADMIN_PERMS if role == "admin" else perms}
            save_json(USER_PATH, users)
            self.json({"ok": True})
            return
        if parsed.path == "/api/users/delete":
            admin = self.require_admin()
            if not admin:
                return
            data = self.body_json()
            user = data.get("user", "")
            users = load_json(USER_PATH, {})
            if user == admin:
                self.json({"error": "Admin o'zini o'chira olmaydi"}, HTTPStatus.BAD_REQUEST)
                return
            if user in users:
                users[user]["enabled"] = False
                save_json(USER_PATH, users)
            self.json({"ok": True})
            return
        if parsed.path == "/api/users/change_password":
            user = self.require_user()
            if not user:
                return
            data = self.body_json()
            old_pass = data.get("old_pass", "")
            new_pass = data.get("new_pass", "")
            if not old_pass or not new_pass or len(new_pass) < 6:
                self.json({"error": "Ma'lumotlar to'liq emas"}, HTTPStatus.BAD_REQUEST)
                return
            users = load_json(USER_PATH, {})
            rec = users.get(user)
            if not rec:
                self.json({"error": "Foydalanuvchi topilmadi"}, HTTPStatus.NOT_FOUND)
                return
            if hash_password(old_pass, rec.get("salt", "")) != rec.get("password", ""):
                self.json({"error": "Joriy parol noto'g'ri"}, HTTPStatus.FORBIDDEN)
                return
            rec["salt"] = secrets.token_hex(8)
            rec["password"] = hash_password(new_pass, rec["salt"])
            users[user] = rec
            save_json(USER_PATH, users)
            self.json({"ok": True})
            return
        if parsed.path == "/api/users/update":
            if not self.require_admin():
                return
            data = self.body_json()
            user = re.sub(r"[^\w.@-]+", "", data.get("user", "")).strip()
            users = load_json(USER_PATH, {})
            if not user or user not in users:
                self.json({"error": "Hodim topilmadi"}, HTTPStatus.NOT_FOUND)
                return
            role = data.get("role", users[user].get("role", "user"))
            role = role if role in ROLES else "user"
            perms = [p for p in data.get("perms", users[user].get("perms", DEFAULT_PERMS)) if p in ADMIN_PERMS]
            rec = users[user]
            rec.update({
                "role": role,
                "role_label": ROLE_LABELS.get(role, "Foydalanuvchi"),
                "enabled": bool(data.get("enabled", rec.get("enabled", True))),
                "full_name": data.get("full_name", rec.get("full_name", "")),
                "position": data.get("position", rec.get("position", "")),
                "phone": data.get("phone", rec.get("phone", "")),
                "lang": data.get("lang", rec.get("lang", "uz")),
                "post_code": data.get("post_code", rec.get("post_code", "")),
                "perms": ADMIN_PERMS if role == "admin" else perms,
            })
            password = data.get("pass", "")
            if password:
                rec["salt"] = secrets.token_hex(8)
                rec["password"] = hash_password(password, rec["salt"])
            users[user] = rec
            save_json(USER_PATH, users)
            self.json({"ok": True})
            return
        if parsed.path == "/api/artifacts":
            if not self.require_perm("export"):
                return
            data = self.body_json()
            report_id = data.get("report", "")
            item = next((r for r in load_json(INDEX_PATH, {"reports": []})["reports"] if r["id"] == report_id), None)
            if not item:
                self.json({"error": "Hisobot topilmadi"}, HTTPStatus.NOT_FOUND)
                return
            job_id = "artifacts_" + report_id + "_" + str(int(time.time() * 1000))
            JOBS[job_id] = {"status": "navbatda", "data": {}}
            threading.Thread(target=run_artifact_job, args=(job_id, report_id), daemon=True).start()
            self.json({"job_id": job_id})
            return
        if parsed.path == "/api/tolov":
            if not self.require_perm("upload"):
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            form = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            if "source" not in form or not form["source"]["filename"]:
                self.json({"error": "To'lov baza fayli kerak"}, HTTPStatus.BAD_REQUEST)
                return
            tmp = UPLOAD_DIR / ("tolov_" + str(int(time.time() * 1000)))
            tmp.mkdir(parents=True, exist_ok=True)
            source = tmp / safe_name(form["source"]["filename"])
            source.write_bytes(form["source"]["content"])
            try:
                rows = build_tolov_from_source(source)
            except Exception as exc:
                self.json({"error": f"{type(exc).__name__}: {exc}"}, HTTPStatus.INTERNAL_SERVER_ERROR)
                return
            self.json({"ok": True, "payments": rows})
            return
        if parsed.path == "/api/chunk_upload":
            if not self.require_perm("upload"):
                return
            uid   = self.headers.get("X-Upload-Id", "").strip()
            idx   = int(self.headers.get("X-Chunk-Index", "0") or "0")
            total = int(self.headers.get("X-Total-Chunks", "1") or "1")
            if not uid:
                self.json({"error": "X-Upload-Id kerak"}, HTTPStatus.BAD_REQUEST)
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            te = self.headers.get("Transfer-Encoding", "").lower()
            try:
                self.connection.settimeout(90)
                if "chunked" in te:
                    body = b""
                    while True:
                        sz_line = self.rfile.readline().strip()
                        if not sz_line:
                            break
                        chunk_sz = int(sz_line, 16)
                        if chunk_sz == 0:
                            break
                        body += self.rfile.read(chunk_sz)
                        self.rfile.read(2)
                elif length > 0:
                    body = self.rfile.read(length)
                else:
                    body = b""
            except Exception as e:
                self.json({"error": f"Yuklab olishda xato: {e}"}, HTTPStatus.REQUEST_TIMEOUT)
                return
            finally:
                try:
                    self.connection.settimeout(None)
                except Exception:
                    pass
            if not body:
                self.json({"error": "Bo'sh chunk"}, HTTPStatus.BAD_REQUEST)
                return
            cd = CHUNK_DIR / uid
            cd.mkdir(parents=True, exist_ok=True)
            chunk_path = cd / f"{idx:06d}.bin"
            chunk_path.write_bytes(body)
            self.json({"ok": True, "received": idx + 1, "total": total}, cors=True)
            return
        if parsed.path == "/api/chunk_finalize":
            if not self.require_perm("upload"):
                return
            body = self.body_json()
            uid      = body.get("upload_id", "").strip()
            filename = unquote(body.get("filename", "").strip())
            total    = int(body.get("total_chunks", 1))
            dep_uid  = body.get("deposit_upload_id", "").strip()
            dep_name = unquote(body.get("deposit_filename", "").strip())
            if not uid or not filename:
                self.json({"error": "upload_id va filename kerak"}, HTTPStatus.BAD_REQUEST)
                return
            cd = CHUNK_DIR / uid
            chunks = sorted(cd.glob("*.bin"))
            if len(chunks) < total:
                self.json({"error": f"Qismlar yetishmaydi: {len(chunks)}/{total}"}, HTTPStatus.BAD_REQUEST)
                return
            report_date = report_date_from_name(filename)
            dup = duplicate_report(filename, report_date)
            if dup:
                shutil.rmtree(cd, ignore_errors=True)
                self.json({"error": "Bu fayl avval yuklangan"}, HTTPStatus.CONFLICT)
                return
            job_id = report_date.strftime("%Y%m%d_") + str(int(time.time() * 1000))
            tmp = UPLOAD_DIR / job_id
            tmp.mkdir(parents=True, exist_ok=True)
            source = tmp / safe_name(filename)
            with open(source, "wb") as fh:
                for ch in chunks:
                    fh.write(ch.read_bytes())
            shutil.rmtree(cd, ignore_errors=True)
            deposit = None
            if dep_uid and dep_name:
                dcd = CHUNK_DIR / dep_uid
                dchunks = sorted(dcd.glob("*.bin"))
                if dchunks:
                    deposit = tmp / safe_name(dep_name)
                    with open(deposit, "wb") as fh:
                        for ch in dchunks:
                            fh.write(ch.read_bytes())
                    shutil.rmtree(dcd, ignore_errors=True)
            JOBS[job_id] = {"status": "navbatda", "data": None}
            threading.Thread(target=run_job, args=(job_id, source, deposit, report_date), daemon=True).start()
            self.json({"job_id": job_id}, cors=True)
            return
        if parsed.path == "/api/reports_bulk":
            if not self.require_perm("upload"):
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            form = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            raw_sources = form.get("sources") or form.get("source") or []
            sources = raw_sources if isinstance(raw_sources, list) else [raw_sources]
            sources = [x for x in sources if x and x.get("filename")]
            if not sources:
                self.json({"error": "Asos fayllar kerak"}, HTTPStatus.BAD_REQUEST)
                return
            deposit_item = form.get("deposit")
            job_ids = []
            skipped = []
            for idx, src in enumerate(sources):
                report_date = report_date_from_name(src["filename"])
                dup = duplicate_report(src["filename"], report_date)
                if dup:
                    skipped.append({"filename": src["filename"], "date": fmt_date(report_date), "reason": "Bu fayl avval yuklangan"})
                    continue
                job_id = report_date.strftime("%Y%m%d_") + str(int(time.time() * 1000)) + f"_{idx}"
                tmp = UPLOAD_DIR / job_id
                tmp.mkdir(parents=True, exist_ok=True)
                source = tmp / safe_name(src["filename"])
                source.write_bytes(src["content"])
                deposit = None
                if isinstance(deposit_item, dict) and deposit_item.get("filename"):
                    deposit = tmp / safe_name(deposit_item["filename"])
                    deposit.write_bytes(deposit_item["content"])
                JOBS[job_id] = {"status": "navbatda", "data": None}
                threading.Thread(target=run_job, args=(job_id, source, deposit, report_date), daemon=True).start()
                job_ids.append(job_id)
            self.json({"job_ids": job_ids, "count": len(job_ids), "skipped": skipped})
            return
        if parsed.path == "/api/deposit":
            if not self.require_perm("upload"):
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            form = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            report_id = (form.get("report_id") or {}).get("content", b"").decode("utf-8", errors="ignore").strip()
            if not report_id:
                self.json({"error": "report_id kerak"}, HTTPStatus.BAD_REQUEST)
                return
            if "deposit" not in form or not form["deposit"]["filename"]:
                self.json({"error": "Depozit fayl kerak"}, HTTPStatus.BAD_REQUEST)
                return
            archive = load_json(INDEX_PATH, {"reports": []})
            if not isinstance(archive, dict):
                archive = {"reports": []}
            reports = archive.get("reports") or []
            item = next((r for r in reports if isinstance(r, dict) and r.get("id") == report_id), None)
            if not item:
                self.json({"error": "Hisobot topilmadi"}, HTTPStatus.NOT_FOUND)
                return
            report_dir = Path(item.get("dir", ""))
            if not report_dir.exists():
                self.json({"error": "Hisobot papkasi topilmadi"}, HTTPStatus.NOT_FOUND)
                return
            deposit_path = report_dir / safe_name(form["deposit"]["filename"])
            deposit_path.write_bytes(form["deposit"]["content"])
            item["deposit"] = str(deposit_path)
            save_json(INDEX_PATH, archive)
            job_id = "deposit_" + str(int(time.time() * 1000))
            JOBS[job_id] = {"status": "navbatda"}
            def _rebuild_with_deposit(jid=job_id, it=dict(item), dp=deposit_path):
                j = ensure_job(jid)
                try:
                    j["status"] = "Depozit qayta ishlanmoqda"
                    src = Path(it["source"])
                    rd = datetime.strptime(it["date"], "%d.%m.%Y")
                    data = build_dashboard(it["id"], src, dp, rd)
                    rdir = Path(it["dir"])
                    data_path = rdir / "dashboard.json"
                    existing = load_json(data_path, {})
                    data["files"] = existing.get("files", {})
                    save_json(data_path, data)
                    j.update({"status": "tayyor", "data": data})
                except Exception as exc:
                    j["status"] = "xatolik"
                    j["error"] = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()}"
            threading.Thread(target=_rebuild_with_deposit, daemon=True).start()
            self.json({"job_id": job_id, "report_id": report_id})
            return
        if parsed.path in ("/api/upload_warehouses", "/api/upload_wr_registry"):
            if not self.require_perm("upload"):
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            form = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            if "file" not in form or not form["file"].get("filename"):
                self.json({"error": "Excel fayl kerak"}, HTTPStatus.BAD_REQUEST)
                return
            dest = DASHBOARD_DIR / safe_name(form["file"]["filename"])
            dest.write_bytes(form["file"]["content"])
            try:
                result = load_warehouse_registry(dest)
            except Exception as exc:
                self.json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
                return
            self.json(result)
            return
        if parsed.path == "/api/upload_avia_awb":
            if not self.require_perm("upload"):
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            form = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            if "file" not in form or not form["file"].get("filename"):
                self.json({"error": "Excel fayl kerak"}, HTTPStatus.BAD_REQUEST)
                return
            dest = DASHBOARD_DIR / safe_name(form["file"]["filename"])
            dest.write_bytes(form["file"]["content"])
            try:
                result = load_avia_awb(dest)
            except Exception as exc:
                self.json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
                return
            self.json(result)
            return
        if parsed.path == "/api/upload_yaroqlilik":
            if not self.require_perm("upload"):
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            form = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            if "file" not in form or not form["file"].get("filename"):
                self.json({"error": "Excel fayl kerak"}, HTTPStatus.BAD_REQUEST)
                return
            try:
                result = parse_yaroqlilik_excel(form["file"]["content"])
                save_yaroqlilik(result)
                self.json(result)
            except Exception as exc:
                self.json({"error": f"{type(exc).__name__}: {exc}"}, HTTPStatus.INTERNAL_SERVER_ERROR)
            return
        if parsed.path == "/api/archive/delete":
            if not self.require_admin():
                return
            data = self.body_json()
            rid = data.get("id", "")
            if not rid:
                self.json({"error": "id kerak"}, HTTPStatus.BAD_REQUEST)
                return
            archive = load_json(INDEX_PATH, {"reports": []})
            if not isinstance(archive, dict):
                archive = {"reports": []}
            archive["reports"] = [r for r in (archive.get("reports") or []) if r.get("id") != rid]
            if archive.get("current_id") == rid:
                archive.pop("current_id", None)
            save_json(INDEX_PATH, archive)
            self.json({"ok": True})
            return
        if parsed.path == "/api/archive/set_current":
            if not self.require_admin():
                return
            data = self.body_json()
            rid = data.get("id", "")
            if not rid:
                self.json({"error": "id kerak"}, HTTPStatus.BAD_REQUEST)
                return
            archive = load_json(INDEX_PATH, {"reports": []})
            if not isinstance(archive, dict):
                archive = {"reports": []}
            archive["current_id"] = rid
            save_json(INDEX_PATH, archive)
            self.json({"ok": True})
            return
        if parsed.path == "/api/reports":
            if not self.require_perm("upload"):
                return
            length = int(self.headers.get("Content-Length", "0") or "0")
            form = parse_multipart(self.headers.get("Content-Type", ""), self.rfile.read(length))
            if "source" not in form or not form["source"]["filename"]:
                self.json({"error": "Asos fayl kerak"}, HTTPStatus.BAD_REQUEST)
                return
            report_date = datetime.strptime(form.get("date", {}).get("content", b"").decode("utf-8") or "1900-01-01", "%Y-%m-%d") if form.get("date", {}).get("content") else report_date_from_name(form["source"]["filename"])
            dup = duplicate_report(form["source"]["filename"], report_date)
            if dup:
                self.json({"error": f"Bu fayl avval yuklangan: {fmt_date(report_date)} / {form['source']['filename']}"}, HTTPStatus.CONFLICT)
                return
            job_id = report_date.strftime("%Y%m%d_") + str(int(time.time() * 1000))
            tmp = UPLOAD_DIR / job_id
            tmp.mkdir(parents=True, exist_ok=True)
            source = tmp / safe_name(form["source"]["filename"])
            source.write_bytes(form["source"]["content"])
            deposit = None
            if "deposit" in form and form["deposit"]["filename"]:
                deposit = tmp / safe_name(form["deposit"]["filename"])
                deposit.write_bytes(form["deposit"]["content"])
            JOBS[job_id] = {"status": "navbatda", "data": {}}
            threading.Thread(target=run_job, args=(job_id, source, deposit, report_date), daemon=True).start()
            self.json({"job_id": job_id})
            return
        self.send_error(404)

    def log_message(self, fmt, *args):
        print(f"[IBK] {self.address_string()} {fmt % args}")


def main():
    global STORE
    ensure_dirs()
    STORE = IBKStore(DB_PATH)
    SESSIONS.update(load_json(SESSION_PATH, {}))
    # Pre-load cached files silently
    try:
        load_warehouse_registry()
    except Exception:
        pass
    try:
        load_avia_awb()
    except Exception:
        pass
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"IBK_Dashboard: http://localhost:{PORT}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()













